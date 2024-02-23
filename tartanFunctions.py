"""
SYNOPSIS
    These are various functions used in TARTAN.

    This file is part of Tartan Systems (TARTAN).

AUTHOR
    Written by Paul Malherbe, <paul@tartan.co.za>

COPYING
    Copyright (C) 2004-2023 Paul Malherbe.

    This program is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with this program. If not, see <https://www.gnu.org/licenses/>.
"""

def importTkinter():
    try:
        import tkinter as tk
        import tkinter.ttk as ttk
        return tk, ttk
    except:
        return None, None

def unbindAllWidgets(widget):
    binds = []
    for bind in widget.winfo_toplevel().bind():
        if "<Key-Alt_L>" in bind:
            binds.append((bind, widget.winfo_toplevel().bind(bind)))
            widget.winfo_toplevel().unbind(bind)
    return binds

def getManager(widget):
    # Window Manager
    try:
        children = widget.winfo_children()
        if not children:
            mgr = "pack"
        else:
            mgr = ""
            for c in children:
                if c.winfo_manager():
                    mgr = c.winfo_manager()
                    break
    except:
        mgr = "pack"
    return mgr

def placeWindow(window, parent=None, place="C", size=None, expose=False):
    # Window Placement
    window.update_idletasks()
    if window.winfo_class().lower() == "tk":
        parent = None
    if parent:
        if parent.winfo_class() != "Toplevel":
            parent = parent.winfo_toplevel()
        parent.update_idletasks()
    if size:
        ww, wh = size
    else:
        ww = window.winfo_reqwidth()
        wh = window.winfo_reqheight()
    if parent:
        try:
            import tkinter as tk
            w = tk.Toplevel()
            w.geometry("1x1+0+0")
            w.update()
            bar = w.winfo_y()
            w.destroy()
        except:
            bar = 24
        wx = int(parent.winfo_x() + ((parent.winfo_width() - ww) / 2))
        wy = int(parent.winfo_y() + ((parent.winfo_height() - wh) / 2)) - bar
    else:
        if place == "L":
            wx = 0
        elif place == "R":
            wx = window.winfo_screenwidth() - ww
        else:
            wx = int((window.winfo_screenwidth() - ww) / 2)
        if place == "M":
            wy = int((window.winfo_screenheight() - wh) / 2)
        else:
            wy = 0
    window.geometry("%dx%d+%d+%d" % (ww, wh, wx, wy))
    if expose and window.state() == "withdrawn":
        window.deiconify()
    window.update_idletasks()

def cutpasteMenu(event):
    # Cut, copy and paste menu
    from TartanClasses import tk, tkfont
    wid = event.widget
    font = tkfont.Font(font=("Arial", 10))
    wid.menu = tk.Menu(wid, tearoff=False, takefocus=0, font=font)
    image = getImage("Cut", (20, 20))
    wid.menu.add_command(label="Cut", image=image, compound="left",
        accelerator="Ctl-X", font=font)
    wid.menu.i1 = image
    image = getImage("Copy", (20, 20))
    wid.menu.add_command(label="Copy", image=image, compound="left",
        accelerator="Ctl-C", font=font)
    wid.menu.i2 = image
    image = getImage("Paste", (20, 20))
    wid.menu.add_command(label="Paste", image=image, compound="left",
        accelerator="Ctl-V", font=font)
    wid.menu.i3 = image
    wid.menu.add_separator()
    wid.menu.add_command(label="Select all")
    wid.menu.entryconfigure("Cut", command=lambda:
        wid.focus_force() or wid.event_generate("<<Cut>>"))
    wid.menu.entryconfigure("Copy", command=lambda:
        wid.focus_force() or wid.event_generate("<<Copy>>"))
    wid.menu.entryconfigure("Paste", command=lambda:
        wid.focus_force() or wid.event_generate("<<Paste>>"))
    wid.menu.entryconfigure("Select all", command=wid.select_all)
    wid.menu.tk_popup(event.x_root + 40, event.y_root + 10, entry="0")
    return "break"

def getPrgPath():
    import os, sys
    epath = os.path.dirname(sys.executable)
    if getattr(sys, "frozen", False):
        return sys._MEIPASS, epath
    else:
        return os.path.dirname(os.path.abspath(__file__)), epath

def showDialog(screen, dtype, title, mess, butt=None, dflt=None):
    try:
        if screen == "text":
            if dtype == "question":
                answer = None
                while answer not in ("yes", "no"):
                    answer = input("\n%s (yes/no): " % mess)
                return answer
            else:
                print(mess)
        else:
            try:
                from TartanClasses import MyMessageBox
                mb = MyMessageBox(screen, dtype, title=title, mess=mess,
                    butt=butt, dflt=dflt)
                return mb.answer
            except:
                raise Exception
    except Exception as err:
        print("%s\n%s" % (mess, err))

def askChoice(screen=None, head="", mess="", butt=None, default=""):
    return showDialog(screen, "choice", head, mess, butt, default)

def showError(screen=None, head="", mess=""):
    showDialog(screen, "error", head, mess)

def showInfo(screen=None, head="", mess=""):
    showDialog(screen, "info", head, mess)

def askQuestion(screen=None, head="", mess="", default="no"):
    return showDialog(screen, "question", head, mess, dflt=default)

def showWarning(screen=None, head="", mess=""):
    showDialog(screen, "warning", head, mess)

def getFontSize(tk=None, width=None, height=None, font=10):
    if not tk:
        tk, ttk = importTkinter()
    if not tk:
        return 800, 600, 10
    plus = True
    while True:
        win = tk.Tk()
        win.withdraw()
        if not height:
            height = int(round(win.winfo_screenheight() * .90, 0))
        if not width:
            width = int(round(height * 1.33333, 0))
        for _ in range(28):
            entry = tk.Entry(win, font=("Courier", font))
            entry.pack()
        win.update_idletasks()
        rh = win.winfo_reqheight()
        win.destroy()
        if plus:
            if rh < height:
                font += 1
            elif rh == height:
                break
            else:
                plus = False
        else:
            if rh <= height:
                break
            font -= 1
    return width, height, font - 1

def loadRcFile(rcfile=None, default=False):
    """
    Function to load a tartanrc file and return a dictionary.

        rcfile - The path of the file to load defaulting to tartanrc.
        default - Whether to return default settings if no file.
    """

    import os, sys
    from pathlib import Path
    from TartanClasses import FITZ
    # Defaults
    if FITZ:
        vwr = [""]
    else:
        win = os.path.join(getPrgPath()[1], "uty", "SumatraPDF.exe")
        vwr = ["/usr/bin/evince", win]
    tardir = os.path.join(Path.home(), "Tartan")
    if not os.path.isdir(tardir):
        os.makedirs(tardir)
    if sys.platform == "win32":
        test1 = os.path.normpath("C:/tartanrc")
        test2 = os.path.normpath("C:/Tartan/tartanrc")
        test3 = os.path.join(tardir, "tartanrc")
    else:
        test1 = os.path.join(Path.home(), ".tartanrc")
        test2 = None
        test3 = os.path.join(tardir, ".tartanrc")
    opts = {
        "dbase": ["SQLite"],
        "dbname": ["tartan"],
        "dbdir": [os.path.join(tardir, "fle")],
        "dbhost": [""],
        "dbport": [""],
        "dbuser": [""],
        "dbpwd": [""],
        "bupdir": [os.path.join(tardir, "bup")],
        "wrkdir": [os.path.join(tardir, "wrk")],
        "upgdir": [os.path.join(tardir, "upg")],
        "vwr": vwr,
        "prn": [""],
        "exp": [""],
        "geo": [0],
        "plc": ["C"],
        "img": ["Y"],
        "cnf": ["Y"],
        "acnf": ["L"],
        "ttip": ["Y"],
        "errs": ["Y"],
        "wrkf": ["D"],
        "wrka": ["N"],
        "mft": ["DejaVu Sans", "Arial"],
        "mfs": [0],
        "dft": ["DejaVu Sans Mono", "Courier New"],
        "dfs": [0],
        "theme": ["clam"],
        "lsc": ["R"],
        "nfg": ["#ffffff"],
        "nbg": ["#a40204"],
        "bsc": ["R"],
        "bfg": ["#ffffff"],
        "bbg": ["#a40204"],
        "ffg": ["#ffffff"],
        "fbg": ["#044e1c"],
        "dfg": ["#000000"],
        "dbg": ["#ffffff"],
        "qfg": ["#000000"],
        "qbg": ["#ffc0cb"],
        "cfg": ["#000000"],
        "cbg": ["#add8e6"],
        "sfg": ["#000000"],
        "sbg": ["#90ee90"]}
    pths = ("dbdir", "bupdir", "wrkdir", "upgdir")
    # Read and load rcfile
    if not rcfile:
        rcfile = os.environ.get("TARTANRC")
        if not rcfile:
            if os.path.isfile(test1):
                rcfile = test1
            elif test2 is not None and os.path.isfile(test2):
                rcfile = test2
            else:
                rcfile = test3
    try:
        rcdic = {"name": rcfile}
        fle = open(rcfile, "r")
        for l in fle:
            if l[0] in ("#", ";"):
                continue
            l = l.strip()
            l = l.replace("[", "").replace("]", "")
            l = l.replace("'", "")
            l = l.split(" = ")
            if l[0].strip() not in opts:
                if l[0].strip() == "dcs":
                    rcdic["lsc"] = l[1].strip()
                    rcdic["bsc"] = l[1].strip()
                continue
            if l[0].strip() and l[0].strip() in pths:
                path = os.path.normpath(l[1].strip())
                if path == ".":
                    rcdic[l[0].strip()] = ""
                else:
                    rcdic[l[0].strip()] = os.path.normpath(l[1].strip())
            else:
                rcdic[l[0].strip()] = l[1].strip()
        fle.close()
        if "dbase" not in rcdic or "dbname" not in rcdic:
            raise Exception
    except:
        if not default:
            return "error"
    but = False
    for o in opts:
        if o not in rcdic:
            if o in ("bfg", "bbg"):
                but = True
                continue
            if len(opts[o]) == 1 or sys.platform != "win32":
                rcdic[o] = opts[o][0]
            else:
                rcdic[o] = opts[o][1]
        if o in ("dfs", "mfs"):
            rcdic[o] = int(rcdic[o])
        if o == "acnf" and rcdic[o].lower() == "y":
            rcdic["acnf"] = "L"
        if o == "errs" and rcdic[o] == "M":
            rcdic[o] = "Y"
    if but:
        rcdic["bfg"] = rcdic["nfg"]
        rcdic["bbg"] = rcdic["nbg"]
    if not rcdic["geo"] or rcdic["geo"] == "0":
        try:
            w, h, f = getFontSize()
        except:
            w, h, f = 800, 600, 10
        m = f
    else:
        g = rcdic["geo"].split("x")
        if not rcdic["dfs"]:
            w, h, f = getFontSize(None, int(g[0]), int(g[1]))
            m = f
        else:
            w, h = (int(g[0]), int(g[1]))
            m = rcdic["mfs"]
            f = rcdic["dfs"]
    rcdic["geo"] = "%sx%s" % (w, h)
    rcdic["mfs"] = m
    rcdic["dfs"] = f
    if not rcdic["theme"]:
        rcdic["theme"] = "clam"
    if os.path.exists(rcfile):
        # Create missing directories
        check = []
        for pth in pths:
            if pth and pth in rcdic and not os.path.isdir(rcdic[pth]):
                if rcdic[pth] not in check:
                    check.append(rcdic[pth])
        if check:
            try:
                for d in check:
                    if d:
                        os.makedirs(d)
            except:
                return "makedir (%s)" % d
    return rcdic

def parsePrg(text):
    a = b = 0
    text = text.strip()
    if " /" in text:
        a = text.index(" /")
    if " -" in text:
        b = text.index(" -")
    if a and b:
        if b > a:
            exe = text.split(" /")[0]
        else:
            exe = text.split(" -")[0]
    elif a:
        exe = text.split(" /")[0]
    elif b:
        exe = text.split(" -")[0]
    else:
        exe = text
    cmd = [exe]
    cmd.extend(text.replace(exe, "").strip().split())
    return exe, cmd

def showException(scrn, path, mess, maxTB=None, xits=None, dbm=None):
    """
    Display an exception

    scrn  = The Screen on which to Display the Exception or None to print
    path  = The place where the log file resides
    mess  = The First Line of the Message
    maxTB = The maximum number of entries to display
    xits  = A command to be executed on exiting the exception frame
    dbm   = The database class
    """
    import getpass, os, platform, sys, time, traceback
    from TartanClasses import ScrollText

    def doSaveExc():
        t = time.localtime()
        t = time.strftime("Date: %Y-%m-%d Time: %H:%M:%S", t)
        host = platform.uname()[1]
        try:
            user = getpass.getuser()
        except:
            user = "Unknown"
        onam = platform.platform()
        fnam = os.path.join(path, "errors.txt")
        if os.path.isfile(fnam):
            flenam = open(fnam, "a")
        else:
            flenam = open(fnam, "w")
        flenam.write(t + "\n\n" + onam + "\n\n")
        flenam.write("Host: %s User: %s" % (host, user) + "\n\n")
        flenam.write(mess + "\n")
        flenam.write("*******************************" + "\n")
        flenam.close()
        if dbm:
            if not dbm.dbopen:
                opened = True
                dbm.openDbase()
            else:
                opened = False
            try:
                dbm.cu.execute("Select ctm_name, sys_msvr, sys_mprt, "\
                    "sys_msec, sys_maut, sys_mnam, sys_mpwd, ctm_email, "\
                    "ver_version from ctlmst, ctlsys, verupd where "\
                    "ctm_cono = 1")
                smtp = dbm.cu.fetchone()
                if not smtp or not smtp[1]:
                    raise Exception
                err = sendMail(smtp[1:7], smtp[7], ["errors@tartan.co.za"],
                    "Version: %s Company: %s Host: %s User: %s" % (smtp[8],
                    smtp[0], host, user), attach=[fnam], wrkdir=path)
                if not err:
                    os.remove(fnam)
            except:
                pass
            if opened:
                dbm.closeDbase()
    try:
        title = str(sys.exc_info()[0])
    except:
        title = str(sys.exc_type)
    cla, exc, trbk = sys.exc_info()
    if type(cla) == str:
        excName = cla
    else:
        excName = cla.__name__
    if type(exc) == str or not exc.__dict__:
        excArgs = exc
    else:
        try:
            excArgs = exc.__dict__["args"]
        except KeyError:
            excArgs = "<no args>"
    if maxTB:
        excTb = traceback.format_tb(trbk, maxTB)
    else:
        excTb = traceback.format_tb(trbk)
    excMess = ""
    for ta in excTb:
        excMess = excMess + ta + "\012"
    mess = str(mess) + "\n\n%s\n%s\n%s" % (excMess, excName, excArgs)
    if path:
        doSaveExc()
    if scrn:
        try:
            if scrn.winfo_toplevel().state() == "withdrawn":
                scrn.winfo_toplevel().deiconify()
            ScrollText(scrn=scrn, title=title, mess=mess)
        except:
            pass
    else:
        print(mess)

def makeArray(col, row, dth, typ="I"):
    """
    A routine to generate an array of lists

    col = Number of Columns
    row = Number of Rows
    dth = Number of Levels
    typ = "D" for Float, "I" for Integer, "S" for String
    """
    import copy
    a = []
    b = []
    c = []
    if typ == "S":
        d = ""
    elif typ == "D":
        d = 0.0
    else:
        d = 0
    for _ in range(col):
        a.append(d)
    for _ in range(row):
        b.append(copy.copy(a))
    if dth < 2:
        c = b
    else:
        for x in range(dth):
            c.append(copy.deepcopy(b))
    return(c)

def getPeriods(mf, conum, period, check=False):
    """
    A routine to extract a financial period's start and end date for a Company

    mf     = The MainFrame object
    conum  = The company number
    period = The period number
    check  = Check if a Year End needs to be run
    """
    from TartanClasses import CCD, SplashScreen, Sql

    sql = Sql(mf.dbm, ["ctlynd", "gentrn"], prog=__name__)
    p = sql.getRec("ctlynd", cols=["cye_start", "cye_end",
        "cye_final"], where=[("cye_cono", "=", conum), ("cye_period",
        "=", period)], limit=1)
    if not p:
        showError(mf.window, "Period Error",
            "Invalid Financial Period %s for Company %s" % (period, conum))
        return (None, None, None)
    if period and check:
        sp = SplashScreen(mf.window,
            "Checking if Period is Up To Date,\n\nPlease Wait....")
        chk1 = sql.getRec("ctlynd", cols=["cye_last"],
            where=[("cye_cono", "=", conum), ("cye_period", "=",
                (period - 1))], limit=1)
        chk2 = sql.getRec("gentrn", cols=["max(glt_capdt)"],
            where=[("glt_cono", "=", conum), ("glt_curdt", "<",
            int(p[0] / 100))], limit=1)
        sp.closeSplash()
        if chk2[0] and chk2[0] > chk1[0]:
            ok = askQuestion(mf.window, "Year End Error",
                """A Year End Must be Performed for Period %s as Transactions Have Been Captured After the last Year End Run for that Period!

If Not Performed, certain Reports and Queries could have Incorrect Balances!

Do You Want to Continue Anyway?""" % (period-1), default="no")
            if ok == "no":
                return (None, None, None)
    s_date = CCD(p[0], "D1", 10.0)
    e_date = CCD(p[1], "D1", 10.0)
    return (s_date, e_date, p[2])

def getModName(pth, prg, key, ext=None, wait=0):
    """
    A routine to return a name for a printer output file which is unique ??

    pth   = The path where the file is to be created
    prg   = The calling module name
    key   = The key of the report e.g. Company Number/Account Number
    ext   = The file extension e.g. pdf or None
    wait  = The number of seconds to pause before returning the name
    """
    import os, time

    if wait:
        time.sleep(wait)
    t = time.localtime()
    tim = "%04i%02i%02i%02i%02i%02i" % (t[0], t[1], t[2], t[3], t[4], t[5])
    key = str(key).replace("/", "-").replace("\\", "-").replace(" ", "_")
    if not ext:
        return os.path.join(pth, "%s_%s_%s" % (prg, key, tim))
    if ext and type(ext) == str:
        return os.path.join(pth, "%s_%s_%s.%s" % (prg, key, tim, ext))
    nam = []
    for e in ext:
        nam.append(os.path.join(pth, "%s_%s_%s.%s" % (prg, key, tim, e)))
    return nam

def doPrinter(mf=None, conum=None, pdfnam=None, splash=True, header=None, repprt=None, repeml=None, fromad=None, attach=None, skip=False):
    """
    This module Views or Prints and/or Mails a Report

     mf     = MainFrame object
     conum  = Company Number
     pdfnam = Report PDF File Name
     splash = Display a Splash Screen
     header = Heading Used as Subject of Mail
     repprt = A list containing the following:
                Y or N widget created
                Output Report to (V)iew, (P)rinter or e(X)port
                Printer Name, view, export or blank
     repeml = A list containing the following:
                Y or N widget created
                Y or N to email the report
                Email addresses in a comma delimited string or a list/tuple
                Email message which can be:
                    A string i.e. a message
                    A List or Tuple which can be:
                        Subject and Message
                        Subject, Message and Attachment
                Whether to print/view the emailed report
     fromad = The from address if not the default system or company one.
     attach = A list of attachments
     skip   = Skip mail errors
     """
    import os, subprocess, time
    from TartanClasses import Dbase, ViewPDF, SplashScreen, Sql

    # Email Document
    sp = None
    if repeml and repeml[1] == "Y" and repeml[2]:
        try:
            dbm = Dbase(mf.rcdic)
            if dbm.err:
                raise Exception("Database Error")
            dbm.openDbase()
            sql = Sql(dbm, tables=["ctlsys", "ctlmst", "emllog"])
            if not conum:
                conum = 1
            smtp = sql.getRec(tables=["ctlmst", "ctlsys"], cols=["sys_msvr",
                "sys_mprt", "sys_msec", "sys_maut", "sys_mnam", "sys_mpwd",
                "ctm_email"], where=[("ctm_cono", "=", conum)], limit=1)
            if not fromad:
                fromad = smtp[6]
            if type(repeml[2]) == str:
                toad = repeml[2].split(",")
            else:
                toad = repeml[2]
            if repeml[3]:
                if type(repeml[3]) in (list, tuple):
                    if repeml[3][0]:
                        subj = repeml[3][0]
                    else:
                        subj = ""
                    if repeml[3][1]:
                        mess = repeml[3][1]
                    else:
                        mess = ""
                    if len(repeml[3]) == 3 and repeml[2]:
                        if not attach:
                            attach = repeml[3][2].split(",")
                        else:
                            attach.extend(repeml[3][2])
                else:
                    subj = ""
                    mess = repeml[3]
            else:
                subj = ""
                mess = ""
            if not subj:
                if type(header) in (tuple, list):
                    if len(header) == 1:
                        subj = header[0]
                    elif len(header) == 2:
                        subj = "%s %s" % tuple(header)
                    elif len(header) == 3:
                        subj = "%s %s %s" % tuple(header)
                elif header:
                    subj = header
                else:
                    subj = "PDF Report"
            if not attach or not attach[0]:
                att = [pdfnam]
            else:
                att = [pdfnam] + copyList(attach)
            for eml in toad:
                ok = False
                while not ok:
                    if splash:
                        sp = SplashScreen(mf.window.focus_displayof(),
                            "E-Mailing the Report to:\n\n%s\n\nPlease Wait.." %
                            eml)
                    err = sendMail(smtp[0:6], fromad, eml, subj, mess,
                        attach=att, wrkdir=mf.rcdic["wrkdir"])
                    if splash:
                        sp.closeSplash()
                    if err:
                        if skip:
                            ok = "SKIPPED"
                            break
                        else:
                            ok = askQuestion(mf.window.focus_displayof(),
                                "E-Mail Error", "Problem Delivering This "\
                                "Message.\n\nTo: %s\nSubject: %s\n\n%s\n\n"\
                                "Would You Like to Retry?" % (toad[0],
                                subj,err))
                            if ok == "yes":
                                ok = False
                            else:
                                ok = "FAILED"
                    else:
                        ok = "OK"
                    if ok:
                        try:
                            sql.insRec("emllog", data=[fromad.strip(), eml,
                            subj, "%04i-%02i-%02i %02i:%02i" %
                            time.localtime()[0:5], ok])
                        except:
                            pass
                        break
        except Exception as err:
            showException(mf.window.focus_displayof(), mf.rcdic["wrkdir"],
                "E-Mail Error\n\n%s" % err)
        try:
            dbm.commitDbase()
            dbm.closeDbase()
        except:
            pass
        if repeml[4].lower() == "n":
            return
    if repprt is None:
        return
    elif repprt[1].lower() == "x" or repprt[2] == "export":
        return
    elif repprt[2] and repprt[2].lower() == "none":
        return
    try:
        if repprt[1].lower() == "v" or repprt[2] == "view":
            # View Document
            chk = pdfnam.split(".")[-1]
            if chk == "svg":
                import webbrowser
                webbrowser.open_new(pdfnam)
            else:
                ViewPDF(mf, pdfnam)
            return
        # Print Document
        if splash:
            sp = SplashScreen(mf.window,
                "Printing the Report\n\nPlease Wait....")
        if repprt[2] == "Default":
            prt = getPrinters(donly=True)
            if not prt:
                raise Exception("No Default Printer")
        else:
            prt = repprt[2]
        prn = mf.rcdic["prn"]
        if prn:
            exe, chk = parsePrg(prn)
            if not os.path.exists(exe):
                prn = None
        if prn:
            cmd = []
            fle = False
            for c in chk:
                if "%p%" in c:
                    c = c.replace("%p%", prt)
                if "%f%" in c:
                    fle = True
                    c = c.replace("%f%", pdfnam)
                cmd.append(c)
            if exe.lower().count("sumatra") and len(cmd) == 1:
                cmd.extend(["-print-to", "%s" % prt, pdfnam])
            elif exe.lower().count("foxit") and len(cmd) == 1:
                cmd.extend(["/t", pdfnam, prt])
            elif cmd[0].endswith("lp") and len(cmd) == 1:
                cmd.extend(["-d", prt, "-o", "media=A4", pdfnam])
            elif cmd[0].endswith("lpr") and len(cmd) == 1:
                cmd.extend(["-P", prt, pdfnam])
            elif not fle:
                cmd.append(pdfnam)
            subprocess.Popen(cmd, stdout=subprocess.PIPE)
        else:
            printPDF(prt, pdfnam)
        if splash and sp:
            sp.closeSplash()
    except Exception as err:
        if splash and sp:
            sp.closeSplash()
        showException(mf.window.focus_displayof(), mf.rcdic["wrkdir"],
            "Output Error %s\n\nVwr: %s\nPrn: %s\n" % (err, mf.rcdic["vwr"],
            mf.rcdic["prn"]))

def textFormat(text, width=30, blong=True):
    """
    Format a text block

    Arguments:
        `text`   -- the string to be reformatted
        `width`  -- the maximum width of formatted text
        `blong`  -- break long words
    """
    if text:
        import textwrap
        return textwrap.wrap(text, width=width, break_long_words=blong)
    else:
        return [""]

def findFile(start=".", name=None, ftyp="f", case="n"):
    """
    A routine used to find a file given the starting path, name and whether
    name refers to a (f)ile or (d)irectory and whether case sensitive or not.
    """
    import os

    if type(start) == str:
        start = [start]
    name = os.path.basename(name)
    for d in start:
        for root, dirs, files in os.walk(d):
            if ftyp == "d":
                for pth in dirs:
                    if case == "y" and os.path.basename(pth) == name:
                        return os.path.join(root, pth)
                    elif os.path.basename(pth).lower() == name.lower():
                        return os.path.join(root, pth)
            else:
                for flenam in files:
                    if case == "y" and flenam == name:
                        return os.path.join(root, flenam)
                    elif flenam.lower() == name.lower():
                        return os.path.join(root, flenam)

def sendMail(server, ex, to, subj, mess="", attach=None, embed=None, check=False, timeout=30, local=None, lnkurl=None, wrkdir="."):
    """
    A routine to email a message, embed files and/or attach files.

        server  = A list of the smtp server details i.e.
                    [host, port, security, auth, username, password]
        ex      = The from address.
        to      = A list of recipients.
        subj    = The subject.
        mess    = The Message string or a tuple having (text, html).
        attach  = A list of attachments.
        embed   = A list of attachments to embed in the body of the message.
        check   = Check if the mail server is available.
        timeout = The number of seconds before timing out defaulting to 30.
        local   = The local hostname as fqdn.
        lnkurl  = An http link to add to the embedded attachments
        wrkdir  = The work directory, defaults to "."
    """
    import mimetypes, os, smtplib
    try:
        from email import encoders as Encoders
        from email.mime.base import MIMEBase
        from email.mime.text import MIMEText
        from email.mime.image import MIMEImage
        from email.mime.multipart import MIMEMultipart
        from email.utils import COMMASPACE, formatdate
    except:
        from email import Encoders
        from email.MIMEBase import MIMEBase
        from email.MIMEText import MIMEText
        from email.MIMEImage import MIMEImage
        from email.MIMEMultipart import MIMEMultipart
        from email.Utils import COMMASPACE, formatdate

    def mysubj(subj):
        words = subj.split()
        subj = ""
        for word in words:
            if not subj:
                subj = word[0].upper() + word[1:].lower()
            elif word.startswith("RCI"):
                subj = "%s %s" % (subj, word)
            else:
                subj = "%s %s" % (subj, word[0].upper() + word[1:].lower())
        return subj

    host, port, secu, auth, unam, upwd = server
    try:
        if secu and int(secu) == 2:
            smtp = smtplib.SMTP_SSL(host, port, local, timeout=timeout)
        else:
            smtp = smtplib.SMTP(host, port, local, timeout=timeout)
        if secu and int(secu) == 1:
            smtp.starttls()
        if auth:
            smtp.login(unam, upwd)
        if check:
            smtp.quit()
            return
    except Exception as err:
        if not check:
            showException(None, wrkdir, "Mail Server (%s %s) "\
                "Invalid or Unavailable\n\n%s" % (host, port, err))
        return err
    if type(to) == str:
        to = [to]
    if attach is None:
        attach = []
    elif type(attach) == str:
        attach = [attach]
    if embed is None:
        embed = []
    elif type(embed) == str:
        embed = [embed]
    msgRoot = MIMEMultipart("mixed")
    msgRoot["From"] = ex
    msgRoot["To"] = COMMASPACE.join(to)
    msgRoot["Date"] = formatdate(localtime=True)
    try:
        int(subj[:3])
        subj = subj[4:]
    except:
        pass
    subj = subj.replace("'", "").replace("(", " ").replace(")", " ")
    if not mess:
        if attach:
            mess = "Attached please find:\n\n%s" % subj
        else:
            mess = subj
    msgRoot["Subject"] = mysubj(subj)
    msgRelated = MIMEMultipart("related")
    msgRoot.attach(msgRelated)
    msgAlternative = MIMEMultipart("alternative")
    msgRelated.attach(msgAlternative)
    if type(mess) in (list, tuple):
        msgText = MIMEText(mess[0], "plain", "utf-8")
        mess = mess[1]
    else:
        msgText = MIMEText(mess, "plain", "utf-8")
    msgAlternative.attach(msgText)
    if lnkurl and lnkurl[0] == "Y":
        if not mess:
            mess = "<p><a href=\"%s\">Visit %s</a></p>" % (lnkurl[1], lnkurl[1])
        else:
            mess += "<a href=\"%s\">Visit %s</a></p>" % (lnkurl[1], lnkurl[1])
    if embed:
        seq = 1
        for flenam in embed:
            name = os.path.basename(flenam)
            ctype, encoding = mimetypes.guess_type(flenam)
            maintype, subtype = ctype.split("/", 1)
            if maintype == "image":
                try:
                    fp = open(flenam, "rb")
                    msgImage = MIMEImage(fp.read(), _subtype=subtype, name=name)
                    fp.close()
                    msgImage.add_header("Content-ID", "<image%s>" % seq)
                    msgRelated.attach(msgImage)
                    if lnkurl and lnkurl[0] == "N":
                        mess = mess + "<a href=\"%s\">" % lnkurl[1]
                    mess = mess + "<img src=\"cid:image%s\">" % seq
                    if lnkurl and lnkurl[0] == "N":
                        mess = mess + "</a>"
                    mess = mess + "<br>"
                    seq += 1
                except:
                    pass
    msgText = MIMEText(mess.replace("\n", "<br>"), "html", "utf-8")
    msgAlternative.attach(msgText)
    if attach:
        for flenam in attach:
            if os.path.isfile(os.path.normpath(flenam)):
                name = os.path.basename(flenam)
                ctype, encoding = mimetypes.guess_type(flenam)
                if ctype:
                    maintype, subtype = ctype.split("/", 1)
                else:
                    maintype, subtype = ("application", "octet-stream")
                part = MIMEBase(maintype, subtype, name=name)
                part.set_payload(open(os.path.normpath(flenam),"rb").read())
                Encoders.encode_base64(part)
                part.add_header("Content-Disposition",
                    'attachment; filename="%s"' % os.path.basename(flenam))
                msgRoot.attach(part)
    try:
        smtp.sendmail(ex, to, msgRoot.as_string())
        smtp.quit()
        return
    except Exception as err:
        try:
            smtp.quit()
        except:
            pass
        return err

def mthendDate(date):
    """
    Function to return the month-end date of a date
    """
    import calendar

    yy = int(date / 10000)
    mm = int((date % 10000) / 100)
    dy = calendar.monthrange(yy, mm)[1]
    return int((yy * 10000) + (mm * 100) + dy)

def projectDate(date, period, typ="days"):
    """
    Function to return a date which is the result of a date +- period

    typ can be days, months or years
    """
    import datetime

    yy = int(date / 10000)
    mm = int(date / 100) % 100
    dd = date % 100
    mthend = False
    if (typ == "days" and period == 365) or typ in ("months", "years"):
        if date == mthendDate(date):
            mthend = True
    if typ == "days":
        date = datetime.date(yy, mm, dd)
        ndate = date + datetime.timedelta(days=period)
        yy = ndate.year
        mm = ndate.month
        dd = ndate.day
    elif typ == "months":
        mm += period
        while mm > 12:
            mm -= 12
            yy += 1
        while mm < 1:
            mm += 12
            yy -= 1
    else:
        yy += period
    date = (yy * 10000) + (mm * 100) + dd
    if mthend:
        return mthendDate(date)
    else:
        return date

def paymentDate(base, stday, terms, trdt):
    """
    Function to return the due date of a transaction
    """
    if not terms:
        paydt = trdt
    elif base == "D":
        paydt = projectDate(trdt, terms)
    else:
        styr = int(trdt / 10000)
        stmt = int(trdt / 100) % 100
        if stday in (0, 30):
            stday = 31
        if trdt % 100 > stday:
            stmt += 1
            if stmt > 12:
                styr += 1
                stmt = 1
        lmed = mthendDate((styr * 10000) + (stmt * 100) + 1)
        if stday == 31 or stday > lmed % 100:
            stdt = lmed
        else:
            stdt = (styr * 10000) + (stmt * 100) + stday
        pyyr = int(stdt / 10000)
        pymt = int(stdt / 100) % 100
        if terms in (30, 31):
            pymt += 1
            if pymt > 12:
                pyyr += 1
                pymt = 1
            lmed = mthendDate((pyyr * 10000) + (pymt * 100) + 1)
            paydt = (pyyr * 10000) + (pymt * 100) + stday
            if stday == 31 or paydt > lmed:
                paydt = lmed
        else:
            paydt = projectDate(stdt, terms)
    return paydt

def dateDiff(date1, date2, ptype="years"):
    """
    Function to return the differance between two dates.
        date1 = Starting Date
        date2 = Ending Date
        ptype = years, months or days
    """
    import datetime, time

    def M(d):
        return ((d.year * 12) + d.month)

    def D(d1, d2):
        return (d1.toordinal() - d2.toordinal())

    def I(ymd):
        return datetime.date(*time.strptime(ymd, "%Y%m%d")[:3])

    date1 = I(str(date1))
    date2 = I(str(date2))
    if ptype.lower() == "days":
        days = D(date2, date1)
        if days <= 0:
            return 0
        return days
    months = M(date2) - M(date1)
    if ptype.lower() == "months":
        return months
    if ptype.lower() == "years":
        if date2.day < date1.day:
            months -= 1
        return int(months / 12)

def getUnderline(widget=None, blist=None, text=None):
    """
    blist = dictionary name to use
    text  = text to find underline for
    """
    if not text:
        return text, 0
    used = []
    if widget:
        # Toplevel bindings
        binds = set()
        binds |= set(widget.bind())
        binds |= set(widget.winfo_toplevel().bind())
        for bind in binds:
            if "<Key-Alt_L>" in bind:
                c = bind.split("<Key-Alt_L>")[1].upper()
                if c not in used:
                    used.append(c)
    elif blist:
        # Menu bindings
        used = blist
    if "_" in text:
        start = text.index("_")
        text = text.replace("_", "")
    else:
        start = 0
    pos = -1
    while pos == -1:
        for p in range(start, len(text)):
            c = text[p]
            # If the label is Exit, try and make x the hot key
            if text == "Exit" and c == "E" and "X" not in used:
                continue
            if ord(c.upper()) in range(65, 91) and c.upper() not in used:
                if not widget:
                    used.append(c.upper())
                pos = p
                break
        if pos == -1:
            more = False
            for x in range(65, 91):
                if chr(x) not in used:
                    text = "%s %s" % (text, chr(x))
                    more = True
                    break
            if not more:
                break
    return text, pos

def getPrinters(wrkdir=".", donly=False):
    """
    Function to return avalaible printers.
        donly = True, only return the default printer
        donly = False, return all printers
    """
    import subprocess, sys
    data = []
    dflt = None
    try:
        if sys.platform == "win32":
            import win32print
            # Get the default printer
            dflt = win32print.GetDefaultPrinter()
            if not donly:
                if dflt:
                    data.append(dflt)
                from win32print import EnumPrinters
                lst = EnumPrinters(2)
                for l in lst:
                    if l[2] not in data:
                        data.append(l[2])
        else:
            try:
                # Use pycups
                import cups
                conn = cups.Connection()
                # Get the default printer
                dflt = conn.getDefault()
                if not donly:
                    if dflt:
                        data.append(dflt)
                    lst = conn.getPrinters()
                    for l in lst:
                        if lst[l]["printer-state"] == 3 and l not in data:
                            data.append(l)
            except:
                # Use lpstat
                # Get the default printer
                proc = subprocess.Popen("lpstat -d", shell=True, bufsize=0,
                    stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE, close_fds=True)
                prt = proc.stdout.readline()
                if type(prt) == bytes:
                    prt = prt.decode("utf-8")
                dflt = prt.strip().split(":")
                if len(dflt) != 2:
                    dflt = None
                else:
                    dflt = dflt[1].strip()
                if not donly:
                    if dflt:
                        data.append(dflt)
                    proc = subprocess.Popen("lpstat -a", shell=True, bufsize=0,
                        stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE, close_fds=True)
                    lst = proc.stdout.readlines()
                    for l in lst:
                        if type(l) == bytes:
                            l = l.decode("utf-8")
                        l = l.rstrip().replace('"', "").replace("'", "")
                        if l.count("accepting requests"):
                            p = l.split()
                            if p[0].rstrip() not in data:
                                data.append(p[0].rstrip())
    except Exception as err:
        data = []
        dflt = None
        showException(None, wrkdir, "Get Printer Error\n\n%s" % err)
    if donly:
        return dflt
    elif data:
        return data
    return ["None"]

def removeFunctions(nam, dec=0):
    """removes all aggregate details from a column name
        nam = the column detail
        dec = decimal places"""
    if not nam.count("("):
        return nam
    nam = nam.lower().replace("avg(", "")
    nam = nam.lower().replace("count(", "")
    nam = nam.lower().replace("max(", "")
    nam = nam.lower().replace("min(", "")
    nam = nam.lower().replace("round(", "")
    nam = nam.lower().replace("stddev(", "")
    nam = nam.lower().replace("sum(", "")
    nam = nam.lower().replace("variance(", "")
    nam = nam.replace("),%s)" % dec, "")
    nam = nam.replace("), %s)" % dec, "")
    nam = nam.replace(")", "")
    return(nam.strip())

def getTrn(dbm, sys, cdt=None, jon=None, whr=None, odr=None, neg=True, zer="Y", lim=None):
    """
    This function returns a list of column names and a list of lists of data.

    sys = crs - Creditor's Ledger
          drs - Debtor's Ledger
          mem - Member's Ledger
    cdt = A cut-off curdt period where any transactions zeroed before this
          period will not be returned.
    jon = An addition to the join statement e.g. "cra_curdt <= 200612"
    whr = A list of tuples of where statements e.g. [("drt_cono", "=", 7)]
    odr = An order replacement.
    neg = True or False to include or exclude negative transaction balances
    zer = Y, A, N to include or exclude zero transaction balances
    lim = The number of records to return, defaults to all
    """
    from TartanClasses import ASD, Sql
    if sys == "crs":
        tab = ["crstrn", "crsage"]
        sql = Sql(dbm, tab, prog=__name__)
        col = copyList(sql.crstrn_col)
        join = "left outer join crsage on cra_cono=crt_cono and "\
            "cra_acno=crt_acno and cra_type=crt_type and cra_ref1=crt_ref1"
        if cdt:
            if whr:
                whr.append(("crt_curdt", "<=", cdt))
            else:
                whr = [("crt_curdt", "<=", cdt)]
            join = "%s and cra_curdt <= %s" % (join, cdt)
        if jon:
            join = "%s and %s" % (join, jon)
    elif sys == "drs":
        tab = ["drstrn", "drsage"]
        sql = Sql(dbm, tab, prog=__name__)
        col = copyList(sql.drstrn_col)
        join = "left outer join drsage on dra_cono=drt_cono and "\
            "dra_chain=drt_chain and dra_acno=drt_acno and "\
            "dra_type=drt_type and dra_ref1=drt_ref1"
        if cdt:
            if whr:
                whr.append(("drt_curdt", "<=", cdt))
            else:
                whr = [("drt_curdt", "<=", cdt)]
            join = "%s and dra_curdt <= %s" % (join, cdt)
        if jon:
            join = "%s and %s" % (join, jon)
    elif sys == "mem":
        tab = ["memtrn", "memage"]
        sql = Sql(dbm, tab, prog=__name__)
        col = copyList(sql.memtrn_col)
        join = "left outer join memage on mta_cono=mlt_cono and "\
            "mta_memno=mlt_memno and mta_type=mlt_type and "\
            "mta_refno=mlt_refno"
        if cdt:
            if whr:
                whr.append(("mlt_curdt", "<=", cdt))
            else:
                whr = [("mlt_curdt", "<=", cdt)]
            join = "%s and mta_curdt <= %s" % (join, cdt)
        if jon:
            join = "%s and %s" % (join, jon)
    else:
        return None, None
    grp = ""
    for c in col:
        grp = "%s%s, " % (grp, c)
    grp = grp[:-2]
    if sys == "crs":
        col.append("max(cra_curdt)")
        col.append("sum(cra_amnt)")
        if not odr:
            odr = "crt_trdt, crt_type, crt_ref1"
    elif sys == "drs":
        col.append("max(dra_curdt)")
        col.append("sum(dra_amnt)")
        if not odr:
            odr = "drt_trdt, drt_type, drt_ref1"
    else:
        col.append("max(mta_curdt)")
        col.append("sum(mta_amnt)")
        if not odr:
            odr = "mlt_trdt, mlt_type, mlt_refno"
    recs = sql.getRec(tab[0], join=join, cols=col, where=whr, group=grp,
        order=odr, limit=lim)
    col[-2] = "curdt"
    col[-1] = "paid"
    col.append("balance")
    dat = []
    if recs and lim == 1:
        recs = [recs]
    tot = 0
    for rec in recs:
        if sys == "crs":
            amt = rec[col.index("crt_tramt")]
        elif sys == "drs":
            amt = rec[col.index("drt_tramt")]
        else:
            amt = rec[col.index("mlt_tramt")]
        if rec[col.index("curdt")] is None:
            rec[col.index("curdt")] = 0
        pay = rec[col.index("paid")]
        if pay is None:
            pay = 0.0
        if cdt:
            if sys == "crs":
                ccc = ["sum(cra_amnt)"]
                whr = [("cra_cono", "=", rec[0]), ("cra_acno", "=", rec[1]),
                    ("cra_atyp", "=", rec[2]), ("cra_aref", "=", rec[3]),
                    ("cra_curdt", ">", cdt)]
            elif cdt and sys == "drs":
                ccc = ["sum(dra_amnt)"]
                whr = [("dra_cono", "=", rec[0]), ("dra_chain", "=", rec[1]),
                    ("dra_acno", "=", rec[2]), ("dra_atyp", "=", rec[3]),
                    ("dra_aref", "=", rec[4]), ("dra_curdt", ">", cdt)]
            elif cdt:
                ccc = ["sum(mta_amnt)"]
                whr = [("mta_cono", "=", rec[0]), ("mta_memno", "=", rec[1]),
                    ("mta_atyp", "=", rec[2]), ("mta_aref", "=", rec[3]),
                    ("mta_curdt", ">", cdt)]
            mor = sql.getRec(tab[1], cols=ccc, where=whr, limit=1)
            if mor[0]:
                pay = float(ASD(pay) + ASD(mor[0]))
        rec[col.index("paid")] = pay
        bal = float(ASD(amt) - ASD(pay))
        tot = float(ASD(tot) + ASD(bal))
        rec.append(bal)
        if bal or zer == "A":
            dat.append(rec)
        elif zer == "Y" and (not cdt or rec[col.index("curdt")] >= cdt):
            dat.append(rec)
    if not neg and tot < 0:
        return col, []
    return col, dat

def chkGenAcc(mf, coy, acc, ctl=True, pwd=True):
    """
    This function returns None if the account allows postings or is not an
    integrated control else if pwd it asks for a Password and returns an
    Error Message if not correct otherwise return desc, ind, vat and type.

    coy    - The company number.
    acc    - The general ledger account number.
    ctl    - Whether to check if it is an integrated control account.
    pwd    - Allow overriding restrictions if the password is known.
    """
    from TartanClasses import GetCtl, PwdConfirm, Sql

    gc = GetCtl(mf)
    # Get Company Master Record
    ctlmst = gc.getCtl("ctlmst", coy)
    if not ctlmst:
        return "Invalid Company Number"
    # Check for Company Control Records
    ctlctl = gc.getCtl("ctlctl", coy)
    if not ctlctl:
        return "Missing Control Records"
    # Check if Valid Account
    sql = Sql(mf.dbm, "genmst", prog=__name__)
    rec = sql.getRec("genmst", cols=["glm_desc", "glm_ind", "glm_vat",
        "glm_type"], where=[("glm_cono", "=", coy), ("glm_acno", "=", acc)],
        limit=1)
    if not rec:
        return "Invalid Account Number"
    if not ctl:
        return rec
    # Check if Postings Allowed
    if rec[1] == "N":
        if pwd:
            cf = PwdConfirm(mf, conum=coy, system="GEN", code="NoPost")
            if not cf.pwd or cf.flag == "no":
                return "Postings Disallowed for %s" % acc
            return rec
    # Check If Account Is a Control Record
    mod = []
    for x in range(0, len(ctlmst["ctm_modules"].rstrip()), 2):
        mod.append(ctlmst["ctm_modules"][x:x+2])
    ctlacc = []
    # Load Ass Control
    if mod.count("AR"):
        ctl = gc.getCtl("assctl", coy)
        if not ctl:
            return "Missing Control Record - assctl"
        if ctl["cta_glint"] == "Y":
            if "ass_sls" not in ctlctl:
                return "Missing Control - ass_sls"
            ctlacc.append(ctlctl["ass_sls"])
            sql = Sql(mf.dbm, "assgrp", prog=__name__)
            ass = sql.getRec("assgrp", cols=["asg_assacc",
                "asg_depacc"], where=[("asg_cono", "=", coy)])
            for a in ass:
                ctlacc.extend(a)
    # Load Bkm Controls
    if mod.count("BK"):
        ctl = gc.getCtl("bkmctl", coy)
        if not ctl:
            return "Missing Control Record - bkmctl"
        if ctl["cbk_glint"] == "Y":
            if "bkm_ctl" not in ctlctl:
                return "Missing Control - bkm_ctl"
            ctlacc.append(ctlctl["bkm_ctl"])
    # Load Crs Control
    if mod.count("CR"):
        ctl = gc.getCtl("crsctl", coy)
        if not ctl:
            return "Missing Control Record - crsctl"
        if ctl["ctc_glint"] == "Y":
            if "crs_ctl" not in ctlctl:
                return "Missing Control - crs_ctl"
            ctlacc.append(ctlctl["crs_ctl"])
    # Load Drs Control
    if mod.count("DR"):
        ctl = gc.getCtl("drsctl", coy)
        if not ctl:
            return "Missing Control Record - drsctl"
        if ctl["ctd_glint"] == "Y":
            if "drs_ctl" not in ctlctl:
                return "Missing Control - drs_ctl"
            ctlacc.append(ctlctl["drs_ctl"])
    # Load Lon Control
    if mod.count("LN"):
        ctl = gc.getCtl("lonctl", coy)
        if not ctl:
            return "Missing Control Record - lonctl"
        if ctl["cln_glint"] == "Y":
            if "lon_ctl" not in ctlctl:
                return "Missing Control - lon_ctl"
            ctlacc.append(ctlctl["lon_ctl"])
    # Load Mem Control
    if mod.count("ML"):
        ctl = gc.getCtl("memctl", coy)
        if not ctl:
            return "Missing Control Record - memctl"
        if ctl["mcm_glint"] == "Y":
            if "mem_ctl" not in ctlctl:
                return "Missing Control - mem_ctl"
            ctlacc.append(ctlctl["mem_ctl"])
    # Load Rca Control
    if mod.count("RC"):
        ctl = gc.getCtl("rcactl", coy)
        if not ctl:
            return "Missing Control Record - rcactl"
        if ctl["cte_glint"] == "Y":
            ctlacc.extend([ctlctl["rca_own"], ctlctl["rca_tnt"],
                ctlctl["rca_dep"]])
    # Load Rtl Integration
    if mod.count("RT"):
        ctl = gc.getCtl("rtlctl", coy)
        if not ctl:
            return "Missing Control Record - rtlctl"
        if ctl["ctr_glint"] == "Y":
            sql = Sql(mf.dbm, "rtlprm", prog=__name__)
            rtl = sql.getRec("rtlprm", cols=["rtp_rtlacc"],
                where=[("rtp_cono", "=", coy)])
            for a in rtl:
                ctlacc.append(a[0])
    # Load Sln Control
    if mod.count("SL"):
        ctl = gc.getCtl("wagctl", coy)
        if not ctl:
            return "Missing Control Record - wagctl"
        if ctl["ctw_glint"] == "Y":
            if "wag_slc" not in ctlctl:
                return "Missing Control - wag_slc"
            ctlacc.append(ctlctl["wag_slc"])
    if acc in ctlacc:
        if pwd:
            cf = PwdConfirm(mf, conum=coy, system="GEN", code="CtlAcc")
            if not cf.pwd or cf.flag == "no":
                return "%s Is A Control Account" % acc
        else:
            return "%s Is an Integrated Control Account" % acc
    return rec

def chkAggregate(fld):
    """
    Use this function to check if fld is a sql aggregate function.
    """
    for typ in ("avg(","count(","max(","min(","sum("):
        if fld.lower().count(typ):
            return True

def callModule(mf, df, mod, coy=None, period=None, user=None, args=None, wait=True, ret=None):
    """
    Use this funtion to call another module and pass arguments to it.

    coy    - A list of company number and name.
    period - The financial period tuple.
    user   - The user name.
    args   - Any arguments to send to the called module.
    wait   - Wait for window to be destroyed.
    ret    - A variable to return to the calling module.
    """
    if df:
        state = df.disableButtonsTags()
        df.setWidget(df.mstFrame, state="hide")
    opts = {"mf": mf}
    if coy is not None:
        opts["conum"] = coy[0]
        opts["conam"] = coy[1]
    if period is not None:
        opts["period"] = period
    if user is not None:
        opts["capnm"] = user
    if args is not None:
        opts["args"] = args
    if wait:
        opts["wait"] = True
    obj = runModule(mod, **opts)
    if df:
        df.setWidget(df.mstFrame, state="show")
        df.enableButtonsTags(state=state)
        df.focusField(df.frt, df.pag, df.col)
    if not obj:
        return
    if not ret:
        return True
    try:
        if type(ret) in (list, tuple):
            dat = {}
            for r in ret:
                dat[r] = getattr(obj, r)
            return dat
        else:
            return getattr(obj, ret)
    except:
        return

def runModule(mod, **popt):
    import importlib
    try:
        from tartanWork import pkgs
        pkg = pkgs[mod[:2]]
        com = importlib.import_module("..%s" % mod, "%s.subpkg" % pkg)
        exe = getattr(com, mod)
        # Return object for callModule
        return exe(**popt)
    except Exception as err:
        print("Exception", err)

def doChkCatChg(mf, cono, memno, nxtdt):
    from TartanClasses import Sql

    sql = Sql(mf.dbm, ["memctc", "memmst", "memcat"])
    # Get age and years as a member
    if type(memno) in (list, tuple):
        dob, doe = memno
    else:
        mst = sql.getRec("memmst", where=[("mlm_cono", "=", cono),
            ("mlm_memno", "=", memno)], limit=1)
        if not mst:
            return
        dob = mst[sql.memmst_col.index("mlm_dob")]
        doe = mst[sql.memmst_col.index("mlm_entry")]
    if not dob or not doe:
        return
    age = dateDiff(dob, nxtdt, "years")
    mem = dateDiff(doe, nxtdt, "years")
    # Get current category
    cat = sql.getRec("memcat", where=[("mlc_cono", "=", cono),
        ("mlc_memno", "=", memno), ("mlc_type", "=", "B")], limit=1)
    if not cat:
        return
    ocode = cat[sql.memcat_col.index("mlc_code")]
    start = cat[sql.memcat_col.index("mlc_start")]
    end = cat[sql.memcat_col.index("mlc_end")]
    last = cat[sql.memcat_col.index("mlc_last")]
    # Get current category limits and next code (ncode)
    rec = sql.getRec("memctc", where=[("mcc_cono", "=", cono),
        ("mcc_type", "=", "B"), ("mcc_code", "=", ocode)], limit=1)
    if not rec:
        return
    desc = rec[sql.memctc_col.index("mcc_desc")]
    freq = rec[sql.memctc_col.index("mcc_freq")]
    glac = rec[sql.memctc_col.index("mcc_glac")]
    ncode = rec[sql.memctc_col.index("mcc_ncode")]
    age_l = rec[sql.memctc_col.index("mcc_age_l")]
    and_s = rec[sql.memctc_col.index("mcc_and_s")]
    or_s = rec[sql.memctc_col.index("mcc_or_s")]
    # See if limits have been reached and if so return new code details
    ret = None
    while ncode:
        if not age_l and not or_s:
            break
        if not and_s and age_l and age > age_l:
            chg = True
        elif and_s and age_l and age >= age_l and mem >= and_s:
            chg = True
        elif or_s and mem >= or_s:
            chg = True
        else:
            chg = False
        if chg:
            ret = (ocode, start, end, last, freq, desc, glac, ncode)
        chk = sql.getRec("memctc", where=[("mcc_cono", "=", cono),
            ("mcc_type", "=", "B"), ("mcc_code", "=", ncode)], limit=1)
        ncode = chk[sql.memctc_col.index("mcc_ncode")]
        age_l = chk[sql.memctc_col.index("mcc_age_l")]
        and_s = rec[sql.memctc_col.index("mcc_and_s")]
        or_s = rec[sql.memctc_col.index("mcc_or_s")]
    return ret

def httpDownload(url, dest=None, check=False):
    import requests
    try:
        data = requests.get(url)
        if dest is None:
            return data.content.decode("utf-8").rstrip()
        with open(dest, "wb") as file:
            file.write(data.content)
        return True
    except:
        return

def doDrawTable(fpdf, rr, ppad=1, spad=1, cw=None, ld=None, font=True):
    """
    Draw a pdf table

        fpdf    - A fpdf instance.
        rr      - A dictionary of the layout options of the table as follows:
                        "margins": ((left, right), (top, bottom))
                        "repeat":  (across, below)  # repeat the table
                        "rows":    (
                                 (x, y, w, h, fill, text, Centre),
                                 (x, y, (w, h, fill, (text x qty)), qty))

                  Example: {
                        "margins": ((10, 85), (19.5, 27)),
                        "repeat": (1, 3),
                        "rows": (
                            (10, 19.5, (
                                (10, 1.5, .8, "Column1"),
                                (10, 1.5, .8, "Column2"),
                                (50, 1.5, .8, "Column3"),
                                (10, 1.5, .8, "Column4"))),
                            (10, 21, (
                                (10, 1.5, .9, ("Row1","Row2","Row3","Row4")),
                                (10, 1.5),
                                (50, 1.5),
                                (10, 1.5), 4)))}

                  Notes:
                    You can add True or False after the text to draw a
                        centred string.
                    You can add True or False after the alignment to
                        draw the string if bold format.

        ppad    - Y padding for drawString.
        spad    - Space padding for repeats.
        cw      - The width of a character in mm i.e. fpdf.get_string_width()
        ld      - The line depth e.g. 4mm
        font    - Whether or not to set a font.
    """
    table = []
    last = 0
    cols = []
    rows = []
    rest = []
    if type(font) in (list, tuple):
        dflt = list(font)
    elif ppad == 1:
        dflt = ["helvetica", "B", 10]
    else:
        dflt = ["helvetica", "B", 12]
    if font:
        fpdf.setFont(dflt[0], dflt[1], dflt[2], default=True)
    for row in rr["rows"]:
        if type(row[2]) in (list, tuple):
            # Mutiple columns
            if len(row) == 4:
                qty = row[3]
            else:
                qty = 1
            cv = row[0]                     # x
            tc = [[cv]]                     # total cols
            rv = row[1]                     # y
            tr = [[rv]] * qty               # total rows
            rt = []
            for q in range(qty):
                cv = row[0]                 # x
                tc = [[cv]]                 # total cols
                rt = []
                for n, r in enumerate(row[2]):
                    if len(row[2]) != n+1:
                        cv += r[0]          # add col width
                        tc.append([cv])
                    if len(r) == 4 and type(r[3]) in (list, tuple):
                        ap = list(r[:3])
                        ap.append(r[3][q])
                        rt.append(ap)
                    else:
                        rt.append(list(r))
                cols.extend(tc)
                rest.extend(rt)
            rv = row[1]                     # y
            tr = []                         # total rows
            for q in range(qty):
                tr.append([rv])             # total rows
                for n, r in enumerate(row[2]):
                    if len(row[2]) != n+1:
                        tr.append([rv])
                rv += r[1]                  # add row height
            rows.extend(tr)
        else:
            cols.append([row[0]])
            rows.append([row[1]])
            rest.append(list(row[2:]))
    for x in range(1, rr["repeat"][0]):
        for n, c in enumerate(cols):
            cols[n] = c + [c[-1] + rr["margins"][0][1] + 1]
    d = (rr["margins"][1][1] - rr["margins"][1][0])
    for n, r in enumerate(rows):
        nxt = r
        for x in range(1, rr["repeat"][1]):
            last = (nxt[-1] + d) + spad
            nxt = nxt + [last]
        rows[n] = nxt
    for n, r in enumerate(rows):
        table.append([cols[n]] + [rows[n]] + rest[n])
    for r in table:
        if len(r) > 4 and r[4]:
            fpdf.set_fill_color(int(255 * r[4]))
            fill = 1
        else:
            fill = 0
        for x in r[0]:                      # x
            for y in r[1]:                  # y
                if len(r) > 6 and r[6]:
                    align = "C"
                else:
                    align = ""
                if len(r) > 7 and r[7]:
                    style = dflt[:]
                else:
                    style = [dflt[0], "", dflt[2]]
                xx = round(x * cw, 2)
                yy = round(y * ld, 2)
                if len(r) > 5:
                    txt = r[5]
                    fpdf.drawText(x=xx, y=yy, w=r[2] * cw, h=r[3] * ld,
                        align=align, border="TLRB", fill=fill, txt=txt,
                        font=style)
                else:
                    if fill:
                        style = "DF"
                    else:
                        style = "D"
                    fpdf.rect(xx, yy, r[2] * cw, r[3] * ld, style=style)
    return last, table

def getDeposit(mf, conum, date, owner, code, acno):
    """
    Calculate the deposit plus interest on rented premises:

    conum - The company number
    owner - The owner's code
    code  - the premise's code
    acno  - The tenant's account number
    """
    from TartanClasses import ASD, Sql

    sql = Sql(mf.dbm, ["rcaint", "rcatnt"], prog=__name__)
    tnt = sql.getRec("rcatnt", where=[("rtu_cono", "=", conum),
        ("rtu_owner", "=", owner), ("rtu_code", "=", code), ("rtu_acno",
        "=", acno), ("rtu_trdt", "<=", date), ("rtu_mtyp", "=", 2)],
        order="rtu_trdt")
    if not tnt:
        return
    coi = sql.rcaint_col
    cot = sql.rcatnt_col
    dep = 0
    recs = []
    for n, t in enumerate(tnt):
        start = t[cot.index("rtu_trdt")]
        dep = float(ASD(dep) + ASD(t[cot.index("rtu_tramt")]))
        rte = sql.getRec("rcaint", where=[("rci_cono", "=", conum),
            ("rci_date", "<=", start)], order="rci_date")
        if not rte:
            rte = [[conum, 0, 0, 0, 0]]
        recs.append([start, t[cot.index("rtu_tramt")], rte[-1]])
        if (n + 1) == len(tnt):
            end = date
        else:
            end = tnt[n + 1][cot.index("rtu_trdt")]
        rat = sql.getRec("rcaint", where=[("rci_cono", "=", conum),
            ("rci_date", ">", start), ("rci_date", "<", end)],
            order="rci_date")
        for r in rat:
            recs.append([r[coi.index("rci_date")], 0, r])
    if not recs:
        return
    if recs[-1][2][0] < date:
        recs.append([date, 0, []])
    bal = 0
    dep = 0
    dat = []
    begin = True
    for n, a in enumerate(recs):
        cap = float(ASD(0) - ASD(a[1]))
        dep = float(ASD(dep) + ASD(cap))
        if begin:
            begin = False
            start = a[0]
            bal = cap
            dat.append([start, cap, 0, 0, 0, bal, 0])
            continue
        cdate = recs[n][0]
        bankr = recs[n-1][2][3]
        commr = recs[n-1][2][4]
        invtr = float(ASD(bankr) - ASD(commr))
        days = dateDiff(start, cdate, ptype="days")
        commi = round(bal * commr * days / 365.0 / 100.0, 2)
        invti = round(bal * invtr * days / 365.0 / 100.0, 2)
        bal = float(ASD(bal) + ASD(invti) + ASD(cap))
        dat.append([cdate, cap, days, invtr, invti, bal, commi])
        start = cdate
    return dat

def genAccNum(name, seq, size=7):
    valid = list(map(chr, list(range(65, 91))))
    valid.extend(list(map(chr, list(range(97, 123)))))
    valid.append("&")
    data = ""
    for c in name:
        if c in valid:
            data = data + c
    if len(data) < (size - 2):
        data = data + ("0" * ((size - 2) - len(data)))
    acno = ""
    for c in range(0, (size - 2)):
        acno = (acno + data[c]).upper()
    acno = acno.strip()
    text = "%s%0" + str((size - len(acno))) + "d"
    return text % (acno, seq)

def getNextCode(sql, table, column, where=None, start=1, last=999):
    nxt = sql.getRec(tables=table, cols=[column], where=where, order=column)
    code = 1
    for n in range(start, last):
        if [n] not in nxt:
            code = n
            break
    return code

def getGreens(text, needed, keep=None):
    if not text:
        return None, None, None, "No Greens Entered"
    if keep is None:
        keep = []
    greens = {}
    rinks = 0
    errs = None
    grns = text.split(",")
    for n, g in enumerate(grns):
        if not g:
            errs = True
            break
        if not n:
            # First Green
            first = g[0]
        greens[g[0]] = []
        if len(g) == 1:
            for x in range(1, 7):
                greens[g[0]].append(x)
                rinks += 1
        elif len(g) == 2 and g[1] == "7":
            for x in range(1, 8):
                greens[g[0]].append(x)
                rinks += 1
        else:
            for r in g[1:]:
                try:
                    r = int(r)
                    if r not in greens[g[0]]:
                        greens[g[0]].append(r)
                        rinks += 1
                except:
                    pass
    if errs:
        return None, None, None, "Invalid Entry"
    if rinks < needed:
        return None, None, None, "Rinks Entered %s, %s Needed" % (rinks, needed)
    # End rinks
    endrks = []
    for g in greens:
        if 1 in greens[g]:
            endrks.append("%s%s" % (g, 1))
        if 7 in greens[g]:
            endrks.append("%s%s" % (g, 7))
        elif 6 in greens[g]:
            endrks.append("%s%s" % (g, 6))
    # Remove rinks not needed starting with end rinks
    rem = rinks - needed
    chk = [0, 6, 5, 4, 0, 3, 2, 1]
    num = 7
    keys = list(greens.keys())
    keys.sort()
    while rem:
        for g in keys:
            rnk = "%s%s" % (g, num)
            if num in greens[g] and rnk not in keep:
                greens[g].remove(num)
                if rnk in endrks:
                    endrks.remove("%s%s" % (g, num))
                    if num == 7:
                        endrks.append("%s6" % g)
                rem -= 1
            if not rem:
                break
        num = chk[num]
    return greens, first, endrks, None

def payeTables(sql, year):
    txa = sql.getRec("wagtxa", where=[("wta_year", "=", year)],
        limit=1)
    txr = sql.getRec("wagtxr", where=[("wtr_year", "=", year)],
        order="wtr_from desc")
    if not txa or not txr:
        return None, None
    rates = []
    for rate in txr:
        if not rate[1]:
            rates.append([rate[1], rate[3], rate[4]])
        else:
            rates.append([rate[1] - 1, rate[3], rate[4]])
    return txa[1:], rates

def getSingleRecords(mf, table, cols, title=None, head=None, where=None, group=None, order=None, selcol=None, ttype="T", dic=None, items=None):
    """
    This is a gui used to select individual records.

    table   The table(s) from which to return the records.
            In the case of multiple tables, only the first table's columns
            will be returned, all other tables must only be used for the
            where statement.
    cols    A list of columns to display in the select view.
    title   The title.
    head    A list of column headings.
    where   A list of where conditions else all records or a list of data to
            use in the selection.
    group   True or False
    order   The order in which the records are to be displayed else
            the order will be cols.
    selcol  The column for keyboard selection.
    ttype   T = The where above is a list of where conditions.
            D = The where above is a list of data. In this case the returned
                columns will be as per the cols list.
    dic     A dictionary of column details as per sql.table_dic. This only
            applies when ttype is 'D'.
    items   A list of ticked items as follows:
                [col, [rec, rec, rec, .....]]
    """
    from TartanClasses import CCD, SChoice, Sql

    if not title:
        title = "Available Records"
    if table is not None and type(table) == str:
        table = [table]
    if ttype == "D":
        recs = where
        col = copyList(cols)
        if not dic:
            sql = Sql(mf.dbm, table, prog=__name__)
            tmp = getattr(sql, "%s_dic" % table[0])
            dic = {}
            for key in tmp:
                if key in col:
                    dic[key] = tmp[key]
        if not head:
            head = [""]
            for c in cols:
                head.append(dic[c][5])
    else:
        sql = Sql(mf.dbm, table, prog=__name__)
        dic = getattr(sql, "%s_dic" % table[0])
        col = getattr(sql, "%s_col" % table[0])
        if not head:
            head = [""]
            for c in cols:
                head.append(dic[c][5])
        if not order:
            odr = ""
            for c in cols:
                if not odr:
                    odr = c
                else:
                    odr = "%s, %s" % (odr, c)
        else:
            odr = order
        if group and type(group) == str:
            grp = group
        elif group:
            grp = ""
            for c in col:
                if not grp:
                    grp = c
                else:
                    grp = "%s, %s" % (grp, c)
        else:
            grp = None
        recs = sql.getRec(tables=table, cols=col, where=where, group=grp,
            order=odr)
    if not recs:
        showError(mf.window, "Error", "No Records Available")
        return []
    sel = []
    lines = []
    types = [("CB", 1)]
    for num, acc in enumerate(recs):
        lin = []
        for c in cols:
            if not num:
                types.append((dic[c][2], dic[c][3]))
            lin.append(CCD(acc[col.index(c)], dic[c][2], dic[c][3]).work)
        lines.append(lin)
    if selcol:
        selc = list(cols).index(selcol)
    else:
        selc = None
    sc = SChoice(mf, scrn=mf.window, titl=title, head=head, typs=types,
        data=lines, rows=None, retn="I", mode="M", selc=selc, items=items)
    mf.window.update()
    if sc.selection:
        for idx in sc.selection:
            sel.append(recs[idx])
    return sel

def getVatRate(sql, cono, code, date=None):
    where = [("vtr_cono", "=", cono), ("vtr_code", "=", code)]
    if date:
        where.append(("vtr_date", "<=", date))
    recs = sql.getRec("ctlvrf", cols=["vtr_rate"], where=where,
        order="vtr_date")
    if recs:
        return recs[-1][0]

def getCost(sql, cono, group, code, loc=None, qty=1, ind="I", recp=False, tot=False, bal=False):
    """
    recp    True or Line of a Sales Document which is a Recipe
                 [document type [I/C/O/W/Q], docno, lineno]
    qty     The number of items
    ind     "L" - Only return Last Cost
            "A" - Only return Average Cost
            "I" - Only return Value Indicator Cost
            ( )   or a tuple of indicators e.g. ("A", "L")
    tot     True - Return total cost
    bal     True - Return Quantity and Cost balances
    """
    from TartanClasses import ASD, CCD
    # Check for Recipe
    if not recp:
        chk = sql.getRec("strmf1", cols=["st1_type"], where=[("st1_cono",
            "=", cono), ("st1_group", "=", group), ("st1_code", "=", code)],
            limit=1)
        if chk[0] == "R" and loc:
            recp = True
    if recp:
        # Recipe
        items = []
        if type(recp) in (list, tuple) and "slsiv3" in sql.tables:
            tab = "slsiv3"
            col = ["si3_rgroup", "si3_rcode", "si3_rqty"]
            whr = [
                ("si3_cono", "=", cono),
                ("si3_rtn", "=", recp[0]),
                ("si3_docno", "=", recp[1]),
                ("si3_line", "=", recp[2])]
            items = sql.getRec(tab, cols=col, where=whr)
        if not items:
            tab = "strrcp"
            col = ["srr_rgroup", "srr_rcode", "srr_rqty"]
            whr = [
                ("srr_cono", "=", cono),
                ("srr_group", "=", group),
                ("srr_code", "=", code),
                ("srr_loc", "=", loc)]
            items = sql.getRec(tab, cols=col, where=whr)
    else:
        # Single Item
        items = [[group, code, qty]]
    # Variables
    acost = 0
    lcost = 0
    scost = 0
    tcost = 0
    csts = []
    for c in ind:
        csts.append(c)
    # Items
    for ggg, ccc, qqq in items:
        ac = 0
        lc = 0
        sc = 0
        tc = 0
        vin = sql.getRec("strmf1", cols=["st1_value_ind"],
            where=[("st1_cono", "=", cono), ("st1_group",
            "=", ggg), ("st1_code", "=", ccc)], limit=1)
        if vin[0] == "N":
            idx = len(csts)
            csts = []
            for _ in range(idx):
                csts.append("N")
        elif "I" in csts:
            csts = [vin[0]]
        elif vin[0] == "S" and "A" in csts:
            csts[csts.index("A")] = vin[0]
        # Cost Prices
        where = [
            ("stt_cono", "=", cono),
            ("stt_group", "=", ggg),
            ("stt_code", "=", ccc)]
        grp = "stt_cono, stt_group, stt_code"
        if loc:
            where.append(("stt_loc", "=", loc))
            grp += ", stt_loc"
        bals = sql.getRec("strtrn", cols=["sum(stt_qty)", "sum(stt_cost)"],
            where=where, group=grp, limit=1)
        if not bals:
            bals = [0, 0]
        # Average Cost
        if bals[0] and qqq == bals[0]:
            ac = round((bals[1] / qty), 2)
            tc = bals[1]
        elif bals[0]:
            ac = round((bals[1] / bals[0]), 2)
        # Last Cost
        whr = where[:]
        whr.append(("stt_type", "in", (1, 3)))
        chk = sql.getRec("strtrn", cols=["stt_qty", "stt_cost"],
            where=whr, order="stt_capdt desc, stt_seq desc")
        for rec in chk:
            q = CCD(rec[0], "SD", 11.2)
            c = CCD(rec[1], "SD", 11.2)
            if q.work and q.work > 0 and c.work:
                lc = round((c.work / q.work), 2)
                break
        # If average cost is zero
        if not ac and lc:
            ac = lc
        if "S" in csts:
            # Standard Cost
            whr = [
                ("stc_cono", "=", cono),
                ("stc_group", "=", ggg),
                ("stc_code", "=", ccc)]
            if loc:
                whr.append(("stc_loc", "=", loc))
            std = sql.getRec("strcst", cols=["stc_cost"], where=whr, limit=1)
            if std:
                sc = std[0]
        if recp:
            if tc:
                acost = float(ASD(acost) + ASD(tc))
                tcost = float(ASD(tcost) + ASD(tc))
            else:
                acost = float(ASD(acost) + ASD(round(ac * qqq, 2)))
            lcost = float(ASD(lcost) + ASD(round(lc * qqq, 2)))
            scost = float(ASD(scost) + ASD(round(sc * qqq, 2)))
        else:
            acost = ac
            lcost = lc
            scost = sc
            tcost = tc
    prcs = []
    vals = []
    if recp:
        bals = [1, 0]
    for flg in csts:
        if flg == "A":
            prcs.append(acost)
            if tcost:
                vals.append(tcost)
            else:
                vals.append(round(qty * acost, 2))
        elif flg in ("Y", "L"):
            prcs.append(lcost)
            vals.append(round(qty * lcost, 2))
        elif flg == "S":
            prcs.append(scost)
            vals.append(round(qty * scost, 2))
        elif flg == "N":
            prcs.append(0)
            vals.append(0)
    if len(csts) == 1:
        if tot:
            if bal:
                return prcs[0], vals[0], bals
            else:
                return prcs[0], vals[0]
        elif bal:
            return prcs[0], bals
        else:
            return prcs[0]
    elif tot:
        if bal:
            return prcs, vals, bals
        else:
            return prcs, vals
    elif bal:
        return prcs, bals
    else:
        return prcs

def getSell(sql, cono, group, code, loc=None, lvl=1, recp=False, ind=None):
    """
    recp   - True or Line of a Sales Document which is a Recipe
                 [document type [I/C/O/W/Q], docno, lineno]
    ind    - None - Price Level (1), Marked Up Price (2) or Last Price (3)
             "P"  - Price Level Only (Can Return Zero)
             "L"  - Last Price Only (Can Return Zero)
    """
    from TartanClasses import CCD
    # Check for Recipe
    if not recp:
        chk = sql.getRec("strmf1", cols=["st1_type"], where=[("st1_cono",
            "=", cono), ("st1_group", "=", group), ("st1_code", "=", code)],
            limit=1)
        if chk[0] == "R":
            recp = True
    # Check Price Record
    prc = sql.getRec("strprc", cols=["stp_price"], where=[("stp_cono",
        "=", cono), ("stp_group", "=", group), ("stp_code", "=", code),
        ("stp_loc", "=", loc), ("stp_level", "=", lvl)], limit=1)
    if prc:
        # Price Level
        return prc[0]
    elif ind == "P":
        return 0
    # Last Selling Price
    lst = 0
    where = [
        ("stt_cono", "=", cono),
        ("stt_group", "=", group),
        ("stt_code", "=", code),
        ("stt_loc", "=", loc)]
    whr = where[:]
    whr.append(("stt_type", "in", (7, 8)))
    whr.append(("stt_sell", "<>", 0))
    chk = sql.getRec("strtrn", cols=["stt_qty", "stt_sell"], where=whr,
        order="stt_capdt desc")
    for rec in chk:
        qty = CCD(rec[0], "SD", 11.2)
        sel = CCD(rec[1], "SD", 11.2)
        if qty.work < 0 and sel.work:
            lst = round((sel.work / qty.work), 2)
            break
    # Check for Automatic Markups
    ctl = sql.getRec("strctl", cols=["cts_automu"], where=[("cts_cono",
        "=", cono)], limit=1)
    mui = ctl[0]
    if mui == "A":
        chk = sql.getRec("strmf1", cols=["st1_value_ind"], where=[("st1_cono",
            "=", cono), ("st1_group", "=", group), ("st1_code", "=", code)],
            limit=1)
        if chk[0] == "S":
            mui = "S"
    if not recp and (mui == "N" or ind == "L"):
        # Not a Recipe and ind=L or No Markups
        return lst
    # Markup Prices
    prc = 0
    if mui in ("A", "L", "S") and not prc:
        # Cost Price
        cst = getCost(sql, cono, group, code, loc, ind=mui, recp=recp)
        if cst:
            # Price Markup
            mkp = getMarkup(sql, cono, group, code, loc, lvl)
            if mkp:
                prc = round((cst * (100 + mkp) / 100.0), 2)
    if not prc:
        return lst
    return prc

def getMarkup(sql, cono, group, code, loc, level):
    mkp = 0
    gpm = sql.getRec("strgmu", cols=["smg_markup"], where=[("smg_cono",
        "=", cono), ("smg_group", "=", group), ("smg_level", "=", level)],
        limit=1)
    if gpm:
        mkp = gpm[0]
    st2 = sql.getRec("strcmu", cols=["smc_markup"], where=[("smc_cono",
        "=", cono), ("smc_group", "=", group), ("smc_code", "=", code),
        ("smc_loc", "=", loc), ("smc_level", "=", level)], limit=1)
    if st2 and st2[0]:
        mkp = st2[0]
    return mkp

def getFileName(path, ptyp="FF", wrkdir=None, check=False):
    import os, sys
    if sys.platform == "win32" or path[:2] not in ("\\\\", "\\", "//"):
        # Normal path
        if os.path.exists(path):
            return os.path.normpath(path)
        else:
            return
    # UNC path
    con = None
    fle = None
    svr = ""
    shr = ""
    pth = ""
    usr = ""
    pwd = ""
    try:
        import socket, stat
        from smb.SMBConnection import SMBConnection
        socket.setdefaulttimeout(5)
        obj = path.replace("/", "|").replace("\\", "|")
        obj = obj.split("|")
        obj[:] = (value for value in obj if value != "")
        svr = obj[0]
        shr = obj[1]
        for p in obj[2:]:
            if not pth:
                pth = p
            else:
                pth = os.path.join(pth, p)
        try:
            con = SMBConnection(usr, pwd, usr, svr, is_direct_tcp=False)
        except:
            con = SMBConnection(usr, pwd, svr, is_direct_tcp=True)
        con.connect(svr)
        att = con.getAttributes(shr, pth)
        if ptyp == "FD" and not att.isDirectory:
            raise Exception("Invalid Directory")
        elif ptyp == "FF" and not att.isNormal:
            raise Exception("Invalid File")
        if check or ptyp == "FD":
            con.close()
            return path
        # Download File
        if wrkdir:
            nam = os.path.join(wrkdir, os.path.basename(pth))
        else:
            nam = os.path.join(os.getcwd(), os.path.basename(pth))
        if not os.path.exists(nam):
            fle = open(nam, "wb")
            con.retrieveFile(shr, pth, fle)
            fle.close()
            os.chmod(nam, stat.S_IRWXU)
        con.close()
        return nam
    except Exception as err:
        if con:
            con.close()
        if fle:
            fle.close()

def b64Convert(typ, txt):
    import base64
    if typ == "encode":
        return base64.encodebytes(txt.encode("utf-8")).decode("utf-8")
    else:
        return base64.decodebytes(txt.encode("utf-8")).decode("utf-8")

def copyList(orig):
    import copy, inspect

    def myCopy(orig, newo=None):
        if newo is None:
            newo = []
        for x in orig:
            if type(x) in (list, tuple):
                newx = []
                myCopy(x, newx)
                if type(x) == tuple:
                    newo.append(tuple(newx))
                else:
                    newo.append(newx)
            else:
                newo.append(x)
        if type(orig) == tuple:
            return tuple(newo)
        else:
            return newo

    meth = False
    deep = False
    for obj in orig:
        if type(obj) in (list, tuple):
            deep = True
            for x in obj:
                if type(x) in (list, tuple):
                    for y in x:
                        if inspect.ismethod(y):
                            meth = True
                elif inspect.ismethod(x):
                    meth = True
        elif inspect.ismethod(obj):
            meth = True
    try:
        if meth:
            return myCopy(orig)
        elif deep:
            return copy.deepcopy(orig)
        else:
            return copy.copy(orig)
    except:
        return orig[:]

def luhnFunc(number):
    def digits_of(n):
        return [int(d) for d in str(n)]
    digits = digits_of(number)
    odd_digits = digits[-1::-2]
    even_digits = digits[-2::-2]
    checksum = 0
    checksum += sum(odd_digits)
    for d in even_digits:
        checksum += sum(digits_of(d*2))
    return checksum % 10 == 0

def internetConnect(host="1.1.1.1", port=53):
    """
    Host: 1.1.1.1 free Domain Name System (DNS) service
    OpenPort: 53/tcp
    Service: domain (DNS/TCP)
    """
    import socket

    try:
        socket.setdefaulttimeout(1)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect((host, port))
        return True
    except:
        return False

def roundup(number, places=0):
    from decimal import Decimal as dec
    from decimal import ROUND_HALF_UP

    if type(number) is float:
        number = str(number)
    place = "1."
    for i in range(places):
        place = "".join([place, "0"])
    return float(dec(number).quantize(dec(place), rounding=ROUND_HALF_UP))

def getColors(style, scheme):
    if scheme == "R":
        color = (
            ("#ffffff", "#a40204"),
            ("#ffffff", "#a40204"),
            ("#ffffff", "#044e1c"),
            ("#000000", "#ffffff"))
    elif scheme == "G":
        color = (
            ("#ffffff", "#00883b"),
            ("#ffffff", "#00883b"),
            ("#000000", "#19f913"),
            ("#000000", "#ffffff"))
    elif scheme == "B":
        color = (
            ("#ffffff", "#0058ff"),
            ("#ffffff", "#0058ff"),
            ("#000000", "#ade9e6"),
            ("#000000", "#ffffff"))
    else:
        # Defaults
        widgets = {
            "label": {
                "label": {
                    "foreground": "",
                    "background": ""}},
            "button": {
                "label": {
                    "foreground": "",
                    "background": ""},
                "map": {
                    "background": {
                        "disabled": "",
                        "pressed": "",
                        "active": ""}}}}
        for widget in widgets:
            wid = "T%s" % widget.capitalize()
            for typ in widgets[widget]:
                if typ == "map":
                    mmm = style.map(wid)
                    for m in mmm:
                        if m in widgets[widget][typ]:
                            for n in mmm[m]:
                                if n[0] in widgets[widget][typ][m]:
                                    widgets[widget][typ][m][n[0]] = n[1]
                    continue
                for clr in widgets[widget][typ]:
                    widgets[widget][typ][clr] = style.lookup(wid, clr)
        color = [
            [widgets["label"]["label"]["foreground"],
            widgets["label"]["label"]["background"]],
            [widgets["button"]["label"]["foreground"],
            widgets["button"]["label"]["background"]],
            [widgets["button"]["label"]["foreground"],
            widgets["button"]["map"]["background"]["active"]],
            [widgets["button"]["label"]["foreground"],
            widgets["button"]["map"]["background"]["disabled"]]]
    if color[3] == color[1]:
        color[3][0] = "#ffffff"
    if not color[2][1]:
        color[2] = color[1][:]
    if not color[3][1]:
        color[3] = ["#ffffff", color[1][1]]
    return color

def chkMod(mod):
    """
    This function is used to check if a module is available.
    """
    try:
        mod = __import__(mod)
        return mod
    except:
        return

def doPublish(scrn, flenam):
    """
    This function is used to display a text, rst, md or html file.
    """
    import io
    from TartanClasses import HTML, ScrollHtml, ScrollText
    try:
        from docutils.core import publish_string
        RST = True
    except:
        RST = False
    try:
        import markdown
        MD = True
    except:
        MD = False
    ext = flenam.split(".")[1].lower()
    try:
        input_file = io.open(flenam, mode="r", encoding="utf8")
        text = input_file.read()
        if ext == "txt":
            ScrollText(scrn=scrn, mess=text, wrap=True)
        elif HTML:
            if ext == "rst" and RST:
                html = publish_string(text, writer_name="html",
                    settings_overrides={"embed_stylesheet": False})
            elif ext == "md" and MD:
                html = markdown.markdown(text)
            elif ext == "html":
                html = text
            else:
                raise Exception
            ScrollHtml(scrn=scrn, mess=html, horizontal="auto")
    except:
        return

def doAutoAge(dbm, system, cono=None, chain=None, acno=None, pbar=None):
    from TartanClasses import ASD, Sql
    tab = "%strn" % system
    age = "%sage" % system
    sql = Sql(dbm, [tab, age],  __name__)
    col = getattr(sql, "%s_col" % tab)
    if system == "crs":
        pfx = "crt"
        sql.delRec(age, where=[("cra_cono", "=", cono),
            ("cra_acno", "=", acno)])
        whr = [
            ("crt_cono", "=", cono),
            ("crt_acno", "=", acno)]
    elif system == "drs":
        pfx = "drt"
        sql.delRec(age, where=[("dra_cono", "=", cono),
            ("dra_chain", "=", chain), ("dra_acno", "=", acno)])
        whr = [
            ("drt_cono", "=", cono),
            ("drt_chain", "=", chain),
            ("drt_acno", "=", acno)]
    else:
        pfx = "mlt"
        sql.delRec(age, where=[("mta_cono", "=", cono),
            ("mta_memno", "=", acno)])
        whr = [
            ("mlt_cono", "=", cono),
            ("mlt_memno", "=", acno)]
    crw = whr[:]
    crw.append(("%s_tramt" % pfx, "<", 0))
    drw = whr[:]
    drw.append(("%s_tramt" % pfx, ">", 0))
    tdt = "%s_trdt" % pfx
    cdt = "%s_curdt" % pfx
    amt = "%s_tramt" % pfx
    cr = sql.getRec(tab, where=crw, order=tdt)     # Credit transactions
    if cr:
        dr = sql.getRec(tab, where=drw, order=tdt) # Debit transactions
        if dr:
            for cno, ctr in enumerate(cr):         # For each credit transaction
                if pbar:
                    pbar.displayProgress()
                ccdt = ctr[col.index(cdt)]         # Current period
                cbal = ctr[col.index(amt)]         # Credit amount
                camt = cbal                        # Amount to allocate
                for dno, dtr in enumerate(dr):     # For each debit transaction
                    dcdt = dtr[col.index(cdt)]     # Current period
                    dbal = dtr[col.index(amt)]     # Debit amount
                    if not dbal:
                        continue
                    damt = float(ASD(dbal) + ASD(camt))
                    if damt < 0:
                        camt = damt
                        damt = 0.00
                    else:
                        camt = 0.00
                    diff = float(ASD(dbal) - ASD(damt))
                    dr[dno][col.index(amt)] = damt
                    if diff:
                        data = [cono]
                        if system == "drs":
                            data.append(chain)
                        data.append(acno)
                        data.append(dtr[col.index("%s_type" % pfx)])
                        if system == "mem":
                            data.append(dtr[col.index("%s_refno" % pfx)])
                        else:
                            data.append(dtr[col.index("%s_ref1" % pfx)])
                        if ccdt < dcdt:
                            data.append(dcdt)
                        else:
                            data.append(ccdt)
                        data.append(ctr[col.index("%s_type" % pfx)])
                        data.append(ctr[col.index("%s_ref1" % pfx)])
                        data.extend([diff, 0])
                        sql.insRec(age, data=data, dofmt=False)
                    if not camt:
                        break
                diff = float(ASD(cbal) - ASD(camt))
                if diff:
                    data = [cono]
                    if system == "drs":
                        data.append(chain)
                    data.append(acno)
                    data.append(ctr[col.index("%s_type" % pfx)])
                    if system == "mem":
                        data.append(ctr[col.index("%s_refno" % pfx)])
                    else:
                        data.append(ctr[col.index("%s_ref1" % pfx)])
                    data.append(ccdt)
                    data.append(ctr[col.index("%s_type" % pfx)])
                    data.append(ctr[col.index("%s_ref1" % pfx)])
                    data.extend([diff, 0])
                    sql.insRec(age, data=data, dofmt=False)

def getImage(name, siz=None, fle=None):
    import base64, io
    from TartanClasses import Image, ImageTk
    from tartanImages import aliases, images
    stk = name.lower().split()[0]
    if stk.count("/"):
        stk = stk.split("/")[0]
    if stk not in images and stk not in aliases:
        return
    if stk in aliases:
        stk = aliases[stk]
    dec = io.BytesIO(base64.b64decode(images[stk]))
    img = Image.open(dec)
    if siz is not None:
        img = img.resize(siz, 0)
    if fle:
        img.save(fle)
    else:
        return ImageTk.PhotoImage(img.convert(mode="RGBA"))

def printPDF(prt, fle, cpy=1):
    import sys
    if sys.platform == "win32":
        import io, fitz, win32con, win32gui, win32print, win32ui
        from PIL import Image, ImageWin
        hdl = win32print.OpenPrinter(prt)
        dev = win32print.GetPrinter(hdl, 2)["pDevMode"]
        dev.PaperSize = 9
        fd = fitz.open(fle)
        for pge in fd:
            # Fitz
            buf = io.BytesIO()
            try:
                rect = pge.mediabox
            except:
                rect = pge.MediaBox
            siz = [int(rect[2]), int(rect[3])]
            mat = fitz.Matrix(4.16667, 4.16667)
            clp = fitz.Rect(0, 0, siz[0], siz[1])
            dst = (0, 0, int(rect[2] * 20), int(rect[3] * -20))
            try:
                lst = pge.get_displaylist()
                pix = lst.get_pixmap(matrix=mat, clip=clp, alpha=False)
                buf.write(pix.tobytes(output="ppm"))
            except:
                lst = pge.getDisplayList()
                pix = lst.getPixmap(matrix=mat, clip=clp, alpha=False)
                buf.write(pix.getImageData(output="ppm"))
            # Win32
            img = Image.open(buf)
            dib = ImageWin.Dib(img)
            if siz[0] < siz[1]:
                dev.Orientation = 1
            else:
                dev.Orientation = 2
            hdc = win32gui.CreateDC("WINSPOOL", prt, dev)
            dcf = win32ui.CreateDCFromHandle(hdc)
            dcf.SetMapMode(win32con.MM_TWIPS)
            for _ in range(cpy):
                dcf.StartDoc(fle)
                dcf.StartPage()
                dib.draw(hdc, dst)
                dcf.EndPage()
                dcf.EndDoc()
            del dcf
        win32print.ClosePrinter(hdl)
    else:
        try:
            import cups
            conn = cups.Connection()
            conn.printFile(prt, fle, fle, options={"copies": str(cpy),
                "media": "A4"})
        except:
            import subprocess
            subprocess.Popen(["/usr/bin/lp", "-d%s" % prt, "-n%s" % cpy,
                "-o", "media=A4", fle], stdout=subprocess.PIPE)

def doWriteExport(**args):
    """
    Write an Export file using the following arguments:
        xtype  - C for csv and X for xlsx or a tuple of both
        name   - The output file name without an extension
        heads  - A list of the Report headings. A heading can be a tuple as:
                    (Text, Column width, Font Size)
        colsh  - A list of lists of Column headings comprising:
                    Text
                    Start Col (Default is current col)
                    End Col   (Default is current col)
                or
                    Text
                    Alignment
        forms  - A list of column format details comprising:
                    Format per CCD
                    Size   per CCD
                    Clickable (True or False) - Default is False (ScrollGrid)
                    Negative colour (ScrollGrid)
                    Ignore for ULINE[D/S] and TOTAL - Default is False
        datas  - The column data comprising:
                    Type of data:
                        "PAGE"   - A new page/sheet
                            ["PAGE", (heading, colsh, forms, name)]
                        "HEAD"   - A heading in column 0 only
                        "BODY"   - A body line of columns of data
                        "BLANK"  - A blank line - skip
                        "ULINED" - A double underline - skip
                        "ULINES" - A sigle underline - skip
                        "TOTAL"  - A total line of columns of data
                    List of values
        ctots  - A list of TOTAL columns to sum using the SUBTOTAL function
        rcdic  - The tartanrc dictionary
        view   - Whether to view the report
        wait   - Whether to wait for the viewer to exit.
    """
    import os, sys
    from TartanClasses import XLSX

    def viewFile(exe, cmd, name, wait):
        if not cmd:
            try:
                if sys.platform == "win32":
                    os.startfile(name)
                else:
                    import subprocess
                    if wait:
                        subprocess.call(["xdg-open", name])
                    else:
                        subprocess.Popen(["xdg-open", name])
            except Exception as err:
                showError(None, "Execution Error",
                    "No Valid Export Application.\n\n%s" % err)
            return
        try:
            import subprocess
            if wait:
                subprocess.call(cmd + [name])
            else:
                subprocess.Popen(cmd + [name])
        except Exception as err:
            showError(None, err, "The Application"\
                "\n\n%s\n\nIs Not Found or Not Accessible" % exe)
            return

    if "view" not in args:
        view = True
    else:
        view = args["view"]
    if view:
        if args["rcdic"] and args["rcdic"]["exp"]:
            exe, cmd = parsePrg(args["rcdic"]["exp"])
            if not os.path.isfile(exe):
                exe = cmd = None
        else:
            exe = cmd = None
    if "wait" not in args:
        wait = False
    else:
        wait = args["wait"]
    if args["xtype"].upper() == "C":
        head = ""
        name = args["name"] + ".csv"
        try:
            flenam = open(name, "w")
        except Exception as err:
            showError(None, "Error", err)
            return
        for valc in args["colsh"][-1]:
            if type(valc) in (list, tuple):
                text = valc[0]
            else:
                text = valc
            if head:
                head = '%s,"%s"' % (head, text)
            else:
                head = '"%s"' % text
        if head:
            flenam.write(head + "\n")
        for valx in args["datas"]:
            if valx[0] != "BODY":
                continue
            if not valx[1] or len(valx[1]) != len(args["forms"]):
                # Blank line or invalid line length
                continue
            line = ""
            for colc, valc in enumerate(valx[1]):
                if type(valc) in (list, tuple):
                    valc = valc[0]
                if args["forms"][colc][0][0].lower() == "d" or \
                   args["forms"][colc][0][1].lower() in ("d", "i", "l"):
                    if line:
                        line = "%s,%s" % (line, valc)
                    else:
                        line = "%s" % valc
                elif line:
                    line = '%s,"%s"' % (line, valc)
                else:
                    line = '"%s"' % valc
            flenam.write(line + "\n")
        # Save the csv file
        flenam.close()
        if view:
            # View the csv file
            viewFile(exe, cmd, name, wait)
        return

    def getLetter(col):
        string = ""
        while col > 0:
            col, remainder = divmod(col - 1, 26)
            string = chr(65 + remainder) + string
        return string

    def cellWrite(sheet, row, col, val, fmt):
        if type(row) in (list, tuple):
            ccl = getLetter(col[0])
            ccc = "%s%s" % (ccl, row[0])
            sheet.merge_cells(start_row=row[0], start_column=col[0],
                end_row=row[1], end_column=col[1])
            sheet.cell(column=col[0], row=row[0], value=val)
        else:
            ccl = getLetter(col)
            ccc = "%s%s" % (ccl, row)
            sheet.cell(column=col, row=row, value=val)
        for key in fmt:
            if key == "ccd":
                if fmt[key][0] in xf_map:
                    sheet[ccc].number_format = xf_map[fmt[key][0]]
            elif key == "font":
                sheet[ccc].font = Font(name=fmt[key])
            elif key == "bold":
                sheet[ccc].font = Font(bold=fmt[key])
            elif key == "align":
                if type(fmt[key]) is str:
                    h, v = fmt[key], "center"
                else:
                    h, v = fmt[key]
                sheet[ccc].alignment = Alignment(horizontal=h, vertical=v)
            elif key == "border":
                stl = Side(border_style=fmt[key][1], color="000000")
                if fmt[key][0] == "T":
                    sheet[ccc].border = Border(top=stl)
                elif fmt[key][0] == "L":
                    sheet[ccc].border = Border(left=stl)
                elif fmt[key][0] == "R":
                    sheet[ccc].border = Border(right=stl)
                elif fmt[key][0] == "B":
                    sheet[ccc].border = Border(bottom=stl)
                elif fmt[key][0] == "LRB":
                    sheet[ccc].border = Border(left=stl, right=stl, bottom=stl)
                elif fmt[key][0] == "TLRB":
                    sheet[ccc].border = Border(top=stl, left=stl, right=stl,
                        bottom=stl)

    def createSheet(fmt, page, text=None):
        if text is None:
            text = "Page %s" % page
        sheet = book.create_sheet(title=text)
        # Main headings
        rowx = 0
        blank = False
        for num, valx in enumerate(args["heads"]):
            if type(valx) in (list, tuple):
                hgt, siz = valx[1:]
                hxf = {"font": ["Ariel", siz, True]}
                hxf["align"] = ["center", "center"]
                cols = len(args["colsh"][-1])
            else:
                hxf = {"font": "Ariel"}
                hxf["bold"] = True
                hxf["align"] = ["left", "center"]
                cols = 0
            if valx and blank:
                # Add a blank line
                if cols:
                    cellWrite(sheet, [rowx+1, rowx+1], [1, cols], "", hxf)
                else:
                    cellWrite(sheet, rowx+1, 1, "", hxf)
                blank = False
                rowx += 1
            if cols:
                cellWrite(sheet, [rowx+1, rowx+1], [1, cols], valx[0], hxf)
                sheet.row_dimensions[num+1].height = (hgt / 20)
            else:
                cellWrite(sheet, rowx+1, 1, valx, hxf)
            rowx += 1
            blank = bool(valx)
        if blank:
            rowx += 1
        # Column headings
        for colx, valx in enumerate(args["colsh"]):
            nc = 0
            for cx, vx in enumerate(valx):
                fff = args["forms"][cx]
                hxf = dict(fmt)
                hxf["bold"] = True
                if type(vx) in (list, tuple) and len(vx) > 1:
                    if len(vx) == 2:
                        hxf["align"] = vx[1]
                    else:
                        hxf["align"] = "center"
                elif fff[0] in alpha:
                    hxf["align"] = "left"
                elif fff[0][0] in ("D", "d"):
                    hxf["align"] = "center"
                else:
                    hxf["align"] = "right"
                if type(vx) in (list, tuple):
                    if len(vx) == 3:
                        cellWrite(sheet, [rowx+1, rowx+1], [vx[1]+1, vx[2]+1],
                            vx[0], hxf)
                        nc = vx[2]
                    else:
                        cellWrite(sheet, rowx+1, nc+1, vx[0], hxf)
                else:
                    cellWrite(sheet, rowx+1, nc+1, vx, hxf)
                nc += 1
                if colx == len(args["colsh"]) - 1:
                    ccl = getLetter(nc)
                    sheet.column_dimensions[ccl].width = (int(fff[1]) + 2)
            if colx != len(args["colsh"]) - 1:
                rowx += 1
        # Freeze the headings
        sheet.freeze_panes = "A%s" % (rowx + 2)
        return sheet, rowx

    if args["xtype"] == "X" and XLSX:
        import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Border, Side, Font, Alignment
        from TartanClasses import tkfont

        # Variables and Formats
        fnt = tkfont.nametofont("TkTextFont")
        fmt = {"font": fnt.configure()["family"]}
        alpha = ("HA", "LA", "La", "NA", "Na", "TX", "UA", "Ua")
        numer = ("CD", "CI", "SD", "SI", "UI", "UD")
        xf_map = {
            "D1": "yyyy-mm-dd",
            "d1": "yyyy-mm-dd",
            "D2": "yyyy-mm",
            "d2": "yyyy-mm",
            "UI": "########0",
            "UL": "#0",
            "CI": "#,##0",
            "SI": "#0",
            "SL": "#0",
            "US": "#0",
            "CD": "#,##0.00",
            "SD": "#0.00",
            "UD": "#0.00"}
        for an in alpha:
            xf_map[an] = "@"
        # Create the workbook and worksheet
        book = Workbook()
        del book["Sheet"]
        page = 1
        if args["datas"][0][0] != "PAGE":
            sheet, rowx = createSheet(fmt, page)
            page += 1
        # Generate the body
        if "ctots" in args and args["ctots"]:
            gtot = {}
            stot = {}
            for tot in args["ctots"]:
                if type(tot) in (list, tuple):
                    stot[tot[0]]
                stot[tot] = []
        for num, row in enumerate(args["datas"]):
            if row[0] == "PAGE":
                args["heads"] = row[1][0]
                args["colsh"] = row[1][1]
                if len(row[1]) > 2:
                    args["forms"] = row[1][2]
                if len(row[1]) == 4:
                    pg = row[1][3]
                else:
                    pg = page
                if len(row[1]) == 4:
                    sheet, rowx = createSheet(fmt, pg, row[1][3])
                else:
                    sheet, rowx = createSheet(fmt, pg)
                page += 1
                continue
            if row[0] in ("ULINES", "ULINED"):
                continue
            rowx += 1
            if row[0] == "BLANK":
                cellWrite(sheet, rowx+1, 1, "", fmt)
                continue
            # Check if next line is an underline
            unl = False
            chk = num + 1
            if chk != len(args["datas"]):
                if args["datas"][chk][0] == "ULINES":
                    unl = "s"
                elif args["datas"][chk][0] == "ULINED":
                    unl = "d"
            # Write columns
            for colx, valx in enumerate(row[1]):
                # Dates
                if row[0] == "BODY":
                    if valx and args["forms"][colx][0].lower() == "d1":
                        # D1 Date conversion
                        valx = datetime.datetime.strptime(str(valx), "%Y%m%d")
                    elif valx and args["forms"][colx][0].lower() == "d2":
                        # D2 Date conversion
                        valx = datetime.datetime.strptime(str(valx), "%Y%m")
                # Sum Totals
                if "ctots" in args and args["ctots"]:
                    if colx in stot:
                        cola = getLetter(colx + 1)
                        if row[0] == "BODY":
                            if not stot[colx]:
                                stot[colx] = ["%s%s" % (cola, rowx + 1),
                                    "%s%s" % (cola, rowx + 1)]
                                if not colx in gtot:
                                    gtot[colx] = ["%s%s" % (cola, rowx + 1), ""]
                            else:
                                stot[colx][1] = "%s%s" % (cola, rowx + 1)
                        elif row[0] == "TOTAL":
                            if stot[colx]:
                                valx = "=SUBTOTAL(9,%s:%s)" % tuple(stot[colx])
                                gtot[colx][1] = "%s%s" % (cola, rowx + 1)
                                stot[colx] = []
                            elif gtot[colx]:
                                valx = "=SUBTOTAL(9,%s:%s)" % tuple(gtot[colx])
                # Format Font, Bold and/or Border
                if type(valx) in (list, tuple):
                    valx, bord = valx
                else:
                    bord = None
                hxf = dict(fmt)
                if row[0] in ("HEAD", "TOTAL"):
                    hxf["bold"] = True
                # Underline & Border
                ccd = list(args["forms"][colx][:])
                if ccd[0] in numer and type(valx) == str and \
                        not valx.count("=SUBTOTAL"):
                    ccd[0] = "NA"
                if unl == "s" and ccd[0] in numer:
                    hxf["border"] = ["B", "thin"]
                elif unl == "d" and ccd[0] in numer:
                    hxf["border"] = ["B", "double"]
                elif bord:
                    hxf["border"] = [bord, "thin"]
                # CCD Format
                if ccd[0] not in numer and not valx:
                    ccd[0] = "NA"
                    valx = ""
                hxf["ccd"] = ccd
                # Alignment
                if ccd[0] in alpha:
                    hxf["align"] = "left"
                elif ccd[0][0] in ("D", "d"):
                    hxf["align"] = "center"
                else:
                    hxf["align"] = "right"
                cellWrite(sheet, rowx+1, colx+1, valx, hxf)
        # Save the spreadsheet
        name = args["name"] + ".xlsx"
        try:
            book.save(filename=name)
            # View the spreadsheet
            if view:
                viewFile(exe, cmd, name, wait)
        except Exception as err:
            showError(None, "Error", err)
# END
# vim:set ts=4 sw=4 sts=4 expandtab:
