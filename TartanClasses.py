"""
SYNOPSIS
    These are various classes used in TARTAN.

    This file is part of Tartan Systems (TARTAN).

AUTHOR
    Written by Paul Malherbe, <paul@tartan.co.za>

COPYING
    Copyright (C) 2004-2022 Paul Malherbe.

    This program is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with this program. If not, see <https://www.gnu.org/licenses/>.
"""
# ========================================================
# Standard Python modules
# ========================================================
import copy, csv, datetime, functools, glob, gzip, math, os, re, shutil
import subprocess, sys, tarfile, tempfile, textwrap, threading, time
import webbrowser
# ========================================================
# TARTAN Standard Functions and Variables e.g. showError
# ========================================================
from tartanFunctions import askQuestion, b64Convert, chkAggregate, chkMod
from tartanFunctions import copyList, dateDiff, doPrinter, doWriteExport
from tartanFunctions import getColors, getCost, getFileName, getFontSize
from tartanFunctions import getImage, getModName, getPeriods, getPrgPath
from tartanFunctions import getPrinters, getSell, getSingleRecords, getTrn
from tartanFunctions import getUnderline, importTkinter, loadRcFile
from tartanFunctions import luhnFunc, makeArray, mthendDate, parsePrg
from tartanFunctions import printPDF, projectDate, removeFunctions, sendMail
from tartanFunctions import showError, showException, showWarning, doPublish
from tartanFunctions import unbindAllWidgets, getManager, placeWindow
from tartanFunctions import cutpasteMenu
import tartanWork

try:
    from tarchg import changes
except:
    changes = None
# ========================================================
# TkinterHtml
# ========================================================
try:
    from tkinterhtml import HtmlFrame
    HTML = True
except:
    HTML = False
# ========================================================
# PyFPDF
# ========================================================
try:
    import fpdf
except:
    print("Missing python fpdf module")
    os._exit(1)
if not fpdf.fpdf.Image:
    print("Missing python-imaging-module")
# ========================================================
# MuPDF
# ========================================================
try:
    import fitz
    FITZ = True
except:
    FITZ = False
# ========================================================
# Excel import and export modules
# ========================================================
try:
    from pyexcel_xls import get_data as getxls
    XLS = True
except:
    XLS = False
try:
    import openpyxl
    XLSX = True
except:
    XLSX = False
# ========================================================
# Openoffice import module
# ========================================================
try:
    from pyexcel_ods import get_data as getods
    ODS = True
except:
    ODS = False
# ========================================================
# Curses Progress Bar
# ========================================================
try:
    from progress.bar import Bar
    TBAR = True
except:
    TBAR = False
# ========================================================
# PyGal imports
# ========================================================
try:
    import pygal
    PYGAL = True
except:
    PYGAL = False
try:
    from svglib.svglib import svg2rlg
    from reportlab.graphics import renderPDF
    CVTSVG = True
except:
    CVTSVG = False
# ========================================================
# OfxTools import
# ========================================================
try:
    from ofxtools.Parser import OFXTree
    OFX = True
except:
    OFX = False
# ========================================================
# Tkinter, ttk and PIL imports
# ========================================================
try:
    tk, ttk = importTkinter()
    if not tk:
        raise Exception
    import tkinter.filedialog as tkfile
    import tkinter.font as tkfont
    try:
        from tkcolorpicker import askcolor
        CPICK = True
    except:
        import tkinter.colorchooser as tkcolor
        CPICK = False
    from PIL import Image, ImageTk

    # Stock images and icons and others
    tkinter_umlauts = [
        "adiaeresis", "Adiaeresis",
        "odiaeresis", "Odiaeresis",
        "udiaeresis", "Udiaeresis",
        "ssharp"]

    class ArrowButton(ttk.Button):
        arrow_layout = lambda self, direc: ([("Button.focus",
            {"children": [("Button.%sarrow" % direc, None)]})])

        def __init__(self, master, **kw):
            direction = kw.pop("direction", "left")
            style = ttk.Style(master)
            style.configure("L.TButton", relief="raised")
            style.layout("L.TButton", self.arrow_layout("left"))
            style.configure("R.TButton", relief="raised")
            style.layout("R.TButton", self.arrow_layout("right"))
            kw["style"] = "L.TButton" if direction == "left" else "R.TButton"
            super().__init__(master, **kw)

    class MyButtonBox(ttk.Frame):
        def __init__(self, parent, row=None, padx=0, pady=0, **kwargs):
            super().__init__(parent, **kwargs)
            if getManager(parent) in ("grid", "place"):
                if row is None:
                    row = parent.grid_size()[1]
                self.grid(row=row, padx=padx, pady=pady, sticky="ew")
            else:
                self.pack(fill="x", side="bottom", padx=padx, pady=pady)
            self.rows = {}

        def addButton(self, text, cmd, row=0, spn=1):
            if row not in self.rows:
                self.rows[row] = 0
            text, pos = getUnderline(widget=self, text=text)
            but = MyButton(self, cmd=cmd, text=text, underline=pos)
            but.grid(column=self.rows[row], row=row, columnspan=spn,
                sticky="ew")
            self.columnconfigure(self.rows[row], weight=1)
            self.rows[row] += spn
            return but

        def focusNext(self, event):
            key = event.keysym
            idx = self.buttons.index(event.widget["text"])
            if key == "Left":
                if not idx:
                    idx = len(self.buttons) - 1
                else:
                    idx -= 1
            elif idx == (len(self.buttons) - 1):
                idx = 0
            else:
                idx += 1
            widgets = self.winfo_children()
            widgets[idx].focus_set()

    class MyButton(ttk.Button):
        def __init__(self, parent, cmd=None, txt=True, img=True, fnt=None, relief=None, **kwargs):
            if "style" not in kwargs:
                kwargs["style"] = "MyButton.TButton"
            kwargs["width"] = -1
            if "image" in kwargs:
                img = False
            if "underline" not in kwargs:
                kwargs["underline"] = -1
            kwargs["takefocus"] = False
            super().__init__(parent, **kwargs)
            self.parent = parent
            style = ttk.Style()
            if fnt:
                if type(fnt) == str:
                    name = fnt
                elif len(fnt) == 2:
                    name = "%s%s" % tuple(fnt)
                else:
                    name = "%s%s%s" % tuple(fnt)
                style.configure("%s.%s" % (name, kwargs["style"]), font=fnt)
                self.configure(style="%s.%s" % (name, kwargs["style"]))
            if relief:
                style.configure(kwargs["style"], relief=relief)
            self.setLabel(kwargs["text"], kwargs["underline"], txt=txt,
                img=img, cmd=cmd)

        def setLabel(self, text, udl=-1, txt=True, img=True, cmd=None):
            self.configure(text=text, underline=udl)
            if img:
                try:
                    hgt = int(self.winfo_reqheight() * .6)
                    if hgt < 1:
                        hgt = 20
                    self.image = getImage(text, (hgt, hgt))
                    if self.image is None:
                        self.image = getImage("yes", (hgt, hgt))
                    if txt:
                        self.configure(compound="left", image=self.image)
                    else:
                        self.configure(image=self.image)
                except:
                    pass
            if not cmd:
                return
            self.cmd = cmd
            self.bind("<Return>", self.execButCmd)
            self.bind("<KP_Enter>", self.execButCmd)
            self.bind("<ButtonRelease-1>", self.execButCmd)
            if udl != -1:
                self.bind("<Alt_L><%s>" % text[udl].upper(), self.execButCmd)
                self.bind("<Alt_L><%s>" % text[udl].lower(), self.execButCmd)
                topl = self.parent.winfo_toplevel()
                topl.bind("<Alt_L><%s>" % text[udl].upper(), self.execButCmd)
                topl.bind("<Alt_L><%s>" % text[udl].lower(), self.execButCmd)

        def execButCmd(self, event=None):
            try:
                if str(self.cget("state")) == "normal":
                    if type(self.cmd) in (list, tuple):
                        self.cmd[0](self.cmd[1])
                    else:
                        self.cmd()
                self.event_generate("<Leave>")
                self.update_idletasks()
            except:
                pass
            return "break"

    class MyCheckButton(ttk.Checkbutton):
        def __init__(self, parent, **kwargs):
            if "style" not in kwargs:
                kwargs["style"] = "MyCheckbutton.TCheckbutton"
            super().__init__(parent, **kwargs)
            self.configure(onvalue="Y", offvalue="N")
            self.event_add("<<mychkbx>>", "<F5>", "<Escape>",
                "<Return>", "<KP_Enter>")

    class MyEntry(ttk.Entry):
        def __init__(self, parent, maxsize=None, cmd=None, aut=False, **kwargs):
            self.parent = parent
            super().__init__(self.parent, **kwargs)
            self.bind("<Button-3><ButtonRelease-3>", self.show_menu)
            self.event_add("<<myentry>>", "<F1>", "<F5>", "<Escape>")
            self.configure(takefocus=False)
            if maxsize:
                vcmd = (self.register(self.maxsize), maxsize, "%s", "%P")
                self.configure(validate="key", validatecommand=vcmd)
            if cmd:
                self.cmd = cmd
                self.bind("<Return>", self.command)
                self.bind("<KP_Enter>", self.command)
            self.lbwid = False
            if aut == "I":
                self.clist = []
                self.hits = []
                self.hitindex = 0
                self.position = 0
                self.bind("<KeyRelease>", self.inline_way)
            elif aut == "L":
                self.clist = []
                self.bind("<Down>", self.down)
                self.bind("<KeyRelease>", self.list_way)

        def select_all(self, *args):
            self.focus_force()
            self.selection_range(0, "end")

        def show_menu(self, event):
            if str(self.cget("state")) == "disabled":
                return
            cutpasteMenu(event)

        def maxsize(self, maxsize, before, after):
            return bool(len(after) <= int(maxsize))

        def inline_way(self, event):
            if not self.get():
                return
            if event.keysym == "BackSpace":
                self.delete(self.index("insert"), "end")
                self.position = self.index("end")
            if event.keysym == "Left":
                if self.position < self.index("end"):
                    self.delete(self.position, "end")
                else:
                    self.position = self.position - 1
                    self.delete(self.position, "end")
            if event.keysym == "Right":
                self.position = self.index("end")
            if event.keysym == "Down":
                self.inline_aut(1)
            if event.keysym == "Up":
                self.inline_aut(-1)
            if len(event.keysym) == 1 or event.keysym in tkinter_umlauts:
                self.inline_aut()

        def inline_aut(self, delta=0):
            if delta:
                self.delete(self.position, "end")
            else:
                self.position = len(self.get())
            hits = []
            for element in self.clist:
                if element.lower().startswith(self.get().lower()):
                    hits.append(element)
            if hits != self.hits:
                self.hitindex = 0
                self.hits = hits
            if hits == self.hits and self.hits:
                self.hitindex = (self.hitindex + delta) % len(self.hits)
            if self.hits:
                self.delete(0, "end")
                self.insert(0, self.hits[self.hitindex])
                self.select_range(self.position, "end")

        def list_way(self, event):
            self.escape()
            if str(self.cget("state")) == "disabled":
                return
            if not self.get() or event.keysym in ("F1", "Return"):
                return
            words = self.comparison()
            if words:
                if len(words) == 1 and words[0] == self.get():
                    return
                self.lf = tk.Toplevel(self, bd=2, relief="raised")
                self.lf.overrideredirect(True)
                wdth = int(self.cget("width"))
                if len(words) > 10:
                    yScroll = ttk.Scrollbar(self.lf, orient="vertical")
                    yScroll.grid(row=0, column=1, sticky="ns")
                    self.lb = tk.Listbox(self.lf, width=wdth, height=10,
                        yscrollcommand=yScroll.set)
                    yScroll["command"] = self.lb.yview
                else:
                    self.lb = tk.Listbox(self.lf, width=wdth, height=len(words))
                self.lb.bind("<Double-Button-1>", self.list_aut)
                self.lb.bind("<KP_Enter>", self.list_aut)
                self.lb.bind("<Return>", self.list_aut)
                self.lb.bind("<Escape>", self.escape)
                self.lb.grid(row=0, column=0, sticky="nsew")
                self.lf.geometry("+%d+%d" % (self.winfo_rootx(),
                    self.winfo_rooty() + self.winfo_height()))
                self.lb.delete(0, "end")
                for w in words:
                    self.lb.insert("end", w)
                self.lbwid = True

        def comparison(self):
            try:
                pattern = re.compile(self.get() + ".*", re.IGNORECASE)
                return [w for w in self.clist if re.match(pattern, w)]
            except:
                pass

        def list_aut(self, event):
            self.delete(0, "end")
            self.insert(0, self.lb.get("active"))
            self.escape()
            self.command()

        def down(self, event):
            if self.lbwid:
                self.lb.focus_set()
                self.lb.selection_set(first="0")

        def escape(self, event=None):
            if self.lbwid:
                self.lb.destroy()
                self.lf.destroy()
                self.lbwid = False
                self.parent.focus_force()
                self.update_idletasks()
                self.icursor("end")
                self.focus_set()

        def command(self, event=None):
            self.escape()
            if type(self.cmd) in (list, tuple):
                self.cmd[0](self.cmd[1])
            elif self.cmd:
                self.cmd()

    class MyFrame(ttk.Frame):
        def __init__(self, parent=None, bg=None, **kwargs):
            if "style" in kwargs:
                name = kwargs["style"]
            else:
                name = "TFrame"
                kwargs["style"] = name
            super().__init__(parent, **kwargs)
            if bg:
                style = ttk.Style()
                style.configure("%s.%s" % (bg, name), background=bg)
                self.configure(style="%s.%s" % (bg, name))

    class MyLabel(ttk.Label):
        def __init__(self, parent=None, color=True, size=None, **kwargs):
            if "style" in kwargs:
                name = kwargs["style"]
            else:
                name = "TLabel"
                kwargs["style"] = name
            super().__init__(parent, **kwargs)
            style = ttk.Style()
            if color and type(color) in (list, tuple):
                name = "%s.%s" % (color[0].capitalize(), name)
                style.configure(name,
                    foreground=color[0],
                    background=color[1])
            elif color:
                name = "CLabel.TLabel"
            self.configure(style=name)
            if size:
                self.configure(width=size)

    class MyMenuButton(tk.Menubutton):
        def __init__(self, parent, font="TkMenuFont", fg=None, bg=None, af=None, ab=None, relief="raised", fill="both", exp="yes", side="left", **kwargs):
            super().__init__(parent, **kwargs)
            if not fg:
                fg = "#000000"
            if not bg:
                bg = "#ffffff"
            if ab is None:
                af = fg
                try:
                    style = ttk.Style()
                    for s in style.map("TButton")["background"]:
                        if s[0] == "active":
                            ab = s[1]
                except:
                    ab = bg
            self.configure(relief=relief, font=font, bg=bg, fg=fg,
                activebackground=ab, activeforeground=af)
            self.pack(fill=fill, expand=exp, side=side)
            if "underline" in kwargs:
                self.bind("<Alt-%s>" % kwargs["text"]
                    [kwargs["underline"]].lower(), self.doPost)
                self.bind("<Alt-%s>" % kwargs["text"]
                    [kwargs["underline"]].upper(), self.doPost)
                topl = parent.winfo_toplevel()
                topl.bind("<Alt-%s>" % kwargs["text"]
                    [kwargs["underline"]].lower(), self.doPost)
                topl.bind("<Alt-%s>" % kwargs["text"]
                    [kwargs["underline"]].upper(), self.doPost)

        def doPost(self, event):
            self.event_generate("<Enter>")
            self.event_generate("<Button-1>")
            self.event_generate("<ButtonRelease-1>")
            self.event_generate("<Leave>")
            self.update_idletasks()

    class MyMessageBox(object):
        def __init__(self, parent, dtype, title, mess, butt=None, dflt=None):
            self.parent = parent
            if self.parent:
                if self.parent.winfo_toplevel().state() == "withdrawn":
                    self.parent.winfo_toplevel().deiconify()
                    self.parent.update_idletasks()
                wrap = int(self.parent.winfo_width() / 2)
                mkw = MkWindow(trans=self.parent, decor=False, remov=False,
                    resiz=False)
                self.msgwin = mkw.newwin
                grabs = self.msgwin.master.grab_current()
                # Save and clear toplevel bindings
                self.topbinds = unbindAllWidgets(self.msgwin.master)
                # Save and disable all buttons
                cnt = 0
                self.butsve = []
                frm = [self.msgwin.master]
                while True:
                    try:
                        for ch in frm[cnt].winfo_children():
                            if ch.winfo_class() == "Toplevel":
                                frm.append(ch)
                            elif ch.winfo_class() == "TFrame":
                                frm.append(ch)
                            elif ch.winfo_class() == "TButton":
                                self.butsve.append((ch, str(ch.cget("state"))))
                                ch.configure(state="disabled")
                        cnt += 1
                    except:
                        break
            else:
                wrap = 800
                mkw = MkWindow(decor=True, resiz=False)
                self.msgwin = mkw.newwin
            self.frame = MyFrame(self.msgwin, bg="black", borderwidth=5,
                relief="ridge", style="MFrame.TFrame")
            tit = MyLabel(self.frame, text=title, anchor="c", relief="raised")
            tit.pack(fill="x", expand="yes", ipadx=2, ipady=5)
            self.binds = []
            label = MyLabel(self.frame, text=mess, padding=5, anchor="w",
                justify="left", color=False, foreground="black",
                background="white", wraplength=wrap)
            try:
                self.tkimg = getImage(dtype)
                if self.tkimg:
                    label.configure(image=self.tkimg, compound="left")
            except:
                pass
            label.pack(fill="both", expand="yes")
            bbox = MyButtonBox(self.frame, row=1)
            self.butts = []
            if dtype in ("error", "info", "warning"):
                butt = (("Close", self.doChoice),)
                dflt = "Close"
            elif dtype == "question":
                butt = (
                    ("Yes", (self.doChoice, "yes")),
                    ("No", (self.doChoice, "no")))
            for num, but in enumerate(butt):
                if type(but[1]) in (int, str):
                    but = list(but)
                    but[1] = (self.doChoice, but[1])
                b = bbox.addButton(but[0], but[1])
                for bind in b.bind():
                    if "<Key-Alt_L>" in bind:
                        self.binds.append(bind)
                ecmd = lambda event, num=num: self.navigate(event, num)
                b.bind("<Left>", ecmd)
                b.bind("<Right>", ecmd)
                if len(but) == 3:
                    ToolTip(b, but[2])
                self.butts.append(b)
                if dflt:
                    b.bind("<Enter>", self.doEnter)
                    b.bind("<Leave>", self.doLeave)
                    if but[0].lower() == dflt.lower():
                        dflt = b
            self.frame.update_idletasks()
            # Message window width and height
            w = self.frame.winfo_reqwidth()
            h = self.frame.winfo_reqheight()
            if self.parent:
                # Centre message window on parent
                if self.parent.state() == "withdrawn":
                    self.parent.deiconify()
                    self.parent.update_idletasks()
                placeWindow(self.msgwin, self.parent, size=(w, h))
            else:
                placeWindow(self.msgwin, place="M", size=(w, h), expose=True)
            # Place message frame in centre of message window
            self.frame.place(anchor="center", relx=0.5, rely=0.5)
            self.frame.update_idletasks()
            if self.frame.winfo_toplevel().state() == "withdrawn":
                self.frame.winfo_toplevel().deiconify()
            # Make message window modal
            self.msgwin.grab_set()
            # Set focus to default button or message window
            if dflt:
                dflt.focus_force()
            else:
                self.msgwin.focus_set()
            if self.parent:
                self.msgwin.wait_window()
                self.parent.update_idletasks()
                if grabs is not None:
                    grabs.grab_set()
            else:
                self.msgwin.mainloop()

        def doEnter(self, event):
            event.widget.focus_set()

        def doLeave(self, event):
            event.widget.focus_set()

        def doChoice(self, answer=None):
            self.answer = answer
            # Clear frame bindings
            for bind in self.binds:
                self.frame.winfo_toplevel().unbind(bind)
            if self.parent:
                # Reinstate toplevel bindings
                for bind in self.topbinds:
                    self.msgwin.master.bind(bind[0], bind[1])
                # Reinstate buttons
                for butt in self.butsve:
                    butt[0].configure(state=butt[1])
            self.msgwin.grab_release()
            self.msgwin.destroy()
            if not self.parent:
                self.msgwin.quit()

        def navigate(self, event, num):
            if event.keysym == "Left":
                if num == 0:
                    self.butts[-1].focus_set()
                else:
                    self.butts[num - 1].focus_set()
            else:
                if num == len(self.butts) - 1:
                    self.butts[0].focus_set()
                else:
                    self.butts[num + 1].focus_set()

    class MyRadioButton(ttk.Radiobutton):
        def __init__(self, parent, cmd=None, **kwargs):
            self.parent = parent
            if "style" not in kwargs:
                kwargs["style"] = "MyRadiobutton.TRadiobutton"
            super().__init__(parent, **kwargs)
            self.event_add("<<myradio>>", "<F5>", "<Escape>")
            self.bind("<Return>", lambda event: event.widget.invoke())
            self.bind("<KP_Enter>", lambda event: event.widget.invoke())
            self.bind("<Left>", self.goLeft)
            self.bind("<Right>", self.goRight)
            if cmd:
                vvv = kwargs["variable"]
                ecmd = lambda var=vvv, opt=cmd[1]: cmd[0](var, opt)
                self.configure(command=ecmd)

        def goLeft(self, event):
            childs = self.parent.winfo_children()
            for x, c in enumerate(childs):
                if c.winfo_class() == "TRadiobutton":
                    break
            if self.focus_get() == childs[x]:
                return
            if self.tk_focusPrev().winfo_class() == "TRadiobutton":
                self.tk_focusPrev().focus_set()
                return "break"

        def goRight(self, event):
            childs = self.parent.winfo_children()
            for x in range(len(childs) - 1, -1, -1):
                if childs[x].winfo_class() == "TRadiobutton":
                    break
            if self.focus_get() == childs[x]:
                return
            if self.tk_focusNext().winfo_class() == "TRadiobutton":
                self.tk_focusNext().focus_set()
                return "break"

    class MyText(tk.Text):
        def __init__(self, parent, **kwargs):
            super().__init__(parent, **kwargs)
            self.configure(bg="white", fg="black", selectforeground="black",
                selectbackground="light gray", highlightbackground="gray",
                relief="flat")
            self.bind("<Button-3><ButtonRelease-3>", self.show_menu)

        def select_all(self, *args):
            self.focus_force()
            self.tag_add("sel", "1.0", "end")

        def show_menu(self, event):
            if str(self.cget("state")) == "disabled":
                return
            cutpasteMenu(event)

    class MyTextView_v(tk.Text):
        def __init__(self, parent, bg=None, fg=None, bd=1, **kwargs):
            super().__init__(parent, **kwargs)
            self.configure(bg="white", fg="black", selectforeground="black",
                selectbackground="light gray", highlightbackground="gray",
                bd=bd, relief="flat")
            self.bind("<Key>", self.checkHeight)
            self.bind("<FocusIn>", self.checkHeight)
            self.bind("<Button-3><ButtonRelease-3>", self.show_menu)
            self.event_add("<<mytxtvv>>", "<F1>", "<F5>", "<Escape>",
                "<Return>", "<KP_Enter>")

        def select_all(self, *args):
            self.focus_force()
            self.tag_add("sel", "1.0", "end")

        def show_menu(self, event):
            if str(self.cget("state")) == "disabled":
                return
            cutpasteMenu(event)

        def checkHeight(self, event):
            text = event.widget.get("1.0", "end").rstrip()
            if event.char.isalpha():
                text += event.char
            else:
                text = text[:-1]
            wrap = textwrap.TextWrapper(width=event.widget["width"],
                break_long_words=False)
            event.widget["height"] = len(wrap.wrap(text))

    class MyTextView_V(tk.Text):
        def __init__(self, parent, bg=None, fg=None, bd=1, **kwargs):
            super().__init__(parent, **kwargs)
            self.configure(bg="white", fg="black", selectforeground="black",
                selectbackground="light gray", highlightbackground="gray",
                bd=bd, relief="flat")
            self.bind("<Button-3><ButtonRelease-3>", self.show_menu)
            self.event_add("<<mytxtvV>>", "<F1>", "<F5>", "<F9>", "<Escape>")

        def select_all(self, *args):
            self.focus_force()
            self.tag_add("sel", "1.0", "end")

        def show_menu(self, event):
            if str(self.cget("state")) == "disabled":
                return
            cutpasteMenu(event)

    class ScrollGrid(object):
        """
        Create a scrollable grid using the following options:

        mf   = Mainframe
        titl = "Title Description" or ["aaaaaa", "bbbbbb"]
        icon = The icon to use for the window
        chgt = The number of rows per cell, default = 1
        tags = [(name, color), (name, color)]
        labs = [(text, width), (text, width), (text, width)]
        cols = [
            [(text, span), (text, span)],                       ### Optional
            [(text, ('SD',10.2)), or (text, 10), ..... ]]
        data = [
            (((label, tag) (label, tag) (label, tag)),
            ((text, tag, span), (text, tag), (text, tag), (text, tag)),)]
        butt = A list of additional buttons e.g. [("hello", cmd)]
        cmds = A list of bind commands
        font = (family, size)
        wait = Wait for main window to be destroyed
        minc = Minimum columns to show
        """
        def __init__(self, **opts):
            if "icon" not in opts:
                opts["icon"] = None
            if "chgt" not in opts:
                opts["chgt"] = 1
            if "wait" not in opts:
                opts["wait"] = True
            self.opts = opts
            self.drawGrid()
            if self.opts["wait"]:
                self.window.wait_window()

        def drawGrid(self):
            # Draw main window
            self.window = MkWindow(icon=self.opts["icon"]).newwin
            ww = int(self.window.winfo_screenwidth())
            wh = int(self.window.winfo_screenheight() * .9)
            # Style and fonts and colours
            style = ttk.Style()
            style.configure("TScrollbar", arrowsize=30)
            if "mf" in self.opts:
                self.nbg = self.opts["mf"].rcdic["nbg"]
                self.nfg = self.opts["mf"].rcdic["nfg"]
            else:
                self.nbg = "white"
                self.nfg = "black"
            if "font" in self.opts:
                ft = self.opts["font"][0]
                fs = self.opts["font"][1]
            elif "mf" in self.opts:
                ft = self.opts["mf"].rcdic["dft"]
                fs = self.opts["mf"].rcdic["mfs"]
            else:
                ft = "Arial"
                fs = 14
            twidth = ww + 1
            while twidth > ww:
                bold = tkfont.Font(font=(ft, fs, "bold"))
                default_font = tkfont.nametofont("TkDefaultFont")
                default_font.configure(size=fs)
                # Calculate label widths
                lwdth = [1, []]
                for lab in self.opts["labs"]:
                    txt = "X" * int(lab[1])
                    lab = MyLabel(self.window, text=txt, font=bold)
                    self.window.update_idletasks()
                    lwdth[0] += lab.winfo_reqwidth()
                    lwdth[1].append(lab.winfo_reqwidth())
                    lab.destroy()
                # Calculate column width and height
                col = self.opts["cols"][len(self.opts["cols"]) - 1][0]
                if type(col[1]) in (list, tuple):
                    tw = int(col[1][1])
                else:
                    tw = int(col[1])
                txt = "X" * tw
                for _ in range(1, self.opts["chgt"]):
                    txt += "\n"
                lab = MyLabel(self.window, text=txt, font=bold)
                self.window.update_idletasks()
                cw, ch = (lab.winfo_reqwidth(), lab.winfo_reqheight())
                lab.destroy()
                if "minc" in self.opts:
                    twidth = lwdth[0] + (cw * self.opts["minc"])
                else:
                    twidth = ww
                fs -= 1
                if fs < 6:
                    break
            # Draw widgets
            # Main Heading
            if type(self.opts["titl"]) in (list, tuple):
                txt = self.opts["titl"][0]
                for t in self.opts["titl"][1:]:
                    txt = "%s\n%s" % (txt, t)
            else:
                txt = self.opts["titl"]
            lb1 = MyLabel(self.window, justify="center", anchor="center",
                padding=2, font=bold, text=txt)
            lb1.pack(fill="x")
            # Container frame
            self.cframe = MyFrame(self.window)
            # Left frames
            lframe = MyFrame(self.cframe)
            lframe.pack(fill="y", side="left")
            spc = ttk.Scrollbar(lframe, orient="horizontal")
            spc.pack(fill="x", side="bottom", expand=False)
            lframeh = MyFrame(lframe)
            lframeh.pack(fill="x")
            lframed = MyFrame(lframe)
            lframed.pack(fill="both", expand="yes")
            # Right frames
            rframe = MyFrame(self.cframe)
            rframe.pack(fill="both", expand="yes", side="left")
            hsb = ttk.Scrollbar(rframe, orient="horizontal",
                command=self._xview)
            hsb.pack(fill="x", side="bottom", expand=False)
            vsb = ttk.Scrollbar(rframe, orient="vertical",
                command=self._yview)
            vsb.pack(fill="y", side="right", expand=False)
            rframeh = MyFrame(rframe)
            rframeh.pack(fill="x")
            self.rframed = MyFrame(rframe)
            self.rframed.pack(fill="both", expand="yes")
            # Label headings canvas
            if len(self.opts["cols"]) == 1:
                hgt = ch
            else:
                hgt = ch * len(self.opts["cols"])
            self.cv1 = tk.Canvas(lframeh, bd=0, highlightthickness=0,
                height=hgt, width=lwdth[0])
            self.cv1.pack(fill="x", anchor="nw")
            # Label details canvas
            self.cv2 = tk.Canvas(lframed, bd=0, highlightthickness=0,
                width=lwdth[0], yscrollcommand=vsb.set)
            self.cv2.pack(fill="both", expand="yes", side="bottom")
            # Column headings canvas
            if len(self.opts["cols"]) == 1:
                self.cv3 = tk.Canvas(rframeh, bd=0, highlightthickness=0,
                    height=ch, xscrollcommand=hsb.set)
            else:
                self.cv3 = tk.Canvas(rframeh, bd=0, highlightthickness=0,
                    height=(ch * 2), xscrollcommand=hsb.set)
            self.cv3.pack(fill="x", anchor="nw")
            # Column details canvas
            self.cv4 = tk.Canvas(self.rframed, bd=0, highlightthickness=0,
                xscrollcommand=hsb.set, yscrollcommand=vsb.set)
            self.cv4.pack(fill="both", expand="yes", anchor="nw")
            self.window.bind("<Configure>", self.set_scrollregion)
            # Populate cells
            tags = {}
            for tag in self.opts["tags"]:
                tags[tag[0]] = tag[1]
            # Label headings
            x1 = 0
            y1 = 0
            for num, lab in enumerate(self.opts["labs"]):
                x2 = x1 + lwdth[1][num]
                if len(self.opts["cols"]) == 1:
                    y2 = y1 + ch
                else:
                    y2 = y1 + (ch * 2)
                rect = self.cv1.create_rectangle(x1, y1, x2, y2, width=2,
                    fill=self.nbg)
                text = self.cv1.create_text(x1 + 10, y1 + int(ch / 2),
                    text=lab[0], anchor="w", font=bold, fill=self.nfg)
                x1 = x2
            # Column headings
            x1 = 0
            y1 = 0
            idx = 0
            if len(self.opts["cols"]) == 2:
                for col in self.opts["cols"][idx]:
                    cc = col[1] * tw            # total width
                    x2 = x1 + (col[1] * cw)     # column span
                    y2 = y1 + ch
                    rect = self.cv3.create_rectangle(x1, y1, x2, y2, width=2,
                        fill=self.nbg)
                    txt = "{:^%i}" % cc
                    txt = txt.format(col[0])
                    text = self.cv3.create_text(x1 + 10, y1 + int(ch / 2),
                        text=txt, font=bold, fill=self.nfg, anchor="w")
                    x1 = x2
                x1 = 0
                y1 = y2
                idx = 1
            for col in self.opts["cols"][idx]:
                x2 = x1 + cw
                y2 = y1 + ch
                rect = self.cv3.create_rectangle(x1, y1, x2, y2, width=2,
                    fill=self.nbg)
                text = self.cv3.create_text(x1 + 10, y1 + int(ch / 2),
                    text=col[0], anchor="w", font=bold, fill=self.nfg)
                x1 = x2
            y1 = 0
            # Label and Column data
            for row, rdat in enumerate(self.opts["data"]):
                x1 = 0
                for num, lab in enumerate(rdat[0]):
                    # Label data
                    x2 = x1 + lwdth[1][num]
                    y2 = y1 + ch
                    if self.opts["chgt"] == 2 and not lab[0].count("\n"):
                        txt = "\n%s" % lab[0]
                    else:
                        txt = lab[0]
                    if lab[1]:
                        tfill, rfill = tags[lab[1]]
                    else:
                        tfill, rfill = None, None
                    rect = self.cv2.create_rectangle(x1, y1, x2, y2, width=2,
                        fill=rfill)
                    text = self.cv2.create_text(x1 + 10, y1 + int(ch / 2),
                        text=txt, font=bold, fill=tfill, anchor="w")
                    x1 = x2
                x1 = 0
                if len(self.opts["cols"]) == 1:
                    idx = 0
                else:
                    idx = 1
                for col, cdat in enumerate(rdat[1]):
                    # Column data
                    if len(cdat) == 3:
                        x2 = x1 + (cdat[2] * cw)
                    else:
                        x2 = x1 + cw
                    y2 = y1 + ch
                    if cdat[1]:
                        tfill, rfill = tags[cdat[1]]
                    else:
                        tfill, rfill = None, None
                    rect = self.cv4.create_rectangle(x1, y1, x2, y2, width=2,
                        fill=rfill)
                    fmt = self.opts["cols"][idx][col][1]
                    if type(fmt) in (list, tuple):
                        f, w = fmt
                        if f[1] in ("D", "I"):
                            if type(cdat[0]) == str:
                                f = "TX"
                            elif f[0] == "S":
                                w -= 1
                        txt = CCD(cdat[0], f, w)
                        if txt.err:
                            txt = cdat[0]
                        else:
                            txt = txt.disp
                    else:
                        txt = cdat[0]
                    if self.opts["chgt"] == 2 and not txt.count("\n"):
                        txt = "\n%s" % t
                    text = self.cv4.create_text(x1 + 10, y1 + int(ch / 2),
                        text=txt, anchor="w", font=bold, fill=tfill)
                    if "cmds" in self.opts:
                        for cmd in self.opts["cmds"]:
                            self.cv4.tag_bind(rect, cmd[0],
                                functools.partial(self._get_cell, cmd[1],
                                    (row, col), txt))
                            self.cv4.tag_bind(text, cmd[0],
                                functools.partial(self._get_cell, cmd[1],
                                    (row, col), txt))
                    x1 = x2
                y1 = y2
            # Arrow and page keys
            for key in ("Left", "Right", "Up", "Down", "Prior", "Next"):
                self.window.bind("<%s>" % key, self._scroll)
            # Close and Other buttons
            _quit = True
            bbox = MyButtonBox(self.window)
            if "butt" in self.opts:
                for but in self.opts["butt"]:
                    bbox.addButton(but[0], but[1])
                    if but[0] in ("Exit", "Quit"):
                        _quit = False
                        self.window.bind("<Escape>", but[1])
            if _quit:
                bbox.addButton("Quit", self._quit)
                self.window.bind("<Escape>", self._quit)
            # Pack frame
            self.cframe.pack(fill="both", expand="yes")
            self.cframe.update_idletasks()
            height = lb1.winfo_reqheight()
            height += self.cv3.bbox("all")[3]
            height += self.cv4.bbox("all")[3]
            height += hsb.winfo_reqheight()
            height += bbox.winfo_reqheight()
            if wh > height:
                self.window.configure(height=height)
                self.window.update()
            # Place window
            self.window.geometry("%dx%d+0+0" % (ww, height))
            self.window.geometry("+0+0")
            if self.window.state() == "withdrawn":
                self.window.deiconify()
            self.window.grab_set()

        def set_scrollregion(self, event):
            self.cv2.configure(scrollregion=self.cv2.bbox("all"))
            b = self.cv3.bbox("all")
            w = self.cv4.bbox("all")[2]
            self.cv3.configure(scrollregion=(b[0], b[1], w, b[3]))
            self.cv4.configure(scrollregion=self.cv4.bbox("all"))

        def _scroll(self, event):
            if event.keysym == "Left":
                self._xview("scroll", -1, "units")
            elif event.keysym == "Right":
                self._xview("scroll", 1, "units")
            elif event.keysym == "Up":
                self._yview("scroll", -1, "units")
            elif event.keysym == "Down":
                self._yview("scroll", 1, "units")
            elif event.keysym == "Prior":
                self._yview("scroll", -1, "pages")
            elif event.keysym == "Next":
                self._yview("scroll", 1, "pages")

        def _xview(self, *args):
            self.cv3.xview(*args)
            self.cv4.xview(*args)

        def _yview(self, *args):
            self.cv2.yview(*args)
            self.cv4.yview(*args)

        def _get_cell(self, *args):
            # execute command with following arguments:
            #   (row, col, (rframed, (x, y)), window)
            plc = (args[3].x, args[3].y)
            args[0](args[1], args[2], (self.rframed, plc), self.window)

        def _quit(self, event=None):
            self.window.destroy()

    class ScrollWindow(MyFrame):
        """
        * Use the 'interior' attribute to place widgets inside the
          scrollable frame
        * Construct and pack/place/grid normally
        * This frame only allows vertical scrolling
        """
        def __init__(self, parent, *args, **kw):
            if "height" not in kw:
                kw["height"] = 200
            super().__init__(parent, *args, **kw)
            # create a canvas object and a vertical scrollbar for scrolling it
            self.vsb = ttk.Scrollbar(self, orient="vertical")
            self.vsb.pack(fill="y", side="right", expand=False)
            self.cnv = tk.Canvas(self, bd=0, height=kw["height"],
                highlightthickness=0, yscrollcommand=self.vsb.set)
            self.cnv.pack(side="left", fill="both", expand=True)
            self.vsb.config(command=self.cnv.yview)
            # reset the view
            self.cnv.xview_moveto(0)
            self.cnv.yview_moveto(0)
            # create a frame inside the canvas which will be scrolled with it
            self.interior = MyFrame(self.cnv)
            self.interior_id = self.cnv.create_window(0, 0,
                window=self.interior, anchor="nw")
            self.interior.bind("<Configure>", self._configure_interior)
            self.cnv.bind("<Configure>", self._configure_canvas)
            if sys.platform == "win32":
                self.bind("<MouseWheel>", self._on_mousewheel)
            else:
                self.bind("<Button-4>", self._on_mousewheel)
                self.bind("<Button-5>", self._on_mousewheel)

        def _configure_interior(self, event):
            # update the scrollbars to match the size of the inner frame
            self.interior.update_idletasks()
            size = (self.interior.winfo_reqwidth(),
                self.interior.winfo_reqheight())
            self.cnv.config(scrollregion="0 0 %s %s" % size)
            if self.cnv.winfo_width() != size[0]:
                # update the canvas's width to fit the inner frame
                self.cnv.config(width=size[0])

        def _configure_canvas(self, event):
            if self.interior.winfo_reqwidth() != self.cnv.winfo_width():
                # update the inner frame's width to fill the canvas
                self.cnv.itemconfigure(self.interior_id,
                    width=self.cnv.winfo_width())

        def _on_mousewheel(self, event):
            if sys.platform == "win32":
                scroll = -1 if event.delta > 0 else 1
            else:
                scroll = -1 if event.num == 4 else 1
            self.cnv.yview_scroll(scroll, "units")

    GUI = True
except Exception as err:
    print(err)
    GUI = False

# =========================================================
def rgb(col):
    # RGB color tuple
    if type(col) == int:
        return int(col / 65536), int(col / 256) % 256, col % 256
    if col[0] == "#":
        col = col.replace("#", "")
        col = (col[0:2], col[2:4], col[4:6])
        return tuple([int(x, 16) for x in col])
# =========================================================

class MainFrame(object):
    def __init__(self, title=None, rcdic=None, xdisplay=True):
        if not rcdic:
            self.rcdic = loadRcFile()
        else:
            self.rcdic = rcdic
        if self.rcdic == "error":
            showError(None, "rcdic Error",
                "There is a Problem Loading the tartanrc File")
        else:
            self.mloop = 0
            self.dbm = None
            geo = self.rcdic["geo"].split("x")
            self.geo = [int(geo[0]), int(geo[1])]
            self.override = b64Convert("decode", "bmF0cmF0")
            if xdisplay:
                if not GUI:
                    print("Missing Tkinter and/or ttk modules")
                    sys.exit()
                if not title:
                    title = "Tartan Systems - Copyright %s 2004-2022 "\
                        "Paul Malherbe" % chr(0xa9)
                self.window = MkWindow(title=title, icon="tartan",
                    size=self.geo, resiz=False).newwin
                self.setThemeFont()
                self.createChildren()
                self.window.config(cursor="arrow")
            else:
                self.childs = False
                self.window = None
                self.body = None

    def setThemeFont(self, butt=True):
        # Set Fonts
        fonts = (
            ("TkDefaultFont", self.rcdic["dft"], self.rcdic["dfs"], "normal"),
            ("TkTextFont", self.rcdic["dft"], self.rcdic["dfs"], "normal"),
            ("TkFixedFont", self.rcdic["dft"], self.rcdic["dfs"], "normal"),
            ("TkMenuFont", self.rcdic["mft"], self.rcdic["mfs"], "normal"),
            ("TkHeadingFont", self.rcdic["dft"], self.rcdic["dfs"], "normal"))
        for font in fonts:
            nf = tkfont.nametofont(font[0])
            if len(font) != 4:
                nf.configure(family=font[1], size=font[2])
            else:
                nf.configure(family=font[1], size=font[2], weight=font[3])
            self.window.option_add(font[0], nf)
        # Set Theme and Styles
        self.style = ttk.Style()
        themes = self.style.theme_names()
        if self.rcdic["theme"] in themes:
            self.style.theme_use(self.rcdic["theme"])
        if butt:
            self.style.configure("MyButton.TButton",
                font=(self.rcdic["mft"], self.rcdic["dfs"]))
        if self.rcdic["theme"] in ("alt", "clam", "classic", "default"):
            self.style.configure("MyButton.TButton",
                foreground=self.rcdic["bfg"],
                background=self.rcdic["bbg"],
                padding=2)
            self.style.map("MyButton.TButton",
                foreground=[
                    ("active", self.rcdic["ffg"]),
                    ("disabled", self.rcdic["dfg"]),
                    ("focus", self.rcdic["ffg"])],
                background=[
                    ("active", self.rcdic["fbg"]),
                    ("disabled", self.rcdic["dbg"]),
                    ("focus", self.rcdic["fbg"])])
        if tk.TkVersion >= 8.5:
            self.style.configure("MyCheckbutton.TCheckbutton")
            self.style.map("MyCheckbutton.TCheckbutton",
                foreground=[
                    ("focus", "white"),
                    ("disabled", "black")],
                background=[
                    ("focus", "gray")],
                indicatorbackground=[
                    ("focus", "white"),
                    ("disabled", "white"),
                    ("pressed", "white")],
                indicatorcolor=[
                    ("focus", "black"),
                    ("pressed", "black"),
                    ("selected", "black")],
                focuscolor=[
                    ("focus", "gray")])
            self.style.configure("MyRadiobutton.TRadiobutton")
            self.style.map("MyRadiobutton.TRadiobutton",
                foreground=[
                    ("focus", "white"),
                    ("disabled", "black")],
                background=[
                    ("focus", "gray")],
                indicatorbackground=[
                    ("focus", "white"),
                    ("disabled", "white"),
                    ("pressed", "white")],
                indicatorcolor=[
                    ("focus", "black"),
                    ("pressed", "black"),
                    ("selected", "black")],
                focuscolor=[
                    ("focus", "gray")])
        self.style.configure("TEntry",
            foreground="black",
            background="white",
            selectforeground="black",
            selectbackground="light gray")
        self.style.map("TEntry",
            foreground=[
                ("active", "black"),
                ("disabled", "black")],
            background=[
                ("active", "white"),
                ("disabled", "white")],
            fieldforeground=[
                ("active", "black"),
                ("disabled", "black")],
            fieldbackground=[
                ("active", "white"),
                ("disabled", "white")])
        self.style.configure("TLabel",
            foreground="black")
        self.style.configure("CLabel.TLabel",
            foreground=self.rcdic["nfg"],
            background=self.rcdic["nbg"])
        self.style.configure("Horizontal.TProgressbar",
            background=self.rcdic["fbg"],
            troughcolor=self.rcdic["ffg"])
        self.style.configure("Treeview",
            background="white",
            foreground="black")
        self.style.map("Treeview",
            foreground=[
                ("selected", self.rcdic["nfg"])],
            background=[
                ("selected", self.rcdic["nbg"])])
        self.window.update_idletasks()

    def createChildren(self):
        # Header
        self.head = MyLabel(self.window, anchor="center", borderwidth=2,
            font=tkfont.Font(font=(self.rcdic["mft"], self.rcdic["mfs"],
            "bold")), padding=0, relief="ridge")
        # Body
        self.body = MyFrame(self.window, borderwidth=2, padding=0,
            relief="ridge")
        self.body.pack_propagate(False)
        # Status
        self.status = MyLabel(self.window, color=False, borderwidth=2,
            background="white", foreground="black", padding=0, relief="ridge")
        self.packChildren()
        self.childs = True

    def packChildren(self, ptyp="pack"):
        if ptyp == "pack":
            self.head.pack(fill="x", side="top")
            self.body.pack(fill="both", expand="yes")
            self.status.pack(fill="x", side="bottom")
            self.window.update_idletasks()
            self.resizeChildren()
        else:
            self.head.pack_forget()
            self.body.pack_forget()
            self.status.pack_forget()

    def resizeChildren(self):
        font = tkfont.Font(font=(self.rcdic["mft"], self.rcdic["mfs"],"bold"))
        self.head.configure(text=self.head.cget("text"), font=font)
        font = tkfont.Font(font=(self.rcdic["dft"], self.rcdic["dfs"],"normal"))
        self.status.configure(text=self.status.cget("text"), font=font)
        placeWindow(self.window, place=self.rcdic["plc"], size=self.geo)

    def destroyChildren(self):
        try:
            for child in self.window.winfo_children():
                child.destroy()
        except:
            pass
        self.window.update_idletasks()
        self.childs = False

    def updateStatus(self, text, bg="white", fg="black"):
        if self.childs:
            fsz = self.rcdic["dfs"]
            while True:
                font = tkfont.Font(font=(self.rcdic["dft"], fsz, "normal"))
                if font.measure(text) > self.status.winfo_width():
                    fsz -= 1
                else:
                    break
            self.status.configure(background=bg, foreground=fg,
                font=font, text=text)
            self.window.update_idletasks()

    def startLoop(self, deicon=True):
        self.mloop += 1
        if self.mloop > 1:
            print("Loop Error", self.mloop)
        if deicon and self.window.state() == "withdrawn":
            self.window.deiconify()
        self.window.mainloop()

    def closeLoop(self):
        self.window.update_idletasks()
        self.mloop -= 1
        self.window.quit()

class MkWindow(object):
    """
    Valid Arguments:
        title   The title of the window
        icon    The icon to use for the window
        size    The size of the window
        resiz   Whether the window is resizable
        trans   The window which this window is transient to
        decor   Whether this window has WM decorations
        frame   Whether this window has a frame
        remov   Whether this window is removable
        hide    Withdraw window
    """
    def __init__(self, **args):
        for arg in ("title", "icon", "size", "resiz", "trans",
                "decor", "frame", "remov", "hide"):
            if arg in args:
                setattr(self, arg, args[arg])
            elif arg == "size":
                self.size = ""
            elif arg == "resiz":
                self.resiz = True
            elif arg == "decor":
                self.decor = True
            elif arg == "frame":
                self.frame = False
            elif arg == "remov":
                self.remov = True
            elif arg == "hide":
                self.hide = True
            else:
                setattr(self, arg, None)
        self.drawWindow()
        self.setAttributes()

    def drawWindow(self):
        if tk._default_root is None:
            # Initial main window
            self.newwin = tk.Tk()
            # Load extra themes if installed
            # Additional themes must be installed in a folder named 'thm'
            # Which must be in the Tartan root folder i.e. where
            # ms0000.py or ms0000.exe resides
            try:
                ign = ("ImageLib", "pkgIndex", "pkgIndex_package")
                tdir = os.path.join(sys.path[0], "thm")
                for root, dirs, files in os.walk(tdir):
                    for fle in files:
                        pth = os.path.join(root, fle)
                        nam = os.path.basename(pth)
                        tst = nam.split(".")
                        if tst[-1] == "tcl" and tst[-2] not in ign:
                            try:
                                self.newwin.tk.eval("source {%s}" % pth)
                            except:
                                continue
            except:
                pass
            # Enable button to show hidden files
            try:
                try:
                    self.newwin.tk.call("tk_getOpenFile", "-foobarbaz")
                except:
                    pass
                self.newwin.tk.call("set",
                    "::tk::dialog::file::showHiddenBtn", "1")
                self.newwin.tk.call("set",
                    "::tk::dialog::file::showHiddenVar", "0")
            except:
                pass
        else:
            self.newwin = tk.Toplevel()
        # Disable the Tab key
        self.newwin.bind_all("<Tab>", lambda event: "break")
        if self.hide:
            self.newwin.withdraw()

    def setAttributes(self):
        if self.trans:
            self.newwin.transient(self.trans)
        if self.decor:
            if self.icon:
                try:
                    self.newwin.tk.call("wm", "iconphoto", self.newwin._w,
                        "-default", getImage(self.icon))
                except:
                    pass
            self.newwin.protocol("WM_DELETE_WINDOW", self.doDestroy)
            if self.title:
                self.newwin.title(self.title)
            else:
                self.newwin.title("Tartan Systems - %s 2004-%s Paul Malherbe"
                    % (chr(169), time.localtime().tm_year))
        else:
            self.newwin.overrideredirect(True)
            self.newwin.update_idletasks()
        if self.size:
            self.newwin.configure(width=self.size[0], height=self.size[1])
        if not self.resiz:
            self.newwin.resizable(0, 0)
        if self.trans:
            self.newwin.transient(self.trans)
        if self.frame:
            self.newfrm = MyFrame(self.newwin)
            self.newfrm.pack(fill="both", expand="yes")

    def doDestroy(self):
        if not self.remov or "TARTANDB" in os.environ:
            return True
        ok = askQuestion(screen=self.newwin, head="Destroy Window",
            mess="Are you Sure you want to Quit?\n\n"
            "It is Possible that Some Data Could Be Lost.\n\n"
            "It is Far Better to Quit via the Menu.", default="no")
        if ok == "no":
            # Re-Focus on first normal input widget, if any
            def traverse(children):
                if self.focus:
                    return True
                for child in children:
                    if child.winfo_class() in ("TEntry", "TRadiobutton"):
                        if str(child.cget("state")) == "normal":
                            child.focus_set()
                            self.focus = True
                            break
                    else:
                        traverse(child.winfo_children())
            self.focus = False
            traverse(self.newwin.winfo_children())
            return True
        try:
            sys.stdout.close()
        except:
            pass
        os._exit(1)

class TartanMenu(object):
    """
    Use this to class to draw menus as follows:

    Create a list of lists of MenuBar Items for example:

        men = [["MM","Control", "Control Routines and exit"],
               ["MM","General", "General Ledger Menu"],
               ["MM","Creditors", "Creditor's Ledger Menu"],
               ["MM","Debtors", "Debtor's Ledger Menu"],
               ["MM","Stores", "Stores Ledger Menu"],
               ["MM","Help", "User's Manual and About","right"]]

        Where the fields are as follows:

        MM = Main Menu:
             Menu Name
             Detail
        CM = Cascade Menu:
             Parent Name
             Menu Name
             Detail

    Create a list of lists of Menu Items for example:

        mod = [["F","Control","tarCfg","Configure Tartan"],
               ["PNN","Control","co1010","Company Records"],
               ["F","Control","sysEnd","Exit Program"],
               ["CY","Control","l -l","Directory"]]

        Where the fields are as follows:

            Type        F    =  Function
                        P123 = Tartan module where:
                                1 = Requiring Company Number (Y/N)
                                2 = Requiring Period Selection (Y/N/L)
                                3 = Pass the User Name to the Module
                        CY   = System Command requiring a terminal session
                        CN   = System Command not requiring a terminal session
                        S    = A standard callable python class
                        U    = A Program without normal arguments only user
            Parent Menu Name
            Program, Function or Command
            Security Level
            Prompt
            Routine number if required for type P above
            If required, Insert into menu before this item on the menu

    Draw the menus as follows:

        menu = TartanMenu({args})

        mf  = Tartan MainFrame
        usr = Name of user
        men = Menus
        mod = Modules
        lvl = Security Level
        cmd = Command to Execute on Selection
        img = Background Image to Display
    """
    def __init__(self, **args):
        for arg in args:
            setattr(self, arg, args[arg])
        self.setVariables()

    def setVariables(self):
        # Create a list of menus and a dictionary of modules
        menus = []
        for m in self.mod:
            if self.lvl is not None and m[3] > self.lvl:
                continue
            if m[1] not in menus:
                menus.append(m[1])
            if m[0][0] == "P":
                mm = "mm_%s" % m[2][:2]
                if mm not in menus:
                    menus.append(mm)
        self.menus = []
        for men in self.men:
            if men[0] == "MM" and men[1] in menus:
                self.menus.append(men)
            elif men[2] in menus:
                self.menus.append(men)

    def drawMenu(self):
        # Clear toplevel bindings
        unbindAllWidgets(self.mf.window)
        if self.img:
            fg = "#ffffff"
            bg = "#a40204"
        else:
            fg = self.mf.rcdic["nfg"]
            bg = self.mf.rcdic["nbg"]
        style = ttk.Style()
        style.configure("Menu.TLabel", foreground=fg, background=bg)
        self.mf.head.configure(text="Tartan Systems Menu (%s)" % self.usr,
            style="Menu.TLabel")
        self.menubar = MyFrame(self.mf.body)
        labs = {}
        mens = {}
        buts = {}
        column = 0
        font = tkfont.Font(font=(self.mf.rcdic["mft"], self.mf.rcdic["mfs"]))
        for m in self.menus:
            if m[0] == "MM":
                if "mbar" not in labs:
                    labs["mbar"] = []
                m[2], pos = getUnderline(blist=labs["mbar"], text=m[2])
                if m[2][pos] not in labs["mbar"]:
                    labs["mbar"].append(m[2][pos])
                m[2] = m[2].split()
                nm = m[2][0].split("'")[0]
                if len(m[2]) == 1:
                    m[2] = m[2][0]
                elif len(m[2]) == 2:
                    m[2] = "%s\n%s" % tuple(m[2])
                else:
                    m[2] = "%s\n%s %s" % tuple(m[2])
                buts[nm] = MyMenuButton(self.menubar, bg=bg, fg=fg,
                    af=bg, ab=fg, font=font, menu=m[1], text=m[2],
                    underline=pos)
                mens[m[1]] = tk.Menu(buts[nm], fg="black", bg="white",
                    activeforeground=fg, activebackground=bg,
                    font=font, tearoff=False)
                buts[nm]["menu"] = mens[m[1]]
                column += 1
            elif m[0] == "CM":
                if m[1] not in labs:
                    labs[m[1]] = []
                mens[m[2]] = tk.Menu(mens[m[1]], fg="black", bg="white",
                    activeforeground=fg, activebackground=bg,
                    font=font, tearoff=False)
                m[3], pos = getUnderline(blist=labs[m[1]], text=m[3])
                if m[3][pos] not in labs[m[1]]:
                    labs[m[1]].append(m[3][pos])
                mens[m[1]].add("cascade", label=m[3], menu=mens[m[2]],
                    font=font, underline=pos)
        for m in self.mod:
            if m[3] > self.lvl:
                continue
            if m[1] not in labs:
                labs[m[1]] = []
            if m[0][0] == "P" and len(m) == 6:
                rtn = int(m[5])
            else:
                rtn = None
            if len(m) == 7:
                idx = 0
                for mm in self.menus:
                    if mm[0] == "MM":
                        continue
                    if m[1] == mm[1]:
                        if m[6] == mm[2]:
                            break
                        idx += 1
            else:
                idx = None
            if m[4] == "Quit":
                mens[m[1]].add_separator()
            m[4], pos = getUnderline(blist=labs[m[1]], text=m[4])
            if m[4][pos] not in labs[m[1]]:
                labs[m[1]].append(m[4][pos])
            cmd = (self.menubar.register(self.cmd), m[0], m[2], m[4], rtn)
            image = getImage(m[2], (20, 20))
            if image:
                try:
                    getattr(self, "image%s" % m[2])
                except:
                    setattr(self, "image%s" % m[2], image)
                if idx is None:
                    mens[m[1]].add("command", label=m[4], command=cmd,
                        font=font, underline=pos, image=getattr(self,
                        "image%s" % m[2]), compound="left")
                else:
                    mens[m[1]].insert(idx, "command", label=m[4], command=cmd,
                        font=font, underline=pos, image=getattr(self,
                            "image%s" % m[2]), compound="left")
            elif idx is None:
                mens[m[1]].add("command", label=m[4], command=cmd, font=font,
                    underline=pos)
            else:
                mens[m[1]].insert(idx, "command", label=m[4], command=cmd,
                    font=font, underline=pos)
        # Calculate optimal font to show all items
        self.mf.window.update_idletasks()
        while self.menubar.winfo_reqwidth() > self.mf.geo[0]:
            self.mf.rcdic["mfs"] = int(self.mf.rcdic["mfs"]) - 1
            font = tkfont.Font(font=(self.mf.rcdic["mft"],
                self.mf.rcdic["mfs"]))
            for but in buts:
                buts[but].configure(font=font)
            self.menubar.update_idletasks()
        self.menubar.pack(anchor="nw", fill="x")
        iwth = self.mf.body.winfo_width()
        ihgt = self.mf.body.winfo_height()
        self.image = tk.Canvas(self.mf.body, width=iwth, height=ihgt)
        if self.img:
            self.tkimg = getImage("tartan", (iwth, ihgt))
            self.image = tk.Canvas(self.mf.body, width=iwth, height=ihgt)
            self.image.create_image(0, 0, image=self.tkimg, anchor="nw")
            self.image.ilist = [self.tkimg]
        else:
            self.image.config(bg=self.mf.rcdic["nbg"])
            self.image.create_text(int(iwth / 2), int(ihgt / 2) - 60,
                text="Tartan Systems", font=("Helvetica", 60), fill="white")
        self.image.pack(fill="both", expand="yes")
        self.mf.updateStatus("Select an Option from the Menu")
        self.menubar.focus_force()
        self.mf.startLoop()

    def closeMenu(self, children=True):
        self.menubar.destroy()
        self.image.destroy()
        self.mf.head.configure(style="CLabel.TLabel")
        self.mf.closeLoop()

class FileDialog(object):
    def __init__(self, **opts):
        for opt in opts:
            if opt == "ftype":
                self.ftype = []
                for typ in opts["ftype"]:
                    self.ftype.append(tuple(typ))
            else:
                setattr(self, opt, opts[opt])
        if "parent" not in opts:
            self.parent = None
            self.root = tk.Tk()
            self.root.withdraw()
        else:
            self.root = self.parent.winfo_toplevel()
        if "title" not in opts:
            self.title = "Select ...."
        if "initd" not in opts:
            self.initd = os.getcwd()
        if "ftype" not in opts:
            self.ftype = []
        if "multi" not in opts:
            self.multi = False

    def setgeometry(self, styp):
        try:
            if styp == "fle":
                dialog = "__tk_filedialog"
            else:
                dialog = "__tk_choosedir"
            if self.parent:
                w0 = int(self.parent.winfo_width() * .8)
                h0 = int(self.parent.winfo_height() * .8)
                x, y = self.parent.winfo_rootx(), self.parent.winfo_rooty()
                x = x + (int(self.parent.winfo_width() - w0) / 2)
                y = y + (int(self.parent.winfo_height() - h0) / 2)
                self.root.tk.call("wm", "geometry",
                    "%s.%s" % (self.parent, dialog),
                    "%dx%d+%d+%d" % (w0, h0, x, y))
            else:
                w = self.root.winfo_screenwidth()
                h = self.root.winfo_screenheight()
                w0, h0 = int(w / 3), int(h / 3)
                x, y = int(w / 2) - int(w0 / 2), int(h / 3) - int(h0 / 2)
                self.root.tk.call("wm", "geometry",
                    ".%s" % dialog,
                    "%dx%d+%d+%d" % (w0, h0, x, y))
        except:
            pass

    def askopenfilename(self):
        self.root.after(10, self.setgeometry, "fle")
        if self.multi:
            sel = tkfile.askopenfilenames(title=self.title,
                parent=self.parent, initialdir=self.initd,
                filetypes=self.ftype)
            if sys.platform == "win32":
                return self.root.tk.splitlist(sel)
            else:
                return sel
        else:
            sel = tkfile.askopenfilename(title=self.title,
                parent=self.parent, initialdir=self.initd,
                filetypes=self.ftype)
            return sel

    def askdirectory(self):
        self.root.after(10, self.setgeometry, "dir")
        sel = tkfile.askdirectory(title=self.title, parent=self.parent,
            initialdir=self.initd, mustexist=True)
        return sel

class SplashScreen(object):
    """
    A splash screen widget
    """
    def __init__(self, scrn=None, text=None):
        self.scrn = scrn
        if not self.scrn:
            self.window = MkWindow(decor=False, remov=False).newwin
            self.frame = MyFrame(self.window, bg="black", borderwidth=5)
        else:
            self.frame = MyFrame(self.scrn, bg="black", borderwidth=5)
        if not text:
            text = """Welcome to TARTAN Systems,


Please enjoy yourself while you wait ..."""
        self.label = MyLabel(self.frame, text=text, font=("Helvetica", 16),
            justify="center")
        self.label.pack(fill="both", expand="yes")
        if self.scrn:
            self.frame.place(anchor="center", relx=0.5, rely=0.5)
        else:
            self.frame.pack()
        self.frame.update()
        if not self.scrn:
            sw = int(self.window.winfo_screenwidth())
            sh = int(self.window.winfo_screenheight())
            rootsize = tuple(int(_) for _ in self.window.geometry().
                split("+")[0].split("x"))
            x = int(sw / 2) - int(rootsize[0] / 2)
            y = int(sh / 2) - int(rootsize[1] / 2)
            self.window.geometry("%dx%d+%d+%d" % (rootsize + (x, y)))
            if self.window.state() == "withdrawn":
                self.window.deiconify()
        self.refreshSplash()

    def refreshSplash(self, text=None):
        if text:
            self.label.configure(text=text)
        self.frame.update()

    def closeSplash(self):
        self.frame.destroy()
        if not self.scrn:
            self.window.destroy()
        else:
            self.scrn.update()

class Dbase(object):
    """
    A class to manipulate databases.

    rcdic  - A dictionary having at least:
                dbase:    The database interface, SQLite or PgSQL
                dbname:   The database name
                dbhost:   The database host for PgSQL
                dbdir:    The directory containing the database for SQLite
                dbport:   The port to be used for PgSQL
                dbuser:   The database user for PgSQL
                dbpwd:    The database user's password for PgSQL
                wrkdir:   A work directory
    screen - A screen on which to display error messages else None
    """
    def __init__(self, rcdic=None, screen=None):
        if not rcdic:
            self.rcdic = loadRcFile()
        else:
            self.rcdic = rcdic
        check = ["dbase", "dbname", "dbhost", "dbdir", "dbport",
            "dbuser", "dbpwd", "wrkdir"]
        for key in self.rcdic:
            setattr(self, key, self.rcdic[key])
            if key in check:
                check.remove(key)
        self.screen = screen
        try:
            self.dbmod = None
            self.mrecs = True
            if check:
                raise Exception("Invalid rcdic")
            if self.dbase == "PgSQL":
                self.dbmod = "psycopg2"
                import psycopg2 as engine
                try:
                    # Typecast the numeric datatype as Float not Decimal
                    pgx = engine.extensions
                    NUM = pgx.new_type(pgx.DECIMAL.values, "NUM",
                        lambda d, c: float(d) if d is not None else None)
                    pgx.register_type(NUM)
                except Exception as err:
                    raise Exception("Typecast Error (%s)" % err)
                self.dbf = "%s"
            elif self.dbase == "SQLite":
                if self.dbdir:
                    self.dbdsn = os.path.join(self.dbdir, self.dbname)
                else:
                    self.dbdsn = self.dbname
                self.dbmod = "sqlite3"
                from sqlite3 import dbapi2 as engine
                if int(engine.sqlite_version.split(".")[1]) < 8:
                    self.mrecs = False
                self.dbf = "?"
            else:
                raise Exception("Invalid Database Type (%s)" % self.dbase)
            self.err = False
            self.dbopen = False
            self.engine = engine
            if not self.setVariables():
                raise Exception("Variables Error")
        except:
            self.err = True
            showException(self.screen, self.wrkdir,
                "DBAPI Error - Database Module (%s)" % self.dbmod)

    def setVariables(self):
        if self.dbase == "PgSQL":
            self.dtm = "timestamp"
            self.key = "varchar"
            self.var = "varchar"
            self.txt = "text"
            self.itg = "int4"
            self.lng = "int8"
            self.dec = "numeric"
            self.flt = "numeric"
            self.ser = "serial"
            self.blb = "varchar"
        elif self.dbase == "SQLite":
            self.dtm = "varchar"
            self.key = "varchar"
            self.var = "varchar"
            self.txt = "blob"
            self.itg = "integer"
            self.lng = "integer"
            self.dec = "float"
            self.flt = "float"
            self.ser = "integer primary key"
            self.blb = "blob"
        else:
            return
        return True

    def checkDbase(self):
        "Check to see if a database exists"
        err = exists = False
        if self.dbase == "PgSQL":
            err = self.openDbase(dbname="postgres", err=False)
            if not err and self.db:
                self.cu.execute("Select count(*) from pg_catalog."
                    "pg_database where datname='%s'" % self.dbname)
                if self.cu.fetchone()[0]:
                    exists = True
        elif os.path.exists(self.dbdsn) and os.stat(self.dbdsn).st_size:
            err = self.openDbase(err=False)
            if not err and self.db and self.checkTable("ffield"):
                exists = True
            else:
                err = False
        try:
            self.closeDbase()
        except:
            pass
        if err:
            return err
        return exists

    def createDbase(self):
        "Create a database if it does not exist"
        try:
            chk = self.checkDbase()
            if chk not in (True, False):
                raise Exception(chk)
            if chk is True:
                raise Exception("Database Already Exists")
            if self.dbase == "PgSQL":
                self.openDbase(dbname="postgres", auto=True)
                self.cu.execute("Create database %s template template0 "
                    "encoding 'utf-8'" % self.dbname)
                self.cu.execute("Select * from pg_catalog.pg_group "
                    "where groname = '%s'" % self.dbname)
                chk = self.cu.fetchall()
                if not chk:
                    self.cu.execute("Create group %s" % self.dbname)
                self.db.commit()
                self.closeDbase()
            else:
                # Create the empty file
                try:
                    os.umask(0000)
                except:
                    pass
                open(self.dbdsn, "w").close()
                self.openDbase()
                if not self.db:
                    return
                self.closeDbase()
        except:
            showException(self.screen, self.wrkdir, "Database %s" % self.dbname)

    def openDbase(self, dbname=None, dbuser=None, dbpwd=None, auto=False, err=True):
        """
        Open a database using the following:

        dbname - The name of the database to open if not self.dbname
        dbuser - The user of the database to open if not self.dbuser
        dbpwd  - The pwd  of the database to open if not self.dbpwd
        auto   - The isolation level for postgresql
        err    - Whether or not to exit on an error
        """
        if not dbname:
            dbname = self.dbname
        try:
            if self.dbase == "PgSQL":
                dsn = "host=%s dbname=%s" % (self.dbhost, dbname)
                if not int(self.dbport):
                    self.dbport = "5432"
                dsn += " port='%s'" % self.dbport
                if self.dbuser:
                    dsn += " user=%s" % self.dbuser
                if self.dbpwd:
                    dsn += " password=%s" % self.dbpwd
                self.db = self.engine.connect(dsn)
            else:
                self.db = self.engine.connect(database=self.dbdsn)
                self.db.text_factory = str
            self.cu = self.db.cursor()
            self.setAutoCommit(auto)
            self.dbopen = True
        except Exception as e:
            self.db = self.cu = None
            self.dbopen = False
            mess = "Database: %s" % dbname
            showException(self.screen, self.wrkdir, mess)
            if not err:
                return e
            sys.exit()

    def commitDbase(self, ask=False, rback=True, mess=None, default="yes"):
        """
        Commit a transaction using the following:

        ask     - Ask for confirmation to commit
        rback   - Whether to perform a rollback if not committing with ask
        mess    - The message to use for confirmation
        default - The default button to highlight, yes or no
        """
        if not ask:
            self.commit = "yes"
        else:
            if not mess:
                mess = "Would you like to COMMIT All Changes?"
            if self.screen is None:
                scrn = "text"
            else:
                scrn = self.screen
            self.commit = askQuestion(scrn, "Commit?", mess, default=default)
        if self.commit == "yes":
            self.db.commit()
        elif rback:
            self.db.rollback()

    def rollbackDbase(self):
        self.db.rollback()

    def setAutoCommit(self, auto=False):
        if self.dbase == "PgSQL":
            self.db.autocommit = auto
        elif auto:
            self.db.isolation_level = None
        else:
            self.db.isolation_level = ""

    def closeDbase(self):
        try:
            self.cu.close()
        except:
            pass
        try:
            self.db.close()
        except:
            pass
        self.dbopen = False

    def dropDbase(self):
        try:
            if self.dbopen:
                self.closeDbase()
            if self.dbase == "PgSQL":
                self.openDbase(dbname="postgres", auto=True)
                self.cu.execute("Drop database %s" % self.dbname)
                self.cu.execute("Drop group %s" % self.dbname)
                self.closeDbase()
            else:
                os.remove(self.dbdsn)
        except:
            showException(self.screen, self.wrkdir,
                "Database %s" % self.dbname)

    def checkTable(self, table):
        try:
            if self.dbase == "PgSQL":
                self.cu.execute("Select relname from pg_class where "
                    "relname = '%s'" % table)
            else:
                self.cu.execute("Select name from SQLITE_MASTER where "
                    "name = '%s' and type = 'table'" % table)
            return self.cu.fetchone()
        except Exception as err:
            showException(self.screen, self.wrkdir,
                "Error Checking Table (%s)\n\n%s" % (table, err))
            os._exit(1)

    def createTable(self, table, drop=False, index=True):
        if self.checkTable(table):
            if not drop:
                ok = askQuestion(self.screen, "Exists",
                    "Table %s Already Exists, Do You Want to Drop It?" % table,
                    default="no")
                if ok == "no":
                    return
            self.dropTable(table)
        # Create the tables
        fd = []
        for dat in tartanWork.tabdic[table]["fld"]:
            fd.append([table] + copy.copy(dat))
        qry = "Create table %s (" % table
        sequence = ""
        for tab, seq, nam, typ, siz, des, hed in fd:
            siz = float(siz)
            qry += nam
            if typ[0].upper() == "D":
                if typ[1] in ("1", "2"):
                    qry += " %s," % self.itg
                elif typ[1] == "T":
                    qry += " %s," % self.lng
            elif typ[0] in ("H", "N", "U") and typ[1].upper() == "A":
                if siz:
                    qry += " %s(%s)," % (self.var, int(siz))
                else:
                    qry += " %s," % self.var
            elif typ == "ID":
                qry += " %s," % self.lng
            elif typ[0] in ("S", "T", "U"):
                if typ[1].upper() == "D":
                    if typ[0] == "S":
                        siz -= 1
                    qry += " %s(%s)," % (self.dec, str(siz).replace(".", ","))
                elif typ[1] == "L":
                    qry += " %s," % self.lng
                elif typ[1] in ("I", "M"):
                    qry += " %s," % self.itg
                elif typ[1] == "S":
                    if typ[0] == "T":
                        qry += " %s," % self.var
                    else:
                        sequence = nam
                        qry += " %s," % self.ser
                elif typ[1] == "X":
                    qry += " %s," % self.txt
            else:
                showError(self.screen, "Invalid",
                    "Invalid Data Type %s in Table %s, Aborting" % (typ, table))
                return
        qry = qry[:-1] + ")"
        self.cu.execute(qry)
        if index:
            # Create the indexes
            ky = []
            for dat in tartanWork.tabdic[table]["idx"]:
                ky.append([table] + copy.copy(dat))
            for key in ky:
                kk = ""
                for f in key[4:]:
                    if not f:
                        continue
                    if not kk:
                        kk = f
                    else:
                        kk += ",%s" % f
                if key[3] == "U":
                    qry = "Create unique index %s_key%s on %s (%s)" % \
                        (table, str(key[2]), table, kk)
                elif key[3] == "N":
                    qry = "Create index %s_key%s on %s (%s)" % \
                        (table, str(key[2]), table, kk)
                self.cu.execute(qry)
        # Set permissions
        if self.dbase == "PgSQL":
            self.cu.execute("Select * from pg_catalog.pg_group where "
                "groname = '%s'" % self.dbname)
            chk = self.cu.fetchall()
            if not chk:
                self.cu.execute("Create group %s" % self.dbname)
            self.cu.execute("Grant ALL on %s to group %s" %
                (table, self.dbname))
            if sequence:
                self.cu.execute("Grant ALL on %s_%s_seq to group %s" %
                    (table, sequence, self.dbname))

    def populateTable(self, table, commit=True):
        # Populate ftable and ffield tables
        for tab in (("ftable", "ft_tabl"), ("ffield", "ff_tabl")):
            qry = "Delete from %s where %s = '%s'" % (tab[0], tab[1], table)
            self.cu.execute(qry)
        fd = []
        for dat in tartanWork.tabdic[table]["fld"]:
            fd.append([table] + copy.copy(dat))
        qry = "Insert into ffield values("+self.dbf+(",%s" % self.dbf)*6+")"
        for f in fd:
            self.cu.execute(qry, f)
        ky = []
        for dat in tartanWork.tabdic[table]["idx"]:
            ky.append([table] + copy.copy(dat))
        qry = "Insert into ftable values("+self.dbf+(",%s" % self.dbf)*13+")"
        for k in ky:
            for x in range(len(k), 14):
                k.append("")
            self.cu.execute(qry, k)
        if commit:
            self.db.commit()

    def dropTable(self, table, frecs=False):
        if self.checkTable(table):
            self.cu.execute("Drop table %s" % table)
        if frecs:
            self.cu.execute(
                "Delete from ffield where ff_tabl = '%s'" % table)
            self.cu.execute(
                "Delete from ftable where ft_tabl = '%s'" % table)

class Sql(object):
    """
    This class is used to select, insert, update and delete records.
        table_col  = A list of all column names
        table_fld  = A string of all column names comma delimited
        table_dic  = A dictionary of columns having:
                     {"colname": [table, seq, type, size, desc, head]}
    """
    def __init__(self, dbm, tables=None, error=True, prog=None):
        """
        dbm    - The database class
        tables - The tables required for the insRec function
        error  - Whether to show an error if a table is missing
        prog   - The calling module name
        """
        self.dbm = dbm
        if prog:
            self.prog = prog
        else:
            self.prog = "unknown"
        self.error = []
        if not tables:
            return
        elif type(tables) is str:
            self.tables = [tables]
        else:
            self.tables = tables
        for table in self.tables:
            setattr(self, table, table)
            self.dbm.cu.execute("Select * from ffield where ff_tabl='%s' "
                "order by ff_seq" % table)
            cols = self.dbm.cu.fetchall()
            if cols and type(cols) in (list, tuple):
                col = []
                fld = ""
                dic = {}
                for n, c in enumerate(cols):
                    col.append(c[2])
                    if not fld:
                        fld = c[2]
                    else:
                        fld = "%s, %s" % (fld, c[2])
                    dic[c[2]] = [table, c[1], c[3], c[4], c[5], c[6]]
                setattr(self, "%s_col" % table, col)
                setattr(self, "%s_fld" % table, fld)
                setattr(self, "%s_dic" % table, dic)
            else:
                self.error.append(table)
        if error and self.error:
            mess = "The Following Tables are Missing\n\n"
            for tab in self.error:
                mess = "%s%s, " % (mess, tab)
            mess = mess[:-2]
            showError(self.dbm.screen, "Error", mess)
        if self.dbm.checkTable("delrec"):
            self.impdbd = True
        else:
            self.impdbd = False

    def sqlRec(self, state=None, limit=None, fetch=False):
        """
        state = A sql statement. It can also be a tuple of the statement and
                the data required for the clause e.g.
                    "Select abc from xyz where a=1 and b=2" or
                    ("Select abc from xyz where a=? and b=?", (1,2)) or
                    ("Delete from xyz where a=? and b=?", (1,2)) or
                    ("Insert into xyz values (?,?,?)", (1,2,3))
            the ? is dependant on the database and could be %s
        limit = The number of records to return. None = All
        fetch = Only fetch from a previous select. True or False
        """
        def retList(ret):
            # Adjust to return [] instead of None and convert to list(s)
            if not ret:
                return []
            elif limit and limit == 1:
                return list(ret)
            else:
                return [list(r) for r in ret]

        try:
            if not state:
                if fetch and limit:
                    ret = self.dbm.cu.fetchmany(limit)
                    return retList(ret)
                raise Exception("Missing sql statement")
            if type(state) == str:
                comm = state.split()
                self.dbm.cu.execute(state)
            else:
                comm = state[0].split()
                if comm[0].lower() == "insert":
                    if self.dbm.mrecs:
                        self.dbm.cu.execute(state[0], state[1])
                    elif type(state[1][0]) is list and len(state[1]) > 1:
                        self.dbm.cu.executemany(state[0], state[1])
                    else:
                        self.dbm.cu.execute(state[0], state[1][0])
                else:
                    self.dbm.cu.execute(state[0], state[1])
            if comm[0].lower() == "select" or comm[0].lower() == "show":
                if fetch:
                    return
                elif limit and limit == 1:
                    ret = self.dbm.cu.fetchone()
                elif limit:
                    ret = self.dbm.cu.fetchmany(limit)
                else:
                    ret = self.dbm.cu.fetchall()
                return retList(ret)
        except:
            try:
                self.dbm.rollbackDbase()
            except:
                pass
            showException(self.dbm.screen, self.dbm.wrkdir,
                "Error Executing SQL Statement\n\n%s\n\nIn module %s" %
                (str(state), self.prog), dbm=self.dbm)
            os._exit(1)

    def insRec(self, table, data=None, unique=False, dofmt=True, xprt=True, pbar=None):
        """
        table   = The table to insert into
        data    = A list or a list of lists of all the column's data
        unique  = Any column which must be unique e.g. drt_ref1
        dofmt   = Whether or not to format the data before inserting
        xprt    = Blank the export flag
        pbar    = ProgressBar object
        """
        if table not in self.tables:
            showError(None, "insRec Error",
                "Table %s Not In Sql Tables\n\nIn module %s" %
                (table, self.prog))
            self.dbm.rollbackDbase()
            sys.exit()
        # Limits for multiple inserts
        if self.dbm.dbase == "PgSQL":
            lmt = 50000
        else:
            lmt = 500
        # Set Variables
        if table == "crstrn":
            unique = "crt_ref1"
        elif table == "drstrn":
            unique = "drt_ref1"
        elif table == "memtrn":
            unique = "mlt_refno"
        else:
            unique = False
        dic = getattr(self, "%s_dic" % table)
        col = getattr(self, "%s_col" % table)
        xfl = None
        seq = None
        idx = None
        fmt = []
        fld = []
        # Generate column formats
        for n, c in enumerate(col):
            if xprt and c.count("_xflag"):
                xfl = c
                idx = n
            elif dic[c][2] == "US":
                seq = c
                if idx is None:
                    idx = n
                continue
            fmt.append(self.dbm.dbf)
            fld.append(c)
        sfmt = "(%s)" % ",".join(map(str, fmt))
        nfld = "(%s)" % ",".join(map(str, fld))
        # Format data
        self.nfmt = ""
        if type(data[0]) not in (list, tuple):
            # Single row
            data = [data]
        else:
            # Not testing for unique and format with multple rows
            unique = False
            dofmt = False
        # Multiple inserts
        def multiples():
            if type(pbar) in (list, tuple):
                if pbar[0] == "t":
                    pbar[1].update()
                elif pbar[0]:
                    pbar[1].displayProgress(onum)
            elif pbar:
                pbar.displayProgress(onum)
            self.sqlRec(("Insert into %s %s values %s" %
                (table, nfld, self.nfmt), self.ndat))
            self.nfmt = ""
            self.ndat = []
        # Create data
        self.ndat = []
        for onum, odat in enumerate(data):
            try:
                tdat = list(odat)
                if xfl and len(tdat) > idx:
                    dat = tdat[:idx] + [""]
                elif xfl:
                    dat = tdat + [""]
                elif seq and len(tdat) > idx:
                    dat = tdat[:idx]
                else:
                    dat = tdat[:len(col)]
                if dofmt:
                    for n, c in enumerate(fld):
                        chk = CCD(dat[n], dic[c][2], dic[c][3])
                        if chk.err:
                            raise Exception(chk.err)
                        dat[n] = chk.work
            except Exception as err:
                showError(None, "insRec Error",
                    """Data and CSV Formats Differ

(%s)

Table %s in Program %s""" % (err, table, self.prog))
                self.dbm.rollbackDbase()
                os._exit(1)
            # Test and Correct for a Column in Unique Keys (e.g. drt_ref1)
            if unique:
                cols = self.sqlRec(state=("Select * from ftable where "
                    "ft_tabl=%s and ft_type = 'U' order by ft_seq" %
                    self.dbm.dbf, (table,)))
                ok = False
                for key in cols:
                    if unique in key[4:]:
                        ok = True
                        break
                if not ok:
                    showError(None, "insRec Error",
                        "Invalid Unique column %s in table %s" %
                        (unique, table))
                    sys.exit()
                typ = dic[unique][2]
                siz = int(dic[unique][3])
                if typ[1].lower() in ("a", "v", "x"):
                    kkk = (65, 90)
                else:
                    kkk = (49, 90)
                cnt = 0
                while True:
                    c = ["count(*)"]
                    w = []
                    for k in key[4:]:
                        if k:
                            w.append((k, "=", dat[col.index(k)]))
                    chk = self.getRec(table, cols=c, where=w, limit=1)
                    if not chk[0]:
                        break
                    u = dat[col.index(unique)]
                    d = str(u).strip()[:siz - 1] + chr(kkk[cnt])
                    d = CCD(d, typ, siz)
                    dat[col.index(unique)] = d.work
                    cnt += 1
            # Insert records
            if not self.dbm.mrecs:
                # Single record insert
                self.sqlRec(("Insert into %s %s values (%s)" %
                    (table, nfld, sfmt), [dat]))
                continue
            # Multiple record inserts
            if not self.nfmt:
                self.nfmt = sfmt
            else:
                self.nfmt += ", %s" % sfmt
            self.ndat += dat
            if len(self.ndat) > lmt:
                # Limit write to a maximum of lmt records at a time
                # This particularly applies to sqlite
                multiples()
        if self.ndat:
            # Multiple records remainder
            multiples()

    def getRec(self, tables, join=None, cols=None, where=None, group=None, order=None, fetch=False, limit=None):
        """
        tables = A string or list of tables
        join   = A join statement
        cols   = A list of columns to return else returns ALL
        where  = A list of tuples which can comprise a combination of:
                    "(", column name, function, data, [data] (and/or) ...
                    functions: = < > <= >= <> in between

                 Examples
                 --------
                 [("glm_acno", "=", 601)]
                 [("glm_acno", "in", (601, 602))]
                 [("glm_acno", "between", 601, 602)]
                 [("(", "glm_acno", "=", 601, "or", "glm_acno", "=", 602, ")")]

                 Note: "and" is the default and therefore will automatically
                    be used to join tuples of statements, "or" must be
                    implicitly stated as in the example above.

        group  = A string of columns to group by e.g. "drt_cono, drt_acno"
        order  = A string of columns to order by e.g. "drt_cono, drt_acno"
        fetch  = Only fetch records from a previous select. True or False
        limit  = The number of records to return e.g. 1 else returns all
        """
        # check for tables
        if not tables:
            showError(None, "getRec Error",
                "The tables Option Must Exist In Module %s" % self.prog)
            sys.exit()
        # convert tables to list if string
        if type(tables) is str:
            tables = [tables]
        # Start the get statement
        get = "Select"
        # If no cols get all columns
        if not cols:
            if len(tables) > 1:
                showError(None, "getRec Error",
                    "No list of cols but more than 1 table")
                sys.exit()
            cols = getattr(self, "%s_col" % tables[0])
        grpreq = False
        grpcol = ""
        for col in cols:
            tb_chk = []
            ag_chk = False
            as_chk = False
            rd_chk = False
            # Check for sum without round for decimal
            if col.count(" as "):
                as_chk = True
                ts1 = col.split(" as ")
            else:
                ts1 = [col]
            if ts1[0].count("sum(") and not ts1[0].count("round("):
                ts2 = ts1[0].split("(")[1].split(")")[0].split()
                for t in ts2:
                    if t in ("+", "-", "*", "/"):
                        continue
                    for tab in tables:
                        try:
                            dic = getattr(self, "%s_dic" % tab)
                            if t in dic and dic[t][2][1].lower() == "d":
                                rd_chk = str(dic[t][3]).split(".")[1]
                        except:
                            tb_chk.append(tab)
                if rd_chk:
                    col = "round(" + ts1[0] + ", %s)" % rd_chk
                    if as_chk:
                        col = col + " as " + ts1[1]
                if tb_chk:
                    print("Tables %s missing in sql class in module %s" % \
                        (tb_chk, self.prog))
            # Continue building the get statement
            get = "%s %s," % (get, col)
            if as_chk:
                col = col.split(" as ")[1].strip()
            else:
                for agg in ("count(", "sum(", "avg(", "max(", "min("):
                    if col.lower().count(agg):
                        grpreq = True
                        ag_chk = True
                        break
            if not ag_chk:
                if not grpcol:
                    grpcol = col
                else:
                    grpcol = "%s, %s" % (grpcol, col)
        get = "%s from" % get[:-1]
        for table in tables:
            get = "%s %s," % (get, table)
        get = get[:-1]
        if join:
            get = "%s %s" % (get, join)
        if where:
            # Selection of records
            whr, dat = self.getSqlWhere(where)
        else:
            # All records
            whr = None
            dat = None
        if whr:
            get = "%s %s" % (get, whr)
        if grpreq and grpcol:
            if not group:
                group = grpcol
            else:
                g1 = group.split()
                gr = []
                for g in g1:
                    gr.append(g.replace(",", ""))
        if group:
            get = "%s group by %s" % (get, group)
        if order:
            get = "%s order by %s" % (get, order)
        if not dat:
            return self.sqlRec(get, fetch=fetch, limit=limit)
        else:
            return self.sqlRec((get, dat), fetch=fetch, limit=limit)

    def updRec(self, table, where=None, data=None, cols=None, dofmt=True):
        """
        table   = The table to update
        where   = A where statement like the getRec one
        data    = A list of data matching the columns in cols or all if no cols
        cols    = A list of columns matching 'data' else None for all data cols
        dofmt   = Whether or not to format the data before updating
        """
        if not where and not data:
            showError(None, "updRec Error", "Missing where and list of data")
            sys.exit()
        if table not in self.tables:
            showError(None, "updRec Error",
                "Table %s Not In Sql Tables\n\nIn module %s" %
                (table, self.prog))
            self.dbm.rollbackDbase()
            sys.exit()
        cmd = ""
        ccc = getattr(self, "%s_col" % table)
        pfx = ccc[0].split("_")[0]
        xfl = "%s_xflag" % pfx
        if self.impdbd and xfl in ccc:
            # Exported database, delete old record and insert new one
            if where:
                # Selected records
                recs = self.getRec(table, where=where)
            else:
                # All records
                recs = self.getRec(table)
            for rec in recs:
                dat = rec[:]
                self.delRec(table, data=rec)
                for num, col in enumerate(ccc):
                    if col == xfl:
                        dat[num] = ""
                    elif cols and col in cols:
                        dat[num] = data[cols.index(col)]
                    else:
                        dat[num] = rec[num]
                self.insRec(table, data=dat)
                return
        if not cols:
            cols = ccc[:]
        if len(data) != len(cols):
            showError(None, "updRec Error",
                "Length of Data not equal to length of Cols (%s)" % table)
            sys.exit()
        dic = getattr(self, "%s_dic" % table)
        for num, col in enumerate(cols):
            if not cmd:
                cmd = "Update %s set %s = %s" % (table, col, self.dbm.dbf)
            else:
                cmd = "%s, %s = %s" % (cmd, col, self.dbm.dbf)
            if dofmt:
                dat = CCD(data[num], dic[col][2], dic[col][3])
                if dat.err:
                    showError(None, "updRec Error",
                        "Invalid Data for %s in table %s\n\n%s" %
                        (dic[col][4], table, dat.err))
                    sys.exit()
                data[num] = dat.work
        dat = copyList(data)
        if where:
            whr, ext = self.getSqlWhere(where)
            dat.extend(ext)
            cmd = "%s %s" % (cmd, whr)
        self.sqlRec((cmd, tuple(dat)))

    def delRec(self, table, cols=None, data=None, where=None):
        """
        table = The table to delete from
        cols  = A list of columns matching 'data' else None for all
        data  = A list of data matching the columns in cols or all columns
        where = A where statement if no data and cols, as per getRec
        """
        if table not in self.tables:
            showError(None, "delRec Error",
                "Table %s Not In Sql Tables\n\nIn module %s" %
                (table, self.prog))
            self.dbm.rollbackDbase()
            sys.exit()
        ccc = getattr(self, "%s_col" % table)
        pfx = ccc[0].split("_")[0]
        xfl = "%s_xflag" % pfx
        if where:
            whr, dat = self.getSqlWhere(where)
            cmd = ("Delete from %s %s" % (table, whr), tuple(dat))
            ddd = ("Select * from %s %s" % (table, whr), tuple(dat))
        elif data:
            # Delete where the data matches
            if not cols:
                cols = ccc[:]
            if len(data) != len(cols):
                showError(None, "delRec Error",
                    "Length of Data not equal to length of Cols")
                sys.exit()
            cmd = ""
            ddd = ""
            for col in cols:
                if not cmd:
                    cmd = "Delete from %s where %s = %s" % (table, col,
                        self.dbm.dbf)
                    ddd = "Select * from %s where %s = %s" % (table, col,
                        self.dbm.dbf)
                else:
                    cmd = "%s and %s = %s" % (cmd, col, self.dbm.dbf)
                    ddd = "%s and %s = %s" % (ddd, col, self.dbm.dbf)
            cmd = (cmd, tuple(data))
            ddd = (ddd, tuple(data))
        else:
            # Delete all records
            cmd = "Delete from %s" % table
            ddd = "Select * from %s" % table
        if self.impdbd and xfl in ccc:
            # Exported database, insert records into delrec table
            recs = self.sqlRec(ddd)
            for rec in recs:
                if rec[ccc.index(xfl)] == "Y":
                    # Exported record
                    self.sqlRec(("Insert into delrec values (%s, %s)" %
                        (self.dbm.dbf, self.dbm.dbf), (table, str(rec))))
        self.sqlRec(cmd)

    def getSqlWhere(self, where):
        whr = ""
        dat = []
        for w in where:
            if len(w) == 1:
                if not whr:
                    whr = "where %s" % w[0]
                else:
                    whr = "%s and %s" % (whr, w[0])
                continue
            flag = None
            for c in w:
                if not flag:
                    if not whr:
                        whr = "where %s" % c
                    else:
                        whr = "%s and %s" % (whr, c)
                    if c in ("(", ")"):
                        flag = "a"
                    else:
                        flag = "f"
                elif flag == "a":
                    whr = "%s %s" % (whr, c)
                    flag = "f"
                elif flag == "f":
                    if c.lower() == "ilike" and self.dbm.dbase != "PgSQL":
                        c = "like"
                    whr = "%s %s" % (whr, c)
                    if c.lower() == "between":
                        flag = "b1"
                    elif c.lower() == "not":
                        continue
                    elif c.lower() in ("in", "not in"):
                        flag = "i"
                    else:
                        flag = "b0"
                elif flag in ("b0", "b1"):
                    whr = "%s %s" % (whr, self.dbm.dbf)
                    if type(c) == float:
                        dat.append(str(c))
                    else:
                        dat.append(c)
                    if flag == "b0":
                        flag = "n"
                    else:
                        flag = "b2"
                elif flag == "b2":
                    whr = "%s and %s" % (whr, self.dbm.dbf)
                    if type(c) == float:
                        dat.append(str(c))
                    else:
                        dat.append(c)
                    flag = "n"
                elif flag == "i":
                    whr = "%s (" % whr
                    for x in c:
                        whr = "%s %s," % (whr, self.dbm.dbf)
                        if type(c) == float:
                            dat.append(str(x))
                        else:
                            dat.append(x)
                    whr = whr[:-1] + ")"
                    flag = "n"
                elif flag in ("n", "x"):
                    whr = "%s %s" % (whr, c)
                    if c in ("(", ")"):
                        continue
                    if flag == "n":
                        flag = "x"
                    else:
                        flag = "f"
        if self.dbm.dbase == "PgSQL":
            whr = whr.replace(" % ", " %% ")
        return whr, dat

class CCD(object):
    """
    Data Types
    ============================================================================
    CB = CheckButton Box
    CD = Currency Decimal Signed i.e. thousands separated by comma
    CI = Currency Integer Signed i.e. thousands separated by comma
    D1 = Date (CCYYMMDD)
    D2 = Current Period Date (CCYYMM)
    DT = Date and Time (DD-MMM-YYYY-HH:MM)
    FF = File or Directory Format
    HA = Normal Alphanumeric Hidden as for Passwords
    ID = Identity Number
    LA = Lower Case Alphanumeric (Must also have 'size' set)
    La = Lower Case Alphanumeric Right Justified (Must also have 'size' set)
    MD = Month and Day (MMDD)
    NA = Normal Alphanumeric (Must also have 'size' set)
    Na = Normal Alphanumeric Right Justified (Must also have 'size' set)
    RB = RadioButton Box
    RW = Raw Data, No formatting or checking
    SD = Signed Decimal (size includes sign and decimals)
    SI = Signed Integer (size includes sign)
    SL = Signed Long Integer (size includes sign)
    Sd = Convert Integer to Signed Decimal (size includes sign and decimals)
    TM = Time (HH:MM)
    TS = Timestamp (CCYY-MM-DD HH:MM:SS)
    TV = Text View - This is only used with TartanDialog for input purposes
    TX = Text or Blob Alphanumeric Variable Length - Storing Text Views etc.
    Tv = Text View - This is only used with TartanDialog for input purposes
    UA = Upper Case Alphanumeric (Must also have 'size' set)
    UD = Unsigned Decimal (size includes decimals)
    UI = Unsigned Integer
    UL = Unsigned Long Integer
    US = Unsigned Serial
    Ua = Upper Case Alphanumeric Right Justified (Must also have 'size' set)
    Ud = Convert Integer to Unsigned Decimal (size includes decimals)
    d1 = Date (CCYYMMDD) Alowing Zero
    d2 = Current Period Date (CCYYMM) Allowing Zero
    tM = Time (HH:MM) Allowing Zero
    """
    def __init__(self, data, types, size=0):
        self.mthnum = {
            "JAN":1, "FEB":2, "MAR":3, "APR":4, "MAY":5, "JUN":6,
            "JUL":7, "AUG":8, "SEP":9, "OCT":10, "NOV":11, "DEC":12}
        # ======================================================================
        # Work around data types from TartanDialog which are 3 characters long
        # ======================================================================
        if len(types) == 3:
            types = types[1:]
        # ======================================================================
        # Check and Radio Button Boxes, Raw Data Types and Timestamp
        # ======================================================================
        if types == "TS" and len(str(data)) == 14:
            self.err = ""
            data = self.work = str(data)
            yr = int(data[:4])
            mt = int(data[4:6])
            dy = int(data[6:8])
            hr = int(data[8:10])
            mm = int(data[10:12])
            ss = int(data[12:14])
            self.disp = "%4i-%02i-%02i %02i:%02i:%02i" % (yr,mt,dy,hr,mm,ss)
            return
        elif types in ("CB", "RB", "RW", "TS"):
            self.err = ""
            self.work = self.disp = self.data = data
            return
        # ======================================================================
        # Work around for None
        # ======================================================================
        if data is None:
            if types[1].lower() in ("a", "v", "w", "x"):
                self.data = ""
            else:
                self.data = "0"
        # ======================================================================
        # Work around for previous format error
        # ======================================================================
        elif types[0] != "T" and types[1].lower() != "a" and data == "None":
            self.data = "0"
        # ======================================================================
        elif types[1].lower() in ("a", "v", "x"):
            if type(data) is str:
                self.data = data.rstrip().replace("\\", "/")
            else:
                self.data = str(data)
        else:
            self.data = str(data).strip()
        self.types = types
        self.size = size
        if types in ("D1","d1","D2","d2","DT","dT","DM","dM","TM","tM"):
            if self.data[-1:] == "L":
                self.data = self.data[:-1]
        elif types in ("Sd", "Ud"):
            self.data = str(float(int(self.data) / 100.0))
        elif types[1] in ("I","L"):
            self.data = self.data.split(".")[0]
            if self.data[-1:] == "L":
                self.data = self.data[:-1]
        self.processData()

    def processData(self):
        self.quo = int(self.size)
        self.rem = int(round((self.size % 1) * 10))
        self.disp = None
        self.work = None
        self.err = ""
        if self.types in ("D1","d1"):
            self.chkD1_Date()
        elif self.types in ("D2","d2"):
            self.chkD2_Date()
        elif self.types in ("DM","dM"):
            self.chkDM_Date()
        elif self.types in ("DT","dT"):
            self.chkDT_Date()
        elif self.types in ("TM","tM"):
            self.chkTM_Time()
        elif self.types[1] in ("A","a"):
            dat = self.data[:self.quo]
            if self.types[1] == "A":
                if not dat:
                    self.work = dat
                    self.disp = " " * self.quo
                else:
                    if self.types[0] == "L":
                        dat = dat.lower()
                    elif self.types[0] == "U":
                        dat = dat.upper()
                    self.work = dat
                    self.disp = dat.ljust(self.quo)
            elif self.types[1] == "a":
                try:
                    if dat[-1:] == "L":
                        raise Exception
                    dat = str(int(dat))
                    if dat[-1:] == "L":
                        dat = dat[:-1]
                    self.work = dat.rjust(self.quo)
                    self.disp = dat.rjust(self.quo)
                except:
                    if self.types[0] == "L":
                        dat = dat.lower()
                    elif self.types[0] == "U":
                        dat = dat.upper()
                    self.work = dat
                    self.disp = dat.ljust(self.quo)
        elif self.types == "UI":
            self.unsignedInteger()
        elif self.types == "US":
            self.unsignedInteger()
        elif self.types in ("CI", "SI"):
            self.signedInteger()
        elif self.types in ("ID", "UL"):
            self.unsignedLong()
        elif self.types == "SL":
            self.signedLong()
        elif self.types in ("UD", "Ud"):
            self.unsignedDecimal()
        elif self.types in ("CD", "SD", "Sd"):
            self.signedDecimal()
        elif self.types == "FF":
            self.pathFormat()
        elif self.types in ("TV", "Tv", "TX"):
            self.work = self.data.strip()
            self.disp = self.data.strip()
        else:
            self.err = "Invalid Input, Retry (%s)" % self.types

    def chkD1_Date(self):
        date = self.data.replace("-", "").replace("/", "")
        if date in ("", "0", "00000000") and (self.types == "d1"):
            self.work = 0
            self.disp = "0000-00-00"
            return
        try:
            if len(date) < 5 or len(date) > 8:
                raise Exception
            yy = int(date[0:(len(date) - 4)])
            mm = int(date[(len(date) - 4):(len(date) - 2)])
            dd = int(date[(len(date) - 2):len(date)])
            if 100 > yy > 30:
                yy += 1900
            elif yy < 100:
                yy += 2000
            datetime.datetime(yy, mm, dd)
            self.work = (yy * 10000) + (mm * 100) + dd
            self.disp = "%04i-%02i-%02i" % (yy, mm, dd)
            if len(self.disp) < self.size:
                self.disp = self.disp + (
                    " " * (int(self.size / 1) - len(self.disp)))
        except:
            self.err = "Invalid Date (D1)"

    def chkD2_Date(self):
        date = self.data.replace("-", "").replace("/", "")
        if date in ("", "0", "000000") and (self.types == "d2"):
            self.work = 0
            self.disp = "0000-00"
            return
        try:
            if len(date) < 3 or len(date) > 6:
                raise Exception
            yy = int(date[0:(len(date) - 2)])
            mm = int(date[(len(date) - 2):len(date)])
            if 100 > yy > 30:
                yy += 1900
            elif yy < 100:
                yy += 2000
            datetime.datetime(yy, mm, 1)
            self.work = (yy * 100) + mm
            self.disp = "%04i-%02i" % (yy, mm)
            if len(self.disp) < self.size:
                self.disp = self.disp + (
                    " " * (int(self.size / 1) - len(self.disp)))
        except:
            self.err = "Invalid Current Date (D2)"

    def chkDM_Date(self):
        date = self.data.replace("-", "")
        if date in ("", "0", "0000") and (self.types == "dm"):
            self.work = 0
            self.disp = "00-00"
            return
        try:
            if len(date) < 3 or len(date) > 4:
                raise Exception
            mm = int(date) / 100
            dd = int(date) % 100
            self.work = int(date)
            self.disp = "%02i-%02i" % (mm, dd)
        except:
            self.err = "Invalid Current Date (DM)"

    def chkDT_Date(self):
        date = self.data.replace(" ", "").replace("-", "").replace(":", "")
        if date in ("", "0", "000000000000") and (self.types == "dT"):
            self.work = 0
            self.disp = "0000-00-00 00:00"
            return
        try:
            if len(date) < 9 or len(date) > 12:
                raise Exception
            yy = int(date[0:(len(date) - 8)])
            mm = int(date[(len(date) - 8):(len(date) - 6)])
            dd = int(date[(len(date) - 6):(len(date) - 4)])
            hh = int(date[(len(date) - 4):(len(date) - 2)])
            tt = int(date[(len(date) - 2):len(date)])
            if 100 > yy > 30:
                yy += 1900
            elif yy < 100:
                yy += 2000
            datetime.datetime(yy, mm, dd, hh, tt)
            self.work = (yy*100000000)+(mm*1000000)+(dd*10000)+(hh*100)+tt
            self.disp = "%04i-%02i-%02i %02i:%02i" % (yy, mm, dd, hh, tt)
            if len(self.disp) < self.size:
                self.disp = self.disp + (
                    " " * (int(self.size / 1) - len(self.disp)))
        except:
            self.err = "Invalid Date (DT)"

    def chkTM_Time(self):
        tim = self.data.replace(":", "").replace(".", "")
        if tim in ("", "0", "0000") and (self.types == "tM"):
            self.work = 0
            self.disp = "00:00"
            return
        if len(tim) > 4:
            # Seconds
            tim = tim[:-2]
        try:
            if len(tim) < 1:
                raise Exception
            tim = int(tim)
            hh = int(tim / 100)
            mm = tim % 100
            self.work = (hh*100)+mm
            self.disp = "%02i:%02i" % (hh, mm)
            if len(self.disp) < self.size:
                self.disp = self.disp + (
                    " " * (int(self.size / 1) - len(self.disp)))
        except:
            self.err = "Invalid Time (%s)" % self.types

    def unsignedInteger(self):
        if not self.data:
            self.data = "0"
        try:
            if self.data.count("-"):
                self.err = "Invalid UnSigned Integer (UI-)"
            else:
                self.work = int(self.data)
                self.disp = self.data.rjust(self.quo)
                if len(self.disp) > self.quo:
                    self.err = "Invalid UnSigned Integer (UIs)"
        except:
            self.err = "Invalid UnSigned Integer (UI)"

    def signedInteger(self):
        if not self.data:
            self.data = "0"
        try:
            s = self.data.find("-")
            if s == (len(self.data) - 1):
                self.data = self.data.replace("-", "")
                self.data = "-" + self.data
            self.work = int(self.data)
            if self.types == "CI":
                self.data = "{:,}".format(self.work)
            s = self.data.find("-")
            if s != -1:
                self.data = self.data.replace("-", "")
                self.data = self.data + "-"
            else:
                self.data = self.data + " "
            self.disp = self.data.rjust(self.quo)
            if len(self.disp) > self.quo:
                self.err = "Invalid Signed Integer (SIs)"
        except:
            self.err = "Invalid Signed Integer (SI)"

    def unsignedLong(self):
        if not self.data:
            self.data = "0"
        try:
            if self.data.count("-"):
                self.err = "Invalid UnSigned Long (UL-)"
            else:
                self.work = int(self.data)
                self.disp = self.data.rjust(self.quo)
                if len(self.disp) > self.quo:
                    self.err = "Invalid UnSigned Long (ULs)"
        except:
            self.err = "Invalid UnSigned Long (UL)"

    def signedLong(self):
        if not self.data:
            self.data = "0"
        try:
            s = self.data.find("-")
            if s == (len(self.data) - 1):
                self.data = self.data.replace("-", "")
                self.data = "-" + self.data
            self.work = int(self.data)
            s = self.data.find("-")
            if s != -1:
                self.data = self.data.replace("-", "")
                self.data = self.data + "-"
            else:
                self.data = self.data + " "
            self.disp = self.data.rjust(self.quo)
            if len(self.disp) > self.quo:
                self.err = "Invalid UnSigned Long (ULs)"
        except:
            self.err = "Invalid Signed Long (SL)"

    def unsignedDecimal(self):
        if not self.data:
            self.data = "0.0"
        try:
            if self.data.count("-"):
                self.err = "Invalid UnSigned Decimal (UD-)"
            else:
                self.work = round(float(self.data), self.rem)
                self.data = str(self.work)
                dec = self.data.split(".")
                if len(dec[0]) > self.quo - (self.rem + 1):
                    self.err = "Invalid Unsigned Decimal (UDs)"
                elif len(dec[1]) > self.rem:
                    self.err = "Invalid Unsigned Decimal (UDd)"
                else:
                    self.data = self.data + ((self.rem - len(dec[1]))*"0")
                    self.disp = self.data.rjust(self.quo)
                    d = len(self.disp) - self.quo
                    if d > 0:
                        self.disp = self.disp[d:]
        except:
            self.err = "Invalid UnSigned Decimal (UD)"

    def signedDecimal(self):
        if not self.data:
            self.data = "0.0"
        try:
            if self.data[-1:] == "-":
                self.data = "-" + self.data[:-1]
            self.work = round(float(self.data), self.rem)
            if not self.work:
                self.work = 0.0
            if self.types == "SD":
                self.data = str(self.work)
            else:
                self.data = "{:,}".format(self.work)
            dec = self.data.split(".")
            if len(dec[0]) > self.quo - (self.rem + 1):
                self.err = "Invalid Signed Decimal (%ss)" % self.types
            elif len(dec[1]) > self.rem:
                self.err = "Invalid Signed Decimal (%sd)" % self.types
            else:
                self.data = self.data + ((self.rem - len(dec[1]))*"0")
                if self.work < 0:
                    self.data = self.data.replace("-", "")
                    self.data = self.data + "-"
                else:
                    self.data = self.data + " "
                self.disp = self.data.rjust(self.quo)
                d = len(self.disp) - self.quo
                if d > 0:
                    self.disp = self.disp[d:]
        except:
            self.err = "Invalid Signed Decimal (%s)" % self.types

    def pathFormat(self):
        if not self.data:
            self.work = self.disp = self.data
            return
        try:
            if self.data.count("/"):
                data = self.data.split("/")
            else:
                data = [self.data]
            work = []
            for dat in data:
                if dat.count("\\"):
                    work.extend(dat.split("\\"))
                else:
                    work.append(dat)
            if sys.platform == "win32":
                if work[0].count(":"):
                    drive = work.pop(0).upper()
                else:
                    drive = os.path.splitdrive(os.getcwd())[0].upper()
            else:
                if not work[0]:
                    del work[0]
                    drive = os.sep
                else:
                    drive = os.getcwd()
            if sys.platform == "win32":
                data = os.path.join(drive, os.sep, *work)
            else:
                data = os.path.join(drive, *work)
            self.work = self.disp = os.path.normpath(data)
            if not os.path.isabs(self.work):
                raise Exception
        except:
            self.err = "Invalid Path (%s)" % self.types

class ASD(object):
    """
    Used for Adding and Subtracting Floats as follows:

    answer = float(ASD(amount1) + ASD(amount2))
    answer = float(ASD(amount1) - ASD(amount2))
    """
    def __init__(self, amount, decimals=2):
        self.decimals = decimals
        self.multiply = 1.0
        for x in range(decimals):
            self.multiply = self.multiply * 10
        self.amt = int(round((float(amount) * self.multiply), 0))

    def __add__(self, other):
        return ASD((self.amt + other.amt) / self.multiply)

    def __sub__(self, other):
        return ASD((self.amt - other.amt) / self.multiply)

    def __float__(self):
        return(round((self.amt / self.multiply), self.decimals))

class DBCreate(object):
    """
    This is used to setup a database and create necessary tables.
    The command line options are as follows:

    -c  = Do as Follows:
            i = Initial setup, i.e. create the database as well
            f = Create ftable and ffield Only
            p = Create and Populate ftable and ffield Only, needs -t
            t = Create ftable, ffield and Tables (Default), needs -t
            g = Grant Privileges to group for all tables.
            u = Create Database User Only
    -d  = Drop existing database, needs -ci
    -i  = Create indexes, defaults to True
    -l  = Display a splash screen on this frame
    -p  = Admin Password
    -t  = String of comma separated Table Names
    -u  = Admin User
    -v  = Tartan Version if -ci
    -w  = Commit transactions, True/False (True)
    -x  = Close the database on completion, True/False (True)
    """
    def __init__(self, dbm=None, opts=None):
        if not dbm:
            sys.exit()
        else:
            self.dbm = dbm
        for dflt in [("put", "t"), ("usr", ""), ("pwd", ""), ("tab", None),
                ("cmt", True), ("xit", True), ("spl", False), ("idx", True),
                ("ver", None), ("drp", False)]:
            setattr(self, dflt[0], dflt[1])
        if opts:
            for o, v in opts:
                if o == "-c":
                    self.put = v.lower()
                if o == "-d":
                    self.drp = True
                elif o == "-i":
                    self.idx = v
                elif o == "-l":
                    self.spl = v
                elif o == "-p":
                    self.pwd = v
                elif o == "-t":
                    self.tab = v.split(",")
                elif o == "-u":
                    self.usr = v
                elif o == "-v":
                    self.ver = v
                elif o == "-w":
                    self.cmt = v
                elif o == "-x":
                    self.xit = v
        if self.drp and self.put != "i":
            self.drp = False
        self.setVariables()
        if self.doCreate():
            if self.ver:
                t = time.localtime()
                sysdt = ((t[0] * 10000) + (t[1] * 100) + t[2])
                self.sql.sqlRec("Insert into verupd values('%s', %s)" %
                    (self.ver, sysdt))
            if self.dbm.dbase == "SQLite":
                self.dbm.commitDbase()
                self.dbm.cu.execute("PRAGMA JOURNAL_MODE=DELETE")
                self.dbm.cu.execute("PRAGMA SYNCHRONOUS=FULL")
            if self.cmt:
                self.dbm.commitDbase()
        if self.xit:
            self.dbm.closeDbase()

    def setVariables(self):
        if self.put != "i" and not self.dbm.dbopen:
            self.dbm.openDbase()
        self.sql = Sql(self.dbm)

    def doCreate(self):
        if self.tab:
            chk = self.doCheckTables(self.tab)
            if chk:
                print("Details for Table(s) %s Do(es) Not Exist" % chk)
                return False
        if self.put == "i":
            chk = self.dbm.checkDbase()
            if chk not in (True, False):
                os._exit(1)
            elif chk is True:
                if self.drp:
                    answer = "yes"
                else:
                    answer = askQuestion(self.spl, "Drop Database?",
                        "The Database Already Exists, Do You Want to Drop It?",
                        default="no")
                if answer == "yes":
                    self.dbm.dropDbase()
                else:
                    return False
            self.dbm.createDbase()
            self.dbm.openDbase(dbuser=self.usr, dbpwd=self.pwd, auto=True)
        if self.dbm.dbase == "SQLite":
            self.dbm.commitDbase()
            self.dbm.cu.execute("PRAGMA JOURNAL_MODE=OFF")
            self.dbm.cu.execute("PRAGMA SYNCHRONOUS=OFF")
        if self.spl:
            sp = SplashScreen(self.spl,
                "Creating Format and Data Tables\n\nPlease Wait...")
        if self.put == "i":
            self.createUser(self.usr, self.pwd)
        if self.put in ("i", "f"):
            for table in ("ffield", "ftable"):
                self.dbm.createTable(table, drop=True, index=self.idx)
            for table in ("ffield", "ftable"):
                self.dbm.populateTable(table, commit=False)
        elif self.put == "g" and self.dbm.dbase == "PqSQL":
            tbs = self.sql.sqlRec("Select ft_tabl from ftable group by ft_tabl")
            for tb in tbs:
                self.sql.sqlRec("Grant ALL on %s to group %s" % (tb[0],
                    self.dbm.dbname))
                seq = self.sql.sqlRec("Select ff_name from ffield where "\
                    "ff_tabl = '%s' and ff_type = 'US'" % tb[0])
                if seq:
                    self.sql.sqlRec("Grant ALL on %s_%s_seq to group %s" %
                        (tb[0], seq[0][0], self.dbm.dbname))
        elif self.put == "u":
            self.createUser(self.newusr, self.newpwd)
        if self.put in ("i", "p", "t"):
            alltab = []
            for tab in list(tartanWork.tabdic.keys()):
                if tab not in ("ffield", "ftable"):
                    alltab.append(tab)
            if self.tab:
                tables = []
                for tab in self.tab:
                    if tab.strip() in alltab:
                        tables.append(tab.strip())
            else:
                tables = alltab
            if not tables:
                if self.spl:
                    sp.closeSplash()
                showError(None, "Error",
                    "No Valid Tables Specified (%s)" % self.tab)
                return True
            if self.spl:
                sp.closeSplash()
                pg = ProgressBar(self.spl, mxs=len(tables),
                    typ="Creating & Populating Tables ... Please Wait")
            for num, table in enumerate(tables):
                if self.spl:
                    pg.displayProgress(num)
                if self.put in ("i", "t"):
                    self.dbm.createTable(table, drop=True, index=self.idx)
                self.dbm.populateTable(table, commit=False)
                if table == "ctlpwu" and self.put == "i":
                    sql = Sql(self.dbm, table, prog=__name__)
                    sql.insRec(table, data=["admin", "System Administrator",
                        "", "", "", 0, "", "", 9])
            if self.spl:
                pg.closeProgress()
        elif self.spl:
            sp.closeSplash()
        return True

    def doCheckTables(self, tables):
        for tab in tables:
            if tab not in tartanWork.tabdic:
                return tab

    def createUser(self, user, pwd):
        try:
            if self.dbm.dbase == "PgSQL":
                use = self.sql.sqlRec(("Select usename from pg_user where "\
                    "usename = %s", (user,)), limit=1)
                if not use:
                    if pwd:
                        cmd = "Create user %s with password '%s' "\
                            "in group %s" % (user, pwd, self.dbm.dbname)
                    else:
                        cmd = "Create user %s "\
                            "in group %s" % (user, self.dbm.dbname)
                    self.sql.sqlRec(cmd)
        except:
            showException(self.spl, self.dbm.wrkdir,
                "(Error Creating User %s)" % user, dbm=self.dbm)
            sys.exit()

class TartanDialog(object):
    """
    Dialog for TARTAN Financial Systems

    Use this class to create a Dialog Screen by passing the folowing
    parameters to it:

    mf     =  The mainframe object

    screen =  The frame to place the widget on, defaults to mf

    tops   =  True to display a Top floating window, defaults to False

    title  =  The list/tuple of titles of the capture screen, None for None

    tags   =  A list/tuple of tuples of tags on the notebook pages, if
              applicable, having:
                name    = the tag name
                comd    = the command to execute when enabling else None
                on      = Tuple showing when to enable the tag (frt, pag, pos),
                          this can also be tuples of tuples,
                          for the end of a Frame use (frt, pag, 0)
                off     = Tuple showing when to disable the tag (frt, pag, pos),
                          this can also be tuples of tuples,
                          for the end of a Frame use (frt, pag, 0)
                unl     = Whether to have an underlined letter (default True)

    eflds  =  A list/tuple of lists/tuples of entry fields to place in a frame
              or notebook page containing the following:
                place   = (where,page,row,col) to place the field as follows:
                            where   = T for top and C for Array of Columns
                            page    = 0   = On the Base Frame or
                                      1-9 = On a notebook page 1-9
                            row     = The row
                            col     = The column
                            ent     = The entry field start pos (Only top)
                type    = 3 fields as follows:
                            first   = (I)nput, (O)utput
                            second  = as per TartanClasses CCD class
                                      e.g. N,U,S,U,H for normal, upper,
                                      signed, unsigned or Hidden
                                                       or
                                      @ to use a column from the ffield table
                            third   = as per TartanClasses class e.g. A,I,L,D
                                      for alpha, integer, long or decimal
                                                       or
                                      the column name in the ffield table
                            Note: If a field from ffield is used then the type
                                  will be taken from the ffield data and the
                                  following fields will default to the ffield
                                  values:
                                      size
                                      text
                                      prompt
                          OR
                            ("[I,O]CB","xyz") for Checkbutton where:
                                I,O = (I)nput or (O)utput
                                CB  = Checkbutton indicator
                                xyz = the text of the button
                          OR
                            ("[I,O]RB",(xyz)) for Radiobutton where:
                                I,O = (I)nput or (O)utput
                                RB  = Radiobutton indicator
                                xyz = the list or tuple of options
                          OR
                            [I,O][TV,Tv] for TextViews where:
                                I,O = (I)nput or (O)utput
                                TV= Textview F9 terminating
                                Tv= Textview Return terminating
                size    = The size of the entry field in characters as follows:
                            Entry Fields:
                                Either a tuple of (display size, entry size) or
                                a single value denoting both sizes. The sizes
                                are either an integer or float where the
                                quotient is the total display size
                                e.g. 13.2 means the display size, including the
                                sign, would be 13 chrs.  i.e. 999999999.99-
                            Textview Fields:
                                Either a tuple of (display size, entry height)
                                or a single value denoting display size with a
                                height of 1 line.
                text    = Heading or Label
                prompt  = The entry field prompt. Append '(noesc)' to ignore
                          the Escape key i.e. No Backup Field or Exit Frame.
                default = Only if 'where' above = C:
                            r to repeat the previous line's column
                            i to increment the previous line's Integer
                            p to increment the previous line's Period
                          Else if 'where' above = T:
                            Normal value, '' for no default
                          Else if type above = RNV:
                            The String or Integer
                clear   = Clear the field when escaping to back up Y/N
                cr      = The function to perform for the <CR> key else None
                          It must return one of the following (default = None):
                            'bu'          = Back up to the previous field
                            'cl'          = Clear line and focus on 1st column
                            'ff[1-99]'    = Force focus to fieldnumber [1-99]
                            'nf'          = Do not auto focus on next field
                            'nd'          = End the frame normally
                            'nc'          = End the frame with no confirm
                            'ok' or None  = All fine
                            'rf'          = Re-focus on same field
                            'rt'          = Return True
                            'sk[1-99]'    = Skip the next [1-99] fields
                            'text'        = An error message
                            'xt'          = Exit the frame
                f1      = The record selection function to perform as:
                            a_name = {
                                 'stype':  "R",
                                 'title':  Selection Title
                                 'tables': ('genbal',),
                                 'cols':   name, type size heading e.g.
                                           (('glo_acno', '', 0, 'Acc-Num'),
                                            ('glo_trdt', '', 0, 'Date'),
                                            ('glo_cyr', '', 0, 'Balance')),
                                           if type is a tuple ("XX", gltrtp)
                                           then display XXX type from list.
                                 'extra':  A tuple of tuples of additional
                                           data e,g. ((0, "All Companies"),)
                                 'where':  'glm_cono=%s' % self.conum,
                                 'whera':  [['T', 'glo_acno', 0, 0]],
                                              (the type of frame T or C,
                                               the field name,
                                               the column number,
                                               the column page)
                                 'autoc':  True/False to override default
                                 'group':  The fields to group on
                                 'order':  The fields to order on
                                 'wtype':  The where is a select statement
                                 'index':  The index of the item to return
                                 'comnd':  The command to execute after
                                           the select statement exits
                                 'sort':   If the columns can be sorted,
                                 'size':   The default screen size,
                                 'butt':   Additional Buttons and functions}
                     OR = A selection of single records
                            a_name = {
                                 'stype':  "S",
                                 'title':  Selection Title
                                 'head':   A list of column headings
                                 'tables': ('genbal',),
                                            The table(s) from which to return
                                            the records.  In the case of
                                            multiple tables, only the first
                                            table's columns will be returned,
                                            all other tables must only be used
                                            for the where statement
                                 'cols':   A list/tuple of columns to display
                                           ('glo_acno', 'glo_trdt', 'glo_cyr')
                                 'where':  A list of where conditions else all
                                           records or a list of data to use in
                                           the selection
                                 'group':  True or False
                                 'order':  The fields to order on
                                 'selcol': The column for keyboard selection
                                 'ttype':  T = The where above is a list of
                                               where conditions
                                           D = The where above is a list of
                                               data. In this case the returned
                                               columns will be as per the cols
                                               list
                                 'dic':    A dictionary of column details as
                                           per sql.table_dic. This only applies
                                           when ttype is 'D'
                     OR = The choice selection function to perform as:
                            a_name = {
                                 'stype':  "C",
                                 'titl':   Selection Title
                                 'head':   Column Headings, None for None
                                 'data':   (('aaaa','aaa'),('bbbb','bbb')),
                                 'retn':   "I"=index, "D"=data or None
                                 'index':  The index of the item to return,
                                 'butt':   Additional Buttons and functions}
                     OR = The file/directory selection dialog to perform as:
                            a_name = {
                                 'stype':  "F",
                                 'types':  "fle or dir",
                                 'initd':  The base directory,
                                 'ftype':  (('Zip Files', '*.zip'),),
                                 'multi':  True/False for multiples}
                            Note: ftype is a tuple of tuples for file selection
                                  and a string for directory selection
                     OR = The calendar date selection:
                            a_name = {
                                 'stype':  "D"}
                     OR = A function in the root module to perform:
                            a_name = {
                                 'stype':  "M",
                                 'func':   the function to execute}
                f5      = The function to perform to delete a record
                            returns a tuple indicating the next focus position
                            i.e. (frt, pag, col) or None for the default next
                            focus position
                valid   = A tuple of tuples of type and data for validation e.g.
                            ('efld',)
                            ('notblank',)
                            ('notzero',)
                            ('between', 5, 2)
                            ('idno',)
                            ('in', ("Y",'N'))
                            ('dir',)
                            ('file',)
                            ('path',)
                label   = A list only where place[0] above = C, this will
                            be a label on the left of the first column. The
                            list contains the following: (heading, rows) else
                            None
                tooltip = A string to display as a tooltip.

    rows   =  A tuple giving the number of rows per page e.g. (5,5) (max = 15)

    butt   =  A tuple of tuples of buttons to create at the bottom of the frame
              as follows:
                titl = The buttons title
                srec = The record selection function as for cols
                comm = The command to execute
                       Note: Only one of srec or comm can be used not both
                norm = The normal state of the button 0=off, 1=on, 2=ignore
                on   = Tuple showing when to enable button (frt, pag, pos),
                       this can also be tuples of tuples,
                       for the end of a Frame use (frt, pag, 0)
                off  = Tuple showing when to disable button (frt, pag, pos),
                       this can also be tuples of tuples,
                       for the end of a Frame use (frt, pag, 0)
                tip  = The tooltip
                row  = A row to place the button on.
                spn  = Columns to span.

    tend   =  A list of lists of functions to execute on finishing top frames
                by page having: (function, cnf)

    txit   =  A list of functions to execute on exiting top frames by page

    cend   =  A list of lists of functions to execute on finishing a col line
                by page having: (function, cnf)

    cxit   =  A list of functions to execute on exiting a col frames by page

    view   =  An entry field asking whether you would like to view the report
              It is a tuple (csv, dft) where:
                csv is whether to have an export facility:
                    Y - Yes
                    C - CSV only
                    X - XLS only
                    N - No
                dft is the default e.g. ("(V)iew", "(P)rint" or "(X)port")
              A list self.repprt is created having 3 fields as follows:
                1) Y or N output widget created
                2) Query widget contents
                3) Printer widget contents

    mail   =  An entry field asking whether you would like to email the report
              It is a tuple (flag, prnt, dflt, desc) where:
                flag is one of the following:
                   "Y" = Address Field (noblank) - Entered Address
                   "N" = No Address Field (default N) - Address from account
                   "B" = Address Field (efld) - Combination of both
                prnt  is "Y" or "N" to choose to view/print emailed docs or not
                dflt is the default answer
                desc is the description to replace the default description
              A list self.repeml is created having up to 5 fields as follows:
                1) Y or N widget created
                2) Y or N to email the report
                3) Email addresses is a comma delimited string or a list/tuple
                4) Email message which can be:
                     A string i.e. a message
                     A List or Tuple which can be:
                         Subject and Message
                         Subject, Message and Attachment
                5) Whether to print/view the emailed report

    focus  =  Whether to automatically focus on the first screen's first column.
                 The default is True

    vport  =  Create a label in a viewport to display the message

    clicks =  A function to execute when an Entry Label field label is clicked

    Notes:  The 'self.pos' and 'pos' variable is the colEntry field number (0-?)
            The 'self.idx' and 'idx' variable is the pos in the row (0-?)
            The 'self.col' and 'col' variable is the column number (1-?)
            The mail and view options only apply to the ('T',0) frame

    In the function executed by the widget you can do the following:
            Verify the data
            Update tables
            Quit the program
            Continue by returning a value as per above
    """
    def __init__(self, mf, **args):
        self.mf = mf
        options = ("screen", "tops", "title", "tags", "eflds", "rows", "butt",
            "tend", "txit", "cend", "cxit", "view", "mail", "save", "focus",
            "vport", "clicks")
        # Set defaults
        for option in options:
            if option == "tops":
                self.tops = False
            elif option == "focus":
                self.focus = True
            elif option == "eflds":
                self.eflds = []
            elif option == "vport":
                self.vport = False
            elif option == "clicks":
                self.clicks = False
            else:
                setattr(self, option, None)
        # Set passed options
        for arg in args:
            if arg == "title":
                if type(args["title"]) == str:
                    self.title = args["title"]
                elif len(args["title"]) == 1:
                    self.title = args["title"][0]
                else:
                    self.title = "%-s - %s" % tuple(args["title"][:2])
            elif arg == "butt" and args["butt"]:
                self.butt = []
                for but in args["butt"]:
                    self.butt.append(list(but))
            else:
                setattr(self, arg, args[arg])
        # Window and title
        self.window = self.mf.window
        if not self.tops and not self.mf.head.cget("text"):
            self.mf.head.configure(text=self.title)
        # Save window bindings
        self.svebind = []
        for bind in self.window.bind():
            if "<Key-Alt_L>" in bind:
                self.svebind.append((bind, self.window.bind(bind)))
                self.window.unbind(bind)
        # Set master frame
        if self.screen:
            self.master = self.screen
        else:
            self.master = self.mf.body
        # Set all variables
        if self.setVariables():
            # Draw all widgets
            self.drawScreens()
            # Focus to the first field
            if self.focus and self.first:
                self.focusField(self.first[0], self.first[1], 1)
        self.master.update_idletasks()

    def setVariables(self):
        self.first = []             # the first field details
                                    # i.e. (frt, page, row, col)
        self.last = [[0,0]]         # last position visited by frt by page
        self.skip = [[]]            # columns skipped by frame 'sk'
        self.topq = [0]             # top number of columns by page
        self.colq = [0]             # col number of columns by page
        self.topf = [[]]            # eflds tops data by page
        self.colf = [[]]            # eflds cols data by page
        self.topLabel = [{}]        # dic of top label fields by page by col
        self.topEntry = [{}]        # dic of top entry fields by page by col
        self.topCompl = [{}]        # dic of top entry widget by page by col
        self.topListl = [{}]        # dic of top entry list by page by col
        self.colLabel = [{}]        # dic of col label fields by page by row
        self.colEntry = [{}]        # dic of col entry fields by page by col
        self.colCompl = [{}]        # dic of col entry widget by page by col
        self.colListl = [{}]        # dic of col entry list by page by col
        self.t_work = []            # list of arrays of top work by page by
                                    # row 0 by idx
        self.t_disp = []            # list of arrays of top disp by page by
                                    # row 0 by idx
        self.c_work = []            # list of arrays of col work by page by
                                    # row by idx
        self.c_disp = []            # list of arrays of col disp by page by
                                    # row by idx
        self.tsiz = [[0,0,0]]       # largest label, entry field and also the
                                    # maximum width by top page
        self.pag = None             # Clearing for doKeys() problem
        self.pgs = 0                # Last page number
        self.topz = False           # Whether or not there is a top page 0
        self.colz = False           # Whether or not there is a col page 0
        self.topv = [[]]            # top validators by page by col
        self.colv = [[]]            # col validators by page by col
        self.rpad = 3               # Default padding between Radiobuttons
        self.has_focus = []         # List of widgets which can focus
        self.hidden = {}            # Hidden widgets
        self.lastnbpage = None      # Last Notebook Page
        self.bindings = []          # Toplevel bindings
        # ======================================================================
        # Checking and Adding Ouput and Mail Widget Requirements
        # ======================================================================
        if not self.view:
            self.repprt = ["N", "N", ""]
        else:
            self.repprt = ["Y", "", ""]
        self.repeml = ["N", "N", "", "", "Y"]
        if self.mail:
            try:
                gc = GetCtl(self.mf)
                ctlsys = gc.getCtl("ctlsys", error=False)
                if ctlsys and ctlsys["sys_msvr"]:
                    server = [ctlsys["sys_msvr"], ctlsys["sys_mprt"],
                        ctlsys["sys_msec"], ctlsys["sys_maut"],
                        ctlsys["sys_mnam"], ctlsys["sys_mpwd"]]
                    if sendMail(server, "", "", "", check=True,
                            errwid=self.mf.body,
                            wrkdir=self.mf.rcdic["wrkdir"]):
                        self.repeml = ["Y", "", "", "", "Y"]
                    else:
                        showError(self.mf.body, "Error",
                            "Mail Server Invalid or Unavailable")
            except:
                pass
        if self.view or self.mail:
            if self.eflds and tuple(self.eflds[0][0][:2]) != ("T", 0):
                showError(self.master, "View or Mail Option",
                    "These Options Require a (T)op Zero Page")
                sys.exit()
            row = 0
            col = 0
            for fld in self.eflds:
                if fld[0][0] == "T" and fld[0][1] == 0:
                    row = fld[0][2]
                    col = fld[0][3]
                else:
                    break
            self.eflds = list(self.eflds)
            if self.repprt[0] == "Y":
                # View/Print/Export dialog
                prts = getPrinters(wrkdir=self.mf.rcdic["wrkdir"])
                data = prts[:]
                if data:
                    dflt = data[0]
                else:
                    prts = data = ["None"]
                    dflt = "None"
                prt = {
                    "stype": "C",
                    "titl": "Available Printers",
                    "data": data}
                rvs = [("View","V"),("Print","P"),("None","N")]
                mes = """Select what to do with the generated report.

View   - The generated report will be opened in the default
         pdf viewer.

Print  - The report will be printed on the selected printer."""
                if self.view[0].lower() in ("y", "c", "x"):
                    if self.view[0].lower() in ("y", "c"):
                        rvs.append(("Export CSV","C"))
                    if XLSX and self.view[0].lower() in ("y", "x"):
                        rvs.append(("Export XLSX","X"))
                    mes = """%s

Export - The report in the selected format will be opened
         in the default spreadsheet application.""" % mes
                row += 1
                self.eflds.append((("T",0,row,col),("IRB",rvs),0,"Output",
                    "Output Method",self.view[1].upper(),"Y",self.setView,
                    None,None, None,None,mes))
                row += 1
                self.eflds.append((("T",0,row,col),"INA",(30,50),"Printer Name",
                    "",dflt,"Y",self.setPrtr,prt,None,("in", prts)))
            if self.repeml[0] == "Y":
                # Email dialog
                yns = (("Yes","Y"),("No","N"))
                if len(self.mail) == 4:
                    dflt = self.mail[2]
                    desc = self.mail[3]
                else:
                    desc = "Email Document"
                if len(self.mail) == 3:
                    dflt = self.mail[2]
                else:
                    dflt = "N"
                if dflt != "S":
                    row += 1
                    self.eflds.append((("T",0,row,col),("IRB",yns),0,
                        desc,"",dflt,"N",self.setMail,None,None,None))
                if self.mail[0].lower() in ("b", "y"):
                    # Email addresses
                    data = []
                    sql = Sql(self.mf.dbm, ["telmst", "telcon"],
                        prog=__name__)
                    tdm = sql.getRec("telmst", cols=["tdm_name",
                        "tdm_email"], where=[("tdm_email", "<>", "")],
                        order="tdm_name")
                    for t in tdm:
                        data.append([t[0], "", "", t[1]])
                    tdc = sql.getRec("telcon", cols=["tdc_name",
                        "tdc_contact", "tdc_desig", "tdc_email"],
                        where=[("tdc_email", "<>", "")],
                        order="tdc_name, tdc_contact")
                    if tdc:
                        data.extend(tdc)
                    adr = {
                        "stype": "C",
                        "titl": "Available Addresses",
                        "head": (
                            "Name", "Contact", "Designation", "Address"),
                        "typs": (("NA",0,"Y"),("NA",0),("NA",0),("NA",0)),
                        "data": data,
                        "mode": "M",
                        "comnd": self.getAddr}
                    row += 1
                    a = [("T",0,row,col),"ITX",50,"E-Mail Address","",
                        "","N",self.setAddr,adr,None]
                    if self.mail[0].lower() == "y":
                        a.append(("email", False))
                    else:
                        a.append(("email", True))
                    a.extend([None,"E-Mail Addresses, Comma Separated"])
                    self.eflds.append(tuple(a))
                # Email message and additional attachments
                row += 1
                self.eflds.append((
                    ("T",0,row,col),("IRB",yns),0,"E-Mail Message","",
                        "N","N",self.setMess,None,None,None))
                if self.mail[1].lower() == "y":
                    # View/Print emailed document
                    row += 1
                    self.eflds.append((("T",0,row,col),("IRB",yns),0,
                        "View/Print Emailed Document","","N","Y",
                        self.setPrint,None,None,None))
        if not self.eflds:
            return
        # ======================================================================
        num = 1
        pos = []
        for col in self.eflds:
            col = list(col)
            if col[1][1] == "@":
                # Load details from ffield
                sql = Sql(self.mf.dbm, "ffield", prog=__name__)
                seq = sql.ffield_col
                ffld = sql.getRec("ffield",
                    where=[("ff_name", "=", col[1][2:])], limit=1)
                if not ffld:
                    showError(self.master, "Field Error",
                        "The Column Name %s Is Invalid" % col[1][2:])
                    sys.exit()
                col[1] = "%s%s" % (col[1][0], ffld[seq.index("ff_type")])
                if not col[2]:
                    col[2] = float(ffld[seq.index("ff_size")])
                if col[0][0] == "T":
                    if not col[3]:
                        col[3] = ffld[seq.index("ff_desc")]
                    if len(col) == 4:
                        pass
                    elif not col[4]:
                        col[4] = ffld[seq.index("ff_desc")]
                else:
                    if not col[3]:
                        col[3] = ffld[seq.index("ff_head")]
                    if len(col) == 4:
                        pass
                    elif not col[4]:
                        col[4] = ffld[seq.index("ff_desc")]
            if type(col[1]) == str:
                if len(col) > 10 and col[10]:
                    # Entry validation
                    col[1] = col[1].replace("Tv", "TX")
                elif col[1][1:] == "TX":
                    col[1] = col[1].replace("TX", "Tv")
                    if type(col[2]) in (list, tuple):
                        col[2] = col[2][0]
            if len(col) > 4 and not col[4]:
                # Duplicate text field if blank
                col[4] = col[3]
            if type(col[2]) in (int, float):
                # Make size a tuple
                if col[1][1:] in ("TV", "Tv"):
                    col[2] = (col[2], 1)
                else:
                    col[2] = (col[2], col[2])
            if not self.first:
                self.first = col[0]
            if col[0][1] < self.pgs:
                showError(self.master, "Page Number",
                    "The Page Number (%s) is Out of Sequence" % str(col[0]))
                sys.exit()
            elif col[1][0] == "R":
                showError(self.master, "Old Radio Button %s" % str(col[0]),
                    "Radio Buttons MUST Have a Tuple of Choices")
                sys.exit()
            elif type(col[1]) in (list, tuple) and len(col) > 9 and col[10]:
                showError(self.master, "Radio Button %s" % str(col[0]),
                    "Radio Buttons CAN NOT Have a Validation")
                sys.exit()
            elif len(col) not in (4, 11, 12, 13):
                showError(self.master, "Field Details",
                    "This Field Has Invalid Details %s" % str(col[0]))
                sys.exit()
            elif col[0][1] > self.pgs:
                while col[0][1] > self.pgs:     # Added for Blank Page
                    num = 1
                    self.pgs += 1               # Changed for Blank Page
                    self.last.append([0,0])
                    self.skip.append([])
                    self.topq.append(0)
                    self.topf.append([])
                    self.topLabel.append({})
                    self.topEntry.append({})
                    self.topCompl.append({})
                    self.topListl.append({})
                    self.colq.append(0)
                    self.colf.append([])
                    self.colLabel.append({})
                    self.colEntry.append({})
                    self.colCompl.append({})
                    self.colListl.append({})
                    self.tsiz.append([0,0,0])
                    self.topv.append([])
                    self.colv.append([])
            if col[0][0] == "T":
                if col[0][1] == 0:
                    self.topz = True
                if not pos or col[0][1] != pos[0]:
                    pos = [col[0][1], 0]
                if col[0][2] not in (pos[1], pos[1] + 1):
                    showError(self.master, "Field Details",
                        "This Row, %s, is Not Sequential" % str(col[0]))
                    sys.exit()
                pos[1] = col[0][2]
                if col[0][3] == 0:
                    if len(col[0]) == 5:
                        t = col[0][4]
                    else:
                        t = len(col[3])
                    if type(col[1]) in (list, tuple):
                        d = 0
                        for x in range(0, len(col[1][1])):
                            d = d + self.rpad + len(col[1][1][x][0])
                    else:
                        d = int(col[2][0] / 1)
                    if t > self.tsiz[self.pgs][0]:
                        self.tsiz[self.pgs][0] = t
                    if d > self.tsiz[self.pgs][1]:
                        self.tsiz[self.pgs][1] = d
                try:
                    self.topq[self.pgs] += 1
                except:
                    self.topq[self.pgs] = 1
                self.topf[self.pgs].append(col)
                if len(col) > 10:
                    self.topv[self.pgs].append(col[10])
                else:
                    self.topv[self.pgs].append(None)
            elif col[0][0] == "C":
                if col[0][1] == 0:
                    self.colz = True
                try:
                    self.colq[self.pgs] += 1
                except:
                    self.colq[self.pgs] = 1
                self.colf[self.pgs].append(col)
                if len(col) > 10:
                    self.colv[self.pgs].append(col[10])
                else:
                    self.colv[self.pgs].append(None)
                num += 1
        if self.colz and self.pgs:
            showError(self.master, "Col Frame",
                "Cannot Have a Zero Column Page With Notebook")
            sys.exit()
        if self.pgs > 0 and self.pgs > len(self.tags):
            showError(self.master, "Notebook",
                "Not All Notebook Tags Have Been Set")
            sys.exit()
        # ======================================================================
        # Set Number of Rows, Defaulting to 15
        # ======================================================================
        if not self.rows:
            self.rows = []
            for r in range(0, self.pgs+1):
                self.rows.append(15)
        else:
            self.rows = list(self.rows)
        for p in self.colf:
            for c in p:
                if len(c) > 11 and c[11]:
                    self.rows[c[0][1]] = c[11][1]
        # ======================================================================
        # Create t_work, t_disp, c_work and c_disp arrays for all pages
        # ======================================================================
        for x in range(0, self.pgs+1):
            q = self.topq[x]
            a = makeArray(q,1,1)
            self.t_work.append(a)
            a = makeArray(q,1,1,"S")
            self.t_disp.append(a)
            q = self.colq[x]
            r = self.rows[x]
            a = makeArray(q,r,1)
            self.c_work.append(a)
            a = makeArray(q,r,1,"S")
            self.c_disp.append(a)
        # ======================================================================
        # Calculate width of the widest row on the page
        # ======================================================================
        for p in range(0, self.pgs+1):
            row = None
            wid = 0
            for x in self.topf[p]:
                if x[0][3] == 0:
                    txt = self.tsiz[p][0]
                else:
                    if wid < x[0][3]:
                        wid = x[0][3]
                    if len(x[0]) == 5:
                        txt = x[0][4] - x[0][3]
                    else:
                        txt = len(x[3])
                if type(x[1]) in (list, tuple):
                    siz = 0
                    for y in range(0, len(x[1][1])):
                        siz = siz + self.rpad + len(x[1][1][y][0])
                else:
                    siz = int(x[2][0] / 1)
                if row == x[0][2]:
                    wid = wid + txt + siz
                else:
                    wid = txt + siz
                if wid > self.tsiz[p][2]:
                    self.tsiz[p][2] = wid*1
                row = x[0][2]
        # ======================================================================
        if self.butt is not None:
            self.bsiz = 0
            self.brow = 0
            for but in self.butt:
                if but[1] and but[2]:
                    showError(self.master, "Buttons",
                        "Buttons Cannot Have Both 'srec' and 'comm' Commands")
                    sys.exit()
                if self.bsiz < len(but[0]):
                    self.bsiz = len(but[0])
        return True

    def setView(self, frt, pag, r, c, p, i, w):
        self.repprt[1] = w
        if w == "V":
            self.repprt[2] = "view"
            self.loadEntry(frt, pag, p+1, data="")
            if self.mail and self.mail[1].lower() == "y":
                self.topf[pag][-1][4] = "View Emailed Document"
            return "sk1"
        elif w == "N":
            self.repprt[2] = "none"
            self.loadEntry(frt, pag, p+1, data="None")
            return "sk1"
        elif w in ("C", "X"):
            self.repprt[2] = "export"
            self.loadEntry(frt, pag, p+1, data="")
            if self.repeml[0] == "N":
                return "sk1"
            elif self.repeml[0] == "Y":
                self.repeml[1] = "N"
                self.loadEntry(frt, pag, p+2, data="N")
                if self.mail and self.mail[0].lower() in ("b", "d", "y"):
                    self.repeml[2] = ""
                    self.loadEntry(frt, pag, p+3, data="")
                return "nd"
        elif self.mail and self.mail[1].lower() == "y":
            self.topf[pag][-1][4] = "Print Emailed Document"

    def setPrtr(self, frt, pag, r, c, p, i, w):
        self.repprt[2] = w

    def setMail(self, frt, pag, r, c, p, i, w):
        self.repeml[1] = w
        if w == "N":
            self.repeml[2] = ""
            self.repeml[3] = ""
            skp = 1
            if self.mail[0].lower() in ("b", "y"):
                skp += 1
            if self.mail[1].lower() == "y":
                skp += 1
            for x in range(1, skp + 1):
                self.loadEntry(frt, pag, p+x, data="")
            return "sk%s" % skp

    def getAddr(self, frt, pag, r, c, p, i, w):
        if w:
            addr = ""
            for a in w:
                if not addr:
                    addr = a[3]
                else:
                    addr = "%s,%s" % (addr, a[3])
            self.loadEntry(frt, pag, p, data=addr)

    def setAddr(self, frt, pag, r, c, p, i, w):
        self.repeml[2] = w
        if w == "N":
            return "sk1"

    def setMess(self, frt, pag, r, c, p, i, w):
        if w == "N":
            self.repeml[3] = ""
        else:
            tit = "Email Details"
            fle = {
                "stype": "F",
                "types": "fle",
                "ftype": [("All Files", "*.*")]}
            fld = [
                (("T",0,0,0),"ITV",(60,10),"Message","",
                    "","Y",None,None,None,("efld",)),
                (("T",0,1,0),"IFF",60,"Attach","",
                    "","Y",None,fle,None,("file",))]
            but = (
                ("Continue", None, self.emlCont,1,None,None),
                ("Cancel", None, self.emlExit,1,None,None))
            self.dg = TartanDialog(self.mf, title=tit, tops=True, eflds=fld,
                txit=(self.emlExit,), butt=but)
            self.dg.mstFrame.wait_window()
        if self.repprt[2].lower() == "none" and self.mail[1].lower() == "y":
            self.loadEntry(frt, pag, p+1, data="N")
            return "sk1"

    def emlCont(self):
        mess = self.dg.topEntry[0][0].get("1.0", "end")
        self.repeml[3] = ["", mess]
        atat = self.dg.topEntry[0][1].get()
        if atat:
            self.repeml[3].append(atat)
        self.dg.closeProcess()

    def emlExit(self):
        self.repeml[3] = ""
        self.loadEntry("T", 0, 2, data="N")
        self.dg.closeProcess()

    def setPrint(self, frt, pag, r, c, p, i, w):
        self.repeml[4] = w

    def drawScreens(self):
        if self.tops:
            self.mstFrame = MyFrame(self.master, borderwidth=2, relief="ridge")
            tops = MyLabel(self.mstFrame, anchor="center", text=self.title,
                borderwidth=2, relief="raised")
            tops.grid(column=0, sticky="nsew")
        else:
            self.mstFrame = MyFrame(self.master)
        self.frt = ""
        self.nbf = False
        self.selectable = []
        for pag in range((self.pgs + 1)):
            if pag and not self.nbf:
                self.drawNoteBook(pag)
                self.nbf = True
            if self.topf[pag]:
                self.drawTopFrame(pag)
            if self.colf[pag]:
                self.drawColFrame(pag)
        if self.butt:
            self.drawButtons()
        # Check if frame is too large, adjust font until it fits.
        self.dfs = self.mf.rcdic["dfs"]
        while True:
            self.mstFrame.update_idletasks()
            bh = self.mf.body.winfo_height()
            if self.mstFrame.winfo_reqwidth() > int(self.mf.geo[0]) or \
                    self.mstFrame.winfo_reqheight() > int(bh):
                self.mf.rcdic["dfs"] = int(self.mf.rcdic["dfs"]) - 1
                self.mf.setThemeFont(butt=False)
                self.mf.resizeChildren()
            else:
                break
        self.mstFrame.place(anchor="center", relx=0.5, rely=0.5)
        self.mf.window.update_idletasks()
        self.mstFrame.grab_set()

    def drawNoteBook(self, page):
        # Create images for enabled and disabled
        self.nbeimg = getImage("yes", (20, 20))
        self.nbdimg = getImage("no", (20, 20))
        # Create notebook
        self.nb = ttk.Notebook(self.mstFrame, takefocus=False)
        self.nb.grid(column=0, sticky="nswe")
        self.nb.bindtags(("Notebook", self.nb, ".", "all"))
        self.nb.bind("<ButtonRelease-1>", functools.partial(self.chgPage, None))
        self.normaltags = []
        for num, tag in enumerate(self.tags):
            tag = list(tag)
            if len(tag) == 5:
                unl = tag[4]
            else:
                unl = True
            if unl:
                tag[0], pos = getUnderline(widget=self.mstFrame, text=tag[0])
            else:
                pos = -1
            fr = MyFrame(self.nb, borderwidth=2)
            fr.grid(column=0, row=num, sticky="nswe")
            self.nb.add(fr, text=tag[0], underline=pos, padding=0, sticky="n")
            setattr(self.nb, "Page%s" % (num + 1), fr)
            if unl:
                self.nb.winfo_toplevel().bind("<Alt_L><%s>" %
                    tag[0][pos].upper(), functools.partial(self.chgPage, num))
                self.bindings.append("<Key-Alt_L>%s" % tag[0][pos].upper())
                self.nb.winfo_toplevel().bind("<Alt_L><%s>" %
                    tag[0][pos].lower(), functools.partial(self.chgPage, num))
                self.bindings.append("<Key-Alt_L>%s" % tag[0][pos].lower())
            self.disableTag(num)

    def chgPage(self, page, event):
        if page is None and event.widget.identify(event.x, event.y) == "label":
            index = event.widget.index("@%d,%d" % (event.x, event.y))
        else:
            index = page
        if index not in self.normaltags:
            return
        self.selPage(index=index)

    def selPage(self, label=None, index=None, page=None, focus=True):
        # Select notebook page by label text or page number
        if label:
            found = False
            for idx, tag in enumerate(self.tags):
                if tag[0].replace("_", "") == label.replace("_", ""):
                    found = True
                    break
            if not found:
                print("Invalid Notebook Tag %s" % label)
                return
        elif index is None:
            idx = page - 1
        else:
            idx = index
        if self.nb.select() == self.nb.tabs()[idx]:
            # Page already selected
            return
        self.lastnbpage = self.nb.index(self.nb.select())
        self.nb.select(idx)
        self.pag = idx + 1
        cmd = self.tags[idx][1]
        if cmd:
            self.window.focus_set()
            cmd()
        elif self.last[self.pag][1]:
            m = self.colq[self.pag]
            l = self.last[self.pag][1]
            while l > m:
                l -= m
            p = (l - 1) % m
            if l == m and self.colf[self.pag][p][1][0] == "O":
                dirs = "B"
            else:
                dirs = "F"
            if focus:
                self.focusField("C", self.pag, self.last[self.pag][1], ddd=dirs)
        elif self.last[self.pag][0]:
            l = self.last[self.pag][0]
            m = self.topq[self.pag]
            p = l - 1
            if l == m and self.topf[self.pag][p][1][0] in ("O", "OCB", "ORB"):
                dirs = "B"
            else:
                dirs = "F"
            if focus:
                self.focusField("T", self.pag, self.last[self.pag][0], ddd=dirs)
        elif self.topq[self.pag]:
            if focus:
                self.focusField("T", self.pag, 1)
        else:
            if focus:
                self.focusField("C", self.pag, 1)

    def drawTopFrame(self, pag):
        if pag:
            mast = getattr(self.nb, "Page%s" % pag)
        else:
            mast = self.mstFrame
        if not self.tops:
            frame = MyFrame(mast, borderwidth=2, relief="ridge")
        else:
            frame = MyFrame(mast)
        frame.grid(column=0)
        self.lrow = None
        for num, fld in enumerate(self.topf[pag]):
            if num == (self.topq[pag] - 1):
                nxt = None
            else:
                nxt = self.topf[pag][num + 1]
            self.drawTopFlds(frame, num, fld, nxt)
        if self.vport:
            self.vport = MyLabel(frame, anchor="e", font=("Helvetica",
                int(self.mf.rcdic["dfs"]) * 2), color=False, width=15,
                relief="ridge")
            col, row = frame.grid_size()
            self.vport.grid(column=col, row=0, rowspan=row, padx=5,
                sticky="nswe")
        setattr(self, "topPage%s" % pag, frame)

    def setViewPort(self, typ, total):
        bg = "white"
        if not typ:
            fg = "black"
        elif (typ == "C" and total > 0) or (typ != "C" and total < 0):
            fg = "red"
        else:
            fg = "blue"
        if typ:
            self.vport.configure(text=CCD(total, "SD", 13.2).disp,
                background=bg, foreground=fg)
        else:
            self.vport.configure(text="", background=bg, foreground=fg)

    def drawTopFlds(self, frame, col, fld, nxt=None):
        pag = fld[0][1]
        row = fld[0][2]
        self.sme = False
        if row != self.lrow:
            self.ttabl = MyFrame(frame)
            self.ttabl.grid(column=0, row=row, sticky="w")
            self.acc = 0
            self.lrow = row
        elif fld[3] and fld[3] != " ":
            self.sme = True
        if type(fld[1]) in (list, tuple):
            typ = fld[1][0]
        else:
            typ = fld[1]
        if typ[1:] in ("CB", "RB"):
            siz = self.tsiz[pag][1]
        elif typ[1:] in ("TV", "Tv"):
            siz = int(fld[2][0] / 1)
            hgt = fld[2][1]
        else:
            siz = int(fld[2][0] / 1)
            if typ[1:] in ("FF", "TX"):
                lim = 0
            else:
                lim = int(fld[2][1] / 1)
        dsc = fld[0][3]                         # the text starting column
        if not dsc:
            dsc = self.acc
        if len(fld[0]) == 5:
            dse = fld[0][4]                     # the entry starting column
        elif fld[0][3] == 0 and self.acc == 0:
            dse = self.tsiz[pag][0]
        else:
            dse = dsc + len(fld[3])
        if self.sme:
            dse += 1
        if dse < dsc:
            showError(self.mstFrame, "Attach Error",
                "Label '%s' Has Invalid attach paramaters (%s %s)" % \
                (fld[3], dsc, dse))
            sys.exit()
        if typ[0] == "O":                       # Output Field Only
            cr = f1 = None                      # Not Used with Output Field
        else:
            cr = fld[7]                         # Enter Key
            f1 = fld[8]                         # F1 Key
        if dsc != self.acc:
            lpad = dsc - self.acc
        elif self.sme:
            lpad = 1
        else:
            lpad = 0
        if lpad and lpad > 0:
            txt = " " * lpad
        else:
            txt = ""
        if fld[3]:
            txt = txt + fld[3]
        txt = txt.replace("(noesc)", "")
        rpad = dse - self.acc - len(txt) + 1
        if txt and rpad and rpad > 0:
            txt = txt + " " * rpad
        if dse == dsc:
            dse = dsc + len(txt)
        t_dic = self.topLabel[pag]
        if len(txt) > 0:
            lab = MyLabel(self.ttabl, color=False, text=txt)
            lab.grid(row=0, column=dsc, columnspan=(dse - dsc))
            t_dic[col] = lab
            dsc = dse
        e_dic = self.topEntry[pag]              # Entry Fields
        c_dic = self.topCompl[pag]              # Completion Widget
        l_dic = self.topListl[pag]              # Completion List
        if typ[1:] == "CB":
            var = tk.StringVar()
            e_dic[col] = (MyCheckButton(self.ttabl, text=fld[1][1],
                variable=var), var)
            e_dic[col][0].grid(row=0, column=dsc, columnspan=siz)
            e_dic[col][0].bind("<<mychkbx>>", self.doKeys)
            if self.mf.rcdic["ttip"] == "Y" and len(fld) == 13 and fld[12]:
                ToolTip(e_dic[col][0], fld[12])
            self.doCheckButton(e_dic[col], "disabled")
        elif typ[1:] == "RB":
            e_dic[col] = []
            hbox = MyFrame(self.ttabl)
            hbox.grid(row=0, column=dsc, columnspan=siz)
            var = tk.StringVar()
            for num, val in enumerate(fld[1][1]):
                b = MyRadioButton(hbox, variable=var, text=val[0],
                    value=val[1], cmd=(self.execRB, num))
                b.bind("<<myradio>>", self.doKeys)
                if self.clicks and typ[0] == "I":
                    b.bind("<ButtonRelease-1>",
                        functools.partial(self.clicks, (pag, col, b)))
                b.grid(row=0, column=num)
                e_dic[col].append([b, var, (val[1], cr)])
            if self.mf.rcdic["ttip"] == "Y" and len(fld) == 13 and fld[12]:
                ToolTip(hbox, fld[12])
            self.doRadioButton(e_dic[col], "disabled")
        elif typ[1:] in ("TV", "Tv"):
            if typ[1:] == "Tv":
                e_dic[col] = MyTextView_v(self.ttabl, width=siz, height=hgt,
                    wrap="word", fg=self.mf.rcdic["nfg"],
                    bg=self.mf.rcdic["nbg"], bd=3)
            else:
                e_dic[col] = MyTextView_V(self.ttabl, width=siz, height=hgt,
                    wrap="word", fg=self.mf.rcdic["nfg"],
                    bg=self.mf.rcdic["nbg"], bd=3)
            if typ[0] == "I":
                if typ[1:] == "TV":
                    e_dic[col].bind("<<mytxtvV>>", self.doKeys)
                else:
                    e_dic[col].bind("<<mytxtvv>>", self.doKeys)
            e_dic[col].grid(row=0, column=dsc, columnspan=siz)
            c_dic[col] = None
            l_dic[col] = False
            if self.mf.rcdic["ttip"] == "Y" and len(fld) == 13 and fld[12]:
                ToolTip(e_dic[col], fld[12])
            if self.clicks and typ[0] == "I":
                e_dic[col].bind("<ButtonRelease-1>",
                    functools.partial(self.clicks, (pag, col)))
            self.doTextView("T", pag, e_dic[col], "disable")
        else:
            if self.mf.rcdic["acnf"] in ("I", "L") and f1 and \
                    self.checkAutoc(f1, typ):
                l_dic[col] = self.mf.rcdic["acnf"]
            else:
                l_dic[col] = False
            e_dic[col] = MyEntry(self.ttabl, maxsize=lim, cmd=(self.execEntry,
                ("T", pag, cr)), aut=l_dic[col], width=siz)
            if typ[0] == "I":
                e_dic[col].bind("<<myentry>>", self.doKeys)
            if typ[1] == "H":
                e_dic[col].configure(show="*")
            e_dic[col].grid(row=0, column=dsc, columnspan=siz)
            if self.mf.rcdic["ttip"] == "Y" and len(fld) == 13 and fld[12]:
                ToolTip(e_dic[col], fld[12])
            if self.clicks and typ[0] == "I":
                e_dic[col].bind("<ButtonRelease-1>",
                    functools.partial(self.clicks, (pag, col)))
            self.doEntryField("T", pag, col, e_dic[col], False, "disable")
        self.acc = dse + siz

    def drawColFrame(self, pag):
        if pag:
            mast = getattr(self.nb, "Page%s" % pag)
        else:
            mast = self.mstFrame
        frame = MyFrame(mast, borderwidth=2)
        frame.grid(column=0)
        # Create left label heading
        if len(self.colf[pag][0]) > 11 and self.colf[pag][0][11]:
            rows = self.colf[pag][0][11][1]
            text = self.colf[pag][0][11][0]
            siz = int(self.colf[pag][0][2][0] / 1)
            lab = MyLabel(frame, text=text, relief="ridge", size=siz)
            lab.grid(row=0, column=0, sticky="ew")
            self.dolab = True
        else:
            rows = self.rows[pag]
            self.dolab = False
        # Create headings for each column
        for x, y in enumerate(self.colf[pag]):
            if y[1][1].lower() == "d":
                text = str(y[3])
                anchor = "c"
            elif y[1][1].lower() == "s":
                text = "%s " % str(y[3].strip())
                anchor = "e"
            elif y[1][1:].lower() in ("ud", "ui", "ul", "us"):
                text = str(y[3])
                anchor = "e"
            else:
                text = str(y[3])
                anchor = "w"
            siz = int(self.colf[pag][x][2][0] / 1)
            lab = MyLabel(frame, text=text, anchor=anchor, relief="ridge",
                size=siz)
            lab.grid(row=0, column=x+1, sticky="ew")
        for row in range(rows):
            for col, fld in enumerate(self.colf[pag]):
                self.drawColFlds(frame, pag, row, col, fld)
        setattr(self, "colPage%s" % pag, frame)

    def drawColFlds(self, frame, pag, row, col, fld):
        typ = fld[1]
        siz = int(fld[2][0] / 1)
        lim = int(fld[2][1] / 1)
        txt = fld[3].replace("(noesc)", "")
        seq = ((row * self.colq[pag]) + col)
        if typ[0] == "O":                       # Output Field Only
            cr = f1 = None                      # Not Used with Output Field
        else:
            cr = fld[7]                         # Enter Key
            f1 = fld[8]                         # F1 Key
        if self.dolab and col == 0:
            t_dic = self.colLabel[pag]
            t_dic[row] = MyLabel(frame, color=True, text=txt)
            t_dic[row].grid(row=row+1, column=col, sticky="ew")
        e_dic = self.colEntry[pag]
        c_dic = self.colCompl[pag]
        l_dic = self.colListl[pag]
        if typ[1:] in ("TV", "Tv"):
            e_dic[seq] = MyTextView_v(frame, wrap="word",
                fg=self.mf.rcdic["nfg"], bg=self.mf.rcdic["nbg"])
            e_dic[seq].configure(width=siz, height=1)
            if typ[0] == "I":
                e_dic[seq].bind("<<mytxtvv>>", self.doKeys)
            e_dic[seq].grid(row=row+1, column=col+1)
            c_dic[seq] = None
            l_dic[seq] = False
            if self.mf.rcdic["ttip"] == "Y" and len(fld) == 13 and fld[12]:
                ToolTip(e_dic[col], fld[12])
            self.doTextView("C", pag, e_dic[seq], "disable")
        else:
            if self.mf.rcdic["acnf"] in ("I", "L") and f1 and \
                                        self.checkAutoc(f1, typ):
                l_dic[seq] = self.mf.rcdic["acnf"]
            else:
                l_dic[seq] = False
            e_dic[seq] = MyEntry(frame, maxsize=lim, cmd=(self.execEntry,
                ("C", pag, cr)), aut=l_dic[seq], width=siz)
            if typ[0] == "I":
                e_dic[seq].bind("<<myentry>>", self.doKeys)
            if typ[1] == "H":
                e_dic[seq].configure(show="*")
            e_dic[seq].grid(row=row+1, column=col+1)
            if self.mf.rcdic["ttip"] == "Y" and len(fld) == 13 and fld[12]:
                ToolTip(e_dic[seq], fld[12])
            self.doEntryField("C", pag, col, e_dic[seq], False, "disable")

    def checkAutoc(self, f1, typ):
        # If 'autoc' field in f1 dictionary then that is the default else if
        # alphanumeric data field then True else False
        if "stype" in f1 and f1["stype"] in ("C", "R") and \
                                    ("cols" or "typs" in f1):
            if "autoc" in f1:
                return f1["autoc"]
            if typ in ("INA", "INa", "IUA", "IRW"):
                return True

    def chgFieldSize(self, frt, pag, pos, siz):
        if frt == "T":
            self.topf[pag][pos][2] = siz
            self.topEntry[pag][pos].set_max_length(siz)
        else:
            self.colf[pag][pos][2] = siz
            for r in range(self.rows[pag]):
                seq = ((r * self.colq[pag]) + pos)
                self.colEntry[pag][seq].set_max_length(siz)

    def drawButtons(self):
        rows = int(len(self.butt) / 6)
        if len(self.butt) % 6:
            rows += 1
        for but in self.butt:
            if len(but) > 7 and but[7] > rows:
                rows = but[7]
        cols = int(len(self.butt) / rows)
        if len(self.butt) % rows:
            cols += 1
        bbox = MyButtonBox(self.mstFrame)
        for num, but in enumerate(self.butt):
            if but[1]:
                cmd = (self.selectBut, num)
            elif but[2]:
                cmd = but[2]
            if len(but) > 7:
                row = but[7]
                if len(but) == 9:
                    spn = but[8]
                else:
                    spn = 1
            else:
                row = int(num / cols)
                spn = 1
            butts = bbox.addButton(but[0], cmd, row=row, spn=spn)
            binds = butts.bind()
            for bind in binds:
                if "<Key-Alt_L>" in bind:
                    self.bindings.append(bind)
            if self.mf.rcdic["ttip"] == "Y" and len(but) > 6 and but[6]:
                ToolTip(butts, but[6])
            butts.configure(state="disabled")
            setattr(self, "B%s" % num, butts)

    def selectData(self):
        if self.frt == "T":
            opts = self.topf[self.pag][self.idx][8]
        else:
            opts = self.colf[self.pag][self.idx][8]
        rs = self.selectItem(self.pag, opts)
        self.focusField(self.frt, self.pag, self.col)
        if rs is not None:
            self.loadEntry(self.frt, self.pag, self.pos, data=rs)
            self.doInvoke(None, rs)

    def selectBut(self, num):
        opts = self.butt[num][1]
        state = self.disableButtonsTags()
        rs = self.selectItem(self.pag, opts)
        self.enableButtonsTags(state=state)
        self.focusField(self.frt, self.pag, self.col)
        if rs is not None:
            self.doInvoke(None, rs)

    def selectItem(self, pag, opts):
        if opts["stype"] == "C":
            self.rs = self.selChoice(opts)
        elif opts["stype"] == "F":
            return self.selFile(opts)
        elif opts["stype"] == "M" and "func" in opts:
            try:
                self.rs = opts["func"]()
            except:
                self.rs = None
            if self.rs is None:
                return None
            if type(self.rs) == str:
                return self.rs
        elif opts["stype"] == "R":
            if "screen" not in opts:
                opts["screen"] = self.mstFrame
            self.rs = self.selRecord(pag, opts)
        elif opts["stype"] == "S":
            if "screen" not in opts:
                opts["screen"] = self.mstFrame
            if "title" not in opts:
                opts["title"] = None
            if "head" not in opts:
                opts["head"] = None
            if "where" not in opts:
                opts["where"] = None
            if "group" not in opts:
                opts["group"] = None
            if "order" not in opts:
                opts["order"] = None
            if "selcol" not in opts:
                opts["selcol"] = None
            if "ttype" not in opts:
                opts["ttype"] = None
            if "dic" not in opts:
                opts["dic"] = None
            sel = getSingleRecords(self.mf, opts["tables"], opts["cols"],
                opts["title"], opts["head"], opts["where"], opts["group"],
                opts["order"], opts["selcol"], opts["ttype"], opts["dic"])
            if not sel:
                return None
            else:
                return sel
        elif opts["stype"] == "X":
            return self.selColour(opts)
        else:
            return
        if self.rs.selection:
            if "comnd" in opts and opts["comnd"]:
                ok = opts["comnd"](self.frt, pag, self.row, self.col,
                    self.pos, self.idx, self.rs.selection)
                return ok
            if type(self.rs.selection) in (int, str):
                return self.rs.selection
            if "mode" in opts and opts["mode"] == "M":
                d = []
                for s in self.rs.selection:
                    if type(s) in (int, str):
                        d.append(s)
                    elif "index" in opts:
                        d.append(s[opts["index"]])
                    else:
                        d.append(s[0])
                return d
            if "index" in opts:
                return self.rs.selection[opts["index"]]
            return self.rs.selection[0]

    def selChoice(self, opts):
        if "titl" in opts:
            titl = opts["titl"]
        else:
            titl = None
        if "head" in opts:
            head = opts["head"]
        else:
            head = None
        data = opts["data"]
        if "typs" in opts:
            typs = list(opts["typs"])
        else:
            typs = None
        if "mode" in opts:
            mode = opts["mode"]
        else:
            mode = "S"
        if "retn" in opts:
            retn = opts["retn"]
        else:
            retn = "D"
        if "sort" in opts:
            sort = opts["sort"]
        else:
            sort = False
        if "butt" in opts:
            butt = opts["butt"]
        else:
            butt = None
        if mode == "M":
            if "index" in opts:
                selc = opts["index"]
            else:
                selc = 0
            if typs:
                typs.insert(0, ("", 0))
        else:
            selc = None
        sc = SChoice(self.mf, titl=titl, head=head, data=data, typs=typs,
            sort=sort, mode=mode, selc=selc, retn=retn, butt=butt,
            scrn=self.mstFrame)
        self.mf.updateStatus("")
        return sc

    def selRecord(self, pag, opts):
        # Set the Defaults
        lic = {}
        for key in ("screen", "tables", "cols", "extra", "group",
                "title", "order", "wtype", "zero", "butt", "sort"):
            if key in opts:
                lic[key] = opts[key]
            elif key == "screen":
                lic[key] = self.mstFrame
            elif key == "tables":
                lic[key] = []
            elif key == "sort":
                lic[key] = True
            elif key == "zero":
                lic[key] = "N"
            else:
                lic[key] = None
        if not lic["order"]:
            # Default the order to the first column
            lic["order"] = lic["cols"][0][0]
        whr = self.setWhere(pag, opts)
        sr = SRec(self.mf, screen=lic["screen"], tables=lic["tables"],
            title=lic["title"], cols=lic["cols"], where=whr,
            group=lic["group"], order=lic["order"], extra=lic["extra"],
            wtype=lic["wtype"], butt=lic["butt"], sort=lic["sort"],
            zero=lic["zero"])
        return sr

    def setWhere(self, pag, opts):
        if "where" in opts and opts["where"]:
            where = copyList(opts["where"])
        else:
            where = None
        if "whera" in opts and opts["whera"]:
            for o in opts["whera"]:
                s = o[0]
                f = o[1]
                c = int(o[2])
                if len(o) == 4:
                    pag = o[3]
                if s == "T":
                    d = self.t_work[pag][0][c]
                else:
                    r = int((self.last[pag][1] - 1) / self.colq[pag])
                    d = self.c_work[pag][r][c]
                if not where:
                    where = [(f, "=", d)]
                else:
                    where.append((f, "=", d))
        return where

    def selFile(self, opts):
        if "parent" not in opts:
            opts["parent"] = self.mf.body
        if opts["types"] == "dir":
            dialog = FileDialog(**opts)
            sel = dialog.askdirectory()
        else:
            if "initd" not in opts:
                opts["initd"] = None
            if "multi" not in opts:
                opts["multi"] = False
            dialog = FileDialog(**opts)
            sel = dialog.askopenfilename()
            if type(sel) in (list, tuple):
                tmp = copyList(sel)
                sel = ""
                for t in tmp:
                    if not sel:
                        sel = t
                    else:
                        sel = "%s,%s" % (sel, t)
        return sel

    def selColour(self, opts):
        color = self.getWidget().get().rstrip()
        if color:
            if CPICK:
                sel = askcolor(color, parent=self.mstFrame)
            else:
                sel = tkcolor.askcolor(parent=self.mstFrame, initialcolor=color)
        elif CPICK:
            sel = askcolor(parent=self.mstFrame)
        else:
            sel = tkcolor.askcolor(parent=self.mstFrame)
        if sel[1]:
            return sel[1].lower()

    def doKeys(self, event):
        if str(event.widget.cget("state")) == "disabled":
            return True
        if self.frt == "T":
            fld = self.topf[self.pag][self.idx]
        else:
            fld = self.colf[self.pag][self.idx]
        if type(fld[1]) in (list, tuple):
            typ = fld[1][0][0]
        else:
            typ = fld[1][0]
        if typ == "O":
            # Output field only
            return True
        if event.widget.winfo_class() == "TEntry":
            # Remove autocomplete widget if it exists
            event.widget.escape()
        if fld[1][1:] == "Tv":
            if type(fld[2]) in (list, tuple):
                event.widget["height"] = fld[2][1]
            else:
                event.widget["height"] = 1
        if event.keysym == "F1" and fld[8]:
            event.widget.configure(state="disabled")
            self.master.update_idletasks()
            self.master.focus_force()
            self.selectData()
        elif event.keysym == "F5" and fld[9]:
            self.deleteRecord(event.widget)
        elif event.keysym == "F9" and fld[1][1:] == "TV":
            self.execCommand(self.frt, self.pag, fld[7])
        elif event.keysym in ("KP_Enter", "Return") and fld[1][1:] == "Tv":
            self.execCommand(self.frt, self.pag, fld[7])
        elif event.keysym in ("KP_Enter", "Return") and fld[1][0][1:] == "CB":
            self.execCommand(self.frt, self.pag, fld[7])
        elif event.keysym == "Escape" and self.esc:
            if self.idx == 0:
                self.doExitFrame(self.frt, self.pag)
            else:
                self.backUp(self.frt, self.pag)
        return True

    def focusField(self, frt, pag, col, clr=True, err="", ddd="F", tag=True):
        """
        This routine is used to focus on a widget:
            frt = The frame type, T = Top, C = Columnar
            pag = The frame number
            col = The column number
            clr = Clear the frame if first field of first page or
                  Clear the rest of the row if Column Frame - True or False
            err = An error message to display on the status line
            ddd = Direction, F = Forwards, B = Backwards when skipping fields
            tag = Select data in the text view widget
        """
        if frt == "T" and not self.topq[pag]:
            return
        elif frt == "C" and not self.colq[pag]:
            return
        if type(clr) == str:
            clr = bool(clr == "True")
        # Force focus to application
        self.window.focus_force()
        if col == 1 and tuple(self.first[:2]) == (frt, pag) and clr:
            # First field of first screen and page, clear all frames
            for x in range(0, self.pgs+1):
                self.clearFrame("T", x)
                self.clearFrame("C", x)
            self.skip[pag] = []
        for widget in self.has_focus:
            try:
                widget.configure(state="disabled")
            except:
                pass
        self.has_focus = []
        self.frt = frt
        self.pag = int(pag)
        self.col = int(col)
        self.pos = self.col - 1
        self.checkButtonsTags(self.frt, self.pag, self.col)
        if self.frt == "T":
            self.row = int(self.pos / self.topq[self.pag])
            self.idx = self.pos
            flds = self.topf[self.pag][self.idx]
            fldd = self.topEntry[self.pag][self.pos]
            faut = bool(flds[1][0] == "I" and self.topListl[self.pag][self.pos])
            # ==================================================================
            # If column field with Y in [6] clear rest of row
            # ==================================================================
            if len(flds) > 4 and flds[6] == "Y":
                r1 = self.col
                r2 = self.topq[pag]
                for c in range(r1, r2):
                    self.clearEntry(frt, pag, c)
            # ==================================================================
            # Setting the default data in top entry field
            # ==================================================================
            if len(flds) > 4 and flds[6] == "Y":
                self.clearEntry(self.frt, self.pag, self.col)
                work = ""
            else:
                work = self.t_work[self.pag][self.row][self.pos]
            if work:
                dflt = work
            elif len(flds) == 4:
                dflt = ""
            else:
                dflt = flds[5]
            if not dflt:
                dflt = ""   # Clear zero (0) fields
            # ==================================================================
        else:
            self.row = int(self.pos / self.colq[self.pag])
            self.idx = self.pos % self.colq[self.pag]
            flds = self.colf[self.pag][self.idx]
            fldd = self.colEntry[self.pag][self.pos]
            faut = bool(flds[1][0] == "I" and self.colListl[self.pag][self.pos])
            # ==================================================================
            # If column field with Y in [6] clear rest of row
            # ==================================================================
            if len(flds) > 4 and flds[6] == "Y":
                r1 = self.col
                r2 = self.col + (self.colq[pag] - self.idx)
                for c in range(r1, r2):
                    self.clearEntry(frt, pag, c)
            # ==================================================================
            # Setting the default data in col entry field
            # ==================================================================
            if len(flds) > 4 and flds[6] == "Y" and self.idx == 0:
                self.clearLine(self.pag, self.row)
            work = self.c_work[self.pag][int(self.pos / self.colq[
                self.pag])][self.idx]
            if work:
                dflt = work
            elif len(flds) == 4:
                dflt = ""
            elif flds[5] not in ("r","i","p"):
                dflt = flds[5]
            elif self.col > self.colq[self.pag]:
                dflt = self.c_work[self.pag][int(self.pos / self.colq[
                    self.pag]) - 1][self.idx]
                if not dflt or flds[5] == "r":
                    # Repeat previous value
                    pass
                elif flds[5] == "i":
                    # Increment previous value
                    try:
                        dflt = int(dflt) + 1
                    except:
                        dflt = ""
                elif flds[5] == "p":
                    # Increment previous period
                    y = int(dflt / 100)
                    m = (dflt % 100) + 1
                    if m > 12:
                        y += 1
                        m = m - 12
                    dflt = (y * 100) + m
                    if dflt == 1:
                        dflt = ""
            else:
                dflt = ""
            if not dflt:
                dflt = ""   # Clear zero (0) fields
        # ======================================================================
        # Store the last col position by page
        # ======================================================================
        if self.frt == "T":
            self.last[self.pag][0] = self.col
        else:
            self.last[self.pag][1] = self.col
        # ======================================================================
        # Check if Output Field Only and if so, skip it and return
        # ======================================================================
        if flds[1][0] in ("O", "OCB", "ORB"):
            self.doSkipFld(ddd.upper(), dflt)
            return
        # ======================================================================
        # Setting the status line help message
        # ======================================================================
        sufx = ""
        if not err:
            if flds[8]:
                sufx = ", Enter to Continue, <F1> to Select"
            else:
                sufx = ", Enter to Continue"
        elif flds[8]:
            sufx = ", <F1> to Select"
        if flds[9]:
            sufx = sufx + ", <F5> to Delete"
        if flds[1][1:] == "TV":
            sufx = sufx.replace("Enter", "<F9>")
        if flds[4].count("(noesc)"):
            self.esc = False
            text = flds[4].replace("(noesc)", "")
            cnt = sufx.rfind(",")
            if cnt != -1:
                sufx = "%s or%s" % (sufx[:cnt], sufx[cnt+1:])
        else:
            self.esc = True
            text = flds[4]
            if self.idx == 0:
                if frt == "T" and self.txit and self.txit[pag]:
                    sufx = sufx + " or <Esc> to Exit"
                elif self.cxit and self.cxit[pag]:
                    sufx = sufx + " or <Esc> to Exit"
            else:
                sufx = sufx + " or <Esc> to Back Up"
        if err and err != "ok":
            if self.mf.rcdic["errs"] == "Y":
                try:
                    import beepy
                    beepy.beep(1)
                except:
                    try:
                        import winsound
                        winsound.Beep(2500, 1000)
                    except:
                        self.mf.window.bell()
            self.mf.updateStatus("%s, Retry%s" % (err, sufx), "white", "red")
        else:
            self.mf.updateStatus("%s%s" % (text.lstrip(), sufx))
        # ======================================================================
        if type(flds[1]) in (list, tuple):
            if flds[1][0][1:] == "CB":
                self.doCheckButton(fldd, "focus", dflt)
            elif flds[1][0][1:] == "RB":
                self.doRadioButton(fldd, "focus", dflt)
        elif flds[1][1:] in ("TV", "Tv"):
            self.doTextView(self.frt, self.pag, fldd, "focus", dflt, tag)
        else:
            self.doEntryField(self.frt, self.pag, self.idx, fldd, faut,
                "focus", dflt)

    def doSkipFld(self, ddd, dflt=""):
        if ddd == "F":
            self.doInvoke(None, dflt)
        else:
            self.loadEntry(self.frt, self.pag, self.pos, data=dflt)
            if self.col == 1:
                if self.txit and self.txit[self.pag]:
                    self.txit[self.pag]()
                elif self.tend and self.tend[self.pag]:
                    if self.tend[self.pag][0]:
                        self.tend[self.pag][0]()
            else:
                self.col = self.col - 1
                self.focusField(self.frt, self.pag, self.col, ddd=ddd)

    def deleteRecord(self, widget=None):
        # Disable the widget
        if widget:
            self.setWidget(widget, state="disabled")
        ok = self.entryConfirm(mess="Delete This Record (Y/N)")
        if ok == "n":
            self.focusField(self.frt, self.pag, self.col)
        else:
            if self.frt == "C":
                com = self.colf[self.pag][self.idx][9]
            else:
                com = self.topf[self.pag][self.idx][9]
            if com:
                ok = com()
            else:
                ok = "Invalid or Missing Deletion Command"
            if ok == "nf":
                return
            if type(ok) in (list, tuple):
                if len(ok) == 4:
                    showError(self.master, "Deletion Error", ok[3])
                self.focusField(ok[0], ok[1], ok[2])
                return
            if ok and ok != "ok":
                showError(self.master, "Deletion Error", ok)
            if self.frt == "C":
                col = (self.row * self.colq[self.pag]) + 1
                self.focusField(self.frt, self.pag, col)
            else:
                self.focusField(self.first[0], self.first[1], 1)

    def backUp(self, frt, pag):
        if frt == "T":
            flds = self.topf[pag]
            fldd = self.topEntry[pag][self.pos]
        else:
            flds = self.colf[pag]
            fldd = self.colEntry[pag][self.pos]
        if (frt == "T" and self.col == 1) or (frt == "C" and self.idx == 0):
            self.doExitFrame(frt, pag)
            return
        if type(flds[self.idx][1]) in (list, tuple):
            if flds[self.idx][1][0][1:] == "CB":
                self.doCheckButton(fldd, "disabled")
            elif flds[self.idx][1][0][1:] == "RB":
                self.doRadioButton(fldd, "disabled")
        elif flds[self.idx][1][1:] in ("TV", "Tv"):
            self.doTextView(frt, pag, fldd, "disable")
        else:
            self.doEntryField(frt, pag, self.idx, fldd, False, "disable")
        self.idx = self.idx - 1
        self.pos = self.pos - 1
        self.col = self.col - 1
        for s in range(len(self.skip[self.pag]) - 1, -1, -1):
            if self.skip[self.pag][s] == self.col:
                if frt == "T" and self.col == 1:
                    self.doExitFrame(frt, pag)
                    return
                elif frt == "C" and self.idx == 0:
                    self.doExitFrame(frt, pag)
                    return
                else:
                    self.idx = self.idx - 1
                    self.pos = self.pos - 1
                    self.col = self.col - 1
        while flds[self.idx][1][0] in ("O", "OCB", "ORB"):
            if (frt == "T" and self.col == 1) or (frt == "C" and self.idx == 0):
                self.doExitFrame(frt, pag)
                return
            self.idx = self.idx - 1
            self.pos = self.pos - 1
            self.col = self.col - 1
        clr = False
        if frt == "T" and self.topf[pag][self.pos][6] == "Y":
            clr = True
        elif frt == "C" and self.colf[pag][self.idx][6] == "Y":
            clr = True
        self.focusField(frt, pag, self.col, clr=clr)

    def clearFrame(self, frt, pag):
        self.skip[pag] = []
        if pag > 0:
            if frt == "T":
                self.last[pag][0] = 0
            else:
                self.last[pag][1] = 0
        if frt == "T":
            flds = self.topEntry[pag]
        else:
            flds = self.colEntry[pag]
        for x in range(0, len(flds)):
            self.clearEntry(frt, pag, (x+1))

    def clearLine(self, pag, row=None, focus=False):
        self.skip[pag] = []
        if row:
            pos = row * self.colq[pag]
        else:
            pos = int(self.pos / self.colq[pag]) * self.colq[pag]
        for y in range(pos, (pos + self.colq[pag])):
            r = int(y / self.colq[pag])
            c = y - (r*self.colq[pag])
            self.c_work[pag][r][c] = ""
            self.c_disp[pag][r][c] = ""
            fldd = self.colEntry[pag][y]
            if self.colEntry[pag][y].winfo_class() == "Text":
                fldd.configure(state="normal")
                fldd.delete("1.0", "end")
                fldd.configure(state="disabled")
            else:
                fldd.configure(state="normal")
                fldd.delete(0, "end")
                fldd.configure(state="disabled")
        if focus:
            self.focusField("C", pag, pos+1)

    def clearEntry(self, frt, pag, col):
        if frt == "T":
            idx = col - 1
            pos = idx = col - 1
            row = int(pos / self.topq[pag])
            if self.topf[pag][idx][1][2:] in ("I","D","d","0","1","2","3"):
                self.t_work[pag][row][idx] = 0
            else:
                self.t_work[pag][row][idx] = ""
            self.t_disp[pag][row][idx] = ""
            fldd = self.topEntry[pag][pos]
            if type(self.topf[pag][idx][1]) in (list, tuple):
                if self.topf[pag][idx][1][0][1:] == "CB":
                    self.doCheckButton(fldd, "clear")
                elif self.topf[pag][idx][1][0][1:] == "RB":
                    self.doRadioButton(fldd, "clear")
            elif self.topf[pag][idx][1][1:] in ("TV", "Tv"):
                fldd.configure(state="normal")
                fldd.delete("1.0", "end")
                fldd.configure(state="disabled")
            else:
                fldd.configure(state="normal")
                fldd.delete(0, "end")
                fldd.configure(state="disabled")
        else:
            pos = col - 1
            idx = pos % self.colq[pag]
            row = int(pos / self.colq[pag])
            if self.colf[pag][idx][1][2:] in ("I","D","d","0","1","2","3"):
                self.c_work[pag][row][idx] = 0
            else:
                self.c_work[pag][row][idx] = ""
            self.c_disp[pag][row][idx] = ""
            fldd = self.colEntry[pag][pos]
            if self.colf[pag][idx][1][1:] in ("TV", "Tv"):
                fldd.configure(state="normal")
                fldd.delete("1.0", "end")
                fldd.configure(state="disabled")
            else:
                fldd.configure(state="normal")
                fldd.delete(0, "end")
                fldd.configure(state="disabled")

    def loadEntry(self, frt, pag, pos, data=None, zero=False):
        """
        where
            frt  =    Frame Type ('T', 'C')
            pag  =    Page Number
            pos  =    Position (0 to (self.rows * self.[t,c]olq) - 1)
            data =    Data to be loaded in work and disp
            zero =    Display zero amounts
        """
        if frt == "T":
            fld = self.topEntry[pag][pos]
            if data is not None:
                if type(self.topf[pag][pos][1]) in (list, tuple):
                    if self.topf[pag][pos][1][0][1:] == "CB":
                        d = CCD(data, "CB", 0)
                    elif self.topf[pag][pos][1][0][1:] == "RB":
                        d = CCD(str(data), "NA", 1)
                    if d.err or d.work == 0:
                        d.disp = ""
                        d.work = 0
                else:
                    d = CCD(data, self.topf[pag][pos][1],
                        self.topf[pag][pos][2][1])
                    if d.err:
                        d.disp = ""
                        if self.topf[pag][pos][1][2:] in (
                                "I","D","d","0","1","2","3"):
                            d.work = 0
                        else:
                            d.work = ""
                    if d.work == 0 and not zero:
                        d.disp = ""
                self.t_disp[pag][0][pos] = d.disp
                self.t_work[pag][0][pos] = d.work
            if type(self.topf[pag][pos][1]) in (list, tuple):
                if self.topf[pag][pos][1][0][1:] == "CB":
                    self.doCheckButton(fld, "load", data)
                elif self.topf[pag][pos][1][0][1:] == "RB":
                    self.doRadioButton(fld, "load", data)
            else:
                state = fld.cget("state")
                fld.configure(state="normal")
                disp = self.t_disp[pag][0][pos]
                if self.topf[pag][pos][1][1:] == "TV":
                    fld.delete("1.0", "end")
                    fld.insert("1.0", disp.rstrip())
                elif self.topf[pag][pos][1][1:] == "Tv":
                    fld.delete("1.0", "end")
                    fld.insert("1.0", disp.rstrip())
                else:
                    fld.delete(0, "end")
                    fld.insert(0, disp.rstrip())
                fld.update_idletasks()
                fld.configure(state=state)
        else:
            fld = self.colEntry[pag][pos]
            row = int(pos / self.colq[pag])
            idx = int(pos % self.colq[pag])
            if data is not None:
                d = CCD(data, self.colf[pag][idx][1],
                    self.colf[pag][idx][2][1])
                if d.err:
                    d.disp = ""
                    if self.colf[pag][idx][1][2:] in (
                            "I","D","d","0","1","2","3"):
                        d.work = 0
                    else:
                        d = ""
                if d.work == 0 and not zero:
                    d.disp = ""
                self.c_disp[pag][row][idx] = d.disp
                self.c_work[pag][row][idx] = d.work
            state = fld.cget("state")
            fld.configure(state="normal")
            disp = self.c_disp[pag][row][idx]
            if self.colf[pag][idx][1][1:] == "TV":
                fld.delete("1.0", "end")
                fld.insert("1.0", disp.rstrip())
            elif self.colf[pag][idx][1][1:] == "Tv":
                fld.delete("1.0", "end")
                fld.insert("1.0", disp.rstrip())
            else:
                fld.delete(0, "end")
                fld.insert(0, disp.rstrip())
            fld.update_idletasks()
            fld.configure(state=state)
        if data is not None:
            return d

    def getEntry(self, frt, pag, pos, wid=True, txt=True, cr=False):
        """
        where
            frt  =  Frame Type ('T', 'C')
            pag  =  Page Number
            pos  =  Position (0 to (self.rows * self.[t,c]olq) - 1)
            wid  =  Return the widget
            txt  =  Return the widget contents
            cr   =  Leave carriage returns
        """
        if frt == "T":
            fwid = self.topEntry[pag][pos]
            if type(self.topf[pag][pos][1]) in (list, tuple):
                if self.topf[pag][pos][1][0][1:] == "CB":
                    ftxt = self.doCheckButton(fwid, "active")
                elif self.topf[pag][pos][1][0][1:] == "RB":
                    ftxt = self.doRadioButton(fwid, "active")
                else:
                    return
            elif self.topf[pag][pos][1][1:] in ("TV", "Tv"):
                ftxt = fwid.get("1.0", "end")
            else:
                ftxt = fwid.get()
        else:
            fwid = self.colEntry[pag][pos]
            if self.topf[pag][pos][1][1:] in ("TV", "Tv"):
                ftxt = fwid.get("1.0", "end")
            else:
                ftxt = fwid.get()
        if not cr:
            ftxt = ftxt.rstrip()
        if wid and txt:
            return (fwid, ftxt)
        elif wid:
            return fwid
        elif txt:
            return ftxt

    def doKeyPressed(self, frt, pag, pos, data):
        self.focusField(frt, pag, pos+1)
        self.doInvoke(data=data)

    def doInvoke(self, fldd=None, data=None):
        # fldd = (frt, pag, col, cmd)
        # data = text
        if fldd:
            self.frt, self.pag, self.col, cmd = fldd
            self.pos = self.col - 1
            if self.frt == "T":
                self.row = int(self.pos / self.topq[self.pag])
                self.idx = self.pos
            else:
                self.row = int(self.pos / self.colq[self.pag])
                self.idx = int(self.pos % self.colq[self.pag])
        elif self.frt == "T" and len(self.topf[self.pag][self.idx]) > 4:
            cmd = self.topf[self.pag][self.idx][7]
        elif self.frt == "C" and len(self.colf[self.pag][self.idx]) > 4:
            cmd = self.colf[self.pag][self.idx][7]
        else:
            cmd = None
        self.loadEntry(self.frt, self.pag, self.pos, data=data)
        self.execEntry((self.frt, self.pag, cmd))
        return True

    def execRB(self, wid, num):
        self.mf.updateStatus("")
        try:
            txt, cmd = self.topEntry[self.pag][self.pos][num][2]
            self.loadEntry("T", self.pag, self.pos, data=txt)
            self.execCommand("T", self.pag, cmd)
        except:
            pass

    def execEntry(self, *opt):
        self.mf.updateStatus("")
        frt, pag, cmd = opt[0]
        self.execCommand(frt, pag, cmd)

    def execCommand(self, frt, pag, cmd):
        # Take focus away from widget
        self.mf.window.focus_set()
        if frt == "T":
            flds = self.topf[pag][self.idx]
            fldd = self.topEntry[pag][self.pos]
        else:
            flds = self.colf[pag][self.idx]
            fldd = self.colEntry[pag][self.pos]
        if type(flds[1]) in (list, tuple):
            if flds[1][0][1:] == "CB":
                fld = self.doCheckButton(fldd, "active")
                fld = CCD(fld, "CB", 0)
            elif flds[1][0][1:] == "RB":
                fld = self.doRadioButton(fldd, "active")
                fld = CCD(fld, "NA", 1)
        elif flds[1][1:] == "TV":
            text = fldd.get("1.0", "end").rstrip()
            fld = CCD(text, flds[1], flds[2][1])
        elif flds[1][1:] == "Tv":
            text = fldd.get("1.0", "end").rstrip()
            fld = CCD(text, flds[1], flds[2][1])
        else:
            fld = CCD(fldd.get(), flds[1], flds[2][1])
        if fld.err:
            ok = ("no", fld.err, fld.data)
        elif len(flds) > 4:
            ok = self.doValidation(frt, pag, self.idx, fld.work)
        else:
            ok = ("ok",)
        if ok[0] == "no":
            if flds[1][0] in ("O", "OCB", "ORB"):
                print("Invalid Output Data %s for (%s)" % (ok[1:], flds))
                sys.exit()
            else:
                self.focusField(frt, pag, self.col, err=ok[1])
        else:
            if ok[0] == "ex":
                fld.work = fld.disp = ok[1]
            if type(flds[1]) in (list, tuple):
                if flds[1][0][1:] == "CB":
                    self.doCheckButton(fldd, "disabled")
                elif flds[1][0][1:] == "RB":
                    self.doRadioButton(fldd, "disabled")
            elif flds[1][1:] in ("TV", "Tv"):
                self.doTextView(frt, pag, fldd, "disable")
            else:
                if flds[1][1] != "H":
                    self.loadEntry(frt, pag, self.pos, data=fld.work)
                self.doEntryField(frt, pag, self.idx, fldd, False, "disable")
            if frt == "T":
                self.t_work[pag][self.row][self.idx] = fld.work
                self.t_disp[pag][self.row][self.idx] = fld.disp
                if cmd:
                    if cmd == "rt":                 # Return True
                        return True
                    ok = cmd(frt, pag, self.row, self.col, self.pos, self.idx,
                        fld.work)
                    if type(ok) == tuple:
                        col, ok = ok
                    else:
                        col = None
                    if ok in (None, "ok"):          # Continue normally
                        if self.col == self.topq[pag]:
                            self.doEndFrame(frt, pag)
                        else:
                            self.focusField(frt, pag, (self.col + 1))
                    elif ok[:2] == "bu":            # Invoke Esc to backup
                        if len(ok) == 2:
                            tm = 1
                        else:
                            tm = int(ok[2:])
                        for x in range(tm):
                            self.backUp(frt=frt, pag=pag)
                    elif ok[:2] == "sk":            # Skip one or more columns
                        if len(ok) == 2:
                            tm = 1
                        else:
                            tm = int(ok[2:])
                        for x in range(0, tm):
                            self.col += 1
                            self.skip[self.pag].append(self.col)
                        if self.col == self.topq[pag]:
                            self.doEndFrame(frt, pag)
                        else:
                            self.focusField(frt, pag, self.col + 1)
                    elif ok[:2] == "ff":            # Force new column focus
                        ok = ok.split("|")          # ff[col]|Message|Clear
                        col = int(ok[0][2:])
                        if col == 0:
                            col = 1
                        if len(ok) == 1:
                            self.focusField(frt, pag, col)
                        elif len(ok) == 2:
                            self.focusField(frt, pag, col, err=ok[1])
                        else:
                            self.focusField(frt, pag, col, err=ok[1], clr=ok[2])
                    elif ok == "xt":                # Force frame exit
                        self.doExitFrame(frt, pag, xit="y")
                    elif ok == "nc":                # End frame, no confirm
                        self.doEndFrame(frt, pag, "N")
                    elif ok == "nd":                # End frame normally
                        self.doEndFrame(frt, pag)
                    elif ok == "rf":                # Refocus same field
                        self.focusField(frt, pag, self.col)
                    elif ok == "nf":                # No focus to next field
                        pass
                    elif col:
                        self.focusField(frt, pag, col, err=ok)
                    else:
                        self.focusField(frt, pag, self.col, err=ok)
                elif self.col == self.topq[pag]:    # End of frame
                    self.doEndFrame(frt, pag)
                else:                               # Focus to next field
                    self.focusField(frt, pag, (self.col + 1))
            else:
                self.c_work[pag][self.row][self.idx] = fld.work
                self.c_disp[pag][self.row][self.idx] = fld.disp
                if cmd:
                    if cmd in ("nd", "nf", "rt"):
                        return True                 # Return True
                    ok = cmd(frt, pag, self.row, self.col, self.pos,
                        self.idx, fld.work)
                    if type(ok) == tuple:
                        col, ok = ok
                    else:
                        col = None
                    if ok in (None, "ok"):          # Continue normally
                        if self.col % self.colq[pag] == 0:
                            self.doEndFrame(frt, pag)
                        else:
                            self.focusField(frt, pag, (self.col + 1))
                    elif ok[:2] == "bu":            # Invoke Esc to backup
                        if len(ok) == 2:
                            tm = 1
                        else:
                            tm = int(ok[2:])
                        for x in range(tm):
                            self.backUp(frt=frt, pag=pag)
                    elif ok == "cl":                # Clear line
                        self.clearLine(pag, row=self.row)
                        col = 1 + (self.row * self.colq[pag])
                        self.focusField(frt, pag, col)
                    elif ok[:2] == "sk":            # Skip one or more columns
                        if len(ok) == 2:
                            tm = 1
                        else:
                            tm = int(ok[2:])
                        for x in range(0, tm):
                            self.col += 1
                            self.skip[self.pag].append(self.col)
                        if (self.col) % self.colq[pag] == 0:
                            self.doEndFrame(frt, pag)
                        else:
                            self.focusField(frt, pag, self.col + 1)
                    elif ok[:2] == "ff":            # Force new column focus
                        ok = ok.split("|")          # ff[col]|Message|Clear
                        col = int(ok[0][2:])
                        if col == 0:
                            col = 1
                        if len(ok) == 1:
                            self.focusField(frt, pag, col)
                        elif len(ok) == 2:
                            self.focusField(frt, pag, col, err=ok[1])
                        else:
                            self.focusField(frt, pag, col, err=ok[1], clr=ok[2])
                    elif ok == "xt":                  # Force frame exit
                        self.doExitFrame(frt, pag)
                    elif ok == "nc":                  # End frame, no confirm
                        self.doEndFrame(frt, pag, "N")
                    elif ok == "nd":                  # End frame normally
                        self.doEndFrame(frt, pag)
                    elif ok == "rf":                  # Refocus same field
                        self.focusField(frt, pag, self.col)
                    elif ok == "nf":                  # No focus to next field
                        pass
                    elif col:
                        self.focusField(frt, pag, col, err=ok)
                    else:
                        self.focusField(frt, pag, self.col, err=ok)
                elif self.col % self.colq[pag] == 0:  # End of frame
                    self.doEndFrame(frt, pag)
                else:                                 # Focus to next field
                    self.focusField(frt, pag, (self.col + 1))

    def doCheckButton(self, grp, action=None, data=None):
        """
        Check Button actions as follows:

        grp         = The widget details
        action      = The action to be taken as follows:
                        active   = Return the active buttons
                        clear    = Clear all buttons
                        disabled = Disable the buttons
                        focus    = Set and focus the buttons
                        load     = Set the button
        data        = The data to be loaded if action is load
        """
        if action == "active":
            return grp[1].get()
        if action == "clear":
            grp[1].set("N")
            grp[0].configure(state="disabled")
            return
        if action == "disable":
            grp[0].configure(state="disabled")
            return
        if action == "focus":
            if self.pag and self.nbf:
                self.nb.select(self.pag - 1)
            grp[0].configure(state="normal")
            if data:
                grp[1].set(data)
            grp[0].focus_set()
            return
        if action == "load":
            grp[1].set(data)

    def doRadioButton(self, grp, action=None, data=None):
        """
        Radio Button actions as follows:

        grp         = The widget details
        action      = The action to be taken as follows:
                        active   = Return the active button
                        clear    = Clear all buttons
                        disabled = Disable the buttons
                        focus    = Set and focus the button
                        load     = Set the button
        data        = The data to be loaded if action is load
        """
        if action == "active":
            return grp[0][1].get()
        if action == "clear":
            grp[0][1].set(None)
            for b in grp:
                b[0].configure(state="disabled")
        elif action == "disabled":
            for b in grp:
                b[0].configure(state="disabled")
        elif action == "focus":
            if self.pag and self.nbf:
                self.nb.select(self.pag - 1)
            grp[0][1].set(data)
            for b in grp:
                b[0].configure(state="normal")
                if b[2][0] == data:
                    b[0].focus_set()
        elif action == "load":
            grp[0][1].set(data)

    def doTextView(self, frt, pag, vwr, act, data=None, tag=True):
        """
        vwr:      The textview widget
        act:      The action to take as follows:
                    disable = disable the field
                    focus   = set and focus the field
        data:     The default data
        """
        if act == "focus":
            if pag and self.nbf:
                self.nb.select(self.pag - 1)
            vwr.configure(state="normal")
            self.has_focus.append(vwr)
            if data is not None:
                self.loadEntry(frt, pag, self.pos, data=data)
                if tag:
                    vwr.tag_add("sel", "1.0", "end")
            vwr.focus_set()
        elif act == "disable":
            vwr.configure(state="disabled")
            vwr.tag_remove("sel", "1.0", "end")
            if vwr in self.has_focus:
                self.has_focus.remove(vwr)

    def doEntryField(self, frt, pag, idx, wid, aut, act, data=None):
        """
        wid:      The entry widget
        aut:      Autocompletion - True or False
        act:      The action to take as follows:
                    disable = disable the field
                    focus   = set and focus the field
        data:     The default data
        """
        if frt == "T":
            flds = self.topf[pag][idx]
        else:
            flds = self.colf[pag][idx]
        if act == "focus":
            wid.configure(state="normal")
            self.has_focus.append(wid)
            if data is not None:
                self.loadEntry(frt, pag, self.pos, data=data)
            if aut and self.mf.rcdic["acnf"] in ("I", "L"):
                wid.clist = []
                odr = None
                if "index" in flds[8]:
                    pos = flds[8]["index"]
                else:
                    pos = 0
                if flds[8]["stype"] == "C" and flds[8]["data"]:
                    for item in flds[8]["data"]:
                        if type(item) in (tuple, list):
                            element = item[pos]
                        else:
                            element = item
                        if type(element) == str:
                            wid.clist.append(element)
                elif flds[8]["stype"] == "R":
                    if "wtype" in flds[8] and flds[8]["wtype"] == "D":
                        dat = flds[8]["where"]
                    else:
                        col = [flds[8]["cols"][pos][0]]
                        tab = []
                        for t in flds[8]["tables"]:
                            tab.append(t)
                        whr = self.setWhere(pag, flds[8])
                        if "group" in flds[8]:
                            grp = flds[8]["group"]
                        else:
                            grp = None
                        if "order" in flds[8]:
                            odr = flds[8]["order"]
                        else:
                            odr = None
                        sql = Sql(self.mf.dbm, tab, prog=__name__)
                        dat = sql.getRec(tables=tab, cols=col, where=whr,
                            group=grp, order=odr)
                    for d in dat:
                        if type(d[0]) == str:
                            wid.clist.append(d[0].strip())
                        else:
                            wid.clist.append(d[0])
                if not odr:
                    wid.clist.sort()
            wid.focus_set()
            wid.selection_range(0, "end")
        elif act == "disable":
            wid.configure(state="disabled")
            if wid in self.has_focus:
                self.has_focus.remove(wid)

    def doValidation(self, frt, pag, idx, wrk):
        if frt == "T":
            des = self.topf[pag][idx][4].strip()
            val = self.topv[pag][idx]
            if type(self.topf[pag][idx][1]) in (list, tuple):
                if self.topf[pag][idx][1][0][1:] == "CB":
                    return ("ok", )
                # Radio Button
                if wrk:
                    return ("ok", )
                else:
                    return ("no", "Invalid %s" % des)
        else:
            des = self.colf[pag][idx][4]
            val = self.colv[pag][idx]
        if not val:
            return ("ok", )
        if val[0] == "efld":
            if frt == "T":
                typ = self.topf[pag][idx][1]
                siz = self.topf[pag][idx][2][1]
            else:
                typ = self.colf[pag][idx][1]
                siz = self.colf[pag][idx][2][1]
            chk = CCD(wrk, typ, siz)
            if chk.err:
                return ("no", chk.err)
            else:
                return ("ok", )
        if val[0] == "notblank":
            if not wrk:
                return ("no", "Invalid %s" % des)
            elif len(val) == 2 and val[1] == "nospaces" and wrk.count(" "):
                return ("no", "Invalid, Spaces Not Allowed")
            else:
                return ("ok", )
        elif val[0] == "notzero":
            if wrk == 0:
                return ("no", "Invalid %s" % des)
            else:
                return ("ok", )
        elif val[0] == "in":
            if wrk in val[1]:
                return ("ok", )
            else:
                return ("no", "Invalid %s" % des)
        elif val[0] == "notin":
            if wrk in val[1]:
                return ("no", "Invalid %s" % des)
            else:
                return ("ok", )
        elif val[0] == "between":
            if wrk < val[1] or wrk > val[2]:
                return ("no", "Invalid %s (%s-%s)" % (des,val[1], val[2]))
            else:
                return ("ok", )
        elif val[0] == "dir":
            if os.path.isdir(wrk):
                return ("ok", )
            else:
                return ("no", "Invalid %s, Not Found" % des)
        elif val[0] == "fle":
            if not wrk and len(val) == 2 and val[1] == "blank":
                return ("ok", )
            wrk = wrk.split(",")
            err = False
            for w in wrk:
                if not getFileName(w, check=True):
                    err = True
                    break
            if not err:
                return ("ok", )
            else:
                return ("no", "Invalid %s, Not Found" % w)
        elif val[0] == "cmd":
            if not wrk:
                return ("ok", )
            if not getFileName(wrk, check=True):
                return ("no", "Invalid %s" % des)
            return ("ex", wrk)
        elif val[0] == "file":
            if os.path.isfile(wrk):
                return ("ok", )
            else:
                return ("no", "Invalid %s, Not Found" % des)
        elif val[0] == "idno":
            if wrk and not luhnFunc(int(wrk)):
                return ("no", "Invalid ID Number")
            else:
                return("ok")
        elif val[0] == "path":
            if os.path.exists(wrk):
                return ("ok", )
            else:
                return ("no", "Invalid %s, Not Found" % des)
        elif val[0] == "email":
            if wrk:
                eml = wrk.split(",")
                for e in eml:
                    if len(val) == 2:
                        no = self.doCheckEmail(e, val[1])
                    else:
                        no = self.doCheckEmail(e, True)
                    if no:
                        return ("no", "Invalid %s, %s" % (des, no))
            elif len(val) == 2 and not val[1]:
                return ("no", "Missing Address")
            return ("ok", )
        elif val[0] == "chars":
            if len(wrk) > val[1]:
                return ("no", "Invalid Size, Only %s Allowed" % val[1])
            return ("ok", )

    def doCheckEmail(self, address, blank=False):
        """
        Check for problems with the supplied e-mail address.
        """
        if not address:
            if blank:
                return
            return "Address is empty"
        by_at_sign = address.split("@")
        if len(by_at_sign) != 2:
            return "Address must have one and only one @"
        name, hostname = by_at_sign
        if not name:
            return "Name of account is missing"
        if "." not in hostname:
            return "Address is missing . after @"
        for component in hostname.split("."):
            if not component:
                return "Domain name (%s) not fully qualified" % hostname

    def doCheckFields(self, fields=None):
        """
        fields is either None for all or a tuple having:
            Page Type i.e. T or C
            Page Number i.e. 0-9
            Range of Column Numbers on the Page i.e. (0,1,2,3,4,8,9) or
                None for all columns
        """
        # Load current field
        try:
            dat = self.getEntry(self.frt, self.pag, self.pos, wid=False)
            self.loadEntry(self.frt, self.pag, self.pos, data=dat)
        except:
            pass
        # Check fields
        if fields is None:
            for pag in range(0, self.pgs+1):
                for fld in range(0, len(self.topf[pag])):
                    ok = self.doCheckField("T", pag, fld)
                    if ok:
                        return ok
                for fld in range(0, len(self.colf[pag])):
                    ok = self.doCheckField("C", pag, fld)
                    if ok:
                        return ok
        elif fields[2] is None:
            if fields[0] == "T":
                for fld in range(0, len(self.topf[fields[1]])):
                    ok = self.doCheckField("T", fields[1], fld)
                    if ok:
                        return ok
            elif fields[0] == "C":
                for fld in range(0, len(self.colf[fields[1]])):
                    ok = self.doCheckField("C", fields[1], fld)
                    if ok:
                        return ok
        else:
            for col in fields[2]:
                ok = self.doCheckField(fields[0], fields[1], col)
                if ok:
                    return ok
        return ("", 0, 0, "")

    def doCheckField(self, frt, pag, fld):
        if frt == "T":
            if len(self.topf[pag][fld]) == 4:
                # Output field
                return
            wrk = self.t_work[pag][0][fld]
            ok = self.doValidation(frt, pag, fld, wrk)
            if ok[0] == "no":
                return (frt, pag, fld, ok[1])
            if ok[0] == "ex":
                self.loadEntry(frt, pag, fld, data=ok[1])
        elif frt == "C":
            if len(self.colf[pag][fld]) == 4:
                # Output field
                return
            for row in range(0, self.colq[pag]):
                wrk = self.c_work[pag][row][fld]
                ok = self.doValidation(frt, pag, fld, wrk)
                if ok[0] == "no":
                    return (frt, pag, fld, ok[1])
                if ok[0] == "ex":
                    self.loadEntry(frt, pag, fld, data=ok[1])

    def doExitFrame(self, frt, pag, xit="n"):
        self.mf.updateStatus("")
        if frt == "T" and self.txit and self.txit[pag]:
            self.txit[pag]()
        elif frt == "C" and self.cxit and self.cxit[pag]:
            self.cxit[pag]()
        elif xit == "y":
            sys.exit()

    def doEndFrame(self, frt, pag, cnf=None):
        """
        frt = The frame type, T or C
        pag = The frame number, 0 - 9
        cnf = None, "y", "n" or "d"
                  None or "n" = No confirmation required
                  "y"         = Confirmation required, either a Y or N
                  "d"         = Same as "y" but Enter defaults to Y
        """
        self.mf.updateStatus("")
        if self.frt != frt or pag != self.pag:
            self.frt = frt
            self.pag = pag
        self.checkButtonsTags(frt, pag, 0)
        if frt == "T":
            if not self.tend or not self.tend[pag]:
                cmd = None
            else:
                cmd = self.tend[pag][0]
                if not cnf:
                    cnf = self.tend[pag][1]
        elif not self.cend or not self.cend[pag]:
            cmd = None
        else:
            cmd = self.cend[pag][0]
            if cnf is None:
                cnf = self.cend[pag][1]
        if cnf:
            cnf = cnf.lower()
        if cnf in (None, "n"):
            ok = "y"
        else:
            ok = self.entryConfirm(mess="", dflt=cnf)
        if ok == "n":
            for x in range(len(self.skip[self.pag]) - 1, -1, -1):
                if self.col == self.skip[self.pag][x]:
                    self.checkButtonsTags(frt, pag, self.col)
                    self.col = self.col - 1
            self.focusField(frt, pag, self.col, ddd="B")
        elif cmd:
            self.window.focus_set()
            cmd()
        elif frt == "C":
            self.advanceLine(pag)

    def advanceLine(self, pag, focus=True):
        self.skip[pag] = []
        row = int((self.last[pag][1] - 1) / self.colq[pag]) + 1
        if row == self.rows[pag]:
            self.scrollScreen(pag, focus)
        else:
            col = (row * self.colq[pag]) + 1
            if focus:
                self.focusField("C", pag, col)
            else:
                self.mstFrame.focus_set()

    def scrollScreen(self, pag, focus=True):
        for x in range(1, self.rows[pag]):
            for y in range(self.colq[pag]):
                d = self.c_work[pag][x][y]
                p = ((x - 1) * self.colq[pag]) + y
                self.loadEntry("C", pag, p, data=d)
        self.clearLine(pag, row=(self.rows[pag] - 1))
        col = (self.rows[pag] * self.colq[pag]) - (self.colq[pag] - 1)
        if focus:
            self.focusField("C", pag, col)
        else:
            self.mstFrame.focus_set()

    def entryConfirm(self, mess="", dflt=""):
        state = self.disableButtonsTags(tags=False)
        self.answer = tk.StringVar()
        if self.mf.rcdic["cnf"] == "N":
            self.answer.set("y")
        else:
            if not mess:
                mess = "Confirm Entry (Y/N)"
            self.mf.updateStatus(mess, bg="red", fg="white")
            for key in ("N", "Y"):
                self.mf.status.bind("<%s>" % key, self.checkYesNo)
                self.mf.status.bind("<%s>" % key.lower(), self.checkYesNo)
            if dflt == "d":
                self.mf.status.bind("<Return>", self.checkYesNo)
                self.mf.status.bind("<KP_Enter>", self.checkYesNo)
            self.mf.status.focus_set()
            self.mf.status.wait_variable(self.answer)
            if dflt == "d" and self.answer.get() in ("return", "kp_enter"):
                self.answer.set("y")
            for key in ("N", "Y"):
                self.mf.status.unbind("<%s>" % key)
                self.mf.status.unbind("<%s>" % key.lower())
            if dflt == "d":
                self.mf.status.unbind("<Return>")
                self.mf.status.unbind("<KP_Enter>")
        self.mf.updateStatus("")
        self.enableButtonsTags(state=state)
        return self.answer.get()

    def checkYesNo(self, event):
        answer = event.keysym.lower()
        if answer.lower() not in ("y", "n", "return", "kp_enter"):
            return
        self.answer.set(answer)

    def checkButtonsTags(self, frt, pag, col):
        if self.butt:
            if col == 0:
                pass
            elif frt == "T" and self.topq[pag] and col != self.topq[pag]:
                col = col % self.topq[pag]
            elif frt == "C" and self.colq[pag] and col != self.colq[pag]:
                col = col % self.colq[pag]
            for num, but in enumerate(self.butt):
                if but[3] == 2:
                    continue
                state = None
                if not but[4] or not but[5]:
                    if but[3] == 0:
                        state = "disabled"
                    elif but[3] == 1:
                        state = "normal"
                if but[4]:
                    if type(but[4][0]) in (list, tuple):
                        for chk in but[4]:
                            if chk == (frt, pag, col):
                                state = "normal"
                    elif tuple(but[4]) == (frt, pag, col):
                        state = "normal"
                if but[5]:
                    if type(but[5][0]) in (list, tuple):
                        for chk in but[5]:
                            if chk == (frt, pag, col):
                                state = "disabled"
                    elif tuple(but[5]) == (frt, pag, col):
                        state = "disabled"
                if state:
                    getattr(self, "B%s" % num).configure(state=state)
        if self.tags:
            first = None
            for num, tag in enumerate(self.tags):
                if not tag[2] and not tag[3]:
                    continue
                if tag[2]:
                    if type(tag[2][0]) in (list, tuple):
                        on = []
                        for l in tag[2]:
                            on.append(list(l))
                    else:
                        on = [list(tag[2])]
                else:
                    on = []
                if tag[3]:
                    if type(tag[3][0]) in (list, tuple):
                        off = []
                        for l in tag[3]:
                            off.append(list(l))
                    else:
                        off = [list(tag[3])]
                else:
                    off = []
                for flag in on:
                    if flag[1] is None:
                        flag[1] = pag
                    if tuple(flag) == (frt, pag, col):
                        self.enableTag(num)
                        if first is None:
                            first = num + 1
                for flag in off:
                    if flag[1] is None:
                        flag[1] = pag
                    if tuple(flag) == (frt, pag, col):
                        self.disableTag(num)
            if first and self.nb.select() == "":
                self.selPage(page=first)

    def disableTag(self, num):
        self.nb.tab(num, compound="right", image=self.nbdimg)
        setattr(self, "dimage%s" % num, self.nbdimg)
        self.nb.update_idletasks()
        if num in self.normaltags:
            self.normaltags.remove(num)

    def enableTag(self, num):
        self.nb.tab(num, compound="right", image=self.nbeimg)
        setattr(self, "eimage%s" % num, self.nbeimg)
        self.nb.update_idletasks()
        if num not in self.normaltags:
            self.normaltags.append(num)

    def disableButtonsTags(self, tags=True):
        state = {"butt": []}
        if self.butt:
            for num in range(len(self.butt)):
                bt = getattr(self, "B%s" % num)
                state["butt"].append((bt, str(bt["state"])))
                bt.configure(state="disabled")
        if tags and self.tags:
            state["tags"] = self.normaltags
            self.normaltags = []
            for num in range(len(self.tags)):
                self.disableTag(num)
        return state

    def enableButtonsTags(self, state=None):
        if state:
            for but in state["butt"]:
                but[0].configure(state=but[1])
            if "tags" in state:
                self.normaltags = []
                for num in range(len(state["tags"])):
                    self.enableTag(num)
                if self.pag:
                    self.selPage(page=self.pag)

    def getWidget(self, frt=None, pag=None, pos=None):
        if not frt:
            frt = self.frt
        if not pag:
            pag = self.pag
        if not pos:
            pos = self.pos
        if frt == "T":
            return self.topEntry[pag][pos]
        else:
            return self.colEntry[pag][pos]

    def setWidget(self, wid, state=None):
        if type(wid) in (list, tuple):
            if wid[0][0].winfo_class() == "TRadiobutton":
                self.doRadioButton(wid, state)
            else:
                self.doCheckButton(wid, state)
            return
        if state in ("normal", "disabled"):
            wid.configure(state=state)
        elif state == "focus":
            wid.focus_set()
        elif state == "hide" and wid not in self.hidden:
            try:
                wid.update_idletasks()
                geom = wid.winfo_manager()
                info = self.getInfo(wid, geom)
                if geom == "grid":
                    wid.grid_forget()
                elif geom == "pack":
                    wid.pack_forget()
                elif geom == "place":
                    wid.place_forget()
                self.hidden[wid] = (geom, info)
            except:
                pass
        elif state == "show" and wid in self.hidden:
            if self.hidden[wid][0] == "grid":
                wid.grid(self.hidden[wid][1])
            elif self.hidden[wid][0] == "pack":
                wid.pack(self.hidden[wid][1])
            elif self.hidden[wid][0] == "place":
                wid.place(self.hidden[wid][1])
            del self.hidden[wid]
        try:
            wid.winfo_toplevel().update_idletasks()
        except:
            pass

    def getInfo(self, wid, geom):
        # Python 2.7.6 geom_info is broken
        root = wid.winfo_toplevel()
        words = root.splitlist(str(root.tk.call(geom, "info", wid)))
        dics = {}
        for i in range(0, len(words), 2):
            key = words[i][1:]
            value = words[i+1]
            if str(value)[:1] == ".":
                value = root._nametowidget(value)
            dics[key] = value
        return dics

    def closeProcess(self):
        self.mstFrame.destroy()
        # Clear dialog bindings
        for bind in self.bindings:
            try:
                self.window.unbind(bind)
            except:
                pass
        # Restore saved bindings
        for bind in self.svebind:
            try:
                self.window.bind(bind[0], bind[1])
            except:
                pass
        # Restore default font
        self.mf.rcdic["dfs"] = self.dfs
        # Restore default theme
        self.mf.setThemeFont(butt=False)
        self.mf.updateStatus("")

class SRec(object):
    """
    This class is used for selection of a range of valid records.
    The following parameters are allowed:

        mf      :   The MainFrame
        screen  :   Screen to build Enquiry On or None for new window
        title   :   A title to display else None for the default
        tables  :   A list of relevant tables e.g. ["genmst","gentrn"]
        cols    :   A list of columns to be printed, all functions MUST be
                        renamed using the 'as' clause containing:
                            column or 'as' name
                            display code
                            display size
                            heading
                            Function key selection
                    e.g. [["omb_state","UI", 5.0,"Stmnt"]],
                          ["omb_date", "D1", 10.0,"Date"],
                          ["sum(omb_amount) as value","SD",13.2,"Value"],
                          ["omb_recon", "NA", 1.0, "R"]]
                    None = All columns
                    SPECIAL:
                        If display code = ("XX",systp), it uses the values in
                            systp subscripted by the value of 'column'  e.g.
                            ('crt_type',('XX',crtrtp),3,'Typ') produces 'Pur'
                        If display code = ("xx",systp), it uses the desc in
                            systp subscripted by the value of 'column' e.g.
                            ('crt_type',('xx',crtrtp),10,'Description')
                            produces 'Purchases'
        where   :   A list of tuples of where conditions e.g.
                    [("omb_recon", "=", "Y")]
                            OR
                    A list of data if wtype = 'D'
                            OR
                    None = No where statement
        group   :   A string for the 'group' statement
                    e.g. "omb_date, omb_state, omb_recon"
                    None = No group statement
        order   :   A string of the 'order by' statement e.g.
                    "omb_state, omb_date"
                    None = No order statement
        wtype   :   'D' = Use the 'where' option as the data
                    'S' = Use the 'where' option as the select statement
                    None = Normal i.e. Create select statement
        extra   :   Additional data for type 'S'
        zero    :   0, 1 or N
                    Return without display if there are '0' or '1' records to
                    select from else "N" to display irrespectively.
        butt    :   A list of tuples of buttons, in addition to Quit, to draw
                        with their respective commands
        sort    :   Whether the columns can be sorted
        deco    :   Whether the window must be decorated or not, default True
    """
    def __init__(self, mf, **args):
        self.mf = mf
        defaults = {
            "screen": None,
            "title": None,
            "tables": [],
            "cols": None,
            "where": None,
            "group": None,
            "order": None,
            "wtype": None,
            "extra": None,
            "sort": True,
            "zero": "N",
            "butt": None,
            "deco": True}
        for nam in args:
            defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        self.doProcess()

    def doProcess(self):
        fltr = False
        if not self.cols:
            self.cols = []
        self.coll = []
        for col in self.cols:
            col = list(col)
            if len(col) == 4:
                col.append("N")
            if col[4].lower() == "f":
                fltr = True
            self.coll.append(col)
        self.setTableFields()
        if self.wtype == "D":
            self.data = self.where
        else:
            self.data = self.getSelectStatement()
            if self.extra:
                for num, exr in enumerate(self.extra):
                    self.data.insert(num, exr)
        if not self.data:
            self.gt = 0
        else:
            self.gt = len(self.data)
        if not self.gt:
            self.selection = None
        elif self.zero == "1" and self.gt == 1:
            self.selection = self.data
        else:
            sc = SelectChoice(self.screen, self.title, self.cols, self.data,
                sort=self.sort, butt=self.butt, fltr=fltr)
            if self.butt and sc.button:
                sc.button(sc.selection)
            elif sc.selection:
                self.selection = sc.selection[1:]
            else:
                self.selection = None

    def setTableFields(self):
        self.l1 = ""                    # string of all columns
        self.l3 = []                    # list of all columns
        self.cols = []
        if not self.coll:
            for tab in self.tables:
                sql = Sql(self.mf.dbm, tab, prog=__name__)
                col = getattr(sql, "%s_col" % tab)
                dic = getattr(sql, "%s_dic" % tab)
                fld = getattr(sql, "%s_fld" % tab)
                self.l1 = "%s%s," % (self.l1, fld)
                self.l3.extend(col)
                for c in col:
                    self.cols.append([c, dic[c][5], dic[c][3], dic[c][2], None])
        else:
            sql = Sql(self.mf.dbm, "ffield", prog=__name__)
            for col in self.coll:
                if col[0][:4].lower() == "sum(" and col[1][1:] == "D":
                    col[0] = "Round(" + col[0]
                    b = col[0].find(")") + 1
                    d = int(round(col[2] * 10) % 10)
                    col[0] = col[0][:b] + ",%s)" % d + col[0][b:]
                else:
                    d = 0
                self.l1 = "%s%s," % (self.l1, col[0])
                nam = removeFunctions(col[0], d)
                c = nam.split(" as ")
                self.l3.append(c[-1:])
                if not col[1] or not col[2] or not col[3]:
                    dat = sql.getRec("ffield", where=[("ff_name",
                        "=", col[0])], limit=1)
                    if dat:
                        if not col[1]:
                            col[1] = dat[3]
                        if not col[2]:
                            col[2] = dat[4]
                        if not col[3]:
                            col[3] = dat[6]
                if len(col) == 5:
                    sch = col[4]
                else:
                    sch = None
                self.cols.append([col[0], col[3], col[2], col[1], sch])
        self.l1 = self.l1[:-1]

    def getSelectStatement(self):
        sql = Sql(self.mf.dbm, self.tables, prog=__name__)
        cols = []
        for col in self.cols:
            cols.append(col[0])
        if self.group:
            # Check that all columns are in the group statement
            chk = self.group.split(",")
            for col in cols:
                if col.count("("):
                    continue
                if col not in chk:
                    chk.append(col)
            self.group = ""
            for col in chk:
                if not self.group:
                    self.group = col
                else:
                    self.group = "%s,%s" % (self.group, col)
        # If no order statement make first non function column the order col
        if not self.order:
            for col in cols:
                if not col.count("("):
                    self.order = col
        return sql.getRec(tables=self.tables, cols=cols, where=self.where,
            group=self.group, order=self.order)

class SChoice(object):
    """
    This class is used for selection of a range of valid choices.
    The following parameters are allowed:

    deco  = Whether the window must be decorated or not, default False
    titl  = The widget title, None = No Title
    head  = The column headings, None = No Headings
    data  = Either a tuple/list or a tuple/list of tuples/lists
    sort  = Whether columns can be sorted, default is False
    typs  = A list of data types for formatting and searchable column i.e.
                [("UI",2), ("NA",30,"Y"), ("SD",13.2)]
    scrn  = A widget to display the choices on
    retn  = D to return the data selected or I to return the index position of
            the data selected or None to return nothing. Defaults to Data
    mode  = S, M or N for Single, Multiple or No (Display Only) select mode
    butt  = A list of tuples of buttons, in addition to Quit, to draw
            with their respective commands.
    selc  = The column to use for Selection if mode = 'M'
    items = A list showing ticked items if mode = 'M' as follows:
               [col, [item, item, item, ........]]
    """
    def __init__(self, mf, **args):
        self.mf = mf
        defaults = {
            "scrn": None,
            "titl": None,
            "head": None,
            "data": None,
            "sort": False,
            "typs": None,
            "retn": "D",
            "mode": "S",
            "butt": None,
            "selc": None,
            "deco": False,
            "items": None}
        for nam in args:
            defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        if self.scrn is None:
            self.scrn = self.mf.body
        self.selection = None
        if self.data:
            if type(self.data[0]) == tuple:
                pass
            elif type(self.data[0]) == list:
                pass
            else:
                data = []
                for txt in self.data:
                    data.append([txt])
                self.data = list(data)
            if self.mode.upper() == "M":
                data = []
                for dat in self.data:
                    dat = list(dat)
                    if self.items and dat[self.items[0]] in self.items[1]:
                        dat.insert(0, "X")
                    else:
                        dat.insert(0, None)
                    data.append(dat)
                self.data = data
            else:
                self.data = list(self.data)
            if self.head and len(self.head) != len(self.data[0]):
                self.head = list(self.head)
                self.head.insert(0, "")
            size = [0] * len(self.data[0])
            for dat in self.data:
                for x, d in enumerate(dat):
                    try:
                        if len(str(d)) > size[x]:
                            size[x] = len(str(d))
                    except:
                        pass
            cols = []
            search = False
            for n, c in enumerate(self.data[0]):
                if self.head:
                    desc = self.head[n]
                else:
                    desc = ""
                if self.typs:
                    if len(self.typs[n]) == 3:
                        typ, siz, srch = self.typs[n]
                    else:
                        typ, siz = self.typs[n]
                        srch = "N"
                    if not siz:
                        siz = size[n]
                else:
                    typ, siz = ("NA", size[n])
                    srch = "N"
                if srch == "Y":
                    search = True
                if self.mode.upper() == "M" and n == 0:
                    cols.append([str(n), desc, siz, "CB", srch])
                else:
                    cols.append([str(n), desc, siz, typ, srch])
                if self.mode.upper() != "M":
                    if type(c) == tuple or type(c) == list:
                        if len(c) > 3:
                            srch = c[4].upper()
            if self.mode.upper() == "M":
                if self.selc is None:
                    # Last column selectable
                    cols[n][4] = "Y"
                else:
                    # Selected column
                    cols[self.selc + 1][4] = "Y"
            elif not search:
                # Make the last column searchable if no searchable columns set
                cols[n][4] = "Y"
            if len(self.data) > 50:
                # Set decoration mode to True
                self.deco = True
            self.sc = SelectChoice(self.scrn, self.titl, cols, self.data,
                sort=self.sort, butt=self.butt)
            if self.butt and self.sc.button:
                self.sc.button(self.sc.selection)
            elif self.sc.selection:
                if self.mode == "S":
                    if self.retn == "D":
                        self.selection = self.sc.selection[1:]
                    elif self.retn == "I":
                        self.selection = self.sc.selection[0]
                elif self.mode == "M":
                    self.selection = []
                    for sel in self.sc.selection:
                        if self.retn == "D":
                            self.selection.append(sel[1:])
                        elif self.retn == "I":
                            self.selection.append(sel[0])

class SelectChoice(object):
    """
    This class displays a scrolled list of items.

    scrn   - The widget to display this widget on else None.
    titl   - The title of the widget else None.
    cols   - A list of tuples of columns to display as follows:
                Name       - The column name from ffield
                Heading    - The column heading
                Size       - As per CCD e.g. 7.2
                Type       - Valid CCD type or else:
                                "CB" for a check box
                                ("XX",systp) to use the code in systp
                                    subscripted by the value of 'data' for
                                    this column e.g.  ('crt_type','XX',3,'Typ')
                                    produces 'Pur'
                                ("xx",systp) to use the desc in systp
                                    subscripted by the value of 'data' for this
                                    column e.g. ('crt_type','xx',10,'Desc')
                                    produces 'Purchases'
                Searchable
                    Y or N to allow search
                    F      to allow filtering of data (auto includes Y)
    data   - A list of lists of data as per the cols.
    lines  - The maximum number of lines to display.
    sort   - Whether the columns can be sorted, defaults to True
    wait   - Whether to wait until mstFrame is destroyed.
    cmnd   - The command to execute with the selected column(s).
    butt   - A list of tuples of buttons, in addition to Exit, to draw,
                with their respective commands and if a selection must
                be returned. An Exit button will override the default one.
    neww   - Whether to create a new window, defaults to True
    deco   - Whether or not the window must be decorated, default True.
    modal  - Whether or not the window is modal.
    live   - Whether the treeview can have focus, defaults to True.
    posn   - The Initial Row to highlight.
    fltr   - If filtering is available, True, False or a list/tuple having:
                (parent=None, transient=None, modal=False, style=None)
    scrl   - Have Scrollbars True or False.
    styl   - Style to use.
    font   - Font to use.
    escape - Allow Escape.
    colr   - List or tuple denoting the fg and bg colours of the row
             depending on the value of the column e.g.
                 (9, {"X": ("red", "white"), "Y": ("blue", "white")})
             column nine's contents are X or Y
    rowc   - Alternate the colours of every x rows e.g.
             rowc = (2, "black", "white", "black", "grey") or
             rowc = 2
    spos   - When doing a search the following applies:
                 A - Search anywhere in the string.
                 B - Search from the beginning of the string.
    """
    def __init__(self, scrn, titl, cols, data, lines=0, sort=True, wait=True, cmnd=None, butt=None, neww=True, deco=True, modal=True, live=True, posn=None, fltr=False, scrl=True, styl="Treeview", font="TkHeadingFont", spos="B", escape=True, colr=None, rowc=1):
        self.scrn = scrn
        self.ocol = cols
        if titl:
            self.titl = titl
        else:
            self.titl = "Available Selections"
        if cols[0][1] or cols[0][3] == "CB":
            self.headings = True
        else:
            self.headings = False
        self.cols = []
        self.fcol = []
        self.cdet = {}
        self.chek = False
        self.srch = None
        for num, col in enumerate(cols):
            if col[3] == "CB":
                self.chek = True
                continue
            col = list(col)
            if len(col) == 4:
                col.append("N")
            elif col[4] == "F":
                fltr = True
            if col[4] == "Y":
                self.srch = num
            if self.headings:
                self.cols.append(col[1])
                self.cdet[col[1]] = [col[2], col[3], col[4]]
            else:
                self.cols.append(col[0])
                self.cdet[col[0]] = [col[2], col[3], col[4]]
        if self.srch is None:
            self.srch = num
        self.data = data
        if self.chek:
            self.srch -= 1
            self.sort = False
        else:
            self.sort = sort
        self.wait = wait
        self.cmnd = cmnd
        if butt is None:
            self.butt = []
        else:
            self.butt = butt
        self.neww = neww
        self.deco = deco
        self.modal = modal
        self.live = live
        if self.live:
            self.mode = "browse"
        else:
            self.mode = "none"
        self.posn = posn
        self.fltr = fltr
        self.scrl = scrl
        self.styl = styl
        if type(font) == str:
            if font not in (
                    "TkDefaultFont",
                    "TkTextFont",
                    "TkFixedFont",
                    "TkMenuFont",
                    "TkHeadingFont"):
                font = "TkHeadingFont"
            self.font = tkfont.Font(name=font, exists=1)
        else:
            self.font = tkfont.Font(font=tuple(font))
        if self.font.cget("size") > 12:
            self.font.configure(size=12)
        if colr is None:
            self.colr = []
        else:
            self.colr = colr
        if rowc and type(rowc) == int:
            self.rowc = (rowc, "black", "white", "black", "gray92")
        elif rowc:
            self.rowc = rowc
        self.spos = spos
        self.escape = escape
        self.selection = None
        if not lines:
            if len(self.data) < 20:
                self.lines = 20
            elif len(self.data) > 30:
                self.lines = 30
            else:
                self.lines = len(self.data)
        else:
            self.lines = lines
        self.setupWidgets()
        self.buildTree()
        if self.wait:
            self.mstFrame.wait_window()

    def setupWidgets(self, clear=False):
        if not clear:
            if self.neww:
                if self.scrn:
                    self.scrn.winfo_toplevel().config(cursor="watch")
                mkw = MkWindow(trans=self.scrn, decor=self.deco,
                    title=self.titl, remov=False, resize=True)
                if not self.deco:
                    mkw.newwin.configure(bg="black", bd=3)
                    mkw.newwin.attributes("-topmost", False)
                    if self.titl:
                        titl = MyLabel(mkw.newwin, text=self.titl, anchor="c")
                        tf = tkfont.nametofont("TkDefaultFont")
                        nf = tkfont.Font(font=(tf.cget("family"),
                            tf.cget("size"), "bold"))
                        titl.configure(font=nf)
                        titl.pack(fill="x", expand="yes")
                self.window = mkw.newwin
                self.mstFrame = MyFrame(self.window)
                self.mstFrame.pack(fill="both", expand="yes")
            else:
                self.window = self.scrn.winfo_toplevel()
                self.mstFrame = MyFrame(self.scrn)
                self.mstFrame.grid(column=0, sticky="nsew")
        else:
            for child in self.mstFrame.winfo_children():
                child.destroy()
            self.bbox.destroy()
        self.mstFrame.columnconfigure(0, weight=1)
        self.mstFrame.rowconfigure(0, weight=1)
        chgt = self.font.metrics("linespace")
        if self.scrn:
            tlin = int((self.scrn.winfo_reqheight() / chgt) * .8)
        else:
            tlin = int((self.window.winfo_screenheight() / chgt) * .8)
            if self.lines > tlin:
                self.lines = tlin
        if self.headings:
            show = ["headings"]
        else:
            show = []
        if self.chek:
            show.append("tree")
        nams = []
        for num, col in enumerate(self.cols):
            nams.append("%s%s" % (col, num))
        style = ttk.Style()
        style.configure(self.styl, font=self.font, rowheight=int(chgt + 2))
        style.configure("%s.Heading" % self.styl, font=self.font)
        self.tree = ttk.Treeview(self.mstFrame, columns=nams, height=self.lines,
            show=show, selectmode=self.mode, style=self.styl)
        if self.chek:
            self.tree.configure(padding=[-15, 0, 0, 0])
        self.tree.grid(column=0, row=0, sticky="nswe")
        if self.scrl:
            vsb = ttk.Scrollbar(self.mstFrame, orient="vertical",
                command=self.tree.yview)
            vsb.grid(column=1, row=0, sticky="ns")
            hsb = ttk.Scrollbar(self.mstFrame, orient="horizontal",
                command=self.tree.xview)
            hsb.grid(column=0, row=1, sticky="ew")
            self.tree.configure(xscroll=hsb.set, yscroll=vsb.set)
        if not self.wait:
            return
        exits = False
        if self.chek:
            accept = False
        if self.butt is not False:
            butt = copyList(list(self.butt))
            for but in butt:
                if but[0].lower() in ("exit", "quit"):
                    exits = True
                elif self.chek and but[0].lower() == "accept":
                    accept = True
            if self.chek and not accept:
                butt.append(("Accept", self.doSelect))
            if not exits:
                butt.append(("Exit", self.doExit))
            if len(self.cols) > 1 and self.fltr:
                butt.insert(0, ("Filter", self.doFilter))
        self.buttons = []
        if butt:
            self.button = None
            self.bbox = MyButtonBox(self.window)
            for num, but in enumerate(butt):
                b = self.bbox.addButton(but[0], (self.doButton, but))
                self.buttons.append(b)

    def doButton(self, button):
        self.button = None
        if len(button) == 3 and button[2] is True:
            self.selection = [self.iids[self.tree.focus()]]
            self.selection.extend(self.data[self.selection[0]])
            self.button = button[1]
            self.closeProcess()
        else:
            if button in self.butt:
                self.closeProcess()
            if type(button[1]) in (list, tuple):
                button[1][0](button[1][1])
            else:
                button[1]()

    def buildTree(self):
        if self.chek:
            try:
                # Create images for check and uncheck
                self.unchek = getImage("uncheck", siz=(20, 20))
                self.dochek = getImage("check", siz=(20, 20))
                self.tree.heading("#0", image=self.unchek, anchor="e",
                    command=functools.partial(self.doToggle, "all"))
                self.tree.column("#0", anchor="e", width=43, minwidth=43,
                    stretch=False)
            except:
                if self.neww:
                    placeWindow(self.window, parent=self.scrn, expose=True)
                showError(self.window, "Image",
                    "Missing check or uncheck image data.")
                return
        for num, col in enumerate(self.cols):
            if not col:
                continue
            nam = "%s%s" % (col, num)
            if self.headings:
                siz, typ, sch = self.cdet[self.cols[num]]
                if type(typ) in (list, tuple):
                    typ = typ[0]
                if typ[0].lower() == "d":
                    text = str(col)
                    anchor = "c"
                elif typ[0].lower() == "s":
                    text = "%s " % col
                    anchor = "e"
                elif typ.lower() in ("ud", "ui", "ul", "us"):
                    text = str(col)
                    anchor = "e"
                else:
                    text = str(col)
                    anchor = "w"
                if self.sort:
                    self.tree.heading(nam, text=text, anchor=anchor,
                        command=lambda c=nam: self.doSort(c, False))
                else:
                    self.tree.heading(nam, text=text, anchor=anchor)
            width = self.font.measure("X" * (int(self.cdet[col][0]) + 1))
            self.tree.column(nam, width=width, minwidth=width, stretch=True)
        if self.rowc:
            self.tree.tag_configure("A", foreground=self.rowc[1],
                background=self.rowc[2])
            self.tree.tag_configure("B", foreground=self.rowc[3],
                background=self.rowc[4])
            self.cchg = []
            for x in range(0, len(self.data), (self.rowc[0] * 2)):
                for y in range(self.rowc[0]):
                    self.cchg.extend([x + y])
        self.iids = {}
        start = None
        last = None
        for num, item in enumerate(self.data):
            iid = self.addItem(num, item)
            if not start or num == self.posn:
                start = iid
            last = iid
        self.tree.tag_configure("font", font=self.font)
        if self.colr:
            for c in self.colr[1]:
                self.tree.tag_configure("%s" % c, foreground=self.colr[1][c][0],
                    background=self.colr[1][c][1])
        if self.escape:
            self.tree.bind("<Escape>", self.doExit)
        if self.scrl:
            self.tree.bind("<Prior>", self.doScroll)
            self.tree.bind("<Home>", self.doScroll)
            self.tree.bind("<Next>", self.doScroll)
            self.tree.bind("<End>", self.doScroll)
        if not self.chek and self.live is True:
            self.tree.bind("<Double-Button-1>", self.doSelect)
            self.tree.bind("<KP_Enter>", self.doSelect)
            self.tree.bind("<Return>", self.doSelect)
        if self.live:
            self.altkey = False
            self.tree.bind("<Key>", self.showEntry)
            # Create a search entry widget
            self.toSearch = tk.StringVar()
            self.entry = MyEntry(self.tree, textvariable=self.toSearch)
            self.toSearch.trace_variable("w", self.doSearch)
            self.entry.bind("<Up>", self.anotherOne)
            self.entry.bind("<Down>", self.anotherOne)
            self.entry.bind("<Escape>", self.hideEntry)
            self.entry.bind("<Return>", self.hideEntry)
            self.entry.bind("<KP_Enter>", self.hideEntry)
            # Set inital selection and focus
            if start:
                self.tree.selection_set(start)
                self.tree.focus(start)
        self.tree.update_idletasks()
        if start is not None and (self.live or self.posn is not None):
            self.tree.see(start)
        elif last:
            # Scroll to last entry
            self.tree.see(last)
        if self.neww:
            self.window.update_idletasks()
            sw = self.window.winfo_screenwidth() * .95
            sh = self.window.winfo_reqheight()
            if self.window.winfo_reqwidth() > sw:
                placeWindow(self.window, parent=self.scrn, size=(sw, sh),
                    expose=True)
            else:
                placeWindow(self.window, parent=self.scrn, expose=True)
        if self.modal:
            self.window.wait_visibility()
            self.window.grab_set()
            self.tree.focus_force()

    def addItem(self, num, item):
        item = list(item)
        if self.chek:
            chk = item.pop(0)
            if chk == "X":
                img = self.dochek
            else:
                img = self.unchek
        for idx, col in enumerate(self.cols):
            dat = item[idx]
            siz, typ, sch = self.cdet[self.cols[idx]]
            if type(typ) in (list, tuple):
                if typ[0] == "XX":
                    item[idx] = typ[1][int(dat)-1][0]
                elif typ[0] == "xx":
                    item[idx] = typ[1][int(dat)-1][1]
                elif typ[0] == "RR":
                    if dat:
                        item[idx] = typ[1][dat]
                    else:
                        item[idx] = dat
            else:
                ccd = CCD(dat, typ, siz)
                item[idx] = ccd.disp
            item[idx] = item[idx].replace("\n", " \n")
        if self.chek:
            if self.rowc and num in self.cchg:
                tags = ["T%s" % num, "A", "font"]
            elif self.rowc:
                tags = ["T%s" % num, "B", "font"]
            else:
                tags = ["T%s" % num, "font"]
            iid = self.tree.insert("", "end", image=img, values=item,
                tags=tags)
            self.tree.tag_bind("T%s" % num, "<ButtonRelease-1>",
                functools.partial(self.doToggle, iid))
            self.tree.tag_bind("T%s" % num, "<KP_Enter>",
                functools.partial(self.doToggle, iid))
            self.tree.tag_bind("T%s" % num, "<Return>",
                functools.partial(self.doToggle, iid))
        else:
            if self.colr and item[self.colr[0]] in self.colr[1]:
                tags = ["font", item[self.colr[0]]]
            elif self.rowc and num in self.cchg:
                tags = ["font", "A"]
            elif self.rowc:
                tags = ["font", "B"]
            else:
                tags = ["font"]
            iid = self.tree.insert("", "end", values=item, tags=tags)
        self.iids[iid] = num
        return iid

    def doToggle(self, *args):
        def set_image(wid, iid):
            img = wid.item(iid, "image")
            if type(img) == tuple:
                img = img[0]
            if img == str(self.unchek):
                wid.item(iid, image=self.dochek)
            else:
                wid.item(iid, image=self.unchek)
        if args[0] == "all":
            wid = self.tree
            img = wid.heading("#0", option="image")[0]
            if img == str(self.unchek):
                wid.heading("#0", image=self.dochek)
            else:
                wid.heading("#0", image=self.unchek)
            children = wid.get_children()
            for child in children:
                set_image(wid, child)
        else:
            set_image(args[1].widget, args[0])

    def showEntry(self, event):
        if event.keysym == "Alt_L":
            self.altkey = True
        elif self.altkey:
            self.altkey = False
        elif event.char == " " or event.char.isalnum():
            self.entry.place(anchor="ne", relx=1)
            self.entry.insert("end", event.char)
            self.entry.focus_set()

    def hideEntry(self, event):
        self.entry.delete(0, "end")
        self.entry.place_forget()
        if event.keysym == "Return" and not self.chek:
            self.doSelect()
        else:
            self.tree.focus_set()

    def doSearch(self, *args):
        pattern = self.toSearch.get()
        if len(pattern) > 0:
            self.executeSearch(pattern)

    def anotherOne(self, event):
        item = self.tree.selection()[0]
        children = self.tree.get_children("")
        for num, child in enumerate(children):
            if child == item:
                break
        if event.keysym == "Up" and num > 0:
            num -= 1
        elif num < (len(children) - 1):
            num += 1
        self.executeSearch(self.toSearch.get(), start=children[num],
            direction=event.keysym)

    def executeSearch(self, pattern, start=None, direction=None):
        children = list(self.tree.get_children())
        if direction and direction == "Up":
            children.reverse()
        for child in children:
            if start and child != start:
                continue
            start = None
            text = self.tree.item(child, "values")[self.srch]
            if text.lower().count(pattern.lower()):
                self.tree.selection_set(child)
                self.tree.update_idletasks()
                self.tree.focus(child)
                self.tree.see(child)
                return True

    def doSelect(self, event=None):
        if self.chek:
            data = []
            children = self.tree.get_children()
            for num, child in enumerate(children):
                img = self.tree.item(child, "image")
                if type(img) == tuple:
                    img = img[0]
                if img == str(self.dochek):
                    dat = [num]
                    dat.extend(self.tree.item(child, "values"))
                    data.append(dat)
        else:
            item = self.tree.selection()[0]
            data = [[self.tree.index(item)]]
            data[0].extend(self.tree.item(item, "values"))
            self.posn = data[0][0]
        self.selection = []
        for dat in data:
            fdat = [dat[0]]
            for idx, col in enumerate(self.cols):
                det = self.cdet[col]
                if type(det[1]) in (list, tuple):
                    typ = dat[idx+1]
                    for i, t in enumerate(det[1][1]):
                        if det[1][0] == "XX" and typ == t[0]:
                            break
                        if det[1][0] == "xx" and typ == t[1]:
                            break
                    tmp = CCD(i+1, "UI", 2).work
                else:
                    tmp = CCD(dat[idx+1], det[1], det[0]).work
                fdat.append(tmp)
            if not self.chek:
                self.selection = fdat
            else:
                self.selection.append(fdat)
        if self.cmnd:
            self.cmnd(self.selection)
        else:
            self.closeProcess()

    def doScroll(self, event):
        iids = self.tree.get_children()
        if event.keysym in ("Home", "End"):
            if event.keysym == "Home":
                iid = iids[0]
                self.tree.yview("moveto", 0)
            else:
                iid = iids[-1]
                self.tree.yview("moveto", 1)
        else:
            self.tree.update_idletasks()
            try:
                idx = iids.index(self.tree.focus())
            except:
                idx = 0
            if event.keysym == "Next":
                idx += self.lines
                if idx >= len(iids):
                    idx = len(iids) - 1
            else:
                idx -= self.lines
                if idx < 0:
                    idx = 0
            iid = iids[idx]
        if self.live:
            self.tree.selection_set(iid)
        self.tree.focus(iid)

    def doSort(self, col, descending):
        """Sort tree contents when a column is clicked on."""
        # grab values to sort
        data = [(self.tree.set(child, col), child) for child in \
            self.tree.get_children("")]
        # reorder data
        data.sort(reverse=descending)
        for indx, item in enumerate(data):
            if not indx:
                start = item[1]
            self.tree.move(item[1], "", indx)
        # switch the heading so that it will sort in the opposite direction
        self.tree.heading(col,
            command=lambda col=col: self.doSort(col, not descending))
        if self.live:
            self.tree.selection_set(start)
        self.tree.focus(start)

    def doFilter(self, event=None):
        for b in self.buttons:
            b.configure(state="disabled")
        title = "Filter Selection"
        fcol = []
        scol = []
        seq = 0
        for num, col in enumerate(self.ocol):
            if len(col) > 4 and col[4].lower() in ("y", "f"):
                fcol.append((seq, num))
                scol.append(col)
                seq += 1
        if type(self.fltr) in (list, tuple):
            self.window.withdraw()
            diag = SimpleDialog(parent=self.fltr[0], trans=self.fltr[1],
                modal=self.fltr[2], style=self.fltr[3], cols=scol,
                title=title)
            diag.sframe.wait_window()
            self.window.deiconify()
        else:
            diag = SimpleDialog(parent=self.window, cols=scol, title=title)
            diag.sframe.wait_window()
        for b in self.buttons:
            b.configure(state="normal")
        tcol = copyList(fcol)
        ldic = {}
        if diag.data:
            for col in tcol:
                if diag.data[col[0]]:
                    ldic["flt%s" % col[0]] = re.compile(
                        str(diag.data[col[0]]), re.IGNORECASE)
                else:
                    fcol.remove(col)
        else:
            self.tree.focus_set()
        olddata = copyList(self.data)
        newdata = []
        for dat in olddata:
            matches = []
            for col in fcol:
                matches.append(ldic["flt%s" % col[0]].search(str(dat[col[1]])))
            if None not in matches:
                newdata.append(dat)
        if not newdata:
            self.doExit()
        elif len(newdata) == 1:
            self.selection = [0] + list(newdata[0])
            self.closeProcess()
        else:
            self.data = copyList(newdata)
            self.setupWidgets(clear=True)
            self.buildTree()

    def doExit(self, event=None):
        self.selection = None
        self.closeProcess()

    def closeProcess(self):
        if self.neww:
            if self.scrn:
                self.scrn.update()
                self.scrn.winfo_toplevel().config(cursor="arrow")
            self.window.destroy()
        else:
            self.mstFrame.destroy()

class ScrollHtml(object):
    """
    A scrolled html widget
    """
    def __init__(self, **args):
        defaults = {
            "butt": None,
            "height": 40,
            "horizontal": True,
            "mess": None,
            "scrn": None,
            "vertical": True}
        for nam in defaults:
            if nam in args:
                setattr(self, nam, args[nam])
            else:
                setattr(self, nam, defaults[nam])
        self.setVariables()
        self.doProcess()
        self.frame.wait_window()
        if self.scrn:
            self.root.destroy()
        else:
            self.window.destroy()

    def setVariables(self):
        if self.butt:
            self.butt.insert(0, ("Continue", self.execShCmd))
        else:
            self.butt = [("Continue", self.execShCmd)]

    def doProcess(self):
        if self.scrn:
            self.window = self.scrn.winfo_toplevel()
            self.root = MyFrame(self.scrn, borderwidth=5, relief="ridge")
            if getManager(self.scrn) == "grid":
                self.root.grid(column=0)
            else:
                self.root.pack(fill="both", expand="yes")
        else:
            trans, resiz, decor = None, True, True
            win = MkWindow(trans=trans, resiz=resiz, decor=decor, frame=True)
            self.window = win.newwin
            if self.window.state() == "withdrawn":
                self.window.deiconify()
            self.window.grab_set()
            self.root = win.newfrm
        bbox = MyButtonBox(self.root)
        self.binds = []
        for but in self.butt:
            b = bbox.addButton(but[0], but[1])
            pos = b.cget("underline")
            if pos != -1:
                self.binds.append(b.cget("text")[pos])
        self.frame = HtmlFrame(self.root,
            vertical_scrollbar=self.vertical,
            horizontal_scrollbar=self.horizontal,
            fontscale=1.5)
        for key in ("Left", "Right", "Up", "Down", "Prior", "Next"):
            self.window.bind("<%s>" % key, self._scroll)
        self.frame.set_content(self.mess)
        self.frame.pack(fill="both", expand="yes")

    def _scroll(self, event):
        if event.keysym == "Left":
            self.frame.html.xview("scroll", -1, "units")
        elif event.keysym == "Right":
            self.frame.html.xview("scroll", 1, "units")
        elif event.keysym == "Up":
            self.frame.html.yview("scroll", -1, "units")
        elif event.keysym == "Down":
            self.frame.html.yview("scroll", 1, "units")
        elif event.keysym == "Prior":
            self.frame.html.yview("scroll", -1, "pages")
        elif event.keysym == "Next":
            self.frame.html.yview("scroll", 1, "pages")

    def execShCmd(self, *args):
        if args and self.butt[args[0]][1]:
            self.butt[args[0]][1]()
        else:
            self.closeProcess()

    def closeProcess(self):
        self.frame.destroy()
        for c in self.binds:
            self.window.unbind("<Key-Alt_L>%s" % c.lower())
            self.window.unbind("<Key-Alt_L>%s" % c.upper())
        for key in ("Left", "Right", "Up", "Down", "Prior", "Next"):
            self.window.unbind("<%s>" % key)

class ScrollText(object):
    """
    A scrolled text widget
    """
    def __init__(self, **args):
        defaults = {
            "butt": None,
            "height": 40,
            "horizontal": True,
            "font": None,
            "mess": None,
            "select": False,
            "scrn": None,
            "title": "",
            "vertical": True,
            "width": 80,
            "wrap": False,
            "wait": True}
        for nam in defaults:
            if nam in args:
                setattr(self, nam, args[nam])
            else:
                setattr(self, nam, defaults[nam])
        self.setVariables()
        self.doProcess()
        if self.wait:
            self.frame.wait_window()

    def setVariables(self):
        if self.butt:
            self.butt.insert(0, ("Continue", self.execStCmd))
        elif self.butt is None:
            self.butt = [("Continue", self.execStCmd)]
        else:
            self.butt = []

    def doProcess(self):
        if self.scrn:
            self.window = self.scrn.winfo_toplevel()
            self.frame = MyFrame(self.scrn, borderwidth=5, relief="ridge")
            if getManager(self.scrn) == "grid":
                self.frame.grid(column=0)
            else:
                self.frame.pack()
        else:
            trans, resiz, decor = None, True, True
            win = MkWindow(trans=trans, title=self.title,
                resiz=resiz, decor=decor, frame=True)
            self.window = win.newwin
            self.frame = win.newfrm
        bbox = MyButtonBox(self.frame)
        self.binds = []
        for but in self.butt:
            b = bbox.addButton(but[0], but[1])
            pos = b.cget("underline")
            if pos != -1:
                self.binds.append(b.cget("text")[pos])
        self.text = MyText(self.frame, font=self.font, width=self.width,
            height=self.height)
        if self.vertical:
            vert = ttk.Scrollbar(self.frame, orient="vertical")
            vert.config(command=self.text.yview)
            self.text.config(yscrollcommand=vert.set)
            vert.pack(fill="y", expand="yes", side="right")
        if self.horizontal:
            horiz = ttk.Scrollbar(self.frame, orient="horizontal")
            horiz.config(command=self.text.xview)
            self.text.config(xscrollcommand=horiz.set)
            horiz.pack(fill="x", expand="yes", side="bottom")
        if self.mess:
            self.text.insert("1.0", self.mess)
        if self.wrap:
            self.text.configure(wrap="word")
        else:
            self.text.configure(wrap="none")
        if self.select:
            self.text.configure(state="normal")
        else:
            self.text.configure(state="disabled")
        self.text.pack(anchor="center", fill="both", expand="yes")
        if self.scrn:
            self.scrn.update_idletasks()
            self.frame.grab_set()
        else:
            placeWindow(self.window, expose=True)
            self.window.wait_visibility()
            self.window.grab_set()
        self.text.focus_set()

    def execStCmd(self, *args):
        if args and self.butt[args[0]][1]:
            self.butt[args[0]][1]()
        else:
            self.closeProcess()

    def closeProcess(self):
        for c in self.binds:
            self.window.unbind("<Key-Alt_L>%s" % c.lower())
            self.window.unbind("<Key-Alt_L>%s" % c.upper())
        self.frame.destroy()
        if not self.scrn:
            self.window.destroy()

class ProgressBar(object):
    """
    A progress bar widget where:

        scrn -  The parent screen.
        inn  -  Pack the progressbar in this progressbar's window.
        typ  -  The type of progressbar
                    P = Printing the Report
                    G = Generating the Report (Pulse Step)
                    F = Updating Files
                    or a tuple of (typ and text)
        mxs  -  With typ P and F this is the maximum number of records.
        esc  -  Create a Quit button to terminate the job
    """
    def __init__(self, scrn, **args):
        self.scrn = scrn
        defaults = {
            "inn": None,
            "typ": "P",
            "mxs": 0,
            "esc": False}
        for nam in defaults:
            if nam in args:
                if nam == "typ":
                    text = ""
                    if type(args[nam]) in (list, tuple):
                        self.typ = args[nam][0]
                        text = args[nam][1]
                    else:
                        self.typ = args[nam]
                else:
                    setattr(self, nam, args[nam])
            else:
                setattr(self, nam, defaults[nam])
                if nam == "typ":
                    text = ""
        if not self.mxs:
            self.typ = "G"
        if self.typ == "G":
            mode = "indeterminate"
        else:
            mode = "determinate"
        if self.typ in ("G", "P") and not text:
            text = "Generating the Report ... Please Wait"
        elif self.typ == "F" and not text:
            text = "Updating Files ... Please Wait"
        elif not text:
            text = self.typ
            self.typ = "P"
        self.quit = False
        if self.inn:
            self.esc = False
            self.pbframe = MyFrame(self.inn.pbframe, padding=2)
            self.pbframe.pack(fill="x", expand="yes")
        else:
            self.pbframe = MyFrame(self.scrn, padding=2, relief="ridge")
            self.pbframe.place(anchor="center", relx=0.5, rely=0.5)
        self.txtlab = MyLabel(self.pbframe, text=text, anchor="c", width=60,
            borderwidth=2, relief="raised")
        self.txtlab.pack(anchor="n", fill="x", expand="yes")
        self.pbar = ttk.Progressbar(self.pbframe, mode=mode,
            length=self.txtlab.winfo_width())
        self.pbar.pack(fill="x", expand="yes", pady=0)
        if self.esc:
            if self.typ == "G":
                self.pbt = threading.Thread(target=self.pbar.start, args=())
            else:
                self.pbar.configure(maximum=self.mxs)
                self.pbt = threading.Thread(target=None, args=())
            self.pbt.start()
            but = MyButton(self.pbframe, cmd=self.quitProgress, text="Quit",
                underline=0)
            but.pack(anchor="n", fill="x", expand="yes")
        elif self.typ == "G":
            self.pbar.start()
        else:
            self.pbar.configure(maximum=self.mxs)

    def displayProgress(self, value=0):
        if self.typ == "G":
            self.pbar.step()
        else:
            self.pbar.configure(value=value + 1)
        self.scrn.update()

    def closeProgress(self):
        if self.esc and self.pbt.is_alive() is False:
            self.pbt.join()
        self.pbframe.destroy()
        self.pbframe.update()

    def quitProgress(self):
        self.quit = True

class ShowEmail(object):
    def __init__(self, parent=False, message=None, embed=None, attach=None):
        mkw = MkWindow(decor=True, trans=parent, size=(800, -1))
        self.window = mkw.newwin
        if message or embed:
            self.window.configure(width=824, height=600)
        vbox = MyFrame(self.window)
        vbox.pack(fill="both", expand="yes")
        if message.rstrip():
            if embed:
                height = 300
            else:
                height = 600
            sw1 = ScrollWindow(vbox, height=height)
            sw1.pack(fill="both", expand="yes")
            text = MyText(sw1.interior, wrap="word")
            text.delete("1.0", "end")
            text.insert("1.0", message)
            text.configure(state="disabled")
            text.pack(fill="both", expand="yes")
        if embed:
            if message.rstrip():
                height = 300
            else:
                height = 600
            sw2 = ScrollWindow(vbox, height=height)
            sw2.pack(fill="both", expand="yes")
            for r, image in enumerate(embed):
                pi = Image.open(image)
                ti = ImageTk.PhotoImage(image=pi)
                lb = MyLabel(sw2.interior)
                lb.img = ti
                lb.config(image=ti, anchor="n")
                lb.pack(fill="both", expand="yes")
        if attach:
            hed = MyLabel(vbox, color=True, text="Attachment: %s" % attach,
                relief="raised")
            hed.pack(fill="x", expand="yes")
        butt = MyButton(vbox, cmd=self.doClose, text="Close")
        butt.pack()
        placeWindow(self.window, parent, expose=True)
        self.window.wait_visibility()
        self.window.grab_set()
        self.window.wait_window()

    def doClose(self, *args):
        self.window.destroy()

class ShowImage(object):
    def __init__(self, vbox, flenam, wrkdir=None, msiz=400, crop=False):
        self.vbox = vbox
        self.crop = crop
        flenam = getFileName(flenam, wrkdir=wrkdir)
        if flenam:
            pilimg = Image.open(flenam)
            width = int(pilimg.size[0])
            height = int(pilimg.size[1])
            relate = float(height / (width * 1.0))
            if width > height:
                width = msiz
                height = int(width * relate)
            else:
                height = msiz
                width = int(height / relate)
            self.currentImage = {}
            image = pilimg.resize((width, height), Image.ANTIALIAS)
            self.currentImage["data"] = image
            self.frm = MyFrame(self.vbox, width=width, height=height)
            if self.vbox:
                mgr = getManager(self.vbox)
                if mgr == "grid":
                    self.frm.grid(column=0, sticky="nsew")
                    self.frm.grid_propagate(False)
                else:
                    self.frm.pack(fill="both", expand="yes")
                    self.frm.pack_propagate(False)
            else:
                self.frm.pack(fill="both", expand="yes")
                self.frm.pack_propagate(False)
            self.cnv = tk.Canvas(self.frm)
            if self.crop:
                self.cnv.bind("<ButtonPress-1>", self.on_mouse_down)
                self.cnv.bind("<B1-Motion>", self.on_mouse_drag)
                self.cnv.bind("<Button-3>", self.on_right_click)
                self.but = MyButton(self.vbox, text="Crop/Print")
                self.but.bind("<ButtonRelease-1>", self.doCrop)
                image = image.convert("L")
            image = ImageTk.PhotoImage(image=image)
            self.currentImage["photo"] = image
            self.cnv.xview_moveto(0)
            self.cnv.yview_moveto(0)
            self.cnv.create_image(0, 0, image=image, anchor="nw", tags="img")
            self.cnv.pack(side="left", fill="both", expand="yes")
            self.cnv.ilist = [image]
            if self.crop:
                self.but.pack(fill="x", expand="yes")
                self.item = None
                self.lvar = tk.BooleanVar()
                self.lvar.set(True)
                self.frm.wait_variable(self.lvar)
        else:
            self.cnv = None

    def on_mouse_down(self, event):
        if self.item is None:
            self.anchor = (self.cnv.canvasx(event.x), self.cnv.canvasy(event.y))

    def on_mouse_drag(self, event):
        bbox = self.anchor + (self.cnv.canvasx(event.x),
            self.cnv.canvasy(event.y))
        if self.item is None:
            self.item = self.cnv.create_rectangle(bbox,outline="red",width=2)
        else:
            self.cnv.coords(self.item, *bbox)

    def on_right_click(self, event):
        found = self.cnv.find_all()
        for iid in found:
            if self.cnv.type(iid) == "rectangle":
                self.cnv.delete(iid)
        self.item = None

    def doCrop(self, event):
        if self.item:
            box = tuple((int(round(v)) for v in self.cnv.coords(self.item)))
            self.roi = self.currentImage["data"].crop(box)  # crop region
        else:
            self.roi = self.currentImage["data"]            # whole image
        self.lvar.set(False)

    def destroyImage(self):
        if self.cnv:
            self.cnv.destroy()
            if self.crop:
                self.but.destroy()
            self.frm.destroy()

class TartanConfig(object):
    """
    This class is used to create and modify the tartanrc file
    """
    def __init__(self, mf, **args):
        self.mf = mf
        defaults = {
            "rcfile": None,
            "rcdic": None,
            "dbskp": False,
            "level": 3}
        for nam in args:
            defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        self.setVariables()
        self.drawDialog()
        self.mf.startLoop()
        self.mf.setThemeFont()

    def setVariables(self):
        if not self.rcdic:
            self.rcdic = loadRcFile(rcfile=self.rcfile, default=True)
            self.rcfile = os.path.normpath(self.rcdic["name"])
        self.edit = False

    def drawDialog(self):
        pth = {
            "stype":  "F",
            "types":  "dir"}
        if sys.platform == "win32":
            typ = "*.exe"
        else:
            typ = "*"
        fle = {
            "stype":  "F",
            "types":  "fle",
            "ftype":  (("Executable", "%s" % typ),)}
        self.loadFonts()
        aft = {
            "stype": "C",
            "titl": "Available Fonts",
            "data": self.aft}
        fft = {
            "stype": "C",
            "titl": "Available Fonts",
            "data": self.fft}
        self.loadThemes()
        thm = {
            "stype": "C",
            "titl": "Available Themes",
            "data": self.thm}
        fgc = {
            "stype": "X",
            "title": "Select Colour",
            "initc": "foreground"}
        bgc = {
            "stype": "X",
            "title": "Select Colour",
            "initc": "background"}
        self.rcfile = os.path.normpath(self.rcdic["name"])
        self.dbdir = os.path.normpath(self.rcdic["dbdir"])
        r1s = (
            ("SQLite", "S"),
            ("PostgreSQL", "P"))
        r2s = (("Left", "L"), ("Centre", "C"), ("Right", "R"))
        r3s = (("Yes", "Y"), ("No", "N"))
        r4s = (("In-Line", "I"), ("List", "L"), ("None", "N"))
        r5s = (("Trash", "T"), ("Delete", "D"), ("Keep", "K"))
        r6s = (
            ("Red", "R"),
            ("Green", "G"),
            ("Blue", "B"),
            ("Custom", "C"),
            ("Default", "D"))
        tag = (
            ("DataBase",None,("T",1,1),("T",3,0)),
            ("General",None,("T",1,1),("T",3,0)),
            ("Dialog",None,("T",1,1),("T",3,0)))
        fld = (
            (("T",0,0,0),"IFF",65,"Configuration File","",
                self.rcfile,"N",self.doRcFile,None,None,("notblank",),None,
                "This is the main configuration file for a Tartan "\
                "Installation. There can be as many of these as you "\
                "desire, each one accessing a different set of choices. "\
                "This file is loaded by the ms0000 module and the default "\
                "file can be altered with the -r option. Please refer "\
                "to the manual for more detailed information."),
            (("T",1,0,0),("IRB",r1s),0,"Database Engine","",
                "S","N",self.doEngine,None,None,None,None,
                "Use SQLite for single-user "\
                "installations and PgSQL for multi-user installations.\n\n"\
                "Single-user: Only one user will be accessing the database "\
                "at any time.\n\nMulti-User: Many users could be accessing "\
                "the database at any time."),
            (("T",1,1,0),"ILA",30,"Database Name","",
                self.rcdic["dbname"],"N",self.doName,None,None,("notblank",)),
            (("T",1,2,0),"IFF",50,"Files Directory","",
                "","N",self.doDir,pth,None,("notblank",),None,
                "The Directory where the Database will be located."),
            (("T",1,3,0),"INA",30,"Host Name","",
                "","N",None,None,None,("notblank",)),
            (("T",1,4,0),"IUI",4,"Port Number","",
                "","N",None,None,None,("efld",)),
            (("T",1,5,0),"INA",30,"Administrator","",
                self.rcdic["dbuser"],"N",self.doAdmin,None,None,("efld",),None,
                "The Database Administrator."),
            (("T",1,6,0),"IHA",30,"Password","",
                self.rcdic["dbpwd"],"N",self.doPwd,None,None,None,None,
                "The Database Administrator's Password."),
            (("T",2,0,0),"IFF",50,"Backup Path","",
                "","N",self.doDir,pth,None,("notblank",),None,
                "The Directory where all Backup Archives are stored."),
            (("T",2,1,0),"IFF",50,"Work Path","",
                "","N",self.doDir,pth,None,("notblank",),None,
                "The Directory where all Work Files are stored."),
            (("T",2,2,0),"IFF",50,"Upgrade Path","",
                "","N",self.doDir,pth,None,("notblank",),None,
                "The Directory where System Upgrades are stored."),
            (("T",2,3,0),"IFF",50,"PDF Viewer","",
                self.rcdic["vwr"],"N",self.doVwr,fle,None,("cmd",),None,
                "The Path to an External Program used to view PDF files. "\
                "This is only required if pymupdf is not installed."),
            (("T",2,4,0),"IFF",50,"Print Command","",
                self.rcdic["prn"],"N",None,fle,None,("efld",),None,
                "The Path and Options to an External Program used to "\
                "print PDF files. Use %p% for the Printer and %f% for "\
                "the File. These parameters will be replaced with the "\
                "Actual items when activated. If this is left blank the "\
                "Default application will be used, where possible."),
            (("T",2,5,0),"IFF",50,"Spreadsheet Reader","",
                self.rcdic["exp"],"N",self.doExport,fle,None,None,None,
                "The Path to an External Program used to read Spreadsheets. "\
                "If this is left blank the default application will be "\
                "used, where possible."),
            (("T",2,6,0),"INA",10,"Screen Geometry","",
                self.rcdic["geo"],"N",self.doGeometry,None,None,("notblank",),
                None,"The screen geometry, enter 0 to get the default"),
            (("T",2,7,0),("IRB",r2s),0,"Screen Placement","",
                self.rcdic["plc"],"N",None,None,None,None),
            (("T",2,8,0),("IRB",r3s),0,"Show Tartan Image","",
                self.rcdic["img"].upper(),"N",None,None,None,None),
            (("T",2,9,0),("IRB",r3s),0,"Enforce Confirm","",
                self.rcdic["cnf"].upper(),"N",None,None,None,None),
            (("T",2,10,0),("IRB",r4s),0,"Auto-completion","",
                self.rcdic["acnf"].upper(),"N",None,None,None,None),
            (("T",2,11,0),("IRB",r3s),0,"Tool-tips","",
                self.rcdic["ttip"].upper(),"N",None,None,None,None),
            (("T",2,12,0),("IRB",r3s),0,"Error Alarm","",
                self.rcdic["errs"].upper(),"N",None,None,None,None),
            (("T",2,13,0),("IRB",r5s),0,"Work Files","",
                self.rcdic["wrkf"].upper(),"N",None,None,None,None),
            (("T",2,14,0),("IRB",r3s),0,"Automatically Apply","",
                self.rcdic["wrka"].upper(),"N",None,None,None,None),
            (("T",3,0,0),"INA",30,"Menu Font Name","",
                self.rcdic["mft"],"N",self.doFtNam,aft,None,("in",self.aft),
                None,"The Font to be used for the Menu"),
            (("T",3,0,0),"IUI",2,"Size","Menu Font Size",
                self.rcdic["mfs"],"N",self.doFtSiz,None,None,("notzero",),
                None,"The Size of the Menu Font"),
            (("T",3,1,0),"INA",30,"Default Font Name","",
                self.rcdic["dft"],"N",self.doFtNam,fft,None,("in",self.fft),
                None,"The Font to be used for All Other Widgets"),
            (("T",3,1,0),"IUI",2,"Size","Default Font Size",
                self.rcdic["dfs"],"N",self.doFtSiz,None,None,("notzero",),
                None,"The Size of the Default Font"),
            (("T",3,2,0),"INA",30,"Theme","",
                self.rcdic["theme"],"N",self.doTheme,thm,None,("in",self.thm),
                None,"The Theme to be used"),
            (("T",3,3,0),("IRB",r6s),0,"Label Colour Scheme","",
                self.rcdic["lsc"],"N",self.doScheme,None,None,None,
                None,"The Colour Scheme to be used. Default is the "\
                "selected theme's colour scheme."),
            (("T",3,4,0),"INA",20,"Normal Label      FG","",
                self.rcdic["nfg"],"N",self.doColour,fgc,None,("notblank",),
                None,"Normal Label Foreground Colour"),
            (("T",3,4,0),"INA",20,"BG","",
                self.rcdic["nbg"],"N",self.doColour,bgc,None,("notblank",),
                None,"Normal Label Background Colour"),
            (("T",3,4,0),"OTv",10," "),
            (("T",3,5,0),("IRB",r6s),0,"Button Colour Scheme","",
                self.rcdic["bsc"],"N",self.doScheme,None,None,None,
                None,"The Colour Scheme to be used. Default is the "\
                "selected theme's colour scheme."),
            (("T",3,6,0),"INA",20,"Normal Button     FG","",
                self.rcdic["bfg"],"N",self.doColour,fgc,None,("notblank",),
                None,"Normal Label and Button Foreground Colour"),
            (("T",3,6,0),"INA",20,"BG","",
                self.rcdic["bbg"],"N",self.doColour,bgc,None,("notblank",),
                None,"Normal Label and Button Background Colour"),
            (("T",3,6,0),"OTv",10," "),
            (("T",3,7,0),"INA",20,"Focus Button      FG","",
                self.rcdic["ffg"],"N",self.doColour,fgc,None,("notblank",),
                None,"Focussed Button Foreground Colour"),
            (("T",3,7,0),"INA",20,"BG","",
                self.rcdic["fbg"],"N",self.doColour,bgc,None,("notblank",),
                None,"Focussed Button Background Colour"),
            (("T",3,7,0),"OTv",10," "),
            (("T",3,8,0),"INA",20,"Disable Button    FG","",
                self.rcdic["dfg"],"N",self.doColour,fgc,None,("notblank",),
                None,"Disabled Button Foreground Colour"),
            (("T",3,8,0),"INA",20,"BG","",
                self.rcdic["dbg"],"N",self.doColour,bgc,None,("notblank",),
                None,"Disabled Button Background Colour"),
            (("T",3,8,0),"OTv",10," "),
            (("T",3,9,0),"INA",20,"Booking Query     FG","",
                self.rcdic["qfg"],"N",self.doColour,fgc,None,("notblank",),
                None,"Booking Query Foreground Colour"),
            (("T",3,9,0),"INA",20,"BG","",
                self.rcdic["qbg"],"N",self.doColour,bgc,None,("notblank",),
                None,"Booking Query Background Colour"),
            (("T",3,9,0),"OTv",10," "),
            (("T",3,10,0),"INA",20,"Booking Confirmed FG","",
                self.rcdic["cfg"],"N",self.doColour,fgc,None,("notblank",),
                None,"Booking Confirm Foreground Colour"),
            (("T",3,10,0),"INA",20,"BG","",
                self.rcdic["cbg"],"N",self.doColour,bgc,None,("notblank",),
                None,"Booking Confirmed Background Colour"),
            (("T",3,10,0),"OTv",10," "),
            (("T",3,11,0),"INA",20,"Booking Settled   FG","",
                self.rcdic["sfg"],"N",self.doColour,fgc,None,("notblank",),
                None,"Booking Settled Foreground Colour"),
            (("T",3,11,0),"INA",20,"BG","",
                self.rcdic["sbg"],"N",self.doColour,bgc,None,("notblank",),
                None,"Booking Settled Background Colour"),
            (("T",3,11,0),"OTv",10," "))
        but = (
            ("Save", None, self.doSave, 1, ("T",1,1), ("T",0,1)),
            ("Quit", None, self.doQuit, 1, None, None))
        tnd = (
            (self.doEnd, "n"),
            (self.doEnd, "n"),
            (self.doEnd, "n"),
            (self.doEnd, "y"))
        txt = (
            self.doExit,
            self.doExit,
            self.doExit,
            self.doExit)
        self.df = TartanDialog(self.mf, eflds=fld, tags=tag, butt=but,
            tend=tnd, txit=txt, clicks=self.doClick, focus=False)
        if self.dbskp:
            self.df.loadEntry("T", 0, 0, data=self.rcfile)
            self.df.doKeyPressed("T", 0, 0, self.rcfile)
            self.dbase = self.rcdic["dbase"]
            self.df.disableTag(0)
            self.doEnd()
        else:
            self.df.focusField("T", 0, 1)

    def doRcFile(self, frt, pag, r, c, p, i, w):
        if self.level < 3 and w != self.rcfile:
            return "Invalid File"
        if not os.path.isdir(os.path.dirname(w)):
            return "Invalid Directory (%s)" % os.path.dirname(w)
        self.rcfile = w
        self.edit = True

    def doClick(self, *opts):
        if not self.edit or opts == (0, 0):
            return
        self.df.focusField("T", opts[0][0], opts[0][1] + 1)

    def doEngine(self, frt, pag, r, c, p, i, w):
        if w == "P":
            txt = "psycopg2"
            mod = "psycopg2"
        else:
            txt = "pysqlite"
            mod = "sqlite3"
        if not chkMod(mod):
            return "Database Driver (%s) Missing" % txt
        self.dbase = w
        if self.dbase == "P":
            self.df.topv[1][2] = ("efld",)
            self.df.topv[1][3] = ("notblank",)
            self.df.topv[1][5] = ("notblank",)
            self.df.topv[1][6] = ("notblank",)
            self.df.loadEntry(frt, pag, p+2, data="")
        else:
            self.df.topv[1][2] = ("notblank",)
            self.df.topv[1][3] = ("efld",)
            self.df.topv[1][5] = ("efld",)
            self.df.topv[1][6] = ("efld",)
            self.df.loadEntry(frt, pag, p+2, data=self.rcdic["dbdir"])
            self.df.loadEntry(frt, pag, p+3, data="")

    def doName(self, frt, pag, r, c, p, i, w):
        if self.dbase == "P":
            self.df.loadEntry(frt, pag, p+1, data="")
            self.df.loadEntry(frt, pag, p+2, data="localhost")
            return "sk1"

    def doDir(self, frt, pag, r, c, p, i, w):
        if not os.path.isdir(w):
            ok = askQuestion(screen=self.mf.body, head="Missing Directory",
                mess="Directory %s Does Not Exist, Would You Like to Create "\
                "it Now?" % w, default="yes")
            if ok == "no":
                return "Invalid Directory"
            try:
                os.makedirs(w)
            except:
                return "Cannot Create %s Directory" % w
        if self.dbase == "S" and pag == 1:
            for x in range(1, 5):
                self.df.loadEntry(frt, pag, p+x, data="")
            return "nd"

    def doAdmin(self, frt, pag, r, c, p, i, w):
        if self.dbase == "P" and not w:
            return "Invalid Administrator"

    def doPwd(self, frt, pag, r, c, p, i, w):
        if not w:
            return "Invalid Password"

    def doVwr(self, frt, pag, r, c, p, i, w):
        def getDefaultApp():
            dflt = None
            try:
                if sys.platform == "win32":
                    import shlex
                    import winreg
                    class_root = winreg.QueryValue(winreg.HKEY_CLASSES_ROOT,
                        ".pdf")
                    with winreg.OpenKey(winreg.HKEY_CLASSES_ROOT,
                            r"{}\shell\open\command".format(class_root)) as key:
                        command = winreg.QueryValueEx(key, "")[0]
                        return shlex.split(command)[0]
                else:
                    import os
                    fle = os.path.join(os.getenv("HOME"),
                            ".config/mimeapps.list")
                    assoc = None
                    if  os.path.isfile(fle):
                        with open(fle, "r") as opf:
                            for lin in opf:
                                if lin.count("Added Associations"):
                                    added = True
                                    continue
                                elif lin.count("Default Associations"):
                                    added = False
                                    dflt = True
                                    continue
                                if lin.count("application/pdf"):
                                    if added:
                                        assoc = lin.split("=")[1].split(";")
                                    elif dflt:
                                        dflt = lin.split("=")[1]
                                        break
                    if not dflt and assoc:
                        dflt = assoc[0]
                return dflt
            except:
                return
        if not w and not getDefaultApp() and not FITZ:
            return "Missing pymupdf and No Default"

    def doExport(self, frt, pag, r, c, p, i, w):
        if w and not os.path.isfile(w):
            return "Invalid Exporter"

    def doGeometry(self, frt, pag, r, c, p, i, w):
        if w == "0":
            w, h, f = getFontSize(tk)
        else:
            geo = w.lower().split("x")
            if len(geo) != 2:
                return "Invalid Geometry"
            sw = int(self.mf.window.winfo_screenwidth() * .95)
            sh = int(self.mf.window.winfo_screenheight() * .95)
            if int(geo[0]) > sw:
                return "The Maximum Allowable Width of %s is Exceeded" % sw
            if int(geo[1]) > sh:
                return "The Maximum Allowable Height of %s is Exceeded" % sh
            w, h, f = getFontSize(tk, int(geo[0]), int(geo[1]))
        self.df.dfs = f
        self.mf.geo = [w, h]
        self.rcdic["geo"] = "%sx%s" % (w, h)
        self.rcdic["mfs"] = self.mf.rcdic["mfs"] = f
        self.rcdic["dfs"] = self.mf.rcdic["dfs"] = f
        self.df.loadEntry(frt, pag, p, data=self.rcdic["geo"])
        self.df.loadEntry("T", 3, 1, data=self.rcdic["mfs"])
        self.df.loadEntry("T", 3, 3, data=self.rcdic["dfs"])
        self.mf.setThemeFont()
        self.mf.resizeChildren()

    def doFtNam(self, frt, pag, r, c, p, i, w):
        if p == 0:
            self.rcdic["mft"] = self.mf.rcdic["mft"] = w
        else:
            self.rcdic["dft"] = self.mf.rcdic["dft"] = w
        self.mf.setThemeFont()
        self.mf.resizeChildren()

    def doFtSiz(self, frt, pag, r, c, p, i, w):
        mw, mh, ms = getFontSize(tk, self.mf.geo[0], self.mf.geo[1])
        if w > ms:
            return "Maximum Font Size for this Geometry is %s" % ms
        if p == 1:
            self.rcdic["mfs"] = self.mf.rcdic["mfs"] = w
        else:
            self.df.dfs = w
            self.rcdic["dfs"] = self.mf.rcdic["dfs"] = w
        self.mf.setThemeFont()
        self.mf.resizeChildren()

    def doTheme(self, frt, pag, r, c, p, i, w):
        self.rcdic["theme"] = self.mf.rcdic["theme"] = w
        self.mf.setThemeFont()

    def doScheme(self, frt, pag, r, c, p, i, w):
        if w == "C":
            return
        color = getColors(self.mf.style, w)
        if p == 5:
            # Label
            for x in range(1):
                idx = (x * 3) + 1
                for y in range(2):
                    if not color[x][y]:
                        if y:
                            color[x][y] = "#ffffff"
                        else:
                            color[x][y] = "#000000"
                    self.df.doKeyPressed(frt, pag, p+idx+y, data=color[x][y])
            return "ff9"
        else:
            # Button
            for x in range(3):
                idx = (x * 3) + 1
                for y in range(2):
                    if not color[x+1][y]:
                        if y:
                            color[x+1][y] = "#ffffff"
                        else:
                            color[x+1][y] = "#000000"
                    self.df.doKeyPressed(frt, pag, p+idx+y, data=color[x+1][y])
            return "ff20"

    def doColour(self, frt, pag, r, c, p, i, w):
        if not w:
            return "rf"
        try:
            self.df.window.winfo_rgb(w)
        except:
            return "Invalid Colour"
        if p == 6:
            self.rcdic["nfg"] = self.mf.rcdic["nfg"] = w
        elif p == 7:
            self.rcdic["nbg"] = self.mf.rcdic["nbg"] = w
        elif p == 10:
            self.rcdic["bfg"] = self.mf.rcdic["bfg"] = w
        elif p == 11:
            self.rcdic["bbg"] = self.mf.rcdic["bbg"] = w
        elif p == 13:
            self.rcdic["ffg"] = self.mf.rcdic["ffg"] = w
        elif p == 14:
            self.rcdic["fbg"] = self.mf.rcdic["fbg"] = w
        elif p == 16:
            self.rcdic["dfg"] = self.mf.rcdic["dfg"] = w
        elif p == 17:
            self.rcdic["dbg"] = self.mf.rcdic["dbg"] = w
        elif p == 19:
            self.rcdic["qfg"] = self.mf.rcdic["qfg"] = w
        elif p == 20:
            self.rcdic["qbg"] = self.mf.rcdic["qbg"] = w
        elif p == 22:
            self.rcdic["cfg"] = self.mf.rcdic["cfg"] = w
        elif p == 23:
            self.rcdic["cbg"] = self.mf.rcdic["cbg"] = w
        elif p == 25:
            self.rcdic["sfg"] = self.mf.rcdic["sfg"] = w
        elif p == 26:
            self.rcdic["sbg"] = self.mf.rcdic["sbg"] = w
        self.mf.setThemeFont()
        self.setColours()

    def doEnd(self):
        if self.df.pag == 0:
            self.rcdic = loadRcFile(self.rcfile, default=True)
            self.loadAllFields()
            self.df.focusField("T", 1, 1)
        elif self.df.pag == 3:
            self.doSave()
        else:
            self.df.selPage(self.df.tags[self.df.pag][0])

    def doExit(self):
        if self.df.pag in (0, 1):
            self.doQuit()
        elif self.df.pag == 2 and self.dbskp:
            self.df.focusField(self.df.frt, self.df.pag, self.df.col)
        else:
            self.df.selPage(self.df.tags[self.df.pag - 2][0])

    def loadAllFields(self):
        self.rcdic["dfs"] = self.mf.rcdic["dfs"]
        # DataBase
        if self.rcdic["dbase"] == "PgSQL":
            dbase = "P"
        else:
            dbase = "S"
        self.df.loadEntry("T",1,0,dbase)
        self.df.loadEntry("T",1,1,self.rcdic["dbname"])
        self.df.loadEntry("T",1,2,self.rcdic["dbdir"])
        self.df.loadEntry("T",1,3,self.rcdic["dbhost"])
        self.df.loadEntry("T",1,4,self.rcdic["dbport"])
        self.df.loadEntry("T",1,5,self.rcdic["dbuser"])
        self.df.loadEntry("T",1,6,self.rcdic["dbpwd"])
        if dbase == "P":
            self.df.topv[1][2] = ("efld",)
            self.df.topv[1][3] = ("notblank",)
            self.df.topv[1][5] = ("notblank",)
            self.df.topv[1][6] = ("notblank",)
        else:
            self.df.topv[1][2] = ("notblank",)
            self.df.topv[1][3] = ("efld",)
            self.df.topv[1][5] = ("efld",)
            self.df.topv[1][6] = ("efld",)
        # General
        self.df.loadEntry("T",2,0,self.rcdic["bupdir"])
        self.df.loadEntry("T",2,1,self.rcdic["wrkdir"])
        self.df.loadEntry("T",2,2,self.rcdic["upgdir"])
        self.df.loadEntry("T",2,3,self.rcdic["vwr"])
        self.df.loadEntry("T",2,4,self.rcdic["prn"])
        self.df.loadEntry("T",2,5,self.rcdic["exp"])
        self.df.loadEntry("T",2,6,self.rcdic["geo"])
        self.df.loadEntry("T",2,7,self.rcdic["plc"])
        self.df.loadEntry("T",2,8,self.rcdic["img"].upper())
        self.df.loadEntry("T",2,9,self.rcdic["cnf"].upper())
        self.df.loadEntry("T",2,10,self.rcdic["acnf"].upper())
        self.df.loadEntry("T",2,11,self.rcdic["ttip"].upper())
        self.df.loadEntry("T",2,12,self.rcdic["errs"].upper())
        self.df.loadEntry("T",2,13,self.rcdic["wrkf"].upper())
        self.df.loadEntry("T",2,14,self.rcdic["wrka"].upper())
        # Font and Theme
        self.df.loadEntry("T",3,0,self.rcdic["mft"])
        self.df.loadEntry("T",3,1,self.rcdic["mfs"])
        self.df.loadEntry("T",3,2,self.rcdic["dft"])
        self.df.loadEntry("T",3,3,self.rcdic["dfs"])
        self.df.loadEntry("T",3,4,self.rcdic["theme"])
        self.df.loadEntry("T",3,5,self.rcdic["lsc"])
        self.df.loadEntry("T",3,6,self.rcdic["nfg"])
        self.df.loadEntry("T",3,7,self.rcdic["nbg"])
        self.df.loadEntry("T",3,8,"PREVIEW")
        self.df.loadEntry("T",3,9,self.rcdic["bsc"])
        self.df.loadEntry("T",3,10,self.rcdic["bfg"])
        self.df.loadEntry("T",3,11,self.rcdic["bbg"])
        self.df.loadEntry("T",3,12,"PREVIEW")
        self.df.loadEntry("T",3,13,self.rcdic["ffg"])
        self.df.loadEntry("T",3,14,self.rcdic["fbg"])
        self.df.loadEntry("T",3,15,"PREVIEW")
        self.df.loadEntry("T",3,16,self.rcdic["dfg"])
        self.df.loadEntry("T",3,17,self.rcdic["dbg"])
        self.df.loadEntry("T",3,18,"PREVIEW")
        self.df.loadEntry("T",3,19,self.rcdic["qfg"])
        self.df.loadEntry("T",3,20,self.rcdic["qbg"])
        self.df.loadEntry("T",3,21,"PREVIEW")
        self.df.loadEntry("T",3,22,self.rcdic["cfg"])
        self.df.loadEntry("T",3,23,self.rcdic["cbg"])
        self.df.loadEntry("T",3,24,"PREVIEW")
        self.df.loadEntry("T",3,25,self.rcdic["sfg"])
        self.df.loadEntry("T",3,26,self.rcdic["sbg"])
        self.df.loadEntry("T",3,27,"PREVIEW")
        self.setColours()

    def setColours(self):
        for pos in range(6, 9, 3):
            fg, bg = self.df.t_work[3][0][pos], self.df.t_work[3][0][pos+1]
            if not fg or not bg:
                continue
            self.df.topEntry[3][pos+2].configure(foreground=fg, background=bg)
        for pos in range(10, 28, 3):
            fg, bg = self.df.t_work[3][0][pos], self.df.t_work[3][0][pos+1]
            if not fg or not bg:
                continue
            self.df.topEntry[3][pos+2].configure(foreground=fg, background=bg)
        self.df.window.update_idletasks()

    def doSave(self):
        frt, pag, col, mes = self.df.doCheckFields()
        if mes:
            self.df.selPage(self.df.tags[pag - 1][0])
            self.df.focusField(frt, pag, (col+1), err=mes)
        else:
            self.doWrite()

    def doWrite(self):
        rcf = os.path.dirname(self.rcfile)
        dbd = self.df.t_work[1][0][2]
        self.bupdir = self.df.t_work[2][0][0]
        self.wrkdir = self.df.t_work[2][0][1]
        self.upgdir = self.df.t_work[2][0][2]
        for name in (rcf, dbd, self.bupdir, self.wrkdir, self.upgdir):
            name = self.checkCreateDir(name)
        cc = []
        if self.df.t_work[1][0][0] == "P":
            cc.append(["dbase", "PgSQL"])
        elif self.df.t_work[1][0][0] == "S":
            cc.append(["dbase", "SQLite"])
        cc.extend([
            ["dbname", self.df.t_work[1][0][1]],
            ["dbdir", os.path.normpath(self.df.t_work[1][0][2])],
            ["dbhost", self.df.t_work[1][0][3]],
            ["dbport", self.df.t_work[1][0][4]],
            ["dbuser", self.df.t_work[1][0][5]],
            ["dbpwd", self.df.t_work[1][0][6]],
            ["bupdir", os.path.normpath(self.bupdir)],
            ["wrkdir", os.path.normpath(self.wrkdir)],
            ["upgdir", os.path.normpath(self.upgdir)]])
        if self.df.t_work[2][0][3]:
            cc.append(["vwr", os.path.normpath(self.df.t_work[2][0][3])])
        else:
            cc.append(["vwr", ""])
        if self.df.t_work[2][0][4]:
            cc.append(["prn", os.path.normpath(self.df.t_work[2][0][4])])
        else:
            cc.append(["prn", ""])
        cc.extend([
            ["exp", self.df.t_work[2][0][5]],
            ["geo", self.df.t_work[2][0][6]],
            ["plc", self.df.t_work[2][0][7]],
            ["img", self.df.t_work[2][0][8]],
            ["cnf", self.df.t_work[2][0][9]],
            ["acnf", self.df.t_work[2][0][10]],
            ["ttip", self.df.t_work[2][0][11]],
            ["errs", self.df.t_work[2][0][12]],
            ["wrkf", self.df.t_work[2][0][13]],
            ["wrka", self.df.t_work[2][0][14]],
            ["mft", self.df.t_work[3][0][0]],
            ["mfs", self.df.t_work[3][0][1]],
            ["dft", self.df.t_work[3][0][2]],
            ["dfs", self.df.t_work[3][0][3]],
            ["theme", self.df.t_work[3][0][4]],
            ["lsc", self.df.t_work[3][0][5]],
            ["nfg", self.df.t_work[3][0][6]],
            ["nbg", self.df.t_work[3][0][7]],
            ["bsc", self.df.t_work[3][0][9]],
            ["bfg", self.df.t_work[3][0][10]],
            ["bbg", self.df.t_work[3][0][11]],
            ["ffg", self.df.t_work[3][0][13]],
            ["fbg", self.df.t_work[3][0][14]],
            ["dfg", self.df.t_work[3][0][16]],
            ["dbg", self.df.t_work[3][0][17]],
            ["qfg", self.df.t_work[3][0][19]],
            ["qbg", self.df.t_work[3][0][20]],
            ["cfg", self.df.t_work[3][0][22]],
            ["cbg", self.df.t_work[3][0][23]],
            ["sfg", self.df.t_work[3][0][25]],
            ["sbg", self.df.t_work[3][0][26]]])
        try:
            f = open(self.rcfile, "w")
            for a, b in cc:
                f.write("['%s'] = '%s'\n" % (a, b))
                if a == "bupdir":
                    self.bupdir = b
                elif a == "wrkdir":
                    self.wrkdir = b
                elif a == "upgdir":
                    self.upgdir = b
                elif a == "vwr":
                    self.vwr = b
                elif a == "prn":
                    self.prn = b
                elif a == "exp":
                    self.exp = b
                else:
                    setattr(self, a, str(b))
            f.close()
        except Exception as err:
            showError(self.mf.body, "Access Error", "The rcfile %s "\
                "Cannot be Created.\n\n%s." % (self.rcfile, err))
            return
        showWarning(self.mf.body, "Configuration", "Please Note that Some "\
            "Changes will Only take effect once Tartan is Restarted.")
        self.closeProcess()

    def doQuit(self):
        self.rcfile = None
        self.closeProcess()

    def closeProcess(self):
        self.df.closeProcess()
        self.mf.closeLoop()

    def checkCreateDir(self, name):
        path = os.path.normpath(name)
        try:
            os.makedirs(path)
        except:
            pass
        return path

    def loadFonts(self):
        sp = SplashScreen(self.mf.body,
            "Loading List of Fixed Fonts\n\n      Please Wait...")
        fonts = list(tkfont.families())
        fonts.sort()
        self.aft = []
        self.fft = []
        for font in fonts:
            if font.lower().count("emoji"):
                continue
            if font.lower().count("unifont upper"):
                continue
            try:
                ft = tkfont.Font(font=(font, 10, "normal"))
                if ft.metrics()["fixed"] and font not in self.fft:
                    self.fft.append(font)
                if font not in self.aft:
                    self.aft.append(font)
            except:
                continue
        sp.closeSplash()

    def loadThemes(self):
        self.thm = list(self.mf.style.theme_names())
        self.thm.sort()

class AgeAll(object):
    """
    This class is used to Age Debtors, Creditors and Members Transactions where:

    system:     "drs", "crs" or "mem"
    agetyp:     N = Normal  - Selective allocations are allowed.
                H = History - Same as Normal but including paid-up items.
                O = Oldest  - The transaction is automatically allocated against
                              the oldest outstanding transactions, any
                              unallocated amount remains unallocated.
                C = Current - The transaction retains its balance, i.e. not
                              allocated.
    agekey:     drs - [conum, chain, acno, atyp, aref, agedt, amt, dis]
                crs - [conum, acno, atyp, aref, agedt, amt, dis]
                mem - [conum, memno, atyp, aref, agedt, amt, dis]
    """
    def __init__(self, mf, **args):
        self.mf = mf
        defaults = {
            "system": None,
            "agetyp": None,
            "agekey": None}
        for nam in args:
            defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        self.cancel = False
        if self.agetyp == "C":
            return
        self.cono = self.agekey[0]
        if self.system.lower() == "drs":
            self.achn = self.agekey[1]
            self.acno = self.agekey[2]
            idx = 3
        else:
            self.acno = self.agekey[1]
            idx = 2
        flds = ("atyp", "aref", "aged", "aamt", "adis")
        for x in range(idx, len(self.agekey)):
            setattr(self, flds[x - idx], self.agekey[x])
        self.doGetData()
        if self.data and self.agetyp == "O":
            self.doAgeOldest()
        elif self.data:
            self.doAgeTrans()

    def doGetData(self):
        if self.system.lower() == "crs":
            col = ["crt_ref1", "crt_type", "crt_curdt", "crt_tramt"]
            whr = [("crt_cono", "=", self.cono), ("crt_acno", "=",
                self.acno)]
        elif self.system.lower() == "drs":
            col = ["drt_ref1", "drt_type", "drt_curdt", "drt_tramt"]
            whr = [("drt_cono", "=", self.cono), ("drt_chain", "=",
                self.achn), ("drt_acno", "=", self.acno)]
        else:
            col = ["mlt_refno", "mlt_type", "mlt_curdt", "mlt_tramt"]
            whr = [("mlt_cono", "=", self.cono), ("mlt_memno", "=",
                self.acno)]
        get, recs = getTrn(self.mf.dbm, self.system.lower(), whr=whr)
        self.data = []
        for rec in recs:
            bal = rec[get.index("balance")]
            if self.agetyp != "H" and not bal:
                continue
            dat = []
            for c in col:
                dat.append(rec[get.index(c)])
            dat.append(rec[get.index("paid")])
            dat.append(bal)
            self.data.append(dat)

    def doAgeOldest(self):
        unal = float(ASD(self.aamt) + ASD(self.adis))
        for dat in self.data:
            trbal = float(ASD(dat[5]) + ASD(unal))
            if trbal < 0:
                unal = trbal
                trbal = 0
            elif trbal > 0:
                unal = 0
            else:
                unal = 0
                trbal = 0
            alloc = float(ASD(dat[5]) - ASD(trbal))
            if alloc:
                self.doAlloc(dat[2], dat[1], dat[0], alloc)
        alloc = float(ASD(self.aamt) + ASD(self.adis) - ASD(unal))
        if alloc:
            self.doAllocAge(alloc)

    def doAgeTrans(self):
        unal = float(ASD(self.aamt) + ASD(self.adis))
        self.ageT = AgeTrans(self.mf, self.system.lower(), self.data, unal)
        self.cancel = self.ageT.ccl
        if self.cancel:
            return
        for tr in self.ageT.data:
            if tr[6]:
                self.doAlloc(tr[2], tr[1], tr[0], tr[6])
        totl = float(ASD(self.aamt) + ASD(self.adis))
        alloc = float(ASD(totl) + ASD(self.ageT.total.work))
        if alloc:
            self.doAllocAge(alloc)

    def doAllocAge(self, alloc):
        if not self.adis:
            da = 0.00
            aa = alloc
        elif self.adis and self.adis < alloc:
            da = float(ASD(self.adis) - ASD(alloc))
            aa = 0.00
        else:
            da = self.adis
            aa = float(ASD(alloc) - ASD(self.adis))
        if da:
            self.doAlloc(self.aged, 6, self.aref, da)
        if aa:
            self.doAlloc(self.aged, self.atyp, self.aref, aa)

    def doAlloc(self, cdt, typ, ref, amt):
        if cdt < self.aged:
            cdt = self.aged
        if self.system.lower() == "crs":
            data = self.agekey[:2] + [typ, ref, cdt, self.atyp,
                self.aref, amt, 0]
            sql = Sql(self.mf.dbm, "crsage", prog=__name__)
            sql.insRec("crsage", data=data)
        elif self.system.lower() == "drs":
            data = self.agekey[:3] + [typ, ref, cdt, self.atyp,
                self.aref, amt, 0]
            sql = Sql(self.mf.dbm, "drsage", prog=__name__)
            sql.insRec("drsage", data=data)
        else:
            data = self.agekey[:2] + [typ, ref, cdt, self.atyp,
                self.aref, amt, 0]
            sql = Sql(self.mf.dbm, "memage", prog=__name__)
            sql.insRec("memage", data=data)

class AgeTrans(object):
    """
    Class for the Distribution of an Amount Over Transactions

    Usage: AgeTrans(mf, system, data, total)

        system = crs, drs or mem
        data = List of lists of all available transactions comprising:
             [Date, Reference, Type, Amount, Paid, Balance]
        total = Total Amount to be Distributed (Decimal Value)
        xits = Create an Exit Button to leave the balance unallocated if True

    The class returns as follows:
        self.data = List of lists comprising:
            [Date, Reference, Type, Amount, Paid, Balance, Allocated]
        self.total = Total Amount Not Yet Allocated
    """
    def __init__(self, mf, system, data, total, xits=True):
        self.mf = mf
        self.system = system
        self.data = data
        self.total = CCD(float(ASD(0) - ASD(total)), "SD", 13.2)
        self.xits = xits
        self.setVariables()
        if self.total.work:
            while not self.xit and not self.ccl:
                self.doAllocate()
                if not self.total.work:
                    break
        else:
            self.doAllocate()

    def setVariables(self):
        if self.system == "crs":
            typ = ("XX", tartanWork.crtrtp)
        elif self.system == "drs":
            typ = ("XX", tartanWork.drtrtp)
        elif self.system == "mem":
            typ = ("XX", tartanWork.mltrtp)
        self.cols = (
            (0, "Reference",   9,   "Na", None),
            (1, "Typ",         3,   typ,  None),
            (2, "Cur-Dte",     7,   "D2", None),
            (3, "    Amount", 13.2, "SD", None),
            (4, " Allocated", 13.2, "SD", None),
            (5, "   Balance", 13.2, "SD", None))
        self.work = []
        self.xit = False
        self.ccl = False
        self.posn = 0
        for num, rec in enumerate(self.data):
            self.data[num].append(0)
            self.work.append(rec[:6])

    def doAllocate(self):
        # Scrolled Selection
        self.lb = SelectChoice(self.mf.window, "Available Transactions",
            self.cols, self.work, sort=False, wait=False, cmnd=self.enterAll,
            posn=self.posn, escape=False)
        # Allocation Fields
        hbox = MyFrame(self.lb.window)
        hbox.pack(fill="x", expand="yes")
        hbox.columnconfigure(0, weight=1)
        lab = MyLabel(hbox, text="Allocated")
        lab.grid(row=0, column=0, sticky="ew")
        self.allAmt = MyEntry(hbox, width=13, maxsize=13, cmd=self.enterAmt)
        self.allAmt.configure(state="disabled")
        self.allAmt.grid(row=0, column=1, sticky="ew")
        lab = MyLabel(hbox, text="Unallocated")
        lab.grid(row=0, column=2, sticky="ew")
        self.allBal = MyEntry(hbox, width=13, maxsize=13)
        self.allBal.insert(0, self.total.disp)
        self.allBal.configure(state="disabled")
        self.allBal.grid(row=0, column=3, sticky="ew")
        bbox = MyButtonBox(self.lb.window)
        self.allBut = bbox.addButton(text="Apply", cmd=self.enterAmt)
        self.allBut.configure(state="disabled")
        if self.xits:
            xitBut = bbox.addButton(text="Exit", cmd=self.exitAge)
        canBut = bbox.addButton(text="Cancel", cmd=self.cancelAll)
        # Button Tooltips
        ToolTip(self.allBut, "Apply this Allocation")
        if self.xits:
            ToolTip(xitBut, "Leave the Balance as Unallocated")
        ToolTip(canBut, "Cancel All Allocations")
        # Main Loop
        placeWindow(self.lb.window, self.mf.window)
        self.lb.mstFrame.wait_window()

    def enterAll(self, data):
        self.pos = data[0]
        self.posn = self.lb.posn
        self.allAmt.configure(state="normal")
        self.allBut.configure(state="normal")
        bal = self.work[self.pos][5]
        if bal > self.total.work:
            bal = self.total.work
        wrk = CCD(bal, "SD", 13.2)
        self.allAmt.delete(0, "end")
        self.allAmt.insert(0, wrk.disp)
        self.allAmt.focus_set()
        self.allAmt.selection_range(0, "end")

    def enterAmt(self):
        wrk = CCD(self.allAmt.get(), "SD", 13.2)
        if wrk.err:
            showError(self.lb.mstFrame, "Error", "Invalid Allocation Amount")
            wrk = CCD(0, "SD", 13.2)
        self.data[self.pos][6] = float(ASD(self.data[self.pos][6]) + \
            ASD(wrk.work))
        recs = float(ASD(self.work[self.pos][4]) + ASD(wrk.work))
        bal = float(ASD(self.work[self.pos][3]) - ASD(recs))
        self.work[self.pos][4] = recs
        self.work[self.pos][5] = bal
        tot = float(ASD(self.total.work) - ASD(wrk.work))
        self.total = CCD(tot, "SD", 13.2)
        self.closeProcess()

    def exitAge(self):
        self.xit = True
        self.closeProcess()

    def cancelAll(self):
        ok = askQuestion(self.lb.window, head="Cancel",
            mess="Are You Certain You Want to Cancel This Allocation?")
        if ok == "yes":
            self.ccl = True
            self.closeProcess()

    def closeProcess(self):
        self.lb.closeProcess()

class Balances(object):
    """
    This class is used to return various balances from Debtors, Creditors,
        Stores, Assets and Members tables:

    doAllBals       - Returns Total, Current, 30, 60, 90 and 120+ Balances with
                        or without transactions
    doCrsDrsHist    - Returns the Last 12 Months Sale and Payment Balances
    doStrBals       - Returns
                        Current Month
                            [O/Bal Qty and Cost]
                            [Movements]
                            [C/Bal Qty and Cost]
                        Year to Date
                            [O/Bal Qty and Cost]
                            [Movements]
                            [C/Bal Qty and Cost]
                        Average Cost, Last Cost & Selling Price
                        If trans=True the transactions are also returned.
    doStrOrds       - Returns the Qty on Hand, Qty on Order and Qty Back Order
    doStrHist       - Returns the Last 12 Months Receipt and Issue Qty and Value
    doAssBals       - Returns the Opening Book Values for a Financial Period

    system          - ASS, CRS, DRS, MEM or STR
    conum           - Company number
    curdt           - Current period, YYYYMM
    keys            - Assets is the group and code
                      Creditors is the acno
                      Debtors is a list containing (chain, acno)
                      Members is the member number
                      Stores is a list containing:
                          if len(keys) == 1: rep
                          if len(keys) == 2: rep, period
                          if len(keys) == 3: group, code, loc
                          if len(keys) == 4: group, code, loc, period
    """
    def __init__(self, mf, system, conum, curdt, keys):
        self.mf = mf
        self.system = system.upper()
        self.conum = conum
        self.curdt = curdt
        if not keys:
            showError(self.mf.window, "Keys Error",
                "You Must have a Key Argument")
            sys.exit()
        else:
            if self.system == "CRS" and len(keys) == 1:
                self.sql = Sql(self.mf.dbm, "crstrn", prog=__name__)
                self.acno = keys[0]
            elif self.system == "DRS" and len(keys) == 2:
                self.sql = Sql(self.mf.dbm, "drstrn", prog=__name__)
                self.chain, self.acno = keys
            elif self.system == "MEM" and len(keys) == 1:
                self.sql = Sql(self.mf.dbm, "memtrn", prog=__name__)
                self.memno = keys[0]
            elif self.system == "STR" and len(keys) in (1, 2, 3, 4):
                self.sql = Sql(self.mf.dbm, ["strgrp", "strmf2", "strtrn",
                    "strpom", "strpot", "slsiv1", "slsiv2"], prog=__name__)
                self.group = None
                self.code = None
                self.loc = None
                period = None
                self.rep = None
                if len(keys) == 1:
                    self.rep = keys[0]
                elif len(keys) == 2:
                    self.rep, period = keys
                elif len(keys) == 3:
                    self.group, self.code, self.loc = keys
                elif len(keys) == 4:
                    self.group, self.code, self.loc, period = keys
                if period:
                    if period[0] == "P":
                        s, e, f = getPeriods(self.mf, self.conum, period[1])
                        if not s:
                            self.start = self.curdt
                        else:
                            self.start = int(s.work / 100)
                    else:
                        self.start = period[1]
                else:
                    self.start = self.curdt
            elif self.system == "ASS" and len(keys) == 2:
                self.sql = Sql(self.mf.dbm, "asstrn", prog=__name__)
                self.group, self.code = keys
                self.start = self.curdt
            else:
                showError(self.mf.window, "Parameter Error",
                    "Invalid Parameters %s %s" % (self.system, str(keys)))
                sys.exit()

    def doAllBals(self, trans="N"):
        """
        trans:
            "Y" = returns all transactions except those zeroed previously
            "N" = does not return transactions
            "A" = returns all transactions
        """
        ages = [0,0,0,0,0]
        if self.system == "CRS":
            whr = [("crt_cono", "=", self.conum), ("crt_acno", "=", self.acno)]
            w = copyList(whr)
            w.append(("crt_curdt", "<", self.curdt))
            obal = self.sql.getRec("crstrn", cols=["sum(crt_tramt)"],
                where=w, limit=1)
            if not obal or not obal[0]:
                obal = 0.0
            else:
                obal = obal[0]
            cbal = 0.0
            col, trns = getTrn(self.mf.dbm, self.system.lower(), cdt=self.curdt,
                whr=whr, zer=trans)
        elif self.system == "DRS":
            whr = [("drt_cono", "=", self.conum), ("drt_chain", "=",
                self.chain), ("drt_acno", "=", self.acno)]
            w = copyList(whr)
            w.append(("drt_curdt", "<", self.curdt))
            obal = self.sql.getRec("drstrn",
                cols=["round(sum(drt_tramt), 2)"], where=w, limit=1)
            if not obal or not obal[0]:
                obal = 0.0
            else:
                obal = obal[0]
            cbal = 0.0
            col, trns = getTrn(self.mf.dbm, self.system.lower(), cdt=self.curdt,
                whr=whr, zer=trans)
        elif self.system == "MEM":
            whr = [("mlt_cono","=",self.conum), ("mlt_memno","=",self.memno)]
            w = copyList(whr)
            w.append(("mlt_curdt", "<", self.curdt))
            obal = self.sql.getRec("memtrn",
                cols=["round(sum(mlt_tramt), 2)"], where=w, limit=1)
            if not obal or not obal[0]:
                obal = 0.0
            else:
                obal = obal[0]
            cbal = 0.0
            col, trns = getTrn(self.mf.dbm, self.system.lower(), cdt=self.curdt,
                whr=whr, zer=trans)
        else:
            return
        if trns:
            for trn in trns:
                cbal = float(ASD(cbal) + ASD(trn[col.index("balance")]))
                if self.system == "CRS":
                    age = self.curdt - trn[col.index("crt_curdt")]
                elif self.system == "DRS":
                    age = self.curdt - trn[col.index("drt_curdt")]
                else:
                    age = self.curdt - trn[col.index("mlt_curdt")]
                while age > 88:
                    age = age - 88
                if age < 0:
                    age = 0
                if age > 4:
                    age = 4
                ages[age] = float(ASD(ages[age])+ASD(trn[col.index("balance")]))
        if trans == "N":
            return obal, cbal, ages
        else:
            w = copyList(whr)
            if self.system == "CRS":
                w.append(("crt_curdt", "=", self.curdt))
                cmth = self.sql.getRec("crstrn", cols=["count(*)"],
                    where=w, limit=1)
            elif self.system == "DRS":
                w.append(("drt_curdt", "=", self.curdt))
                cmth = self.sql.getRec("drstrn", cols=["count(*)"],
                    where=w, limit=1)
            else:
                w.append(("mlt_curdt", "=", self.curdt))
                cmth = self.sql.getRec("memtrn", cols=["count(*)"],
                    where=w, limit=1)
            return obal, cbal, ages, [col, trns, cmth[0]]

    def doCrsDrsHist(self):
        this = [0,0]
        hist = [[0,0,0,0,0,0,0,0,0,0,0,0,0],[0,0,0,0,0,0,0,0,0,0,0,0,0]]
        start = self.curdt - 199
        if (start % 100) == 13:
            start = start + 100 - 12
        if self.system == "CRS":
            col = ["crt_curdt", "crt_type", "round(sum(crt_tramt), 2)",
                "round(sum(crt_taxamt), 2)"]
            whr = [("crt_cono", "=", self.conum), ("crt_acno", "=", self.acno),
                ("crt_curdt", "between", start, self.curdt)]
            grp = "crt_curdt, crt_type"
            bals = self.sql.getRec("crstrn", cols=col, where=whr,
                group=grp, order=grp)
        elif self.system == "DRS":
            col = ["drt_curdt", "drt_type", "round(sum(drt_tramt), 2)",
                "round(sum(drt_taxamt), 2)"]
            whr = [("drt_cono", "=", self.conum), ("drt_chain", "=",
                self.chain), ("drt_acno", "=", self.acno),
                ("drt_curdt", "between", start, self.curdt)]
            grp = "drt_curdt, drt_type"
            bals = self.sql.getRec("drstrn", cols=col, where=whr,
                group=grp, order=grp)
        for bal in bals:
            if bal[1] in (1, 4):
                x = 0
            elif bal[1] in (2, 5):
                x = 1
            else:
                continue
            this[x] = float(ASD(this[x]) + ASD(bal[2]) - ASD(bal[3]))
            mth = self.curdt - bal[0]
            while mth > 88:
                mth = mth - 88
            if mth > 11:
                mth = 12
            hist[x][mth] = float(ASD(hist[x][mth]) + ASD(bal[2]) - ASD(bal[3]))
        return this, hist

    def doStrBals(self, start=None, trans="N"):
        if not start:
            start = self.start
        if self.loc:
            where = [("stt_cono", "=", self.conum), ("stt_group", "=",
                self.group), ("stt_code", "=", self.code), ("stt_loc", "=",
                self.loc)]
        else:
            where = [("stt_cono", "=", self.conum), ("stt_group", "=",
                self.group), ("stt_code", "=", self.code)]
        # Current Period
        whr = copyList(where)
        whr.append(("stt_curdt", "<", self.curdt))
        obal = self.sql.getRec("strtrn", cols=["round(sum(stt_qty), 2)",
            "round(sum(stt_cost), 2)"], where=whr, limit=1)
        if not obal:
            obal = [0, 0]
        if not obal[0]:
            obal[0] = 0
        if not obal[1]:
            obal[1] = 0
        whr = copyList(where)
        whr.append(("stt_curdt", "between", self.curdt, self.curdt))
        grp = "stt_type"
        move = self.sql.getRec("strtrn", cols=["stt_type",
            "round(sum(stt_qty), 2)", "round(sum(stt_cost), 2)",
            "round(sum(stt_sell), 2)"], where=whr, group=grp, order=grp)
        whr = copyList(where)
        whr.append(("stt_curdt", "<=", self.curdt))
        cbal = self.sql.getRec("strtrn", cols=["round(sum(stt_qty), 2)",
            "round(sum(stt_cost), 2)"], where=whr, limit=1)
        if not cbal:
            cbal = [0, 0]
        if not cbal[0]:
            cbal[0] = 0
        if not cbal[1]:
            cbal[1] = 0
        data = [obal, move, cbal]
        # Year to Date
        whr = copyList(where)
        whr.append(("stt_curdt", "<", start))
        obal = self.sql.getRec("strtrn", cols=["round(sum(stt_qty), 2)",
            "round(sum(stt_cost), 2)"], where=whr, limit=1)
        if not obal:
            obal = [0, 0]
        if not obal[0]:
            obal[0] = 0
        if not obal[1]:
            obal[1] = 0
        whr = copyList(where)
        whr.append(("stt_curdt", "between", start, self.curdt))
        move = self.sql.getRec("strtrn", cols=["stt_type",
            "round(sum(stt_qty), 2)", "round(sum(stt_cost), 2)",
            "round(sum(stt_sell), 2)"], where=whr, group=grp, order=grp)
        if not move:
            move = []
        whr = copyList(where)
        whr.append(("stt_curdt", "<=", self.curdt))
        cbal = self.sql.getRec(tables="strtrn", cols=["round(sum(stt_qty), 2)",
            "round(sum(stt_cost), 2)"], where=whr, limit=1)
        if not cbal:
            cbal = [0, 0]
        if not cbal[0]:
            cbal[0] = 0
        if not cbal[1]:
            cbal[1] = 0
        data.extend([obal, move, cbal])
        # Average and Last Cost Prices
        cost = getCost(self.sql, self.conum, self.group, self.code,
            loc=self.loc, ind="AL")
        data.extend(cost)
        # Last selling price
        data.append(getSell(self.sql, self.conum, self.group, self.code,
            self.loc, ind="L"))
        if trans == "Y":
            # Transactions
            whr = copyList(where)
            whr.append(("stt_curdt", "between", start, self.curdt))
            odr = "stt_trdt, stt_type, stt_ref1"
            data.append((self.sql.getRec("strtrn", where=whr,
                order=odr), self.sql.strtrn_dic))
        return data

    def doStrOrds(self):
        # Closing Balance
        whr = [("stt_cono", "=", self.conum)]
        if self.loc:
            whr.append(("stt_loc", "=", self.loc))
        whr.extend([
            ("stt_group", "=", self.group),
            ("stt_code", "=", self.code),
            ("stt_curdt", "<=", self.curdt)])
        cbal = self.sql.getRec("strtrn", cols=[
            "round(sum(stt_qty), 2)"], where=whr, limit=1)
        if not cbal or not cbal[0]:
            data = [CCD(0, "SD", 12.2)]
        else:
            data = [CCD(cbal[0], "SD", 12.2)]
        # Orders on Suppliers
        whr = [("pom_cono", "=", self.conum)]
        if self.loc:
            whr.append(("pom_loc", "=", self.loc))
        whr.extend([
            ("pom_date", "<=", mthendDate(self.curdt * 100)),
            ("pom_delno", "<>", "cancel"),
            ("pom_deldt", "=", 0),
            ("pot_cono=pom_cono",),
            ("pot_ordno=pom_ordno",),
            ("pot_group", "=", self.group),
            ("pot_code", "=", self.code)])
        oord = self.sql.getRec(tables=["strpom", "strpot"], cols=[
            "round(sum(pot_qty), 2)"], where=whr, limit=1)
        if not oord or not oord[0]:
            data.append(CCD(0, "SD", 12.2))
        else:
            data.append(CCD(oord[0], "SD", 12.2))
        # Orders by Customers
        whr = [
            ("si1_cono", "=", self.conum),
            ("si1_rtn", "in", ("O", "W")),
            ("si1_date", "<=", mthendDate(self.curdt * 100)),
            ("si1_invno", "=", ""),
            ("si2_cono=si1_cono",),
            ("si2_rtn=si1_rtn",),
            ("si2_docno=si1_docno",),
            ("si2_group", "=", self.group),
            ("si2_code", "=", self.code)]
        if self.loc:
            whr.append(("si2_loc", "=", self.loc))
        bord = self.sql.getRec(tables=["slsiv1", "slsiv2"], cols=[
            "round(sum(si2_qty), 2)"], where=whr, limit=1)
        if not bord or not bord[0]:
            data.append(CCD(0, "SD", 12.2))
        else:
            data.append(CCD(bord[0], "SD", 12.2))
        return data

    def doStrHist(self, start=None, more=None):
        if not start:
            start = self.start
        this = [[0,0],[0,0],[0,0,0]]
        hist = [
            [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],
            [0,0],[0,0],[0,0],[0,0],[0,0],[0,0]],
            [[0,0],[0,0],[0,0],[0,0],[0,0],[0,0],
            [0,0],[0,0],[0,0],[0,0],[0,0],[0,0]],
            [[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],
            [0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0],[0,0,0]]]
        whr = [("stt_cono", "=", self.conum)]
        if self.rep:
            whr.append(("stt_rep", "=", self.rep))
        else:
            whr.extend([("stt_group", "=", self.group),
                ("stt_code", "=", self.code)])
            if self.loc:
                whr.append(("stt_loc", "=", self.loc))
        whr.append(("stt_curdt", "between", start, self.curdt))
        if more:
            for state in more:
                whr.append(state)
        grp = "stt_curdt, stt_type"
        bals = self.sql.getRec("strtrn", cols=["stt_curdt",
            "stt_type", "round(sum(stt_qty), 2)", "round(sum(stt_cost), 2)",
            "round(sum(stt_sell), 2)"], where=whr, group=grp, order=grp)
        for bal in bals:
            if bal[1] in (1, 3, 5):
                x = 0
            elif bal[1] in (2, 4, 6, 7, 8):
                x = 1
            else:
                continue
            this[x][0] = float(ASD(this[x][0]) + ASD(bal[2]))
            this[x][1] = float(ASD(this[x][1]) + ASD(bal[3]))
            if bal[1] in (7, 8):
                this[2][0] = float(ASD(this[2][0]) + ASD(bal[2]))
                this[2][1] = float(ASD(this[2][1]) + ASD(bal[3]))
                this[2][2] = float(ASD(this[2][2]) + ASD(bal[4]))
            mth = self.curdt - bal[0]
            while mth > 88:
                mth = mth - 88
            if mth > 11:
                mth = 11
            hist[x][mth][0] = float(ASD(hist[x][mth][0]) + ASD(bal[2]))
            hist[x][mth][1] = float(ASD(hist[x][mth][1]) + ASD(bal[3]))
            if bal[1] in (7, 8):
                hist[2][mth][0] = float(ASD(hist[2][mth][0]) + ASD(bal[2]))
                hist[2][mth][1] = float(ASD(hist[2][mth][1]) + ASD(bal[3]))
                hist[2][mth][2] = float(ASD(hist[2][mth][2]) + ASD(bal[4]))
        return this, hist

    def doAssBals(self, start=None, end=None, trans="N"):
        # trans: N - No Transactions, C - Current Year, H - All
        if not start:
            start = self.start
        w1 = [("ast_cono", "=", self.conum), ("ast_group", "=", self.group),
            ("ast_code", "=", self.code)]
        w2 = copyList(w1)
        w2.append(("ast_curdt", "<", start))
        w3 = copyList(w2)
        w3.append(("ast_mtyp", "in", (3, 5)))
        if self.sql.getRec("asstrn", where=w3):
            # Sold or Written Off Asset, Ignore
            return
        w3 = copyList(w2)
        w3.append(("ast_mtyp", "in", (1, 2)))
        cap = self.sql.getRec("asstrn",
            cols=["round(sum(ast_amt1), 2)"], where=w3, limit=1)
        if not cap[0]:
            cap[0] = 0
        w3 = copyList(w2)
        w3.append(("ast_mtyp", "=", 4))
        dep = self.sql.getRec("asstrn",
            cols=["round(sum(ast_amt1), 2)", "round(sum(ast_amt2), 2)"],
            where=w3, limit=1)
        if not dep[0]:
            dep[0] = 0
        if not dep[1]:
            dep[1] = 0
        cbl = float(ASD(cap[0]) + ASD(dep[0]))
        rbl = float(ASD(cap[0]) + ASD(dep[1]))
        w3 = copyList(w1)
        if not end:
            w3.append(("ast_curdt", ">=", start))
        else:
            w3.append(("ast_curdt", "between", start, end))
        mov = self.sql.getRec("asstrn", cols=["ast_mtyp",
            "round(sum(ast_amt1), 2)", "round(sum(ast_amt2), 2)"], where=w3,
            group="ast_mtyp", order="ast_mtyp")
        col = self.sql.asstrn_col
        if trans == "H":
            trn = self.sql.getRec("asstrn", where=w1,
                order="ast_date, ast_type")
            return cap[0], dep[0], dep[1], cbl, rbl, mov, [trn, col]
        elif trans == "C":
            trn = self.sql.getRec("asstrn", where=w3,
                order="ast_date, ast_type")
            return cap[0], dep[0], dep[1], cbl, rbl, mov, [trn, col]
        else:
            return cap[0], dep[0], dep[1], cbl, rbl, mov

class Batches(object):
    """
    This class is used to maintain batch records.
    """
    def __init__(self, mf, coy, nam, per, sys, rtn, multi=None, glint=None):
        self.mf = mf
        self.conum = coy
        self.conam = nam
        self.per = per
        self.sys = sys.upper()
        self.rtn = rtn
        self.multi = multi
        self.glint = glint
        self.setVariables()

    def setVariables(self):
        self.sql = Sql(self.mf.dbm, ["ctlbat", "ctlctl"], prog=__name__)
        t = time.localtime()
        self.sysdt = ((t[0] * 10000) + (t[1] * 100) + t[2])
        t = str(self.sysdt)
        t = (int(t[0:4]),int(t[4:6]), int(t[6:8]),0,0,0,0,0,0)
        self.curdt = ((t[0] * 100) + t[1])
        self.s_per = int(self.per[1][0] / 100)
        self.e_per = int(self.per[2][0] / 100)
        self.bankctl = False
        if self.sys == "ASS":
            self.ttp = copyList(tartanWork.artrtp)
            self.nam = self.ttp[self.rtn-1][1]
            if self.rtn in (1, 2) and self.glint == "Y":
                self.bankctl = True
        elif self.sys == "CRS":
            self.ttp = copyList(tartanWork.crtrtp)
            self.nam = self.ttp[self.rtn-1][1]
            if self.rtn in (2, 5) and self.glint == "Y":
                self.bankctl = True
        elif self.sys == "DRS":
            self.ttp = copyList(tartanWork.drtrtp)
            self.nam = self.ttp[self.rtn-1][1]
            if self.rtn in (2, 5) and self.glint == "Y":
                self.bankctl = True
        elif self.sys == "GEN":
            self.ttp = copyList(tartanWork.gltrtp)
            self.nam = self.ttp[self.rtn-1][1]
            if self.rtn in (2, 6, 7):
                self.bankctl = True
        elif self.sys == "LON":
            self.ttp = copyList(tartanWork.lntrtp)
            self.nam = self.ttp[self.rtn-1][1]
            if self.rtn in (1, 2) and self.glint == "Y":
                self.bankctl = True
        elif self.sys == "MEM":
            self.ttp = copyList(tartanWork.mltrtp)
            self.nam = self.ttp[self.rtn-1][1]
            if self.rtn in (2, 5) and self.glint == "Y":
                self.bankctl = True
        elif self.sys == "RTL":
            self.ttp = copyList(tartanWork.rttrtp)
            self.nam = self.ttp[self.rtn-1][1]
            if self.rtn in (2, 3) and self.glint == "Y":
                self.bankctl = True
        elif self.sys == "SLN":
            self.ttp = copyList(tartanWork.sltrtp)
            self.nam = self.ttp[self.rtn-1][1]
            if self.rtn in (2, 4) and self.glint == "Y":
                self.bankctl = True
        elif self.sys == "STR":
            self.ttp = copyList(tartanWork.sttrtp)
            self.nam = self.ttp[self.rtn-1][1]
        if self.bankctl:
            acc = self.sql.getRec("ctlctl", cols=["ctl_code"],
                where=[("ctl_cono", "=", self.conum), ("ctl_code", "like",
                "bank_%")])
            if len(acc) == 1:
                self.ckCtl(acc[0][0])

    def doBatchHeader(self):
        tit = "Batch Header"
        btm = {
            "stype": "R",
            "tables": ("ctlbat",),
            "cols": (
                ("btm_batno", "", 0, "Bat-Num"),
                ("btm_rtyp", ("xx", self.ttp), 20, "Type"),
                ("btm_curdt", "", 0, "Curr-Dt"),
                ("btm_multi", "", 0, "M")),
            "where": [
                ("btm_cono", "=", self.conum),
                ("btm_styp", "=", self.sys),
                ("btm_rtyp", "=", self.rtn),
                ("btm_curdt", "between", self.s_per, self.e_per),
                ("btm_ind", "=", "N")],
            "group": "btm_batno",
            "zero": "0"}
        ctl = {
            "stype": "R",
            "tables": ("ctlctl", "genmst"),
            "cols": (
                ("ctl_code", "", 0, "Ctl-Code"),
                ("ctl_conacc", "", 0, "G/L-Num"),
                ("glm_desc", "", 30, "Description")),
            "where": (
                ("ctl_cono", "=", self.conum),
                ("ctl_code", "like", "bank_%"),
                ("glm_cono=ctl_cono",),
                ("glm_acno=ctl_conacc",))}
        r1s = (("No", "N"), ("Yes", "Y"))
        fld = [
            (("T",0,0,0),"INa",7,"Batch Number","",
                "","N",self.doBatNum,btm,None,("notblank",)),
            (("T",0,1,0),"ID2",7,"Current Period (CCYYMM)","",
                self.curdt,"N",self.doCurdt,None,None,("efld",)),
            (("T",0,2,0),"ISI",6,"Number of Entries","",
                "","N",self.doQty,None,None,("efld",)),
            (("T",0,3,0),"ISD",13.2,"Value of Entries","",
                "","N",self.doVal,None,None,("efld",)),
            (("T",0,4,0),("IRB",r1s),0,"Multiple Date Allocations","",
                "N","N",self.doMulti,None,None,None)]
        if self.bankctl:
            fld.append(
                (("T",0,5,0),"INA",10,"Bank Control","Bank Control Record",
                    "","N",self.doCtl,ctl,None,("efld",)))
        tnd = ((self.doHeadEnd, "y"), )
        txt = (self.doHeadExit, )
        self.b1 = TartanDialog(self.mf, tops=True, title=tit, eflds=fld,
            tend=tnd, txit=txt)
        self.mf.startLoop()

    def doBatNum(self, frt, pag, r, c, p, i, w):
        self.batno = w
        self.acc = None
        self.ctl = ""
        whr = [
            ("btm_cono", "=", self.conum),
            ("btm_batno", "=", self.batno),
            ("btm_styp", "=", self.sys),
            ("btm_rtyp", "=", self.rtn)]
        self.bat = self.sql.getRec("ctlbat", where=whr, limit=1)
        bc = self.sql.ctlbat_col
        if not self.bat:
            self.new = True
        elif self.bat[bc.index("btm_ind")] == "Y":
            return "This Batch Has Aready Been Balanced"
        elif self.bankctl and not self.bat[bc.index("btm_ctl")]:
            return "This Batch has No Bank Control"
        else:
            cdt = self.bat[self.sql.ctlbat_col.index("btm_curdt")]
            if cdt < self.s_per or cdt > self.e_per:
                return "This Batch Is Out Of The Financial Period"
            self.new = False
            end = self.sql.ctlbat_col.index("btm_ctl")
            if self.bankctl:
                end += 1
            for n, d in enumerate(self.bat[5:end]):
                self.b1.loadEntry(frt, pag, p+n+1, data=d)
            self.ckCtl(self.bat[self.sql.ctlbat_col.index("btm_ctl")])
            return "sk1"

    def doCurdt(self, frt, pag, r, c, p, i, w):
        if w < self.s_per or w > self.e_per:
            return "Invalid, Not In Financial Period"
        if self.rtn == 7:
            self.b1.loadEntry(frt, pag, p+1, data=0)
            self.b1.loadEntry(frt, pag, p+2, data=0)
            return "sk2"

    def doQty(self, frt, pag, r, c, p, i, w):
        if not self.new:
            if w != self.bat[self.sql.ctlbat_col.index("btm_trno")]:
                cf = PwdConfirm(self.mf, conum=self.conum, system="MST",
                    code="ChgBatch")
                if cf.flag == "no":
                    return "Invalid Change of Batch Quantity"

    def doVal(self, frt, pag, r, c, p, i, w):
        if not self.new:
            if w != self.bat[self.sql.ctlbat_col.index("btm_trval")]:
                cf = PwdConfirm(self.mf, conum=self.conum, system="MST",
                    code="ChgBatch")
                if cf.flag == "no":
                    return "Invalid Change of Batch Value"
            return "nd"
        if self.sys not in ("ASS","CRS","DRS","GEN","LON","MEM","RTL","SLN"):
            if not self.bankctl:
                return "nd"
            elif self.acc:
                self.b1.loadEntry(frt, pag, p+2, self.ctl)
                return "nd"
            else:
                return "sk1"
        if self.multi:
            self.b1.loadEntry(frt, pag, p+1, data=self.multi.upper())
            if not self.bankctl:
                return "nd"
            elif self.acc:
                self.b1.loadEntry(frt, pag, p+2, self.ctl)
                return "nd"
            else:
                return "sk1"

    def doMulti(self, frt, pag, r, c, p, i, w):
        if w == "Y":
            cf = PwdConfirm(self.mf, conum=self.conum, system="MST",
                code="MultiDate")
            if cf.flag == "no":
                return "Invalid Multiple Date Allocations"
        if not self.bankctl:
            return "nd"
        elif self.acc:
            self.b1.loadEntry(frt, pag, p+1, self.ctl)
            return "nd"

    def doCtl(self, frt, pag, r, c, p, i, w):
        if w[:4] != "bank":
            self.acc = None
            self.ctl = ""
        else:
            self.ckCtl(w)
        if not self.acc:
            return "Invalid Bank Control Code"

    def ckCtl(self, ctl):
        if not ctl:
            self.acc = None
            self.ctl = ""
            return
        self.ctl = ctl
        gc = GetCtl(self.mf)
        ctlctl = gc.getCtl("ctlctl", self.conum)
        if not ctlctl:
            self.acc = None
        else:
            if gc.chkRec(self.conum, ctlctl, [self.ctl]):
                self.acc = None
            else:
                self.acc = ctlctl[self.ctl]

    def doHeadEnd(self):
        self.curdt = self.b1.t_work[0][0][1]
        self.trno = self.b1.t_work[0][0][2]
        self.trval = self.b1.t_work[0][0][3]
        self.multi = self.b1.t_work[0][0][4]
        self.getTotals()
        if self.new:
            self.sql.insRec("ctlbat", data=[self.conum, self.sys, self.rtn,
                self.batno, self.sysdt, self.curdt, self.trno, self.trval,
                self.multi, self.ctl, "N"])
        else:
            self.sql.updRec("ctlbat", cols=["btm_trno", "btm_trval"],
                data=[self.trno, self.trval], where=[("btm_cono", "=",
                self.conum), ("btm_styp", "=", self.sys), ("btm_rtyp", "=",
                self.rtn), ("btm_batno", "=", self.batno), ("btm_curdt", "=",
                self.curdt)])
        self.mf.dbm.commitDbase()
        self.b1.closeProcess()
        self.mf.closeLoop()

    def getTotals(self):
        if self.sys == "ASS":
            tab = "asstrn"
            col = ["count(*)", "round(sum(ast_amt1), 2)",
                "round(sum(ast_vat), 2)"]
            whr = [("ast_cono", "=", self.conum), ("ast_type", "=", self.rtn),
                ("ast_batch", "=", self.batno), ("ast_curdt", "between",
                self.s_per, self.e_per), ("ast_batind", "=", "N")]
        elif self.sys == "CRS":
            tab = "crstrn"
            col = ["count(*)", "round(sum(crt_tramt), 2)"]
            whr = [("crt_cono", "=", self.conum), ("crt_type", "=", self.rtn),
                ("crt_batch", "=", self.batno), ("crt_curdt", "between",
                self.s_per, self.e_per), ("crt_batind", "=", "N")]
        elif self.sys == "DRS":
            tab = "drstrn"
            col = ["count(*)", "round(sum(drt_tramt), 2)"]
            whr = [("drt_cono", "=", self.conum), ("drt_type", "=", self.rtn),
                ("drt_batch", "=", self.batno), ("drt_curdt", "between",
                self.s_per, self.e_per), ("drt_batind", "=", "N")]
        elif self.sys == "GEN":
            tab = "gentrn"
            col = ["count(*)", "round(sum(glt_tramt), 2)"]
            whr = [("glt_cono", "=", self.conum)]
            if self.rtn == 7:
                whr.append(("glt_type", "in", (2, 4, 6)))
            else:
                whr.append(("glt_type", "=", self.rtn))
            whr.extend([("glt_batch", "=", self.batno), ("glt_curdt",
                "between", self.s_per, self.e_per)])
            if self.rtn != 4:
                whr.append(("glt_batind", "=", "N"))
        elif self.sys == "LON":
            tab = "lontrn"
            col = ["count(*)", "round(sum(lnt_tramt), 2)"]
            whr = [("lnt_cono", "=", self.conum), ("lnt_type", "=", self.rtn),
                ("lnt_batch", "=", self.batno), ("lnt_curdt", "between",
                self.s_per, self.e_per), ("lnt_batind", "=", "N")]
        elif self.sys == "MEM":
            tab = "memtrn"
            col = ["count(*)", "round(sum(mlt_tramt), 2)"]
            whr = [("mlt_cono", "=", self.conum), ("mlt_type", "=", self.rtn),
                ("mlt_batch", "=", self.batno), ("mlt_curdt", "between",
                self.s_per, self.e_per), ("mlt_batind", "=", "N")]
        elif self.sys == "RTL":
            tab = "rtltrn"
            col = ["count(*)", "round(sum(rtt_tramt), 2)"]
            whr = [("rtt_cono", "=", self.conum), ("rtt_type", "=", self.rtn),
                ("rtt_batch", "=", self.batno), ("rtt_curdt", "between",
                self.s_per, self.e_per), ("rtt_batind", "=", "N")]
        elif self.sys == "SLN":
            tab = "wagltf"
            col = ["count(*)", "round(sum(wlt_amt), 2)"]
            whr = [("wlt_cono", "=", self.conum), ("wlt_type", "=", self.rtn),
                ("wlt_batch", "=", self.batno), ("wlt_curdt", "between",
                self.s_per, self.e_per), ("wlt_batind", "=", "N")]
        elif self.sys == "STR":
            tab = "strtrn"
            col = ["count(*)", "round(sum(stt_cost), 2)"]
            whr = [("stt_cono", "=", self.conum), ("stt_type", "=", self.rtn),
                ("stt_batch", "=", self.batno), ("stt_curdt", "between",
                self.s_per, self.e_per), ("stt_batind", "=", "N")]
        sql = Sql(self.mf.dbm, tab, prog=__name__)
        a = sql.getRec(tables=tab, cols=col, where=whr, limit=1)
        if a[0]:
            self.batqty = CCD(a[0], "SI", 6).work
            if self.sys == "ASS":
                self.batval = CCD(float(ASD(a[1]) + ASD(a[2])), "SD", 13.2).work
            else:
                self.batval = CCD(a[1], "SD", 13.2).work
        else:
            self.batqty = CCD(0, "SI", 6).work
            self.batval = CCD(0, "SD", 13.2).work

    def doHeadExit(self):
        self.batno = None
        self.b1.closeProcess()
        self.mf.closeLoop()

    def doBatchTotal(self, det=False):
        self.qty = CCD(self.batqty, "SI", 6)
        self.val = CCD(self.batval, "SD", 13.2)
        self.qex = CCD(self.trno, "SI", 6)
        self.vex = CCD(self.trval, "SD", 13.2)
        self.qdf = CCD((self.qex.work - self.qty.work), "SI", 6)
        self.vdf = CCD(float(ASD(self.vex.work)-ASD(self.val.work)), "SD", 13.2)
        if det:
            self.sql.updRec("ctlbat", cols=["btm_trno", "btm_trval",
                "btm_ind"], data=[self.qty.work, self.val.work, "Y"],
                where=[("btm_cono", "=", self.conum), ("btm_batno", "=",
                self.batno), ("btm_styp", "=", self.sys), ("btm_rtyp", "=",
                self.rtn), ("btm_curdt", "between", self.s_per, self.e_per)])
            self.flagTrans()
        elif not self.qty.work and not self.val.work:
            self.sql.delRec("ctlbat", where=[("btm_cono", "=", self.conum),
                ("btm_batno", "=", self.batno), ("btm_styp", "=", self.sys),
                ("btm_rtyp", "=", self.rtn), ("btm_curdt", "between",
                self.s_per, self.e_per)])
        elif self.qdf.work or self.vdf.work:
            self.doDispTotal()
            self.doLoadDetail()
            self.b2.mstFrame.wait_window()
        elif self.qty.work or self.val.work:
            self.sql.updRec("ctlbat", cols=["btm_ind"], data=["Y"],
                where=[("btm_cono", "=", self.conum), ("btm_batno", "=",
                self.batno), ("btm_styp", "=", self.sys), ("btm_rtyp", "=",
                self.rtn), ("btm_curdt", "between", self.s_per, self.e_per)])
            self.flagTrans()
        self.mf.dbm.commitDbase()

    def doDispTotal(self):
        tit = ("Batch Totals",)
        fld = (
            (("T",0,0,0),"ONa",7,"Batch Number"),
            (("T",0,1,0),"OSI",6,"Number of Entries Expected"),
            (("T",0,2,0),"OSI",6,"                  Entered"),
            (("T",0,3,0),"OSI",6,"                  Difference"),
            (("T",0,4,0),"OSD",13.2,"Value of Entries  Expected"),
            (("T",0,5,0),"OSD",13.2,"                  Entered"),
            (("T",0,6,0),"OSD",13.2,"                  Difference"))
        but = (("Continue",None,self.doTotExit,1,None,None),)
        self.b2 = TartanDialog(self.mf, tops=True, title=tit, eflds=fld,
            butt=but, tend=None, txit=None)

    def doLoadDetail(self):
        self.b2.loadEntry("T", 0, 0, data=self.batno)
        self.b2.loadEntry("T", 0, 1, data=self.qex.work)
        self.b2.loadEntry("T", 0, 2, data=self.qty.work)
        self.b2.loadEntry("T", 0, 3, data=self.qdf.work)
        self.b2.loadEntry("T", 0, 4, data=self.vex.work)
        self.b2.loadEntry("T", 0, 5, data=self.val.work)
        self.b2.loadEntry("T", 0, 6, data=self.vdf.work)
        self.mf.updateStatus("Batch Does Not Balance")
        self.b2.B0.focus_set()

    def flagTrans(self):
        if self.sys == "ASS":
            tab = "asstrn"
            col = ["ast_batind"]
            whr = [("ast_cono", "=", self.conum), ("ast_type", "=", self.rtn),
                ("ast_batch", "=", self.batno), ("ast_curdt", "between",
                self.s_per, self.e_per), ("ast_batind", "=", "N")]
        elif self.sys == "CRS":
            tab = "crstrn"
            col = ["crt_batind"]
            whr = [("crt_cono", "=", self.conum), ("crt_type", "=", self.rtn),
                ("crt_batch", "=", self.batno), ("crt_curdt", "between",
                self.s_per, self.e_per), ("crt_batind", "=", "N")]
        elif self.sys == "DRS":
            tab = "drstrn"
            col = ["drt_batind"]
            whr = [("drt_cono", "=", self.conum), ("drt_type", "=", self.rtn),
                ("drt_batch", "=", self.batno), ("drt_curdt", "between",
                self.s_per, self.e_per), ("drt_batind", "=", "N")]
        elif self.sys == "GEN":
            tab = "gentrn"
            col = ["glt_batind"]
            whr = [("glt_cono", "=", self.conum), ("glt_type", "=", self.rtn),
                ("glt_batch", "=", self.batno), ("glt_curdt", "between",
                self.s_per, self.e_per), ("glt_batind", "=", "N")]
        elif self.sys == "LON":
            tab = "lontrn"
            col = ["lnt_batind"]
            whr = [("lnt_cono", "=", self.conum), ("lnt_type", "=", self.rtn),
                ("lnt_batch", "=", self.batno), ("lnt_curdt", "between",
                self.s_per, self.e_per), ("lnt_batind", "=", "N")]
        elif self.sys == "MEM":
            tab = "memtrn"
            col = ["mlt_batind"]
            whr = [("mlt_cono", "=", self.conum), ("mlt_type", "=", self.rtn),
                ("mlt_batch", "=", self.batno), ("mlt_curdt", "between",
                self.s_per, self.e_per), ("mlt_batind", "=", "N")]
        elif self.sys == "RTL":
            tab = "rtltrn"
            col = ["rtt_batind"]
            whr = [("rtt_cono", "=", self.conum), ("rtt_type", "=", self.rtn),
                ("rtt_batch", "=", self.batno), ("rtt_curdt", "between",
                self.s_per, self.e_per), ("rtt_batind", "=", "N")]
        elif self.sys == "SLN":
            tab = "wagltf"
            col = ["wlt_batind"]
            whr = [("wlt_cono", "=", self.conum), ("wlt_type", "=", self.rtn),
                ("wlt_batch", "=", self.batno), ("wlt_curdt", "between",
                self.s_per, self.e_per), ("wlt_batind", "=", "N")]
        elif self.sys == "STR":
            tab = "strtrn"
            col = ["stt_batind"]
            whr = [("stt_cono", "=", self.conum), ("stt_type", "=", self.rtn),
                ("stt_batch", "=", self.batno), ("stt_curdt", "between",
                self.s_per, self.e_per), ("stt_batind", "=", "N")]
        sql = Sql(self.mf.dbm, tab, prog=__name__)
        sql.updRec(tab, cols=col, data=["Y"], where=whr)

    def doTotExit(self):
        self.b2.closeProcess()

class FinReport(object):
    """
    This class generates a list comprising:

        Type of line i.e. H, L, G, T, S, U, C, P etc.
        Account Type (B,P,O)
        Print Highlighted (y or n)
        Formfeed (y/n)
        Account Number in the case of 'L'
        Description i.e. Description or Single/Double/Blank for underlines
        Line count
        Number of lines to skip
        List of Amounts (N/A,Lyr x 3,Cmt,Bmt,Ytd,Btd) for type "N"
        List of Monthly Amounts (Balances or Values) x 14 i.e.
            Opening Balance + 12 Monthly Balances + Closing Balance for type 'M'
        Chart Label i.e.
            Whether to include this total in a bar or line chart
        Page Heading i.e. Balance Sheet, Profit and Loss or Other
        Normal sign (+/-)

    Using the report file as follows:

    Num  Field       FF  Size  Description                     Heading
    ---  ----------  --  ----  ------------------------------  -----------
      0  glr_cono    UI   3.0  Company Number                  Coy
      1  glr_repno   UI   3.0  Report Number                   No
      2  glr_seq     UD   7.2  Sequence                        Seq-Num
      3  glr_type    UA   1.0  Type                            T
      4  glr_desc    NA  30.0  Description                     Description
      5  glr_high    UA   1.0  Highlight (y/n)                 H
      6  glr_ffeed   UA   1.0  Form Feed (y/n)                 F
      7  glr_ignore  UA   1.0  Ignore Account Type (y/n)       I
      8  glr_from    UI   7.0  From Account                    From-Ac
      9  glr_to      UI   7.0  To Account                      To-Ac
     10  glr_obal    UA   1.0  Include Opening Balance (y/n)   B
     11  glr_accum   UA   1.0  Accumulate Month Values (y/n)   A
     12  glr_print   UA   1.0  Print Values (y/n/+/-)          P
     13  glr_norm    UA   1.0  Normal Sign (+/-)               S
     14  glr_acbal   UA   1.0  Add/Sub/Ignore Balance (a/s/i)  A
     15  glr_store   UA   1.0  Store Balances (y/n)            S
     16  glr_snum1   UI   2.0  Storage Number 1                N1
     17  glr_snum2   UI   2.0  Storage Number 2                N2
     18  glr_snum3   UI   2.0  Storage Number 3                N3
     19  glr_acstr   UA   1.0  Add/Sub/Ignore Storage (a/s/i)  A
     20  glr_group   UI   3.0  Group Number                    Grp
     21  glr_total   UI   1.0  Total Level                     T
     22  glr_clear   UA   1.0  Clear Total/Storage Level(s)    C
     23  glr_strper  UD   6.2  Percentage of Stored Amount     %-tage
     24  glr_uline   UA   1.0  Underline (b/d/s)               U
     25  glr_cbase   UA   1.0  Calculation Base (p/a/s)        B
     26  glr_ctype   UA   1.0  Calculation Type (+ - * /)      C
     27  glr_camnt   SD  13.2  Percent or Amount               Amount
     28  glr_label   NA  10.0  Chart Label, Space = None       Label
     29 glr_xflag    UA   1.0  Export Flag                     X

    as well as the following parameters:

    mf       = The Mainframe class
    sysdp    = The Department Parameters - (Y or N, Digits and Department(s))
    conum    = The Company to be used
    period   = The Financial Period to be used
    repco    = The report layout company number or 0 for a general report
    curdt    = The current financial period of the report e.g. 200302
    vcode    = The value code to be used:
                V = Values
                B = Budgets
                C = Combination of Values and Budgets
                X = Variance to Budgets
                D = Details
    dcode    = The detail code to be used if vcode = "D"
    varcd    = The value to use for variances, B=Budget, P=Previous Year
    consol   = A tuple or list of Company Numbers to Consolidate or
                "Y" for All or "N" for None
    depart   = The department code to print else 0 for None
    rtype    = The basis for the printing parameter to be based on
                Y = Use the YTD value or
                M = Use the current period value
    """
    def __init__(self, mf, sysdp, conum, period, repco, repno, curdt, **args):
        self.mf = mf
        self.sysdp = sysdp
        self.conum = conum
        self.period = period
        self.repco = repco
        self.repno = repno
        self.curdt = curdt
        defaults = {
            "vcode": "V",
            "dcode": None,
            "varcd": "B",
            "rtype": "Y",
            "consol": "N",
            "depart": 0}
        for nam in args:
            defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        if self.setVariables():
            self.pb = ProgressBar(self.mf.body, typ="G")
            self.mainProcess()
            self.pb.closeProgress()

    def setVariables(self):
        """
        self.pers: Dictionary of current period and 3 previous periods:

        o_dte = Opening balance date
        i_per = Starting Period
        e_per = Ending Period
        d_per = Number of months over 12
        s_per = Starting period in 12 month period
        """
        self.sql = Sql(self.mf.dbm, ["ctlmst", "ctldep", "genbal", "genbud",
            "gendtm", "gendtt", "genmst", "genrpt", "gentrn"], prog=__name__)
        if self.sql.error:
            return
        self.pers = {0: {}, 1: {}, 2: {}, 3: {}}
        # Current Period
        self.pers[0]["o_dte"] = self.period[1][0]
        self.pers[0]["i_per"] = int(self.period[1][0] / 100)
        self.pers[0]["e_per"] = int(self.period[2][0] / 100)
        df = dateDiff(self.period[1][0], self.period[2][0], "months") + 1
        if df > 12:
            self.pers[0]["d_per"] = df - 12
            yr = int(self.pers[0]["i_per"] / 100)
            mt = self.pers[0]["i_per"] % 100
            for _ in range(self.pers[0]["d_per"]):
                mt += 1
                if mt > 12:
                    mt -= 12
                    yr += 1
            self.pers[0]["s_per"] = (yr * 100) + mt
        else:
            self.pers[0]["d_per"] = 0
            self.pers[0]["s_per"] = self.pers[0]["i_per"]
        # Previous Periods
        if self.period[0]:
            if self.period[0] > 2:
                rng = 4
            else:
                rng = self.period[0] + 1
            for p in range(1, rng):
                if p == 1:
                    s = 3
                elif p == 2:
                    s = 2
                else:
                    s = 1
                i_lyr, e_lyr, fin = getPeriods(self.mf, self.conum,
                    self.period[0] - p)
                self.pers[s]["o_dte"] = i_lyr.work
                self.pers[s]["i_per"] = int(i_lyr.work / 100)
                self.pers[s]["e_per"] = int(e_lyr.work / 100)
                df = dateDiff(e_lyr.work, i_lyr.work, "months") + 1
                if df > 12:
                    self.pers[s]["d_per"] = df - 12
                    yr = int(self.pers[s]["i_pyr"] / 100)
                    mt = self.pers[s]["i_per"] % 100
                    for _ in range(self.pers[s]["d_per"]):
                        mt += 1
                        if mt > 12:
                            mt -= 12
                            yr += 1
                    self.pers[s]["s_per"] = (yr * 100) + mt
                else:
                    self.pers[s]["d_per"] = 0
                    self.pers[s]["s_per"] = self.pers[s]["i_per"]
        self.cmth = self.curdt % 100
        self.lines = 1
        self.count = 0
        self.grpind = False
        self.xits = False
        self.clearData()
        self.allFields = []
        self.val_tot = makeArray(8, 10, 1)
        self.mth_tot = makeArray(14, 10, 1)
        self.val_str = makeArray(8, 100, 1)
        self.mth_str = makeArray(14, 100, 1)
        if type(self.consol) == str:
            if self.consol == "Y":
                coy = self.sql.getRec("ctlmst", cols=["ctm_cono"],
                    order="ctm_cono")
                if not coy:
                    return
                self.consol = []
                for c in coy:
                    self.consol.append(c[0])
            else:
                self.consol = []
        self.dname = None
        if self.sysdp[0] == "Y":
            self.alldeps = [0]
            deps = self.sql.getRec("ctldep", cols=["dep_code",
                "dep_name"], where=[("dep_cono", "=", self.conum)],
                order="dep_code")
            for dep in deps:
                self.alldeps.append(dep[0])
                if self.depart and dep[0] == self.depart:
                    self.dname = dep[1]
        return True

    def mainProcess(self):
        if self.vcode == "D":
            self.getDcode()
        self.rpt_lst = self.sql.getRec("genrpt", where=[("glr_cono",
            "=", self.repco), ("glr_repno", "=", self.repno)], order="glr_seq")
        self.rep_num = 0
        if self.rpt_lst and self.rpt_lst[self.rep_num][3] in ("P","B","O"):
            self.s_type = self.rpt_lst[0][3]
            self.atype = self.s_type
            if self.s_type == "P":
                self.stanam = "Profit and Loss"
            elif self.s_type == "B":
                self.stanam = "Balance Sheet"
            elif self.s_type == "O":
                self.stanam = self.rpt_lst[self.rep_num][4]
            if self.dname:
                self.stanam = "%s - %s" % (self.dname, self.stanam)
            while not self.xits:
                self.processRecords()

    def getDcode(self):
        dtm = self.sql.getRec("gendtm", cols=["gdm_desc"],
            where=[("gdm_cono", "=", self.conum), ("gdm_code", "=",
            self.dcode)], limit=1)
        dtt = self.sql.getRec("gendtt", cols=["sum(gdt_value)"],
            where=[("gdt_cono", "=", self.conum), ("gdt_code", "=",
            self.dcode), ("gdt_curdt", "between", self.pers[0]["s_per"],
            self.curdt)], limit=1)
        if not dtt:
            self.ddet = [dtm[0], CCD(0, "SD", 13.2).disp]
        else:
            self.ddet = [dtm[0], CCD(dtt[0], "SD", 13.2).disp]

    def processRecords(self):
        if self.grpind:
            self.grpind = False
        else:
            self.nextRepRecord()
            self.storeRepRecord()
            if self.rpt_dic == {}:
                self.xits = True
        if not self.xits:
            if self.s_type == "H":
                self.doHead()
            elif self.s_type == "L":
                self.doLedger()
            elif self.s_type == "G":
                self.doGroup()
            elif self.s_type == "T":
                self.doTotal()
            elif self.s_type == "S":
                self.doStore()
            elif self.s_type == "U":
                self.doUnderline()
            elif self.s_type == "P":
                self.doPercent()
            elif self.s_type == "C":
                self.doCalc()

    def doHead(self):
        self.pb.displayProgress()
        if self.rpt_dic["glr_ignore"] == "Y":
            if self.atype == "P":
                self.checkAccountType("B")
            else:
                self.checkAccountType("P")
        self.printLine()
        self.clearData()

    def doLedger(self):
        accs = self.readGenmst()
        for acc in accs:
            self.pb.displayProgress()
            self.lstdicGenmst(acc)
            self.s_cono = self.gen_dic["glm_cono"]
            self.s_acno = self.gen_dic["glm_acno"]
            self.s_desc = self.gen_dic["glm_desc"]
            self.accumData()
            if self.s_accum == "Y":
                # Accumulate progressive monthly totals
                for mth in range(1, 13):
                    self.mth[mth] = float(ASD(self.mth[mth]) +
                        ASD(self.mth[mth-1]))
            if self.rtype == "Y":
                amt = self.val[6]
            else:
                amt = self.mth[self.cmth]
            if self.s_print == "+" and amt < 0:
                self.clearData()
                continue
            if self.s_print == "-" and amt >= 0:
                self.clearData()
                continue
            if self.s_print != "N":
                self.printLine()
            self.accumTotals(self.s_acbal)
            if self.s_store == "Y":
                self.storeBalances()
            self.clearData()

    def doGroup(self):
        self.grpind = True
        while self.rpt_dic and self.rpt_dic["glr_type"] == "G" and \
                        self.rpt_dic["glr_group"] == self.s_group:
            accs = self.readGenmst()
            for acc in accs:
                self.pb.displayProgress()
                self.lstdicGenmst(acc)
                self.s_cono = self.gen_dic["glm_cono"]
                self.s_acno = self.gen_dic["glm_acno"]
                self.accumData()
            self.nextRepRecord()
        if self.s_accum == "Y":
            # Accumulate progressive monthly totals
            for mth in range(1, 13):
                self.mth[mth] = float(ASD(self.mth[mth]) + ASD(self.mth[mth-1]))
        if self.rtype == "Y":
            amt = self.val[6]
        else:
            amt = self.mth[self.cmth]
        if self.s_print == "+" and amt < 0:
            pass
        elif self.s_print == "-" and amt >= 0:
            pass
        else:
            if self.s_print != "N":
                self.printLine()
            self.accumTotals(self.s_acbal)
            if self.s_store == "Y":
                self.storeBalances()
        self.clearData()
        self.storeRepRecord()

    def doTotal(self):
        self.pb.displayProgress()
        for x in range(0, 8):
            self.val[x] = self.val_tot[self.s_total][x]
        for x in range(0, 14):
            self.mth[x] = self.mth_tot[self.s_total][x]
        if self.rtype == "Y":
            amt = self.val[6]
        else:
            amt = self.mth[self.cmth]
        if self.s_print == "+" and amt < 0:
            self.clearData()
            return
        elif self.s_print == "-" and amt >= 0:
            self.clearData()
            return
        elif self.s_print != "N":
            self.printLine()
        if self.s_store == "Y":
            self.storeBalances()
        self.clearData()
        if self.s_clear == "Y":
            for x in range(0, self.s_total+1):
                for y in range(0, 8):
                    self.val_tot[x][y] = 0
                for y in range(0, 14):
                    self.mth_tot[x][y] = 0

    def doStore(self):
        self.pb.displayProgress()
        for x in range(0, 8):
            if self.val_str[self.s_snum1][x] == 0:
                self.val[x] = 0
            else:
                self.val[x] = round((self.val_str[self.s_snum1][x] *
                    self.s_strper / 100.0), 2)
        for x in range(0, 13):
            if self.mth_str[self.s_snum1][x] == 0:
                self.mth[x] = 0
            else:
                self.mth[x] = round((self.mth_str[self.s_snum1][x] *
                    self.s_strper / 100.0), 2)
                self.mth[13] = float(ASD(self.mth[13]) + ASD(self.mth[x]))
        if self.rtype == "Y":
            amt = self.val[6]
        else:
            amt = self.mth[self.cmth]
        if self.s_print == "+" and amt < 0:
            self.clearData()
            return
        elif self.s_print == "-" and amt >= 0:
            self.clearData()
            return
        elif self.s_print != "N":
            self.printLine()
        self.accumTotals(self.s_acbal)
        self.clearData()
        if self.s_clear == "Y":
            for x in range(0, 8):
                self.val_str[self.s_snum1][x] = 0
            for x in range(0, 14):
                self.mth_str[self.s_snum1][x] = 0

    def doUnderline(self):
        self.pb.displayProgress()
        if self.rpt_dic["glr_uline"] == "B":
            self.s_desc = "Blank"
        elif self.rpt_dic["glr_uline"] == "S":
            self.s_desc = "Single"
        else:
            self.s_desc = "Double"
        self.printLine()
        self.clearData()

    def doPercent(self):
        self.pb.displayProgress()
        for x in range(0, 8):
            if self.val_str[self.s_snum1][x] == 0 or \
                    self.val_str[self.s_snum2][x] == 0:
                self.val[x] = 0
            else:
                self.val[x] = round((self.val_str[self.s_snum1][x] /
                    self.val_str[self.s_snum2][x]) * 100.0, 2)
        for x in range(0, 14):
            if self.mth_str[self.s_snum1][x] == 0 or \
                    self.mth_str[self.s_snum2][x] == 0:
                self.mth[x] = 0
            else:
                self.mth[x] = round((self.mth_str[self.s_snum1][x] /
                    self.mth_str[self.s_snum2][x]) * 100.0, 2)
        self.printLine()
        self.clearData()

    def doCalc(self):
        self.pb.displayProgress()
        for x in range(0, 8):
            amt1 = self.val_str[self.s_snum2][x]
            if self.s_cbase == "S":
                amt2 = self.val_str[self.s_snum3][x]
            else:
                amt2 = self.s_camnt
            if self.s_ctype == "+":
                self.val[x] = float(ASD(amt1) + ASD(amt2))
            elif self.s_ctype == "-":
                self.val[x] = float(ASD(amt1) - ASD(amt2))
            elif self.s_ctype == "*":
                if self.s_cbase == "P":
                    self.val[x] = round(amt1 * amt2 / 100.0, 2)
                else:
                    self.val[x] = round(amt1 * amt2, 2)
            elif self.s_ctype == "/":
                self.val[x] = round(amt1 / amt2, 2)
        for x in range(0, 14):
            amt1 = self.mth_str[self.s_snum2][x]
            if self.s_cbase == "S":
                amt2 = self.mth_str[self.s_snum3][x]
            else:
                amt2 = self.s_camnt
            if self.s_ctype == "+":
                self.mth[x] = float(ASD(amt1) + ASD(amt2))
            elif self.s_ctype == "-":
                self.mth[x] = float(ASD(amt1) - ASD(amt2))
            elif self.s_ctype == "*":
                if self.s_cbase == "P":
                    self.mth[x] = round(amt1 * amt2 / 100.0, 0)
                else:
                    self.mth[x] = round(amt1 * amt2, 0)
            elif self.s_ctype == "/":
                self.mth[x] = round(amt1 / amt2, 0)
        self.printLine()
        self.clearData()

    def readGenmst(self):
        glfrom = self.rpt_dic["glr_from"]
        glto = self.rpt_dic["glr_to"]
        if self.sysdp[0] == "Y":
            if self.depart:
                deps = [self.depart]
            else:
                deps = copyList(self.alldeps)
        if not self.consol:
            if self.sysdp[0] == "Y":
                accs = []
                for dep in deps:
                    glf = (dep * (10 ** (7 - self.sysdp[1]))) + glfrom
                    if glto:
                        glt = (dep * (10 ** (7 - self.sysdp[1]))) + glto
                        acs = self.sql.getRec("genmst",
                            where=[("glm_cono", "=", self.conum),
                            ("glm_acno", ">=", glf),
                            ("glm_acno", "<=", glt)],
                            order="glm_acno")
                        if acs:
                            accs.extend(acs)
                    else:
                        acs = self.sql.getRec("genmst",
                            where=[("glm_cono", "=", self.conum),
                            ("glm_acno", "=", glf)])
                        if acs:
                            accs.extend(acs)
                if not self.depart:
                    temp = copyList(accs)
                    done = []
                    accs = []
                    for acc in temp:
                        num = acc[1] % (10 ** (7 - self.sysdp[1]))
                        if num in done:
                            continue
                        acc[1] = num
                        accs.append(acc)
                        done.append(num)
            else:
                if not glto:
                    accs = self.sql.getRec("genmst",
                        where=[("glm_cono", "=", self.conum),
                        ("glm_acno", "=", glfrom)])
                else:
                    accs = self.sql.getRec("genmst",
                        where=[("glm_cono", "=", self.conum),
                        ("glm_acno", ">=", glfrom),
                        ("glm_acno", "<=", glto)],
                        order="glm_acno")
                if not accs:
                    accs = []
        else:
            accs = []
            if self.sysdp[0] == "Y":
                for dep in deps:
                    glf = (dep * (10 ** (7 - self.sysdp[1]))) + glfrom
                    if glto:
                        glt = (dep * (10 ** (7 - self.sysdp[1]))) + glto
                        acs = self.sql.getRec("genmst",
                            cols=["glm_acno"], where=[("glm_cono", "in",
                            self.consol), ("glm_acno", ">=", glf),
                            ("glm_acno", "<=", glt)], group="glm_acno",
                            order="glm_acno")
                    else:
                        acs = self.sql.getRec("genmst",
                            cols=["glm_acno"], where=[("glm_cono", "in",
                            self.consol), ("glm_acno", "=", glf)],
                            group="glm_acno")
                    if acs:
                        for ac in acs:
                            accs.append(self.sql.getRec("genmst",
                                where=[("glm_cono", "in", self.consol),
                                ("glm_acno", "=", ac[0])], limit=1))
                temp = copyList(accs)
                done = []
                accs = []
                for acc in temp:
                    num = acc[1] % (10 ** (7 - self.sysdp[1]))
                    if num in done:
                        continue
                    acc[1] = num
                    accs.append(acc)
                    done.append(num)
            else:
                if not glto:
                    acs = self.sql.getRec("genmst",
                        cols=["glm_acno"], where=[("glm_cono", "in",
                        self.consol), ("glm_acno", "=", glfrom)],
                        group="glm_acno")
                else:
                    acs = self.sql.getRec("genmst",
                        cols=["glm_acno"], where=[("glm_cono", "in",
                        self.consol), ("glm_acno", ">=", glfrom),
                        ("glm_acno", "<=", glto)], group="glm_acno",
                        order="glm_acno")
                if acs:
                    for ac in acs:
                        accs.append(self.sql.getRec("genmst",
                            where=[("glm_cono", "in", self.consol),
                            ("glm_acno", "=", ac[0])], limit=1))
        return accs

    def lstdicGenmst(self, acc):
        self.gen_lst = acc
        self.gen_dic = {}
        num = 0
        for num, fld in enumerate(self.sql.genmst_col):
            self.gen_dic[fld] = acc[num]
        if self.rpt_dic["glr_ignore"] == "N":
            self.checkAccountType(self.gen_dic["glm_type"])

    def clearData(self):
        self.val = [0,0,0,0,0,0,0,0]
        self.mth = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]

    def accumData(self):
        if self.rpt_dic["glr_obal"] == "Y" and self.vcode not in ("B", "X"):
            o_lyr, o_cyr = self.getObal()
            self.val[1] = float(ASD(self.val[1]) + ASD(o_lyr[0]))
            self.val[2] = float(ASD(self.val[2]) + ASD(o_lyr[1]))
            self.val[3] = float(ASD(self.val[3]) + ASD(o_lyr[2]))
            if self.varcd == "P":
                self.val[7] = float(ASD(self.val[7]) + ASD(o_lyr[2]))
            self.val[6] = float(ASD(self.val[6]) + ASD(o_cyr))
            self.mth[0] = float(ASD(self.mth[0]) + ASD(o_cyr))
            self.mth[13] = float(ASD(self.mth[13]) + ASD(o_cyr))
        for p in range(1, 4):
            if not self.pers[p]:
                continue
            if self.pers[p]["d_per"] and self.vcode not in ("B", "C", "X"):
                # Previous Financial period exceeds 12 months
                cy = int(self.pers[p]["i_per"] / 100)
                cm = int(self.pers[p]["i_per"] % 100)
                for _ in range(self.pers[p]["d_per"]):
                    curdt = (cy * 100) + cm
                    lyr, cyr, bud = self.getMbal(curdt)
                    self.val[p] = float(ASD(self.val[p]) + ASD(lyr[p-1]))
                    if p == 3 and self.varcd == "P":
                        self.val[5] = float(ASD(self.val[5]) + ASD(bud))
                        self.val[7] = float(ASD(self.val[7]) + ASD(bud))
                    cm += 1
                    if cm > 12:
                        cm -= 12
                        cy += 1
        if self.pers[0]["d_per"]:
            # Current Financial period exceeds 12 months
            cy = int(self.pers[0]["i_per"] / 100)
            cm = int(self.pers[0]["i_per"] % 100)
            for _ in range(self.pers[0]["d_per"]):
                curdt = (cy * 100) + cm
                lyr, cyr, bud = self.getMbal(curdt)
                if self.vcode == "B":
                    self.mth[1] = float(ASD(self.mth[1]) + ASD(bud))
                    self.mth[13] = float(ASD(self.mth[13]) + ASD(bud))
                elif self.vcode == "C":
                    self.mth[1] = float(ASD(self.mth[1]) + ASD(cyr))
                    self.mth[13] = float(ASD(self.mth[13]) + ASD(cyr))
                elif self.vcode == "X":
                    self.mth[1] = float(ASD(self.mth[1]) + ASD(bud))
                    self.mth[13] = float(ASD(self.mth[13]) + ASD(bud))
                    self.mth[1] = float(ASD(self.mth[1]) - ASD(cyr))
                    self.mth[13] = float(ASD(self.mth[13]) - ASD(cyr))
                else:
                    self.val[1] = float(ASD(self.val[1]) + ASD(lyr[0]))
                    self.val[2] = float(ASD(self.val[2]) + ASD(lyr[1]))
                    self.val[3] = float(ASD(self.val[3]) + ASD(lyr[2]))
                    self.val[4] = float(ASD(self.val[4]) + ASD(cyr))
                    self.val[5] = float(ASD(self.val[5]) + ASD(bud))
                    self.val[6] = float(ASD(self.val[6]) + ASD(cyr))
                    self.val[7] = float(ASD(self.val[7]) + ASD(bud))
                    self.mth[1] = float(ASD(self.mth[1]) + ASD(cyr))
                    self.mth[13] = float(ASD(self.mth[13]) + ASD(cyr))
                cm += 1
                if cm > 12:
                    cm -= 12
                    cy += 1
        if self.vcode == "C":
            cut = False
        curdt = self.pers[0]["s_per"]
        for mth in range(1, 13):
            lyr, cyr, bud = self.getMbal(curdt)
            if self.vcode == "B" or (self.vcode == "C" and cut):
                self.mth[mth] = float(ASD(self.mth[mth]) + ASD(bud))
                self.mth[13] = float(ASD(self.mth[13]) + ASD(bud))
            elif self.vcode == "C":
                self.mth[mth] = float(ASD(self.mth[mth]) + ASD(cyr))
                self.mth[13] = float(ASD(self.mth[13]) + ASD(cyr))
            elif self.vcode == "X":
                self.mth[mth] = float(ASD(self.mth[mth]) + ASD(bud))
                self.mth[13] = float(ASD(self.mth[13]) + ASD(bud))
                self.mth[mth] = float(ASD(self.mth[mth]) - ASD(cyr))
                self.mth[13] = float(ASD(self.mth[13]) - ASD(cyr))
            else:
                self.val[1] = float(ASD(self.val[1]) + ASD(lyr[0]))
                self.val[2] = float(ASD(self.val[2]) + ASD(lyr[1]))
                self.val[3] = float(ASD(self.val[3]) + ASD(lyr[2]))
                if curdt == self.curdt:
                    self.val[4] = float(ASD(self.val[4]) + ASD(cyr))
                    self.val[5] = float(ASD(self.val[5]) + ASD(bud))
                self.val[6] = float(ASD(self.val[6]) + ASD(cyr))
                self.val[7] = float(ASD(self.val[7]) + ASD(bud))
                self.mth[mth] = float(ASD(self.mth[mth]) + ASD(cyr))
                self.mth[13] = float(ASD(self.mth[13]) + ASD(cyr))
            if curdt == self.curdt:
                if self.vcode == "C":
                    cut = True
                else:
                    break
            y = int(curdt / 100)
            m = int(curdt % 100)
            m += 1
            if m > 12:
                y += 1
                m -= 12
            curdt = (y * 100) + m

    def getObal(self):
        lyr = [0, 0, 0]
        if not self.consol:
            if self.sysdp[0] == "Y" and not self.depart:
                bas = 10 ** (7 - self.sysdp[1])
                acc = self.s_acno % bas
                if self.period[0]:
                    for p in self.pers:
                        if self.pers[p]:
                            bal = self.sql.getRec("genbal",
                                cols=["round(sum(glo_cyr), 2)"],
                                where=[("glo_cono", "=", self.s_cono),
                                ("glo_acno", "%", bas, "=", acc),
                                ("glo_trdt", "=", self.pers[p]["o_dte"])],
                                limit=1)
                            if not bal[0]:
                                lyr[p-1] = 0
                            else:
                                lyr[p-1] = bal[0]
                cyr = self.sql.getRec("genbal",
                    cols=["round(sum(glo_cyr), 2)"], where=[("glo_cono", "=",
                    self.s_cono), ("glo_acno", "%", bas, "=", acc),
                    ("glo_trdt", "=", self.pers[0]["o_dte"])], limit=1)[0]
            else:
                if self.period[0]:
                    for p in self.pers:
                        if self.pers[p]:
                            bal = self.sql.getRec("genbal",
                                cols=["round(sum(glo_cyr), 2)"],
                                where=[("glo_cono", "=", self.s_cono),
                                ("glo_acno", "=", self.s_acno),
                                ("glo_trdt", "=", self.pers[p]["o_dte"])],
                                limit=1)
                            if not bal[0]:
                                lyr[p-1] = 0
                            else:
                                lyr[p-1] = bal[0]
                cyr = self.sql.getRec("genbal",
                    cols=["round(sum(glo_cyr), 2)"], where=[("glo_cono", "=",
                    self.s_cono), ("glo_acno", "=", self.s_acno), ("glo_trdt",
                    "=", self.pers[0]["o_dte"])], limit=1)[0]
        else:
            if self.sysdp[0] == "Y" and not self.depart:
                bas = 10 ** (7 - self.sysdp[1])
                acc = self.s_acno % bas
                if self.period[0]:
                    for p in self.pers:
                        if self.pers[p]:
                            bal = self.sql.getRec("genbal",
                                cols=["round(sum(glo_cyr), 2)"],
                                where=[("glo_cono", "in", self.consol),
                                ("glo_acno", "%", bas, "=", acc),
                                ("glo_trdt", "=", self.pers[p]["o_dte"])],
                                limit=1)
                            if not bal[0]:
                                lyr[p-1] = 0
                            else:
                                lyr[p-1] = bal[0]
                cyr = self.sql.getRec("genbal",
                    cols=["round(sum(glo_cyr), 2)"], where=[("glo_cono", "in",
                        self.consol), ("glo_acno", "%", bas, "=", acc),
                        ("glo_trdt", "=", self.per[0]["o_dte"])], limit=1)[0]
            else:
                if self.period[0]:
                    for p in self.pers:
                        if self.pers[p]:
                            bal = self.sql.getRec("genbal",
                                cols=["round(sum(glo_cyr), 2)"],
                                where=[("glo_cono", "in", self.consol),
                                ("glo_acno", "=", self.s_acno),
                                ("glo_trdt", "=", self.pers[p]["o_dte"])],
                                limit=1)
                            if not bal[0]:
                                lyr[p-1] = 0
                            else:
                                lyr[p-1] = bal[0]
                cyr = self.sql.getRec("genbal",
                    cols=["round(sum(glo_cyr), 2)"], where=[("glo_cono", "in",
                    self.consol), ("glo_acno", "=", self.s_acno), ("glo_trdt",
                    "=", self.pers[0]["o_dte"])], limit=1)[0]
        if not cyr:
            cyr = 0
        return (lyr, cyr)

    def getMbal(self, curdt):
        if not self.consol:
            if self.sysdp[0] == "Y" and not self.depart:
                bas = 10 ** (7 - self.sysdp[1])
                acc = self.s_acno % bas
                lyr = [0, 0, 0]
                cdt = curdt
                for p in range(3, 0, -1):
                    if self.pers[p]:
                        cdt -= 100
                        bal = self.sql.getRec("gentrn",
                            cols=["round(sum(glt_tramt), 2)"],
                            where=[("glt_cono", "=", self.s_cono),
                            ("glt_acno", "%", bas, "=", acc),
                            ("glt_curdt", "=", cdt)], limit=1)
                        if bal[0]:
                            lyr[p-1] = bal[0]
                cyr = self.sql.getRec("gentrn",
                    cols=["round(sum(glt_tramt), 2)"], where=[("glt_cono",
                    "=", self.s_cono), ("glt_acno", "%", bas, "=", acc),
                    ("glt_curdt", "=", curdt)], limit=1)[0]
                if self.varcd == "P":
                    bud = lyr[2]
                elif self.varcd == "N":
                    bud = 0
                else:
                    bud = self.sql.getRec("genbud",
                        cols=["sum(glb_tramt)"], where=[("glb_cono",
                        "=", self.s_cono), ("glb_acno", "%", bas, "=", acc),
                        ("glb_curdt", "=", curdt)], limit=1)[0]
            else:
                lyr = [0, 0, 0]
                cdt = curdt
                for p in range(3, 0, -1):
                    if self.pers[p]:
                        cdt -= 100
                        bal = self.sql.getRec("gentrn",
                            cols=["round(sum(glt_tramt), 2)"],
                            where=[("glt_cono", "=", self.s_cono),
                            ("glt_acno", "=", self.s_acno),
                            ("glt_curdt", "=", cdt)], limit=1)
                        if bal[0]:
                            lyr[p-1] = bal[0]
                cyr = self.sql.getRec("gentrn",
                    cols=["round(sum(glt_tramt), 2)"], where=[("glt_cono",
                    "=", self.s_cono), ("glt_acno", "=", self.s_acno),
                    ("glt_curdt", "=", curdt)], limit=1)[0]
                if self.varcd == "P":
                    bud = lyr[2]
                elif self.varcd == "N":
                    bud = 0
                else:
                    bud = self.sql.getRec("genbud",
                        cols=["sum(glb_tramt)"], where=[("glb_cono",
                        "=", self.s_cono), ("glb_acno", "=", self.s_acno),
                        ("glb_curdt", "=", curdt)], limit=1)[0]
        elif self.sysdp[0] == "Y" and not self.depart:
            bas = 10 ** (7 - self.sysdp[1])
            acc = self.s_acno % bas
            lyr = [0, 0, 0]
            cdt = curdt
            for p in range(3, 0, -1):
                if self.pers[p]:
                    cdt -= 100
                    bal = self.sql.getRec("gentrn",
                        cols=["round(sum(glt_tramt), 2)"],
                        where=[("glt_cono", "in", self.consol),
                        ("glt_acno", "%", bas, "=", acc),
                        ("glt_curdt", "=", cdt)], limit=1)
                    if bal[0]:
                        lyr[p-1] = bal[0]
            cyr = self.sql.getRec("gentrn",
                cols=["round(sum(glt_tramt), 2)"], where=[("glt_cono", "in",
                self.consol), ("glt_acno", "%", bas, "=", acc),
                ("glt_curdt", "=", curdt)], limit=1)[0]
            if self.varcd == "P":
                bud = lyr[2]
            elif self.varcd == "N":
                bud = 0
            else:
                bud = self.sql.getRec("genbud", cols=["sum(glb_tramt)"],
                    where=[("glb_cono", "in", self.consol), ("glb_acno", "%",
                    bas, "=", acc), ("glb_curdt", "=", curdt)], limit=1)[0]
        else:
            lyr = [0, 0, 0]
            cdt = curdt
            for p in range(3, 0, -1):
                if self.pers[p]:
                    cdt -= 100
                    bal = self.sql.getRec("gentrn",
                        cols=["round(sum(glt_tramt), 2)"],
                        where=[("glt_cono", "in", self.consol),
                        ("glt_acno", "=", self.s_acno),
                        ("glt_curdt", "=", cdt)], limit=1)
                    if bal[0]:
                        lyr[p-1] = bal[0]
            cyr = self.sql.getRec("gentrn",
                cols=["round(sum(glt_tramt), 2)"], where=[("glt_cono", "in",
                self.consol), ("glt_acno", "=", self.s_acno), ("glt_curdt",
                "=", curdt)], limit=1)[0]
            if self.varcd == "P":
                bud = lyr[2]
            elif self.varcd == "N":
                bud = 0
            else:
                bud = self.sql.getRec("genbud", cols=["sum(glb_tramt)"],
                    where=[("glb_cono", "in", self.consol), ("glb_acno", "=",
                    self.s_acno), ("glb_curdt", "=", curdt)], limit=1)[0]
        if not cyr:
            cyr = 0
        if not bud:
            bud = 0
        return (lyr, cyr, bud)

    def printLine(self):
        self.signChkChg()
        if self.vcode == "D":
            val, mth = self.doDcode()
        else:
            val = copyList(self.val)
            mth = copyList(self.mth)
        if self.s_type != "L":
            self.s_acno = 0
        self.allFields.append([self.s_type, self.atype, self.s_high,
            self.s_ffeed, self.s_acno, self.s_desc, self.count, self.lines,
            val, mth, self.s_label, self.stanam, self.s_norm])
        self.signChkChg()
        self.count += 1
        self.lines = 1

    def doDcode(self):
        curdt = self.pers[0]["s_per"]
        val_d = [0,0,0,0,0]
        mth_d = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        val_p = [0,0,0,0,0,0,0,0]
        mth_p = [0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        for mth in range(1, 13):
            lyr, cyr = self.getDbal(curdt)
            for x in range(3):
                val_d[x] = float(ASD(val_d[x]) + ASD(lyr[x]))
            if curdt == self.curdt:
                val_d[3] = cyr
            val_d[4] = float(ASD(val_d[4]) + ASD(cyr))
            if self.s_accum == "Y":
                mth_d[mth] = float(ASD(mth_d[mth-1]) + \
                    ASD(mth_d[mth]) + ASD(cyr))
                mth_d[13] = mth_d[mth]
            else:
                mth_d[mth] = float(ASD(mth_d[mth]) + ASD(cyr))
                mth_d[13] = float(ASD(mth_d[13]) + ASD(cyr))
            if curdt == self.curdt:
                break
            y = int(curdt / 100)
            m = curdt % 100
            m += 1
            if m > 12:
                y += 1
                m = m - 12
            curdt = (y * 100) + m
        if val_d[2] == 0:
            val_p[1] = 0
        else:
            val_p[1] = round((self.val[1] / val_d[2]) * 100, 2)
        if val_d[1] == 0:
            val_p[2] = 0
        else:
            val_p[2] = round((self.val[2] / val_d[1]) * 100, 2)
        if val_d[0] == 0:
            val_p[3] = 0
        else:
            val_p[3] = round((self.val[3] / val_d[0]) * 100, 2)
        if val_d[3] == 0:
            val_p[4] = 0
            val_p[5] = 0
        else:
            val_p[4] = round((self.val[4] / val_d[3]) * 100, 2)
            val_p[5] = round((self.val[5] / val_d[3]) * 100, 2)
        if val_d[4] == 0:
            val_p[6] = 0
            val_p[7] = 0
        else:
            val_p[6] = round((self.val[6] / val_d[4]) * 100, 2)
            val_p[7] = round((self.val[7] / val_d[4]) * 100, 2)
        for x in range(0, 14):
            if mth_d[x] == 0:
                mth_p[x] = 0
            else:
                mth_p[x] = round((self.mth[x] / mth_d[x]) * 100, 2)
        return (val_p, mth_p)

    def getDbal(self, curdt):
        cdt = curdt
        lyr = [0, 0, 0]
        for p in range(1, 4):
            if self.pers[p]:
                cdt -= 100
                bal = self.sql.getRec("gendtt", cols=["gdt_value"],
                    where=[("gdt_cono", "=", self.conum),
                    ("gdt_code", "=", self.dcode),
                    ("gdt_curdt", "=", cdt)], limit=1)
                if bal:
                    lyr[p-1] = bal[0]
        cyr = self.sql.getRec("gendtt", cols=["gdt_value"],
            where=[("gdt_cono", "=", self.conum), ("gdt_code", "=", self.dcode),
            ("gdt_curdt", "=", curdt)], limit=1)
        if not cyr:
            cyr = 0
        else:
            cyr = cyr[0]
        return (lyr, cyr)

    def signChkChg(self):
        if self.s_type in ("C","G","L","S","T") and self.s_norm == "N":
            for x in range(0, 8):
                self.val[x] = float(ASD(0) - ASD(self.val[x]))
            for x in range(0, 14):
                self.mth[x] = float(ASD(0) - ASD(self.mth[x]))

    def accumTotals(self, acbal):
        for x in range(1, 10):
            if acbal == "A":
                for y in range(0, 8):
                    self.val_tot[x][y] = \
                        float(ASD(self.val_tot[x][y]) + ASD(self.val[y]))
                for y in range(0, 14):
                    self.mth_tot[x][y] = \
                        float(ASD(self.mth_tot[x][y]) + ASD(self.mth[y]))
            elif acbal == "S":
                for y in range(0, 8):
                    self.val_tot[x][y] = \
                        float(ASD(self.val_tot[x][y]) - ASD(self.val[y]))
                for y in range(0, 14):
                    self.mth_tot[x][y] = \
                        float(ASD(self.mth_tot[x][y]) - ASD(self.mth[y]))

    def storeBalances(self):
        if self.s_acstr == "A":
            for x in range(0, 8):
                self.val_str[self.s_snum1][x] = \
                    float(ASD(self.val_str[self.s_snum1][x]) + ASD(self.val[x]))
            for x in range(0, 14):
                self.mth_str[self.s_snum1][x] = \
                    float(ASD(self.mth_str[self.s_snum1][x]) + ASD(self.mth[x]))
        elif self.s_acstr == "S":
            for x in range(0, 8):
                self.val_str[self.s_snum1][x] = \
                    float(ASD(self.val_str[self.s_snum1][x]) - ASD(self.val[x]))
            for x in range(0, 14):
                self.mth_str[self.s_snum1][x] = \
                    float(ASD(self.mth_str[self.s_snum1][x]) - ASD(self.mth[x]))

    def nextRepRecord(self):
        self.rpt_dic = {}
        if self.rep_num == len(self.rpt_lst) - 1:
            self.xits = True
        else:
            self.rep_num += 1
            num = 0
            for fld in self.sql.genrpt_col:
                self.rpt_dic[fld] = self.rpt_lst[self.rep_num][num]
                num += 1

    def storeRepRecord(self):
        self.s_acno = 0
        for key in self.rpt_dic:
            k = key.split("_")
            setattr(self, "s_%s" % k[1], self.rpt_dic[key])

    def checkAccountType(self, glm_type):
        if glm_type == self.atype or self.atype == "O":
            return
        self.count = 66
        self.atype = glm_type
        if self.atype == "P":
            self.stanam = "Profit and Loss"
        else:
            self.stanam = "Balance Sheet"
        if self.dname:
            self.stanam = "%s - %s" % (self.dname, self.stanam)

class GetCtl(object):
    """
    Used for extracting control records
    """
    def __init__(self, mf):
        self.mf = mf
        self.setVariables()

    def setVariables(self):
        self.errs = {
            "ctlsys": """Missing ctlsys Record, Please Run

Control -> System Record Maintenance

and Create It.""",
            "assctl": """Missing assctl Record for Company %s, Please Run

Asset's Register -> File Maintenance -> Control Record

and Create It.""",
            "bkmctl": """Missing bkmctl Record for Company %s, Please Run

Bookings Manager -> File Maintenance -> Control Record

and Create It.""",
            "bwlctl": """Missing bwlctl Record for Company %s, Please Run

Bowl's Clubs -> File Maintenance -> Control Record

and Create It.""",
            "cshctl": """Missing cshctl Record for Company %s, Please Run

Cash Analysis -> Control Record

and Create It.""",
            "crsctl": """Missing crsctl Record for Company %s, Please Run

Creditor's Ledger -> File Maintenance -> Control Record

and Create It.""",
            "ctlctl": """Missing ctlctl Record for Company %s, Please Run

General Ledger -> File Maintenance -> Control Accounts

and Create Them.""",
            "ctlmst": """Missing ctlmst Record for Company %s, Please Run

Control -> Company Records Maintenance

and Create It.""",
            "drsctl": """Missing drsctl Record for Company %s, Please Run

Debtor's Ledger -> File Maintenance -> Control Record

and Create It.""",
            "lonctl": """Missing lonctl Record for Company %s, Please Run

Loan's Ledger -> File Maintenance -> Control Record

and Create It.""",
            "memctl": """Missing memctl Record for Company %s, Please Run

Member's Ledger -> File Maintenance -> Control Record

and Create It.""",
            "rcactl": """Missing rcactl Record for Company %s, Please Run

Rental's Ledger -> File Maintenance -> Control Record

and Create It.""",
            "rtlctl": """Missing rtlctl Record for Company %s, Please Run

Rental's Ledger -> File Maintenance -> Control Record

and Create It.""",
            "slsctl": """Missing slsctl Record for Company %s, Please Run

Sales Invoicing -> File Maintenance -> Control Record

and Create It.""",
            "strctl": """Missing strctl Record for Company %s, Please Run

Stores Ledger -> File Maintenance -> Control Record

and Create It.""",
            "wagctl": """Missing wagctl Record for Company %s, Please Run

Salaries and Wages -> File Maintenance -> Control Record

and Create It."""}

    def getCtl(self, table, conum=None, error=True):
        sql = Sql(self.mf.dbm, ["ffield", table], error=error, prog=__name__)
        if sql.error:
            return None
        dics = {}
        if table == "ctlctl":
            ctl = sql.getRec(tables=[table], cols=["ctl_code", "ctl_conacc"],
                where=[("ctl_cono", "=", conum)])
            if not ctl:
                if error:
                    showError(self.mf.body, "Control Record Error",
                        self.errs[table] % conum)
                return None
            for c in ctl:
                dics["%s" % c[0]] = c[1]
            return dics
        else:
            fld = sql.getRec("ffield", cols=["ff_name"],
                where=[("ff_tabl", "=", table)])
            if not fld:
                if error:
                    showError(self.mf.body, "Control Record Error",
                        "Invalid Table %s" % table)
                return None
            col = []
            pos = None
            for f in fld:
                col.append(f[0])
                if f[0].count("_cono"):
                    pos = f[0]
            if not pos:
                recs = sql.getRec(tables=table, cols=col)
            else:
                recs = sql.getRec(tables=table, cols=col,
                    where=[(pos, "=", conum)])
            if not recs:
                if error:
                    if not conum:
                        showError(self.mf.body, "Control Record Error",
                            self.errs[table])
                    else:
                        showError(self.mf.body, "Control Record Error",
                            self.errs[table] % conum)
                return None
            for x, c in enumerate(col):
                dics["%s" % c.lower()] = recs[0][x]
            return dics

    def chkRec(self, conum, ctlctl, ctls):
        sql = Sql(self.mf.dbm, "genmst", prog=__name__)
        if type(ctls) == str:
            ctls = [ctls]
        chk = "ok"
        for key in ctls:
            try:
                if key in ctlctl:
                    acc = ctlctl[key]
                    tst = sql.getRec("genmst", cols=["count(*)"],
                        where=[("glm_cono", "=", conum), ("glm_acno", "=",
                        int(acc))], limit=1)
                    if not tst[0]:
                        chk = "no"
                else:
                    chk = "no"
            except:
                chk = "no"
            if chk == "no":
                break
        if chk == "no":
            showError(self.mf.body, "Control Record Error",
                """Missing Control Record %s for Company %s, Please Run

General Ledger -> File Maintenance -> Control Records

and Create It.""" % (key, conum))
            return key

class LoanInterest(object):
    """
    Used to calculate interest on a loan:

    sys    - The system, L for loans and S for staff loans
    dbm    - The Sql class
    lonrec - The loans record
    update - Update tables, Y or N
    batch  - Batch number
    curdt  - Current date
    tdate  - Transaction date
    refno  - Reference Number, True of False
    glctl  - G/L control accounts, [control, dr_interest, cr_interest]
    capnm  - The name of the data capturer
    """
    def __init__(self, sys, dbm, lonrec, update="N", batch="", curdt=0, tdate=0, refno=False, glctl=None, capnm=""):
        self.sys = sys
        self.dbm = dbm
        self.lonrec = lonrec
        self.update = update
        self.batch = batch
        self.curdt = curdt
        self.tdate = tdate
        if not refno:
            self.refno = "Pending"
        else:
            self.refno = None
        self.glctl = glctl
        self.capnm = capnm
        if self.setVariables():
            if self.pmths:
                self.doFixed()
            else:
                self.doVariable()

    def setVariables(self):
        tabs = ["ctlynd"]
        if self.glctl:
            tabs.append("gentrn")
        if self.sys == "L":
            tabs.extend(["lonctl", "lonmf2", "lontrn", "lonrte"])
        else:
            tabs.extend(["wagctl", "waglmf", "wagltf"])
        self.sql = Sql(self.dbm, tabs, prog=__name__)
        if self.sql.error:
            return
        if self.sys == "L":
            self.cono = self.lonrec[self.sql.lonmf2_col.index("lm2_cono")]
            self.acno = self.lonrec[self.sql.lonmf2_col.index("lm2_acno")]
            self.loan = self.lonrec[self.sql.lonmf2_col.index("lm2_loan")]
            self.start = self.lonrec[self.sql.lonmf2_col.index("lm2_start")]
            self.pmths = self.lonrec[self.sql.lonmf2_col.index("lm2_pmths")]
            self.repay = self.lonrec[self.sql.lonmf2_col.index("lm2_repay")]
            self.lcap = self.lonrec[self.sql.lonmf2_col.index("lm2_lcap")]
            self.inttp = 4
        else:
            self.cono = self.lonrec[self.sql.waglmf_col.index("wlm_cono")]
            self.acno = self.lonrec[self.sql.waglmf_col.index("wlm_empno")]
            self.loan = self.lonrec[self.sql.waglmf_col.index("wlm_loan")]
            self.start = self.lonrec[self.sql.waglmf_col.index("wlm_start")]
            self.srate = self.lonrec[self.sql.waglmf_col.index("wlm_rate")]
            self.inttp = 1
            self.pmths = 0
            self.repay = 0
            self.lcap = 0
        if not self.pmths:
            # Calculate last capitalisation date
            if self.sys == "L":
                ctl = self.sql.getRec("lonctl",
                    where=[("cln_cono", "=", self.cono)], limit=1)
                if not ctl:
                    return
                self.ityp = ctl[self.sql.lonctl_col.index("cln_ityp")]
                self.capb = ctl[self.sql.lonctl_col.index("cln_capb")]
                self.capf = ctl[self.sql.lonctl_col.index("cln_capf")]
                self.lint = ctl[self.sql.lonctl_col.index("cln_last")]
            else:
                ctl = self.sql.getRec("wagctl",
                    where=[("ctw_cono", "=", self.cono)], limit=1)
                if not ctl:
                    return
                self.ityp = "D"
                self.capb = "F"
                self.capf = "A"
                self.lint = ctl[self.sql.wagctl_col.index("ctw_i_date")]
            if not self.lint or self.lint < self.start:
                self.lint = self.start
            if not self.lcap:
                if self.capb == "A":                        # Anniversary
                    nxtd = self.start
                    tmpd = nxtd
                    while tmpd < self.lint:
                        nxtd = tmpd
                        if self.capf == "A":
                            tmpd = projectDate(tmpd, 1, typ="years")
                        else:
                            tmpd = projectDate(tmpd, 6, typ="months")
                    self.lcap = nxtd
                else:                                       # Financial
                    periods = self.sql.getRec("ctlynd",
                        cols=["cye_period", "cye_start", "cye_end"],
                        where=[("cye_cono", "=", self.cono)],
                        order="cye_period")
                    for num, per in enumerate(periods):
                        if self.tdate >= per[1] and self.tdate <= per[2]:
                            break
                    self.lcap = periods[num - 1][2]
                if self.capf == "B":                        # Bi-Annual
                    nxt = projectDate(self.lcap, 6, typ="months")
                    if self.lint >= nxt:
                        self.lcap = nxt
        t = time.localtime()
        self.sysdtw = (t[0] * 10000) + (t[1] * 100) + t[2]
        if self.refno != "Pending":
            refs = self.sql.getRec("lontrn", cols=["lnt_refno"],
                where=[("lnt_cono", "=", self.cono), ("lnt_batch", "=",
                self.batch), ("lnt_type", "=", self.inttp)])
            wrk = []
            for ref in refs:
                try:
                    wrk.append(int(ref[0]))
                except:
                    continue
            if wrk:
                wrk.sort()
                self.refno = wrk[-1] + 1
            else:
                self.refno = 1
        return True

    def doFixed(self):
        # Fixed Period Loan
        cpymth = copyList(tartanWork.mthnam)
        cap = self.sql.getRec("lontrn",
            cols=["round(sum(lnt_tramt), 2)"],
            where=[("lnt_cono", "=", self.cono),
            ("lnt_acno", "=", self.acno),
            ("lnt_loan", "=", self.loan),
            ("lnt_type", "<>", self.inttp),
            ("lnt_trdt", "<=", self.tdate)],
            limit=1)
        if cap[0] is None:
            self.cap = 0
        else:
            self.cap = cap[0]
        for mth in range(self.pmths):
            yy = int(self.start / 10000)
            mm = int(self.start / 100) % 100
            dd = self.start % 100
            mm += (mth + 1)
            while mm > 12:
                mm -= 12
                yy += 1
            if mm == 2:
                if not yy % 4:
                    cpymth[2][2] = 29
                else:
                    cpymth[2][2] = 28
            if dd > cpymth[mm][2]:
                dd = cpymth[mm][2]
            date = (yy * 10000) + (mm * 100) + dd
            date = projectDate(date, -1, typ="days")
            if date > self.tdate:
                break
            trn = self.sql.getRec("lontrn", where=[("lnt_cono",
                "=", self.cono), ("lnt_acno", "=", self.acno),
                ("lnt_loan", "=", self.loan), ("lnt_type", "=", self.inttp),
                ("lnt_trdt", "=", date)], limit=1)
            if not trn:
                rte = self.sql.getRec("lonrte", cols=["lrt_drte",
                    "lrt_crte"], where=[("lrt_cono", "=", self.cono),
                    ("lrt_acno", "=", self.acno), ("lrt_loan", "=",
                    self.loan), ("lrt_start", "<=", date)])[-1]
                if self.cap > 0:
                    rte = rte[0]
                else:
                    rte = rte[1]
                rat = (rte / 1200.0)
                ita = round(self.cap * rat, 2)
                amt = float(ASD(self.repay) - ASD(ita))
                if mth == self.pmths - 1:
                    # Final Month
                    chk = float(ASD(amt) - ASD(self.cap))
                    if chk:
                        ita = float(ASD(ita) + ASD(chk))
                if self.update == "Y":
                    cap = CCD(self.cap, "SD", 13.2)
                    desc = "%s @ %0.2f%s for 1 mth" % (cap.disp, rte, "%")
                    # Loans Transaction
                    self.sql.insRec("lontrn", data=[self.cono, self.acno,
                        self.loan, self.batch, self.inttp, date, self.refno,
                        ita, int(date / 100), desc, "", self.capnm,
                        self.sysdtw, 0])
                    if self.refno != "Pending":
                        self.refno += 1
                    if self.glctl:
                        # General Ledger
                        iti = float(ASD(0) - ASD(ita))
                        if iti < 0:
                            ctl = self.glctl[1]
                        else:
                            ctl = self.glctl[2]
                        vals = ((self.glctl[0], ita), (ctl, iti))
                        for val in vals:
                            whr = [("glt_cono", "=", self.cono),
                                ("glt_acno", "=", val[0]),
                                ("glt_curdt", "=", int(date / 100)),
                                ("glt_trdt", "=", date),
                                ("glt_type", "=", 4),
                                ("glt_refno", "=", self.refno),
                                ("glt_batch", "=", self.batch)]
                            data = self.sql.getRec("gentrn", where=whr,
                                limit=1)
                            if data:
                                col = self.sql.gentrn_col
                                tramt = data[col.index("glt_tramt")]
                                tramt = float(ASD(tramt) + ASD(val[1]))
                                self.sql.updRec("gentrn", cols=["glt_tramt"],
                                    data=[tramt], where=whr)
                            else:
                                data = [self.cono, val[0], int(date / 100),
                                    date, 4, self.refno, self.batch, val[1],
                                    0.00, "Interest on Loans", "", "", 0,
                                    self.capnm, self.sysdtw, 0]
                                self.sql.insRec("gentrn", data=data)
                self.cap = float(ASD(self.cap) - ASD(amt))
            else:
                ita = trn[self.sql.lontrn_col.index("lnt_tramt")]
                amt = float(ASD(self.repay) - ASD(ita))
                self.cap = float(ASD(self.cap) - ASD(amt))

    def doVariable(self):
        if self.lint and self.lint > self.start:
            start = projectDate(self.lint, 1)
        else:
            start = self.start
        # Extract Balances
        self.cap = 0
        self.rin = 0
        if self.sys == "L":
            trns = self.sql.getRec("lontrn", where=[("lnt_cono", "=",
                self.cono), ("lnt_acno", "=", self.acno), ("lnt_loan", "=",
                self.loan), ("lnt_trdt", "<=", self.lint)], order="lnt_trdt")
            for trn in trns:
                typ = trn[self.sql.lontrn_col.index("lnt_type")]
                dte = trn[self.sql.lontrn_col.index("lnt_trdt")]
                amt = trn[self.sql.lontrn_col.index("lnt_tramt")]
                if typ == self.inttp and dte > self.lcap:
                    self.rin = float(ASD(self.rin) + ASD(amt))
                else:
                    self.cap = float(ASD(self.cap) + ASD(amt))
        else:
            trns = self.sql.getRec("wagltf", where=[("wlt_cono", "=",
                self.cono), ("wlt_empno", "=", self.acno), ("wlt_loan", "=",
                self.loan), ("wlt_trdt", "<=", self.lint)], order="wlt_trdt")
            for trn in trns:
                typ = trn[self.sql.wagltf_col.index("wlt_type")]
                dte = trn[self.sql.wagltf_col.index("wlt_trdt")]
                amt = trn[self.sql.wagltf_col.index("wlt_amt")]
                if typ == self.inttp and dte > self.lcap:
                    self.rin = float(ASD(self.rin) + ASD(amt))
                else:
                    self.cap = float(ASD(self.cap) + ASD(amt))
        if self.ityp == "D":
            if self.sys == "L":
                trns = self.sql.getRec("lontrn", cols=["lnt_trdt",
                    "lnt_type", "lnt_tramt"], where=[("lnt_cono", "=",
                    self.cono), ("lnt_acno", "=", self.acno), ("lnt_loan",
                    "=", self.loan), ("lnt_trdt", ">", self.lint)],
                    order="lnt_trdt")
            else:
                trns = self.sql.getRec("wagltf", cols=["wlt_trdt",
                    "wlt_type", "wlt_amt"], where=[("wlt_cono", "=",
                    self.cono), ("wlt_empno", "=", self.acno), ("wlt_loan",
                    "=", self.loan), ("wlt_trdt", ">", self.lint)],
                    order="wlt_trdt")
        else:
            bals = {}
            if self.sys == "L":
                trns = self.sql.getRec("lontrn", cols=["lnt_curdt",
                    "lnt_type", "sum(lnt_tramt)"], where=[("lnt_cono", "=",
                    self.cono), ("lnt_acno", "=", self.acno), ("lnt_loan",
                    "=", self.loan), ("lnt_trdt", ">", self.lint)],
                    group="lnt_curdt, lnt_type", order="lnt_curdt")
            else:
                trns = self.sql.getRec("wagltf", cols=["wlt_curdt",
                    "wlt_type", "sum(wlt_amt)"], where=[("wlt_cono", "=",
                    self.cono), ("wlt_acno", "=", self.acno), ("wlt_loan",
                    "=", self.loan), ("wlt_trdt", ">", self.lint)],
                    group="wlt_curdt, wlt_type", order="wlt_curdt")
            for trn in trns:
                if trn[0] not in bals:
                    bals[trn[0]] = [0, 0]
                if trn[1] == self.inttp:
                    bals[trn[0]][1] = trn[2]
                else:
                    bals[trn[0]][0] = trn[2]
        if self.sys == "L":
            # Extract Rates
            rates = self.sql.getRec("lonrte", cols=["lrt_start",
                "lrt_drte", "lrt_crte"], where=[("lrt_cono", "=", self.cono),
                ("lrt_acno", "=", self.acno), ("lrt_loan", "=", self.loan),
                ("lrt_start", "<=", self.tdate)], order="lrt_start")
        else:
            rates = [[self.start, self.srate, 0]]
        # Extract Capitalisation Dates
        capdt = []
        lcap = self.lcap
        while lcap < self.tdate:
            if self.capf == "A":
                lcap = projectDate(lcap, 1, typ="years")
            else:
                lcap = projectDate(lcap, 6, typ="months")
            if lcap <= self.tdate:
                capdt.append((lcap, True))
            else:
                capdt.append((self.tdate, False))
        # Raise Interest
        for self.ncap in capdt:
            if self.ncap[0] <= start:
                continue
            if self.ityp == "D":
                # Daily
                while start < self.ncap[0]:
                    usetrn = False
                    for trn in trns:
                        if trn[0] <= start or trn[0] > self.ncap[0]:
                            continue
                        trns.remove(trn)
                        endd = trn[0]
                        usetrn = True
                        break
                    if not usetrn:
                        endd = self.ncap[0]
                    lr = rates[0]
                    for rate in rates:
                        if rate[0] > start and rate[0] < endd:
                            self.raiseInterest(start, rate[0], lr[1], lr[2])
                            start = projectDate(rate[0], 1, typ="days")
                        lr = rate
                    for x in range(len(rates) - 1, -1, -1):
                        rate = rates[x]
                        if rate[0] <= endd:
                            break
                    self.raiseInterest(start, endd, rate[1], rate[2])
                    if usetrn:
                        if trn[1] == self.inttp:
                            self.rin = float(ASD(self.rin) + ASD(trn[2]))
                        else:
                            self.cap = float(ASD(self.cap) + ASD(trn[2]))
                    start = projectDate(endd, 1, typ="days")
            else:
                # Monthly
                while start < self.ncap[0]:
                    endd = mthendDate(start)
                    for x in range(len(rates) - 1, -1, -1):
                        rate = rates[x]
                        if rate[0] <= endd:
                            break
                    curdt = int(endd / 100)
                    if curdt in bals:
                        self.cap = float(ASD(self.cap) + ASD(bals[curdt][0]))
                        self.rin = float(ASD(self.rin) + ASD(bals[curdt][1]))
                    self.raiseInterest(start, endd, rate[1], rate[2])
                    start = projectDate(endd, 1, typ="days")
        if self.sys == "L" and self.update == "Y":
            self.sql.updRec("lonmf2", cols=["lm2_lcap"], data=[self.lcap],
                where=[("lm2_cono", "=", self.cono), ("lm2_acno", "=",
                self.acno), ("lm2_loan", "=", self.loan)])

    def raiseInterest(self, fdte, tdte, drte, crte):
        """
        fdte   - The from date
        tdte   - The to date
        drte   - The debit interest rate
        crte   - The credit interest rate
        """
        if self.cap > 0:
            rte = drte
            if self.glctl:
                ctl = self.glctl[1]
        else:
            rte = crte
            if self.glctl:
                ctl = self.glctl[2]
        if self.ityp == "D":
            per = dateDiff(fdte, tdte, "days") + 1
            rin = round((self.cap * rte * per / 36500), 2)
            cap = CCD(self.cap, "SD", 13.2)
            desc = "%s @ %0.2f%s for %s day" % (cap.disp, rte, "%", per)
            if per > 1:
                desc += "s"
        else:
            rin = round((self.cap * rte / 1200), 2)
            cap = CCD(self.cap, "SD", 13.2)
            desc = "%s @ %0.2f%s for 1 mth" % (cap.disp, rte, "%")
        self.rin = float(ASD(self.rin) + ASD(rin))
        # Capitalisation
        if tdte == self.ncap[0] and self.ncap[1]:
            self.cap = float(ASD(self.cap) + ASD(self.rin))
            self.lcap = self.ncap[0]
            self.rin = 0
        # Update Tables
        if rin and self.update == "Y":
            # Loans Transaction
            if self.sys == "L":
                data = [self.cono, self.acno, self.loan, self.batch,
                    self.inttp, tdte, self.refno, rin, int(tdte / 100),
                    desc, "", self.capnm, self.sysdtw, 0]
                self.sql.insRec("lontrn", data=data)
            else:
                data = [self.cono, self.acno, self.loan, self.batch,
                    self.inttp, tdte, self.refno, rin, cap.work, 0,
                    rte, int(tdte / 100), desc, "", self.capnm,
                    self.sysdtw, 0]
                self.sql.insRec("wagltf", data=data)
            if self.refno != "Pending":
                self.refno += 1
            if self.glctl:
                # General Ledger
                iti = float(ASD(0) - ASD(rin))
                vals = ((self.glctl[0], rin), (ctl, iti))
                for val in vals:
                    whr = [("glt_cono", "=", self.cono), ("glt_acno",
                        "=", val[0]), ("glt_curdt", "=", int(tdte / 100)),
                        ("glt_trdt", "=", tdte), ("glt_type", "=", 4),
                        ("glt_refno", "=", self.refno), ("glt_batch",
                        "=", self.batch)]
                    data = self.sql.getRec("gentrn", where=whr, limit=1)
                    if data:
                        tramt = data[self.sql.gentrn_col.index("glt_tramt")]
                        tramt = float(ASD(tramt) + ASD(val[1]))
                        self.sql.updRec("gentrn", cols=["glt_tramt"],
                            data=[tramt], where=whr)
                    else:
                        self.sql.insRec("gentrn", data=[self.cono, val[0],
                            int(tdte / 100), tdte, 4, self.refno, self.batch,
                            val[1], 0.00, "Interest on Loans", "", "", 0,
                            self.capnm, self.sysdtw, 0])

class PrintOrder(object):
    """
    Used to print stores orders using
    the ctlmst, crsmst, strpom and strpot tables
    """
    def __init__(self, mf, conum, conam, docs, **args):
        self.mf = mf
        self.conum = conum
        self.conam = conam
        if type(docs) in (int, str):
            self.docs = [[docs]]
        else:
            self.docs = docs
        defaults = {
            "tname": "purchase_order",
            "repprt": ["N", "V", "view"],
            "repeml": ["N", "N", "", "", "Y"],
            "copy": "n",
            "splash": True}
        for nam in args:
            if nam in ("repprt", "repeml"):
                defaults[nam] = copyList(args[nam])
            else:
                defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        if self.setVariables():
            self.doProcess()

    def setVariables(self):
        self.sql = Sql(self.mf.dbm, ["ctlmst", "ctlmes", "crsmst", "strctl",
            "strpom", "strpot"], prog=__name__)
        if self.sql.error:
            return
        gc = GetCtl(self.mf)
        strctl = gc.getCtl("strctl", self.conum)
        if not strctl:
            return
        self.fromad = strctl["cts_emadd"]
        if self.repeml:
            self.emadd = self.repeml[2]
        else:
            self.emadd = ""
        return True

    def doProcess(self):
        pmc = self.sql.strpom_col
        ptc = self.sql.strpot_col
        cmc = self.sql.crsmst_col
        self.form = DrawForm(self.mf.dbm, self.tname)
        txt = self.form.sql.tpldet_col.index("tpd_text")
        tdc = self.form.sql.tpldet_col
        self.doLoadStatic()
        self.form.doNewDetail()
        for doc in self.docs:
            self.docno = CCD(doc[0], "UI", 9)
            # strpom
            pom = self.sql.getRec("strpom", where=[("pom_cono", "=",
                self.conum), ("pom_ordno", "=", self.docno.work)], limit=1)
            if not pom:
                continue
            for fld in pmc:
                if fld in self.form.tptp:
                    self.form.tptp[fld][1] = pom[pmc.index(fld)]
            if "deliver_to" in self.form.tptp:
                if pom[pmc.index("pom_add1")]:
                    dat = pom[pmc.index("pom_add1")]
                    dat = "%s\n%s" % (dat, pom[pmc.index("pom_add2")])
                    dat = "%s\n%s" % (dat, pom[pmc.index("pom_add3")])
                    dat = "%s\n%s\n " % (dat, pom[pmc.index("pom_add4")])
                    self.deliver = dat
                else:
                    self.deliver = False
            if "message" in self.form.tptp:
                if pom[pmc.index("pom_mess")]:
                    self.message = True
                    self.form.tptp["message"][1] = pom[pmc.index("pom_mess")]
                else:
                    self.message = False
            # crsmst
            crm = self.sql.getRec("crsmst", where=[("crm_cono", "=",
                self.conum), ("crm_acno", "=", pom[pmc.index("pom_acno")])],
                limit=1)
            # If no email address supplied, use orders address else accounts
            eml = crm[cmc.index("crm_ord_email")]
            if not eml:
                eml = crm[cmc.index("crm_acc_email")]
            for fld in cmc:
                if fld in self.form.tptp:
                    d = "%s_C00" % fld
                    self.form.newdic[d][txt] = crm[cmc.index(fld)]
            self.form.account_details("crm", cmc, crm, 0)
            self.form.document_date(pom[pmc.index("pom_date")])
            # strpot
            pot = self.sql.getRec("strpot", where=[("pot_cono", "=",
                self.conum), ("pot_ordno", "=", self.docno.work)],
                order="pot_group, pot_code")
            if not pot:
                continue
            if self.splash:
                sp = SplashScreen(self.mf.body,
                    "Generating Forms\n\n   Please Wait...")
            self.doBody(ptc, pot, tdc)
            self.doTotal(tdc)
            self.doTail(tdc)
            if self.splash:
                sp.closeSplash()
            if self.repeml[1] == "Y" and not self.emadd:
                self.repeml[2] = eml
                self.doPrint()
        if self.repeml[1] == "N" or self.emadd:
            self.repeml[2] = self.emadd
            self.doPrint()

    def doBody(self, ptc, pot, tdc):
        page = 0
        count = 0
        self.total_taxable = 0
        self.total_nontaxable = 0
        self.total_tax = 0
        self.total_value = 0
        for item in pot:
            if not count:
                page += 1
                count = self.doHeader(str(page), tdc)
            if count == self.form.maxlines:
                page = self.doCfwd(page)
                count = self.doHeader(page, tdc)
            ldic = {}
            for cod in self.form.body:
                if cod == "line_value":
                    continue
                if cod == "pot_desc":
                    des = self.form.doSplitText("pot_desc_C00",
                        item[ptc.index(cod)])
                    if not des[-1]:
                        del des[-1]
                else:
                    ldic[cod] = CCD(item[ptc.index(cod)],
                        self.form.tptp[cod][0][1],
                        self.form.tptp[cod][0][2])
            excpri = round((ldic["pot_price"].work * 1), 2)
            incrat = float(ASD(100.0) + ASD(ldic["pot_vatrat"].work))
            incpri = round((ldic["pot_price"].work * incrat / 100.0), 4)
            net = float(ASD(100.0) - ASD(ldic["pot_disper"].work))
            excamt = round((ldic["pot_qty"].work * excpri * net / 100.0), 2)
            incamt = round((ldic["pot_qty"].work * incpri * net / 100.0), 2)
            vatamt = float(ASD(incamt) - ASD(excamt))
            if excamt == incamt:
                self.total_nontaxable = float(ASD(self.total_nontaxable) + \
                    ASD(excamt))
            else:
                self.total_taxable = float(ASD(self.total_taxable) + \
                    ASD(excamt))
            self.total_tax = float(ASD(self.total_tax) + ASD(vatamt))
            self.total_value = float(ASD(self.total_value) + ASD(incamt))
            ldic["line_value"] = CCD(incamt, "SD", 11.2)
            for n, l in enumerate(des):
                if count == self.form.maxlines:
                    page = self.doCfwd(page)
                    count = self.doHeader(page, tdc)
                if n == 0 and len(des) == 1:
                    incl = copyList(self.form.body)
                elif n == 0:
                    incl = ("pot_group", "pot_code")
                elif n + 1 == len(des):
                    incl = copyList(self.form.body)
                    incl.remove("pot_group")
                    incl.remove("pot_code")
                else:
                    incl = []
                for code in self.form.body:
                    seq = "%s_C%02i" % (code, count)
                    if code == "pot_desc":
                        data = l
                    elif code in incl:
                        data = ldic[code].work
                    else:
                        data = "BLANK"
                    self.form.newdic[seq][tdc.index("tpd_text")] = data
                    self.form.doDrawDetail(self.form.newdic[seq])
                count += 1
        for x in range(count, self.form.maxlines):
            for cod in self.form.body:
                seq = "%s_C%02i" % (cod, x)
                self.form.newdic[seq][tdc.index("tpd_text")] = "BLANK"
                self.form.doDrawDetail(self.form.newdic[seq])

    def doTotal(self, tdc):
        for c in self.form.total:
            line = None
            if c in self.form.newdic:
                line = self.form.newdic[c]
            else:
                t = "%s_T00" % c
                if t in self.form.newdic:
                    line = self.form.newdic[t]
            if line:
                self.form.doDrawDetail(line)
            d = "%s_C00" % c
            if d in self.form.newkey:
                line = self.form.newdic[d]
                mrgcod = line[tdc.index("tpd_mrgcod")]
                line[tdc.index("tpd_text")] = getattr(self, "%s" % mrgcod)
                self.form.doDrawDetail(line)

    def doTail(self, tdc):
        for c in self.form.tail:
            line = None
            if c in self.form.newdic:
                line = self.form.newdic[c]
            else:
                t = "%s_T00" % c
                if t in self.form.newdic:
                    line = self.form.newdic[t]
            if line:
                if line[tdc.index("tpd_mrgcod")] == "message":
                    if not self.message:
                        continue
                self.form.doDrawDetail(line)
            d = "%s_C00" % c
            if d in self.form.newdic:
                line = self.form.newdic[d]
                mrgcod = line[tdc.index("tpd_mrgcod")]
                line[tdc.index("tpd_text")] = self.doGetData(mrgcod)
                self.form.doDrawDetail(line, fmat=False)

    def doPrint(self):
        pfx = "PurOrd"
        head = "%s Purchase Order" % self.conam.upper()
        if self.repeml[1] == "Y" and not self.emadd:
            head = "%s %s" % (head, self.docno.work)
            pdfnam = getModName(self.mf.rcdic["wrkdir"], pfx,
                "%s_%s" % (self.conum, self.docno.work), ext="pdf")
        else:
            pdfnam = getModName(self.mf.rcdic["wrkdir"], pfx,
                "%s_all" % self.conum, ext="pdf")
            if len(self.docs) == 1:
                head = "%s %s" % (head, self.docno.work)
            else:
                head += "S"
                for doc in self.docs:
                    head += " %s" % doc[0]
        self.form.output(pdfnam, "F")
        doPrinter(mf=self.mf, conum=self.conum, pdfnam=pdfnam, header=head,
            fromad=self.fromad, repprt=self.repprt, repeml=self.repeml)
        if self.repeml[1] == "Y" and not self.emadd:
            self.form = DrawForm(self.mf.dbm, self.tname)
            self.doLoadStatic()
            self.form.doNewDetail()

    def doLoadStatic(self):
        # ctlmst
        cmc = self.sql.ctlmst_col
        ctm = self.sql.getRec("ctlmst", where=[("ctm_cono", "=",
            self.conum)], limit=1)
        for fld in cmc:
            dat = ctm[cmc.index(fld)]
            if fld in self.form.tptp:
                if fld == "ctm_logo":
                    self.form.letterhead(cmc, ctm, fld, dat)
                    continue
                self.form.tptp[fld][1] = dat
        if "letterhead" in self.form.tptp:
            self.form.letterhead(cmc, ctm, "letterhead", None)
        if "document_type" in self.form.tptp:
            if self.copy.lower() == "c":
                typ = "CANCELLED"
            elif self.copy.lower() == "a":
                typ = "AMENDED"
            elif self.copy.lower() == "y":
                typ = "COPY"
            else:
                typ = ""
            if ctm[cmc.index("ctm_taxno")]:
                typ = "%s TAX" % typ
            self.form.tptp["document_type"][1] = "%s PURCHASE ORDER" % typ

    def doHeader(self, page, tdc):
        self.form.add_page()
        for key in self.form.newkey:
            line = self.form.newdic[key]
            if line[tdc.index("tpd_place")] != "A":
                continue
            nl = copyList(line)
            if nl[tdc.index("tpd_detseq")] == "page_number_C00":
                nl[tdc.index("tpd_text")] = str(page)
            elif nl[tdc.index("tpd_type")] == "T" and \
                        nl[tdc.index("tpd_ttyp")] == "H":
                mrgcod = nl[tdc.index("tpd_mrgcod")]
                if mrgcod and mrgcod == "deliver_to" and not self.deliver:
                    continue
                if mrgcod and self.form.tptp[mrgcod][0][1][0] == "S":
                    txt = nl[tdc.index("tpd_text")]
                    nl[tdc.index("tpd_text")] = "%s " % txt
            elif nl[tdc.index("tpd_type")] == "C":
                mrgcod = nl[tdc.index("tpd_mrgcod")]
                if mrgcod == "deliver_to":
                    if not self.deliver:
                        continue
                    nl[tdc.index("tpd_text")] = self.deliver
                else:
                    nl[tdc.index("tpd_text")] = self.doGetData(mrgcod)
            if key == "document_type_C00":
                self.form.doDrawDetail(nl, fmat=False)
            else:
                self.form.doDrawDetail(nl)
        return 0

    def doCfwd(self, page):
        line = copyList(self.form.cfwd)
        line[5] = "Continued on Page %i" % (page + 1)
        self.form.doDrawDetail(line)
        return page + 1

    def doGetData(self, mrgcod):
        if mrgcod and mrgcod in self.form.tptp and self.form.tptp[mrgcod][1]:
            return self.form.tptp[mrgcod][1]
        return ""

class PrintCharges(object):
    def __init__(self, mf, conum, conam, docs, **args):
        self.mf = mf
        self.conum = conum
        self.conam = conam
        self.docs = docs
        defaults = {
            "tname": "recurring_charges",
            "repprt": ["N", "V", "view"],
            "repeml": ["N", "N", "", "", "Y"],
            "copy": "n",
            "splash": True}
        for nam in args:
            if nam in ("repprt", "repeml"):
                defaults[nam] = copyList(args[nam])
            else:
                defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        if self.setVariables():
            self.doProcess()

    def setVariables(self):
        self.sql = Sql(self.mf.dbm, ["ctlmst", "drsmst", "drsrci"],
            prog=__name__)
        if self.sql.error:
            return False
        gc = GetCtl(self.mf)
        drsctl = gc.getCtl("drsctl", self.conum)
        if not drsctl:
            return
        self.fromad = drsctl["ctd_emadd"]
        if self.repeml:
            self.emadd = self.repeml[2]
        else:
            self.emadd = ""
        return True

    def doProcess(self):
        rcc = self.sql.drsrci_col
        dmc = self.sql.drsmst_col
        self.form = DrawForm(self.mf.dbm, self.tname)
        txt = self.form.sql.tpldet_col.index("tpd_text")
        tdc = self.form.sql.tpldet_col
        self.doLoadStatic()
        self.form.doNewDetail()
        printed = False
        for doc in self.docs:
            self.docno = CCD(doc, "NA", 9)
            # drsrci
            rci = self.sql.getRec("drsrci", where=[("dci_cono", "=",
                self.conum), ("dci_docno", "=", self.docno.work)], limit=1)
            if not rci:
                continue
            if self.splash:
                sp = SplashScreen(self.mf.body,
                    "Generating Forms\n\n   Please Wait...")
            for fld in rcc:
                if fld in self.form.tptp:
                    self.form.tptp[fld][1] = rci[rcc.index(fld)]
            # drsmst
            drm = self.sql.getRec("drsmst", where=[("drm_cono", "=",
                self.conum), ("drm_chain", "=", rci[rcc.index("dci_chain")]),
                ("drm_acno", "=", rci[rcc.index("dci_acno")])], limit=1)
            # Use accounts name else manager
            eml = drm[dmc.index("drm_acc_email")]
            if not eml:
                eml = drm[dmc.index("drm_mgr_email")]
            for fld in dmc:
                if fld in self.form.tptp:
                    d = "%s_C00" % fld
                    self.form.newdic[d][txt] = drm[dmc.index(fld)]
            self.form.account_details("drm", dmc, drm, 0)
            self.form.document_date(rci[rcc.index("dci_date")])
            self.doBody(dmc, drm, rcc, rci, tdc)
            self.doTotal(tdc)
            self.doTail(tdc)
            if self.splash:
                sp.closeSplash()
            if self.repeml[1] == "Y" and not self.emadd:
                self.repeml[2] = eml
                self.doPrint()
            printed = True
        if printed and (self.repeml[1] == "N" or self.emadd):
            self.repeml[2] = self.emadd
            self.doPrint()

    def doBody(self, dmc, drm, rcc, rci, tdc):
        page = 0
        count = 0
        self.total_taxable = 0
        self.total_nontaxable = 0
        self.total_tax = 0
        self.total_value = 0
        ldic = {}
        for cod in self.form.body:
            if cod == "line_value":
                continue
            if cod == "dci_desc":
                des = self.form.doSplitText("dci_desc_C00",
                    rci[rcc.index(cod)])
                if not des[-1]:
                    del des[-1]
            elif cod in dmc:
                ldic[cod] = CCD(drm[dmc.index(cod)],
                    self.form.tptp[cod][0][1],
                    self.form.tptp[cod][0][2])
            else:
                ldic[cod] = CCD(rci[rcc.index(cod)],
                    self.form.tptp[cod][0][1],
                    self.form.tptp[cod][0][2])
        excamt = ldic["dci_charge"].work
        incrat = float(ASD(100.0) + ASD(ldic["dci_vat_rate"].work))
        incamt = round((excamt * incrat / 100.0), 2)
        vatamt = float(ASD(incamt) - ASD(excamt))
        if excamt == incamt:
            self.total_nontaxable = float(ASD(self.total_nontaxable) + \
                ASD(excamt))
        else:
            self.total_taxable = float(ASD(self.total_taxable) + \
                ASD(excamt))
        self.total_tax = float(ASD(self.total_tax) + ASD(vatamt))
        self.total_value = float(ASD(self.total_value) + ASD(incamt))
        ldic["line_value"] = CCD(incamt, "SD", 11.2)
        for n, l in enumerate(des):
            #if count == self.form.maxlines:
            #    page = self.doCfwd(page)
            count = self.doHeader(page, tdc)
            if n == 0 and len(des) == 1:
                incl = copyList(self.form.body)
            elif n + 1 == len(des):
                incl = copyList(self.form.body)
            else:
                incl = []
            for code in self.form.body:
                seq = "%s_C%02i" % (code, count)
                if code == "dci_desc":
                    data = l
                elif code in incl:
                    data = ldic[code].work
                else:
                    data = "BLANK"
                self.form.newdic[seq][tdc.index("tpd_text")] = data
                self.form.doDrawDetail(self.form.newdic[seq])
            count += 1
        for x in range(count, self.form.maxlines):
            for cod in self.form.body:
                seq = "%s_C%02i" % (cod, x)
                self.form.newdic[seq][tdc.index("tpd_text")] = "BLANK"
                self.form.doDrawDetail(self.form.newdic[seq])

    def doTotal(self, tdc):
        for c in self.form.total:
            line = None
            if c in self.form.newdic:
                line = self.form.newdic[c]
            else:
                t = "%s_T00" % c
                if t in self.form.newdic:
                    line = self.form.newdic[t]
            if line:
                self.form.doDrawDetail(line)
            d = "%s_C00" % c
            if d in self.form.newkey:
                line = self.form.newdic[d]
                mrgcod = line[tdc.index("tpd_mrgcod")]
                line[tdc.index("tpd_text")] = getattr(self, "%s" % mrgcod)
                self.form.doDrawDetail(line)

    def doTail(self, tdc):
        for c in self.form.tail:
            line = None
            if c in self.form.newdic:
                line = self.form.newdic[c]
            else:
                t = "%s_T00" % c
                if t in self.form.newdic:
                    line = self.form.newdic[t]
            if line:
                if line[tdc.index("tpd_mrgcod")] == "message":
                    if not self.message:
                        continue
                self.form.doDrawDetail(line)
            d = "%s_C00" % c
            if d in self.form.newdic:
                line = self.form.newdic[d]
                mrgcod = line[tdc.index("tpd_mrgcod")]
                line[tdc.index("tpd_text")] = self.doGetData(mrgcod)
                self.form.doDrawDetail(line, fmat=False)

    def doPrint(self):
        pfx = "Inv"
        head = "%s Invoice" % self.conam.upper()
        if self.repeml[1] == "Y" and not self.emadd:
            head = "%s %s" % (head, self.docno.work)
            pdfnam = getModName(self.mf.rcdic["wrkdir"], pfx,
                "%s_%s" % (self.conum, self.docno.work), ext="pdf")
        else:
            if len(self.docs) == 1:
                head = "%s %s" % (head, self.docno.work)
            else:
                head += "S"
                for doc in self.docs:
                    head += " %s" % doc[0]
            pdfnam = getModName(self.mf.rcdic["wrkdir"], pfx,
                "%s_all" % self.conum, ext="pdf")
        self.form.output(pdfnam, "F")
        doPrinter(mf=self.mf, conum=self.conum, pdfnam=pdfnam, header=head,
            fromad=self.fromad, repprt=self.repprt, repeml=self.repeml)
        if self.repeml[1] == "Y" and not self.emadd:
            self.form = DrawForm(self.mf.dbm, self.tname)
            self.doLoadStatic()
            self.form.doNewDetail()

    def doLoadStatic(self):
        # ctlmst
        cmc = self.sql.ctlmst_col
        ctm = self.sql.getRec("ctlmst", where=[("ctm_cono", "=",
            self.conum)], limit=1)
        for fld in cmc:
            dat = ctm[cmc.index(fld)]
            if fld in self.form.tptp:
                if fld == "ctm_logo":
                    self.form.letterhead(cmc, ctm, fld, dat)
                    continue
                self.form.tptp[fld][1] = dat
        if "letterhead" in self.form.tptp:
            self.form.letterhead(cmc, ctm, "letterhead", None)
        if "document_type" in self.form.tptp:
            if self.copy.lower() == "a":
                typ = "AMENDED"
            elif self.copy.lower() == "y":
                typ = "COPY"
            else:
                typ = ""
            if ctm[cmc.index("ctm_taxno")]:
                typ = "%s TAX" % typ
            self.form.tptp["document_type"][1] = "%s INVOICE" % typ
        self.form.bank_details(cmc, ctm, 0)
        if "terms" in self.form.tptp:
            acc = self.sql.getRec("ctlmes", cols=["mss_detail"],
                where=[("mss_system", "=", "CON"), ("mss_message", "=", 1)],
                limit=1)
            if acc:
                self.form.tptp["terms"][1] = acc[0]
            else:
                del self.form.tptp["terms"]

    def doHeader(self, page, tdc):
        self.form.add_page()
        for key in self.form.newkey:
            line = self.form.newdic[key]
            if line[tdc.index("tpd_place")] != "A":
                continue
            nl = copyList(line)
            if nl[tdc.index("tpd_detseq")] == "page_number_C00":
                nl[tdc.index("tpd_text")] = str(page)
            elif nl[tdc.index("tpd_type")] == "T" and \
                        nl[tdc.index("tpd_ttyp")] == "H":
                mrgcod = nl[tdc.index("tpd_mrgcod")]
                if mrgcod and self.form.tptp[mrgcod][0][1][0] == "S":
                    txt = nl[tdc.index("tpd_text")]
                    nl[tdc.index("tpd_text")] = "%s " % txt
            elif nl[tdc.index("tpd_type")] == "C" and not nl[5]:
                mrgcod = nl[tdc.index("tpd_mrgcod")]
                nl[tdc.index("tpd_text")] = self.doGetData(mrgcod)
            if key == "document_type_C00":
                self.form.doDrawDetail(nl, fmat=False)
            else:
                self.form.doDrawDetail(nl)
        return 0

    def doCfwd(self, page):
        line = copyList(self.form.cfwd)
        line[5] = "Continued on Page %i" % (page + 1)
        self.form.doDrawDetail(line)
        return page + 1

    def doGetData(self, mrgcod):
        if mrgcod and mrgcod in self.form.tptp and self.form.tptp[mrgcod][1]:
            return self.form.tptp[mrgcod][1]
        return ""

class PrintInvoice(object):
    def __init__(self, mf, conum, conam, dtype, docs, **args):
        self.mf = mf
        self.conum = conum
        self.conam = conam
        self.dtype = dtype
        if type(docs) in (int, str):
            self.docs = [[docs]]
        else:
            self.docs = docs
        defaults = {
            "tname": "sales_document",
            "repprt": ["N", "V", "view"],
            "repeml": ["N", "N", "", "", "Y"],
            "dnote": "n",
            "dvals": "n",
            "copy": "n",
            "splash": True}
        for nam in args:
            if nam in ("repprt", "repeml"):
                defaults[nam] = copyList(args[nam])
            else:
                defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        if self.setVariables():
            self.doProcess()

    def setVariables(self):
        self.sql = Sql(self.mf.dbm, ["ctlmst", "ctlmes", "drsmst", "slsiv1",
            "slsiv2"], prog=__name__)
        if self.sql.error:
            return False
        gc = GetCtl(self.mf)
        slsctl = gc.getCtl("slsctl", self.conum)
        if not slsctl:
            return
        self.fromad = slsctl["ctv_emadd"]
        if self.repeml:
            self.emadd = self.repeml[2]
        else:
            self.emadd = ""
        return True

    def doProcess(self):
        s1c = self.sql.slsiv1_col
        s2c = self.sql.slsiv2_col
        dmc = self.sql.drsmst_col
        self.form = DrawForm(self.mf.dbm, self.tname)
        txt = self.form.sql.tpldet_col.index("tpd_text")
        tdc = self.form.sql.tpldet_col
        self.doLoadStatic()
        self.form.doNewDetail()
        printed = False
        for doc in self.docs:
            self.docno = CCD(doc[0], "UI", 9.0)
            # slsiv1
            si1 = self.sql.getRec("slsiv1", where=[("si1_cono", "=",
                self.conum), ("si1_rtn", "=", self.dtype), ("si1_docno", "=",
                self.docno.work)], limit=1)
            si2 = self.sql.getRec("slsiv2", where=[("si2_cono", "=",
                self.conum), ("si2_rtn", "=", self.dtype), ("si2_docno", "=",
                self.docno.work)], order="si2_line")
            if not si1 or not si2:
                continue
            if self.splash:
                sp = SplashScreen(self.mf.body,
                    "Generating Forms\n\n   Please Wait...")
            for fld in s1c:
                if fld in self.form.tptp:
                    self.form.tptp[fld][1] = si1[s1c.index(fld)]
            if "deliver_to" in self.form.tptp:
                if si1[s1c.index("si1_add1")]:
                    dat = si1[s1c.index("si1_add1")]
                    dat = "%s\n%s" % (dat, si1[s1c.index("si1_add2")])
                    dat = "%s\n%s" % (dat, si1[s1c.index("si1_add3")])
                    dat = "%s\n%s\n " % (dat, si1[s1c.index("si1_add4")])
                    self.deliver = dat
                else:
                    self.deliver = False
            if "message" in self.form.tptp:
                if si1[s1c.index("si1_mess")]:
                    self.message = True
                    self.form.tptp["message"][1] = si1[s1c.index("si1_mess")]
                else:
                    self.message = False
            # drsmst
            drm = self.sql.getRec("drsmst", where=[("drm_cono", "=",
                self.conum), ("drm_chain", "=", si1[s1c.index("si1_chain")]),
                ("drm_acno", "=", si1[s1c.index("si1_acno")])], limit=1)
            if self.dtype == "Q":
                # If sales address else accounts
                eml = drm[dmc.index("drm_sls_email")]
                if not eml:
                    eml = drm[dmc.index("drm_acc_email")]
            else:
                # If use accounts address else sales
                eml = drm[dmc.index("drm_acc_email")]
                if not eml:
                    eml = drm[dmc.index("drm_sls_email")]
            for fld in dmc:
                if fld in self.form.tptp:
                    d = "%s_C00" % fld
                    self.form.newdic[d][txt] = drm[dmc.index(fld)]
            self.form.account_details("drm", dmc, drm, 0)
            self.form.document_date(si1[s1c.index("si1_date")])
            self.ttype = self.dtype
            self.doBody(s2c, si2, tdc)
            self.doTotal(tdc)
            self.doTail(tdc)
            if self.dnote.lower() == "y":
                self.ttype = "D"
                if "document_type" in self.form.tptp:
                    dat = self.form.tptp["document_type"][1]
                    self.form.tptp["document_type"][1] = dat.replace("INVOICE",
                        "DELIVERY NOTE")
                self.doBody(s2c, si2, tdc)
                self.doTotal(tdc)
                self.doTail(tdc)
                if "document_type" in self.form.tptp:
                    self.form.tptp["document_type"][1] = dat
            if self.splash:
                sp.closeSplash()
            if self.repeml[1] == "Y" and not self.emadd:
                self.repeml[2] = eml
                self.doPrint()
            printed = True
        if printed and (self.repeml[1] == "N" or self.emadd):
            self.repeml[2] = self.emadd
            self.doPrint()

    def doBody(self, s2c, si2, tdc):
        page = 0
        count = 0
        self.total_taxable = 0
        self.total_nontaxable = 0
        self.total_tax = 0
        self.total_value = 0
        for item in si2:
            if not count:
                page += 1
                count = self.doHeader(str(page), tdc)
            if count == self.form.maxlines:
                page = self.doCfwd(page)
                count = self.doHeader(page, tdc)
            ldic = {}
            for cod in self.form.body:
                if cod == "line_value":
                    continue
                if cod == "si2_desc":
                    des = self.form.doSplitText("si2_desc_C00",
                        item[s2c.index(cod)])
                    if not des[-1]:
                        del des[-1]
                else:
                    ldic[cod] = CCD(item[s2c.index(cod)],
                        self.form.tptp[cod][0][1],
                        self.form.tptp[cod][0][2])
            excpri = round((ldic["si2_price"].work * 1), 2)
            incrat = float(ASD(100.0) + ASD(ldic["si2_vat_rate"].work))
            incpri = round((ldic["si2_price"].work * incrat / 100.0), 4)
            net = float(ASD(100.0) - ASD(ldic["si2_disc_per"].work))
            excamt = round((ldic["si2_qty"].work * excpri * net / 100.0), 2)
            incamt = round((ldic["si2_qty"].work * incpri * net / 100.0), 2)
            vatamt = float(ASD(incamt) - ASD(excamt))
            if excamt == incamt:
                self.total_nontaxable = float(ASD(self.total_nontaxable) + \
                    ASD(excamt))
            else:
                self.total_taxable = float(ASD(self.total_taxable) + \
                    ASD(excamt))
            self.total_tax = float(ASD(self.total_tax) + ASD(vatamt))
            self.total_value = float(ASD(self.total_value) + ASD(incamt))
            ldic["line_value"] = CCD(incamt, "SD", 11.2)
            for n, l in enumerate(des):
                if count == self.form.maxlines:
                    page = self.doCfwd(page)
                    count = self.doHeader(page, tdc)
                if n == 0 and len(des) == 1:
                    incl = copyList(self.form.body)
                    if self.ttype == "D" and self.dvals.lower() == "n":
                        incl.remove("si2_price")
                        incl.remove("line_value")
                elif n == 0:
                    incl = ("si2_group", "si2_code")
                elif n + 1 == len(des):
                    incl = copyList(self.form.body)
                    incl.remove("si2_group")
                    incl.remove("si2_code")
                    if self.ttype == "D" and self.dvals.lower() == "n":
                        incl.remove("si2_price")
                        incl.remove("line_value")
                else:
                    incl = []
                for code in self.form.body:
                    seq = "%s_C%02i" % (code, count)
                    if code == "si2_desc":
                        data = l
                    elif code in incl:
                        data = ldic[code].work
                    else:
                        data = "BLANK"
                    self.form.newdic[seq][tdc.index("tpd_text")] = data
                    self.form.doDrawDetail(self.form.newdic[seq])
                count += 1
        for x in range(count, self.form.maxlines):
            for cod in self.form.body:
                seq = "%s_C%02i" % (cod, x)
                self.form.newdic[seq][tdc.index("tpd_text")] = "BLANK"
                self.form.doDrawDetail(self.form.newdic[seq])

    def doTotal(self, tdc):
        if self.ttype == "D" and self.dvals.lower() == "n":
            return
        for c in self.form.total:
            line = None
            if c in self.form.newdic:
                line = self.form.newdic[c]
            else:
                t = "%s_T00" % c
                if t in self.form.newdic:
                    line = self.form.newdic[t]
            if line:
                self.form.doDrawDetail(line)
            d = "%s_C00" % c
            if d in self.form.newkey:
                line = self.form.newdic[d]
                mrgcod = line[tdc.index("tpd_mrgcod")]
                line[tdc.index("tpd_text")] = getattr(self, "%s" % mrgcod)
                self.form.doDrawDetail(line)

    def doTail(self, tdc):
        for c in self.form.tail:
            line = None
            if c in self.form.newdic:
                line = self.form.newdic[c]
            else:
                t = "%s_T00" % c
                if t in self.form.newdic:
                    line = self.form.newdic[t]
            if line:
                if line[tdc.index("tpd_mrgcod")] == "message":
                    if self.ttype == "D":
                        line[tdc.index("tpd_text")] = "Received By:"
                    elif not self.message:
                        continue
                self.form.doDrawDetail(line)
            d = "%s_C00" % c
            if d in self.form.newdic:
                line = self.form.newdic[d]
                if self.ttype == "D" and d == "message_C00":
                    line[tdc.index("tpd_text")] = " \n \n \n \n "
                else:
                    mrgcod = line[tdc.index("tpd_mrgcod")]
                    line[tdc.index("tpd_text")] = self.doGetData(mrgcod)
                self.form.doDrawDetail(line, fmat=False)

    def doPrint(self):
        if self.dtype in ("C", "D", "I"):
            pfx = "InvCrn"
        elif self.dtype == "O":
            pfx = "SlsOrd"
        elif self.dtype == "W":
            pfx = "WrkOrd"
        elif self.dtype == "Q":
            pfx = "SlsQte"
        if self.dtype == "I":
            head = "%s Invoice" % self.conam.upper()
        elif self.dtype == "C":
            head = "%s Credit Note" % self.conam.upper()
        elif self.dtype == "O":
            head = "%s Sales Order" % self.conam.upper()
        elif self.dtype == "W":
            head = "%s Works Order" % self.conam.upper()
        elif self.dtype == "Q":
            head = "%s Quotation" % self.conam.upper()
        if self.repeml[1] == "Y" and not self.emadd:
            head = "%s %s" % (head, self.docno.work)
            pdfnam = getModName(self.mf.rcdic["wrkdir"], pfx,
                "%s_%s" % (self.conum, self.docno.work), ext="pdf")
        else:
            if len(self.docs) == 1:
                head = "%s %s" % (head, self.docno.work)
            else:
                head += "S"
                for doc in self.docs:
                    head += " %s" % doc[0]
            pdfnam = getModName(self.mf.rcdic["wrkdir"], pfx,
                "%s_all" % self.conum, ext="pdf")
        self.form.output(pdfnam, "F")
        doPrinter(mf=self.mf, conum=self.conum, pdfnam=pdfnam, header=head,
            fromad=self.fromad, repprt=self.repprt, repeml=self.repeml)
        if self.repeml[1] == "Y" and not self.emadd:
            self.form = DrawForm(self.mf.dbm, self.tname)
            self.doLoadStatic()
            self.form.doNewDetail()

    def doLoadStatic(self):
        # ctlmst
        cmc = self.sql.ctlmst_col
        ctm = self.sql.getRec("ctlmst", where=[("ctm_cono", "=",
            self.conum)], limit=1)
        for fld in cmc:
            dat = ctm[cmc.index(fld)]
            if fld in self.form.tptp:
                if fld == "ctm_logo":
                    self.form.letterhead(cmc, ctm, fld, dat)
                    continue
                self.form.tptp[fld][1] = dat
        if "letterhead" in self.form.tptp:
            self.form.letterhead(cmc, ctm, "letterhead", None)
        if "document_type" in self.form.tptp:
            if self.copy.lower() == "a":
                typ = "AMENDED"
            elif self.copy.lower() == "y":
                typ = "COPY"
            else:
                typ = ""
            if ctm[cmc.index("ctm_taxno")]:
                typ = "%s TAX" % typ
            if self.dtype == "I":
                self.form.tptp["document_type"][1] = "%s INVOICE" % typ
            elif self.dtype == "C":
                self.form.tptp["document_type"][1] = "%s CREDIT NOTE" % typ
            elif self.dtype == "O":
                self.form.tptp["document_type"][1] = "%s SALES ORDER" % typ
            elif self.dtype == "W":
                self.form.tptp["document_type"][1] = "%s WORKS ORDER" % typ
            elif self.dtype == "Q":
                self.form.tptp["document_type"][1] = "%s QUOTATION" % typ
        self.form.bank_details(cmc, ctm, 0)
        if "terms" in self.form.tptp:
            acc = self.sql.getRec("ctlmes", cols=["mss_detail"],
                where=[("mss_system", "=", "CON"), ("mss_message", "=", 1)],
                limit=1)
            if acc:
                self.form.tptp["terms"][1] = acc[0]
            else:
                del self.form.tptp["terms"]

    def doHeader(self, page, tdc):
        self.form.add_page()
        for key in self.form.newkey:
            line = self.form.newdic[key]
            if line[tdc.index("tpd_place")] != "A":
                continue
            nl = copyList(line)
            if nl[tdc.index("tpd_detseq")] == "page_number_C00":
                nl[tdc.index("tpd_text")] = str(page)
            elif nl[tdc.index("tpd_type")] == "T" and \
                        nl[tdc.index("tpd_ttyp")] == "H":
                mrgcod = nl[tdc.index("tpd_mrgcod")]
                if mrgcod and mrgcod == "deliver_to" and not self.deliver:
                    continue
                if mrgcod and self.form.tptp[mrgcod][0][1][0] == "S":
                    txt = nl[tdc.index("tpd_text")]
                    nl[tdc.index("tpd_text")] = "%s " % txt
            elif nl[tdc.index("tpd_type")] == "C":
                mrgcod = nl[tdc.index("tpd_mrgcod")]
                if mrgcod == "deliver_to":
                    if not self.deliver:
                        continue
                    nl[tdc.index("tpd_text")] = self.deliver
                else:
                    nl[tdc.index("tpd_text")] = self.doGetData(mrgcod)
            if key == "document_type_C00":
                self.form.doDrawDetail(nl, fmat=False)
            else:
                self.form.doDrawDetail(nl)
        return 0

    def doCfwd(self, page):
        line = copyList(self.form.cfwd)
        line[5] = "Continued on Page %i" % (page + 1)
        self.form.doDrawDetail(line)
        return page + 1

    def doGetData(self, mrgcod):
        if mrgcod and mrgcod in self.form.tptp and self.form.tptp[mrgcod][1]:
            return self.form.tptp[mrgcod][1]
        return ""

class PrintBookingInvoice(object):
    """
    mf: Mainframe object
    conum:  Company number
    conam:  Company name
    dtype:  Document type
    docs:   List of document numbers
    args:
            tname  - The template to use
            repprt - Printing options
            repeml - Emailing options
            copy   - Print a Copy document
            splash - Whether to display a splach screen
    """
    def __init__(self, mf, conum, conam, dtype, docs, **args):
        self.mf = mf
        self.conum = conum
        self.conam = conam
        self.dtype = dtype
        if type(docs) in (int, str):
            self.docs = [[docs]]
        else:
            self.docs = docs
        defaults = {
            "tname": "booking_invoice",
            "repprt": ["N", "V", "view"],
            "repeml": ["N", "N", "", "", "Y"],
            "copy": "n",
            "splash": True}
        for nam in args:
            if nam in ("repprt", "repeml"):
                defaults[nam] = copyList(args[nam])
            else:
                defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        if self.setVariables():
            self.doProcess()

    def setVariables(self):
        self.sql = Sql(self.mf.dbm, ["ctlmst", "bkmmst", "bkmcon", "bkmrtt"],
            prog=__name__)
        if self.sql.error:
            return False
        gc = GetCtl(self.mf)
        bkmctl = gc.getCtl("bkmctl", self.conum)
        if not bkmctl:
            return
        self.fromad = bkmctl["cbk_emadd"]
        if self.repeml:
            self.emadd = self.repeml[2]
        else:
            self.emadd = ""
        return True

    def doProcess(self):
        mcc = self.sql.bkmmst_col
        ccc = self.sql.bkmcon_col
        ttc = self.sql.bkmrtt_col
        self.form = DrawForm(self.mf.dbm, self.tname)
        txt = self.form.sql.tpldet_col.index("tpd_text")
        tdc = self.form.sql.tpldet_col
        self.doLoadStatic()
        self.form.doNewDetail()
        for docno in self.docs:
            if self.splash:
                sp = SplashScreen(self.mf.body,
                    "Generating Forms\n\n   Please Wait...")
            self.docno = CCD(docno, "UI", 9)
            if self.dtype == "I":
                rtn = 2
            else:
                rtn = 6
            doc = self.sql.getRec("bkmtrn", cols=["bkt_number",
                "bkt_date"], where=[("bkt_cono", "=", self.conum),
                ("bkt_type", "=", rtn), ("bkt_refno", "=", "%9s" % docno)],
                group="bkt_number, bkt_refno, bkt_date", limit=1)
            bkno, ivdt = doc
            mst = self.sql.getRec("bkmmst", where=[("bkm_cono", "=",
                self.conum), ("bkm_number", "=", bkno)], limit=1)
            bal = self.sql.getRec("bkmtrn",
                cols=["round(sum(bkt_tramt), 2)"],
                where=[("bkt_cono", "=", self.conum),
                ("bkt_number", "=", bkno)], limit=1)
            self.total_balance = CCD(bal[0], "SD", 13.2).work
            self.cod = mst[self.sql.bkmmst_col.index("bkm_ccode")]
            # bkmcon
            con = self.sql.getRec("bkmcon", where=[("bkc_cono", "=",
                self.conum), ("bkc_ccode", "=", self.cod)], limit=1)
            for fld in ccc:
                if fld in self.form.tptp:
                    d = "%s_C00" % fld
                    self.form.newdic[d][txt] = con[ccc.index(fld)]
                    self.form.tptp[fld][1] = con[ccc.index(fld)]
            # Account details
            if "contact_details" in self.form.tptp:
                dat = "%s %s %s" % (
                    con[ccc.index("bkc_title")],
                    con[ccc.index("bkc_names")],
                    con[ccc.index("bkc_sname")])
                dat = "%1s\n%1s" % (dat, con[ccc.index("bkc_addr1")])
                dat = "%1s\n%1s" % (dat, con[ccc.index("bkc_addr2")])
                dat = "%1s\n%1s" % (dat, con[ccc.index("bkc_addr3")])
                dat = "%1s\n%1s" % (dat, con[ccc.index("bkc_pcode")])
                self.form.tptp["contact_details"][1] = dat
            # Document details
            eml = con[ccc.index("bkc_email")]
            self.form.document_number(docno)
            self.form.document_date(ivdt)
            for fld in mcc:
                if fld in self.form.tptp:
                    self.form.tptp[fld][1] = mst[mcc.index(fld)]
            if "message" in self.form.tptp:
                if mst[mcc.index("bkm_remarks")]:
                    self.message = True
                    self.form.tptp["message"][1] = mst[mcc.index("bkm_remarks")]
                else:
                    self.message = False
            if self.dtype == "I":
                fld = "brt_invno"
            else:
                fld = "brt_crnno"
            rtt = self.sql.getRec("bkmrtt", where=[("brt_cono",
                "=", self.conum), ("brt_number", "=", bkno), (fld, "=",
                self.docno.work)], order="brt_utype, brt_ucode")
            self.doBody(ttc, rtt, tdc)
            self.total_movements = float(ASD(self.total_balance) - \
                ASD(self.total_value))
            self.doTotal(tdc)
            if self.dtype == "I":
                self.doTail(tdc)
            if self.splash:
                sp.closeSplash()
            if self.repeml[1] == "Y" and not self.emadd:
                self.repeml[2] = eml
                self.doPrint()
        if self.repeml[1] == "N" or self.emadd:
            self.repeml[2] = self.emadd
            self.doPrint()

    def doBody(self, ttc, rtt, tdc):
        page = 0
        count = 0
        self.total_taxable = 0
        self.total_nontaxable = 0
        self.total_tax = 0
        self.total_value = 0
        for item in rtt:
            if not count:
                page += 1
                count = self.doHeader(str(page), tdc)
            if count == self.form.maxlines:
                page = self.doCfwd(page)
                count = self.doHeader(page, tdc)
            ldic = {}
            for cod in self.form.body:
                if cod == "line_unit":
                    desc = item[ttc.index("brt_udesc")].strip()
                    room = int(item[ttc.index("brt_uroom")])
                    if room:
                        desc = "%s - Room %s" % (desc, room)
                    ldic[cod] = CCD(desc, self.form.tptp[cod][0][1],
                        self.form.tptp[cod][0][2])
                elif cod != "line_value":
                    ldic[cod] = CCD(item[ttc.index(cod)],
                        self.form.tptp[cod][0][1],
                        self.form.tptp[cod][0][2])
            brt_rbase = item[ttc.index("brt_rbase")]
            if brt_rbase == "A":
                incamt = ldic["brt_quant"].work * ldic["brt_bdays"].work \
                    * ldic["brt_arate"].work
            elif brt_rbase == "C":
                incamt = ldic["brt_bdays"].work * ldic["brt_arate"].work
            else:
                ldic["brt_bdays"].work = ldic["brt_bdays"].disp = ""
                if brt_rbase == "B":
                    incamt = ldic["brt_quant"].work * ldic["brt_arate"].work
                else:
                    incamt = ldic["brt_arate"].work
            vatrat = float(ASD(100.0) + ASD(item[ttc.index("brt_vrate")]))
            excamt = round((incamt * 100.0 / vatrat), 2)
            vatamt = float(ASD(incamt) - ASD(excamt))
            if excamt == incamt:
                self.total_nontaxable = float(ASD(self.total_nontaxable) + \
                    ASD(excamt))
            else:
                self.total_taxable = float(ASD(self.total_taxable) + \
                    ASD(excamt))
            self.total_tax = float(ASD(self.total_tax) + ASD(vatamt))
            self.total_value = float(ASD(self.total_value) + ASD(incamt))
            ldic["line_value"] = CCD(incamt, "SD", 11.2)
            if count == self.form.maxlines:
                page = self.doCfwd(page)
                count = self.doHeader(page, tdc)
            incl = copyList(self.form.body)
            for code in self.form.body:
                seq = "%s_C%02i" % (code, count)
                if code in incl:
                    data = ldic[code].work
                else:
                    data = "BLANK"
                self.form.newdic[seq][tdc.index("tpd_text")] = data
                if code == "brt_bdays" and not data:
                    self.form.doDrawDetail(self.form.newdic[seq], fmat=False)
                else:
                    self.form.doDrawDetail(self.form.newdic[seq])
            count += 1
        if not rtt:
            count = self.doHeader(str(page), tdc)
        for x in range(count, self.form.maxlines):
            for cod in self.form.body:
                seq = "%s_C%02i" % (cod, x)
                self.form.newdic[seq][tdc.index("tpd_text")] = "BLANK"
                self.form.doDrawDetail(self.form.newdic[seq])

    def doTotal(self, tdc):
        for c in self.form.total:
            line = None
            if c in self.form.newdic:
                line = self.form.newdic[c]
            else:
                t = "%s_T00" % c
                if t in self.form.newdic:
                    line = self.form.newdic[t]
            if line:
                self.form.doDrawDetail(line)
            d = "%s_C00" % c
            if d in self.form.newkey:
                line = self.form.newdic[d]
                mrgcod = line[tdc.index("tpd_mrgcod")]
                line[tdc.index("tpd_text")] = getattr(self, "%s" % mrgcod)
                self.form.doDrawDetail(line)

    def doTail(self, tdc):
        for c in self.form.tail:
            line = None
            if c in self.form.newdic:
                line = self.form.newdic[c]
            else:
                t = "%s_T00" % c
                if t in self.form.newdic:
                    line = self.form.newdic[t]
            if line:
                if line[tdc.index("tpd_mrgcod")] == "message":
                    if not self.message:
                        continue
                self.form.doDrawDetail(line)
            d = "%s_C00" % c
            if d in self.form.newdic:
                line = self.form.newdic[d]
                mrgcod = line[tdc.index("tpd_mrgcod")]
                line[tdc.index("tpd_text")] = self.doGetData(mrgcod)
                self.form.doDrawDetail(line, fmat=False)

    def doPrint(self):
        pfx = "InvCrn"
        if self.dtype == "I":
            head = "%s - Invoice" % self.conam.upper()
        elif self.dtype == "C":
            head = "%s - Credit Note" % self.conam.upper()
        if self.repeml[1] == "Y" and not self.emadd:
            pdfnam = getModName(self.mf.rcdic["wrkdir"], pfx,
                "%s_%s" % (self.conum, self.docno.work), ext="pdf")
            head = "%s %s" % (head, self.docno.work)
        else:
            pdfnam = getModName(self.mf.rcdic["wrkdir"], pfx,
                "%s_all" % self.conum, ext="pdf")
            if len(self.docs) == 1:
                head = "%s %s" % (head, self.docno.work)
            else:
                head = "%sS" % head
                for doc in self.docs:
                    head += " %s" % doc.strip()
        self.form.output(pdfnam, "F")
        doPrinter(mf=self.mf, conum=self.conum, pdfnam=pdfnam, header=head,
            fromad=self.fromad, repprt=self.repprt, repeml=self.repeml)
        if self.repeml[1] == "Y":
            self.form = DrawForm(self.mf.dbm, self.tname)
            self.doLoadStatic()
            self.form.doNewDetail()

    def doLoadStatic(self):
        # ctlmst
        cmc = self.sql.ctlmst_col
        ctm = self.sql.getRec("ctlmst", where=[("ctm_cono", "=",
            self.conum)], limit=1)
        for fld in cmc:
            dat = ctm[cmc.index(fld)]
            if fld in self.form.tptp:
                if fld == "ctm_logo":
                    self.form.letterhead(cmc, ctm, fld, dat)
                    continue
                self.form.tptp[fld][1] = dat
        if "letterhead" in self.form.tptp:
            self.form.letterhead(cmc, ctm, "letterhead", None)
        if "document_type" in self.form.tptp:
            if self.copy.lower() == "y":
                typ = "COPY"
            else:
                typ = ""
            if ctm[cmc.index("ctm_taxno")]:
                typ = "%s TAX" % typ
            if self.dtype == "I":
                self.form.tptp["document_type"][1] = "%s INVOICE" % typ
            elif self.dtype == "C":
                self.form.tptp["document_type"][1] = "%s CREDIT NOTE" % typ
        self.form.bank_details(cmc, ctm, 0)

    def doHeader(self, page, tdc):
        self.form.add_page()
        for key in self.form.newkey:
            line = self.form.newdic[key]
            if line[tdc.index("tpd_place")] != "A":
                continue
            nl = copyList(line)
            if nl[tdc.index("tpd_detseq")] == "page_number_C00":
                nl[tdc.index("tpd_text")] = str(page)
            elif nl[tdc.index("tpd_type")] == "T" and \
                        nl[tdc.index("tpd_ttyp")] == "H":
                mrgcod = nl[tdc.index("tpd_mrgcod")]
                if mrgcod and mrgcod == "deliver_to" and not self.deliver:
                    continue
                if mrgcod and self.form.tptp[mrgcod][0][1][0] == "S":
                    txt = nl[tdc.index("tpd_text")]
                    nl[tdc.index("tpd_text")] = "%s " % txt
            elif nl[tdc.index("tpd_type")] == "C":
                mrgcod = nl[tdc.index("tpd_mrgcod")]
                if mrgcod == "deliver_to":
                    if not self.deliver:
                        continue
                    nl[tdc.index("tpd_text")] = self.deliver
                else:
                    nl[tdc.index("tpd_text")] = self.doGetData(mrgcod)
            if key == "document_type_C00":
                self.form.doDrawDetail(nl, fmat=False)
            else:
                self.form.doDrawDetail(nl)
        return 0

    def doCfwd(self, page):
        line = copyList(self.form.cfwd)
        line[5] = "Continued on Page %i" % (page + 1)
        self.form.doDrawDetail(line)
        return page + 1

    def doGetData(self, mrgcod):
        if mrgcod and mrgcod in self.form.tptp and self.form.tptp[mrgcod][1]:
            return self.form.tptp[mrgcod][1]
        return ""

class PrintPayslip(object):
    """
    Used to print payslips

    conum  : Company Number
    conam  : Company Name
    empnos : List of Employees
    rundt  : Run Date
    progs  : Print Progressives (y/n)
    runtp  : c=Copy, o=Original, p=Preview
    tname  : Template Name
    repprt : Printer Name
    repeml : Printer Name
    message: Message to Print
    export : The open file object for Best Upload
    bestac : Standard Bank Best Account Code
    """
    def __init__(self, mf, conum, conam, empnos, rundt, **args):
        self.mf = mf
        self.conum = conum
        self.conam = conam
        self.empnos = empnos
        self.rundt = CCD(rundt, "D1", 10.0)
        defaults = {
            "progs": "n",
            "runtp": "c",
            "tname": "payslip",
            "repprt": ["N", "V", "view"],
            "repeml": ["N", "N", "", "", "Y"],
            "message": None,
            "export": None,
            "bestac": None,
            "splash": True}
        for nam in args:
            if nam in ("repprt", "repeml"):
                defaults[nam] = copyList(args[nam])
            else:
                defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        if self.setVariables():
            self.doProcess()

    def setVariables(self):
        self.sql = Sql(self.mf.dbm, ["ctlmst", "wagedc", "wagmst", "wagtf1",
            "wagtf2"], prog=__name__)
        if self.sql.error:
            return False
        gc = GetCtl(self.mf)
        wagctl = gc.getCtl("wagctl", self.conum)
        if not wagctl:
            return
        self.fromad = wagctl["ctw_emadd"]
        ########################################################################
        # Start date and email details
        ########################################################################
        y = int(self.rundt.work / 10000)
        m = int(self.rundt.work / 100) % 100
        if m < 3:
            y = y - 1
        self.start = (y * 10000) + 301
        if self.repeml:
            self.emadd = self.repeml[2]
        else:
            self.emadd = ""
        self.etotal = 0
        return True

    def doProcess(self):
        self.form = DrawForm(self.mf.dbm, self.tname)
        self.doLoadStatic()
        self.form.doNewDetail()
        for emp in self.empnos:
            self.doPayslip(emp)
        if self.repeml[1] == "N" or self.emadd:
            self.repeml[2] = self.emadd
            self.doPrint()

    def doLoadStatic(self):
        cmc = self.sql.ctlmst_col
        ctm = self.sql.getRec("ctlmst", where=[("ctm_cono", "=",
            self.conum)], limit=1)
        self.conam = ctm[cmc.index("ctm_name")]
        for fld in cmc:
            dat = ctm[cmc.index(fld)]
            if fld in self.form.tptp:
                if fld == "ctm_logo":
                    self.form.letterhead(cmc, ctm, fld, dat)
                    continue
                self.form.tptp[fld][1] = dat
        if "letterhead" in self.form.tptp:
            self.form.letterhead(cmc, ctm, "letterhead", None)
        if "document_type" in self.form.tptp:
            if self.runtp.lower() == "c":
                typ = "COPY PAYSLIP"
            elif self.runtp.lower() == "p":
                typ = "PREVIEW PAYSLIP"
            else:
                typ = "PAYSLIP"
            self.form.tptp["document_type"][1] = typ

    def doPayslip(self, emp):
        tdc = self.form.sql.tpldet_col
        wgm = self.sql.getRec("wagmst", where=[("wgm_cono", "=",
            self.conum), ("wgm_empno", "=", emp)], limit=1)
        wmc = self.sql.wagmst_col
        self.empno = wgm[wmc.index("wgm_empno")]
        w1c = self.sql.wagtf1_col
        wt1 = self.sql.getRec("wagtf1", where=[("wt1_cono", "=",
            self.conum), ("wt1_empno", "=", self.empno), ("wt1_date", "=",
            self.rundt.work)], limit=1)
        if not wt1:
            return
        if self.splash:
            sp = SplashScreen(self.mf.body,
                "Generating Forms\n\n   Please Wait...")
        if "employee_name" in self.form.tptp:
            dat = wgm[wmc.index("wgm_fname")]
            dat = "%s %s" % (dat, wgm[wmc.index("wgm_sname")])
            self.form.newdic["employee_name_C00"][tdc.index("tpd_text")] = dat
        self.ptype = wgm[wmc.index("wgm_ptype")]
        self.btype = wgm[wmc.index("wgm_btype")]
        self.bbranch = wgm[wmc.index("wgm_bbranch")]
        self.bacno = wgm[wmc.index("wgm_bacno")]
        self.bname = wgm[wmc.index("wgm_bname")]
        eml = wgm[wmc.index("wgm_emadd")]
        tmpdic = self.form.newdic.copy()
        for fld in wmc:
            if fld in self.form.tptp:
                d = "%s_C00" % fld
                dat = wgm[wmc.index(fld)]
                tmpdic[d][tdc.index("tpd_text")] = dat
        for fld in w1c:
            if fld in self.form.tptp:
                d = "%s_C00" % fld
                dat = wt1[w1c.index(fld)]
                tmpdic[d][tdc.index("tpd_text")] = dat
        earn = 0
        deduct = 0
        self.total = 0
        earn = self.doLoadDic("E", tdc, tmpdic)
        if "total_earnings" in self.form.tptp:
            tmpdic["total_earnings_C00"][tdc.index("tpd_text")] = earn
        deduct = self.doLoadDic("D", tdc, tmpdic)
        if "total_deductions" in self.form.tptp:
            tmpdic["total_deductions_C00"][tdc.index("tpd_text")] = deduct
        if self.progs.lower() == "y":
            prog = self.doAccumProgs()
            n = -1
            for n, a in enumerate(prog):
                t = "progressives_T%02i" % n
                tmpdic[t][tdc.index("tpd_text")] = a[2]
                c = "progressives_C%02i" % n
                tmpdic[c][tdc.index("tpd_text")] = a[4]
            for x in range(n+1, 20):
                t = "progressives_T%02i" % x
                tmpdic[t][tdc.index("tpd_text")] = "BLANK"
                c = "progressives_C%02i" % x
                tmpdic[c][tdc.index("tpd_text")] = "BLANK"
        if "net_pay" in self.form.tptp:
            tmpdic["net_pay_C00"][tdc.index("tpd_text")] = self.total
        if "payment_method" in self.form.tptp:
            typ = "Method of Payment:"
            if self.ptype == "C":
                typ = "%s Cash" % typ
            elif self.ptype == "E":
                typ = "%s Electronic" % typ
            else:
                typ = "%s Cheque" % typ
            tmpdic["payment_method_C00"][tdc.index("tpd_text")] = typ
        self.form.add_page()
        for key in self.form.newkey:
            nl = tmpdic[key]
            self.form.doDrawDetail(nl)
        if self.ptype == "E" and self.export and self.bestac:
            value = int(round((self.total * 100), 0))
            self.export.write("%1s%4s%06u%07u%019u%1s%1s%011u%-20s%10s"\
                "%-15s%1s\r\n" % (2, self.bestac, self.bbranch,
                int(self.empno), int(self.bacno), "", self.btype,
                value, self.bname, "", self.conam[:15], ""))
            self.etotal = float(ASD(self.etotal) + ASD(self.total))
        if self.splash:
            sp.closeSplash()
        if self.repeml[1] == "Y" and not self.emadd:
            self.repeml[2] = eml
            self.doPrint()

    def doLoadDic(self, typ, tdc, tmpdic):
        if typ == "E":
            txt = "earnings"
        elif typ == "D":
            txt = "deductions"
        amnt = 0
        recs = self.sql.getRec(tables=["wagedc", "wagtf2"], cols=["wt2_code",
            "wt2_taxcode", "ced_desc", "round(sum(wt2_eamt), 2)"],
            where=[("wt2_cono", "=", self.conum), ("wt2_empno", "=",
            self.empno), ("wt2_type", "=", typ), ("wt2_date", "=",
            self.rundt.work), ("ced_cono=wt2_cono",),
            ("ced_type=wt2_type",), ("ced_code=wt2_code",)],
            group="wt2_code, wt2_taxcode, ced_desc", order="wt2_code")
        n = -1
        for n, a in enumerate(recs):
            t = "%s_T%02i" % (txt, n)
            if a[1] == "T":
                tmpdic[t][tdc.index("tpd_text")] = a[2] + " (Notional Only)"
            else:
                tmpdic[t][tdc.index("tpd_text")] = a[2]
            c = "%s_C%02i" % (txt, n)
            tmpdic[c][tdc.index("tpd_text")] = a[3]
            if a[1] != "T":
                amnt = float(ASD(amnt) + ASD(a[3]))
                if typ == "E":
                    self.total = float(ASD(self.total) + ASD(a[3]))
                elif typ == "D":
                    self.total = float(ASD(self.total) - ASD(a[3]))
        for x in range(n+1, 10):
            t = "%s_T%02i" % (txt, x)
            tmpdic[t][tdc.index("tpd_text")] = "BLANK"
            c = "%s_C%02i" % (txt, x)
            tmpdic[c][tdc.index("tpd_text")] = "BLANK"
        return amnt

    def doAccumProgs(self):
        wt2 = self.sql.getRec(tables=["wagtf2", "wagedc"], cols=["wt2_type",
            "wt2_code", "ced_desc", "round(sum(wt2_hrs), 2)",
            "round(sum(wt2_eamt), 2)"], where=[("wt2_cono", "=", self.conum),
            ("wt2_empno", "=", self.empno), ("wt2_date", "between", self.start,
            self.rundt.work), ("ced_cono=wt2_cono",), ("ced_type=wt2_type",),
            ("ced_code=wt2_code",)], group="wt2_type, wt2_code, ced_desc",
            order="wt2_type desc, wt2_code")
        return wt2

    def doPrint(self):
        head = "%s Payslip" % self.conam.upper()
        if self.repeml[1] == "Y" and not self.emadd:
            key = "%s_%s" % (self.conum, self.empno)
        else:
            key = "%s_all" % self.conum
            if len(self.empnos) > 1:
                head += "S"
        head += " for %s" % self.rundt.disp
        self.pdfnam = getModName(self.mf.rcdic["wrkdir"], "payslip", key,
            ext="pdf")
        self.form.output(self.pdfnam, "F")
        doPrinter(mf=self.mf, conum=self.conum, pdfnam=self.pdfnam,
            fromad=self.fromad, header=head, repprt=self.repprt,
            repeml=self.repeml)
        if self.repeml[1] == "Y" and not self.emadd:
            self.form = DrawForm(self.mf.dbm, self.tname)
            self.doLoadStatic()
            self.form.doNewDetail()

class PrintCards(object):
    """
    Print bowls cards using the following options:

    mf     - The Mainframe object.
    conum  - The company number.
    cdes   - The match name.
    game   - The game number.
    datd   - The match date.
    skips  - A list of skips.
    tname  - The template to be used.
    ends   - The number of ends.
    skins  - Whether there are skins.
    sends  - The number of ends per skin.
    """
    def __init__(self, mf, conum, cdes, game, datd, skips, ends=21, skins="N", sends=0, args=None):
        self.mf = mf
        self.conum = conum
        self.cdes = cdes
        self.game = game
        self.datd = datd
        self.skips = skips
        self.ends = ends
        self.skins = skins
        self.sends = sends
        if self.doVariables():
            if args:
                self.tname, self.repprt = args
            else:
                self.doMainProcess()
            self.doPrintCards()

    def doVariables(self):
        self.sql = Sql(self.mf.dbm, ("bwlctl", "tplmst"), prog=__name__)
        if self.sql.error:
            return
        gc = GetCtl(self.mf)
        bwlctl = gc.getCtl("bwlctl", self.conum)
        if not bwlctl:
            return
        self.tplnam = bwlctl["ctb_tplnam"]
        return True

    def doMainProcess(self):
        tit = "Select Score Cards Output"
        tpm = {
            "stype": "R",
            "tables": ("tplmst",),
            "cols": (
                ("tpm_tname", "", 0, "Template"),
                ("tpm_title", "", 0, "Title", "Y")),
            "where": [("tpm_type", "=", "C")],
            "order": "tpm_tname"}
        fld = ((("T",0,0,0),"INA",20,"Template Name","",
            self.tplnam,"N",self.doTplNam,tpm,None,None),)
        tnd = ((self.doEnd,"Y"),)
        txt = (self.doExit,)
        self.df = TartanDialog(self.mf, tops=True, title=tit, eflds=fld,
            tend=tnd, txit=txt, view=("N","V"))
        self.df.mstFrame.wait_window()

    def doTplNam(self, frt, pag, r, c, p, i, w):
        acc = self.sql.getRec("tplmst", where=[("tpm_tname", "=", w),
            ("tpm_type", "=", "C")], limit=1)
        if not acc:
            return "Invalid Template Name"
        self.tname = w

    def doEnd(self):
        self.df.closeProcess()
        self.repprt = self.df.repprt
        showWarning(self.mf.body, "Change Paper",
            "In Order to Print Cards You Need to Change the Paper to "\
            "A6 Cards. Please Do So Now Before Continuing.")

    def doExit(self):
        return "rf"

    def doPrintCards(self):
        self.pdfnam = getModName(self.mf.rcdic["wrkdir"], __name__,
            "comp_cards", ext="pdf")
        self.form = DrawForm(self.mf.dbm, self.tname, foot=False,
            wrkdir=self.mf.rcdic["wrkdir"])
        self.doLoadStatic()
        self.form.doNewDetail()
        tdc = self.form.sql.tpldet_col
        txt = tdc.index("tpd_text")
        fil = tdc.index("tpd_mrg_fill")
        bdr = tdc.index("tpd_mrg_border")
        y1 = tdc.index("tpd_mrg_y1")
        lh = tdc.index("tpd_mrg_lh")
        y2 = tdc.index("tpd_mrg_y2")
        ff = self.form.newdic["%s_C%02i" % (self.form.body[0], 0)]
        y3 = ff[y1]
        mm = round((23 * ff[lh]) / (self.ends + 1), 2)
        y4 = y3 + mm
        for x in range(self.ends + 1):
            if x == self.ends:
                dat = "Total"
                fill = True
            else:
                dat = end = x + 1
                fill = bool(self.skins == "Y" and not end % self.sends)
            for nam in self.form.body:
                fld = "%s_C%02i" % (nam, x)
                if fld in self.form.newdic:
                    self.form.newdic[fld][y1] = y3
                    self.form.newdic[fld][y2] = y4
                    if fld.startswith("ends"):
                        self.form.newdic[fld][txt] = dat
                    if fill:
                        self.form.newdic[fld][fil] = 1
                        self.form.newdic[fld][bdr] = "TLRB"
            y3 = y4
            y4 = y3 + mm
        for grn, skp, opp in self.skips:
            self.form.add_page()
            if "skip_C00" in self.form.newdic:
                self.form.newdic["skip_C00"][txt] = self.doGetName(skp)
            if "opponent_C00" in self.form.newdic:
                self.form.newdic["opponent_C00"][txt] = self.doGetName(opp)
            if "bcg_rink_C00" in self.form.newdic:
                self.form.newdic["bcg_rink_C00"][txt] = grn
            for key in self.form.newkey:
                self.form.doDrawDetail(self.form.newdic[key])
        self.form.output(self.pdfnam, "F")
        doPrinter(mf=self.mf, conum=self.conum, pdfnam=self.pdfnam,
            repprt=self.repprt)

    def doLoadStatic(self):
        tdc = self.form.sql.tpldet_col
        if self.ends < 21:
            for line in self.form.tpldet:
                if line[tdc.index("tpd_repeat")] == 22:
                    line[tdc.index("tpd_repeat")] = self.ends + 1
        if "bcm_name" in self.form.tptp:
            self.form.tptp["bcm_name"][1] = self.cdes
        if "bcg_game" in self.form.tptp:
            self.form.tptp["bcg_game"][1] = self.game
        if "bcg_date" in self.form.tptp:
            self.form.tptp["bcg_date"][1] = self.datd

    def doGetName(self, data):
        if data:
            if data[2]:
                name = "%s. " % data[2].split()[0][0]
            else:
                name = ""
            return "%s%s (%s)" % (name, data[1], data[0])
        else:
            return "* Bye *"

class PrintDraw(object):
    def __init__(self, mf, conum, date, dtim, **args):
        self.mf = mf
        self.conum = conum
        self.date = date
        self.time = dtim
        defaults = {
            "name": None,
            "cdes": None,
            "takings": "y",
            "listing": "n",
            "board": "n",
            "empty": "n",
            "repprt": ["N", "V", "view"]}
        for nam in args:
            if nam in ("repprt", "repeml"):
                defaults[nam] = copyList(args[nam])
            else:
                defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        if self.setVariables():
            self.doProcess()

    def setVariables(self):
        gc = GetCtl(self.mf)
        bwlctl = gc.getCtl("bwlctl", self.conum)
        if not bwlctl:
            return
        self.nstart = bwlctl["ctb_nstart"]
        self.drfmat = bwlctl["ctb_drfmat"]
        self.sql = Sql(self.mf.dbm, ["bwldrm", "bwldrt"], prog=__name__)
        if self.sql.error:
            return False
        bdm = self.sql.getRec("bwldrm", where=[("bdm_cono", "=",
            self.conum), ("bdm_date", "=", self.date), ("bdm_time", "=",
            self.time)], limit=1)
        self.dtype = bdm[self.sql.bwldrm_col.index("bdm_dtype")]
        self.dhist = bdm[self.sql.bwldrm_col.index("bdm_dhist")]
        self.dedit = bdm[self.sql.bwldrm_col.index("bdm_dedit")]
        self.ratem = CCD(bdm[self.sql.bwldrm_col.index("bdm_mrate")], "UD", 6.2)
        self.ratev = CCD(bdm[self.sql.bwldrm_col.index("bdm_vrate")], "UD", 6.2)
        self.dated = CCD(self.date, "D1", 10).disp
        if self.time == "A":
            self.timed = "Afternoon"
        else:
            self.timed = "Morning"
        self.position = ["Skip", "Third", "Second", "Lead"]
        return True

    def doProcess(self):
        col = ["bdt_rink", "bdt_side", "bdt_pos", "bdt_tab", "bdt_name"]
        whr = [("bdt_date", "=", self.date), ("bdt_time", "=", self.time)]
        odr = "bdt_rink, bdt_side, bdt_pos desc"
        recs = self.sql.getRec("bwldrt", cols=col, where=whr, order=odr)
        adraw = []
        mem = 0
        vis = 0
        rink = None
        data = None
        for rec in recs:
            if not rink or (rink and rec[0] != rink):
                if data:
                    while len(data) != 6:
                        data.append([0, ""])
                    adraw.append(data)
                rink = rec[0]
                side = rec[1]
                pos = 4
                data = [rec[0], 0]
            if rec[1] != side:
                while len(data) != 6:
                    data.append([0, ""])
                adraw.append(data)
                side = rec[1]
                pos = 4
                data = [rec[0], 0]
            while pos != rec[2]:
                data.append([0, ""])
                pos -= 1
            if rec[3] > 900000:
                data.append([999, rec[4]])
            else:
                data.append([rec[3], rec[4]])
            if rec[3] < self.nstart:
                mem += 1
            else:
                vis += 1
            pos -= 1
        if data:
            while len(data) != 6:
                data.append([0, ""])
            adraw.append(data)
        self.fpdf = MyFpdf(name=self.name, head=90)
        if self.takings.lower() == "y":
            mem = 0
            vis = 0
            pages = int(len(adraw) / 60)
            if len(adraw) % 60:
                pages += 1
            for page in range(pages):
                first = page * 60
                last = (page + 1) * 60
                if last >= len(adraw):
                    last = len(adraw)
                self.pageHeading("A")
                for x in range(first, last, 2):
                    txt = "%2s %3s %-20s %3s %-20s %3s %-20s %3s %-20s"
                    dat = [adraw[x][0]]
                    for t in adraw[x][2:]:
                        if not t[0]:
                            dat.extend(["", ""])
                        else:
                            dat.extend([t[0], t[1]])
                    self.fpdf.drawText(txt % tuple(dat), border="TLR")
                    dat = [adraw[x+1][0]]
                    for t in adraw[x+1][2:]:
                        if not t[0]:
                            dat.extend(["", ""])
                        else:
                            dat.extend([t[0], t[1]])
                    self.fpdf.drawText(txt % tuple(dat), border="BLR")
            for m in adraw:
                for t in m[2:]:
                    if not t[0]:
                        continue
                    if t[0] < self.nstart:
                        mem += 1
                    else:
                        vis += 1
            self.fpdf.setFont(style="B", size=14)
            self.fpdf.drawText()
            self.fpdf.drawText()
            ul = "                               ---------"
            txt = ul.replace("-", self.fpdf.suc)
            self.fpdf.drawText("%3s Members  @ R%2s         R %7.2f" % \
                (mem, self.ratem.disp, mem*self.ratem.work), h=5)
            self.fpdf.drawText("%3s Visitors @ R%2s         R %7.2f" % \
                (vis, self.ratev.disp, vis*self.ratev.work), h=5)
            self.fpdf.underLine(t="S", txt=txt)
            self.fpdf.drawText("    Total Takings              R %7.2f" % \
                ((mem*self.ratem.work)+(vis*self.ratev.work)), h=5)
            self.fpdf.underLine(t="D", txt=txt)
        if self.listing.lower() == "y":
            tabs = []
            for x in range(0, len(adraw), 2):
                for n, t in enumerate(adraw[x][2:]):
                    if not t[0]:
                        continue
                    tabs.append([t[0], adraw[x][0], "L",
                        self.position[n]])
                for n, t in enumerate(adraw[x+1][2:]):
                    if not t[0]:
                        continue
                    tabs.append([t[0], adraw[x+1][0], "R",
                        self.position[n]])
            tabs.sort()
            pages = int(len(tabs) / 150)
            if len(tabs) % 150:
                pages += 1
            for page in range(pages):
                self.pageHeading("B")
                xaxis = self.fpdf.get_x()
                yaxis = self.fpdf.get_y()
                for row in range(3):
                    self.fpdf.set_y(yaxis)
                    for col in range(50):
                        seq = (page * 150) + (row * 50) + col
                        if seq > (len(tabs) - 1):
                            txt = "                      "
                        else:
                            txt = " %3i  %2s  %s  %-8s " % tuple(tabs[seq])
                        w = self.fpdf.get_string_width(txt)
                        self.fpdf.drawText(txt=txt, x=xaxis, w=w, h=5,
                            border="LRB")
                    xaxis = xaxis + w
        if self.board.lower() == "y":
            greens = {}
            for num in range(0, len(adraw), 2):
                grn = adraw[num][0][0]
                rnk = int(adraw[num][0][1])
                if grn not in greens:
                    greens[grn] = {}
                if rnk not in greens[grn]:
                    greens[grn][rnk] = [[[], [], [], []], [[], [], [], []]]
                for pos, tab in enumerate(adraw[num][2:]):
                    greens[grn][rnk][0][pos] = tab
                if self.drfmat == "A":
                    fmat = (0, 1, 2, 3)
                else:
                    fmat = (3, 2, 1, 0)
                for num, tab in enumerate(adraw[num+1][2:]):
                    greens[grn][rnk][1][fmat[num]] = tab
            keys = list(greens.keys())
            keys.sort()
            for grn in keys:
                self.pageHeading(htyp="C", grn=grn)
                if self.empty == "Y":
                    for x in range(1, 7):
                        if x not in greens[grn]:
                            greens[grn][x] = [[[],[],[],[]], [[],[],[],[]]]
                for rnk in greens[grn]:
                    xaxis = self.fpdf.get_x()
                    yaxis = self.fpdf.get_y()
                    for side in range(2):
                        for pos in range(4):
                            x = xaxis + (side * 130) + (pos * 32.5)
                            if side == 1:
                                x += 15
                            if not greens[grn][rnk][side][pos]:
                                tab = " "
                            else:
                                tab = str(greens[grn][rnk][side][pos][0])
                                if tab == "0":
                                    tab = ""
                            self.fpdf.drawText(tab, x=x, y=yaxis, w=32.5,
                                h=12, font=["Arial", "B", 24], align="C",
                                border="TLR")
                            self.fpdf.setFont("Arial", "B", 10)
                            if not greens[grn][rnk][side][pos]:
                                nam = " \n "
                            else:
                                nam = greens[grn][rnk][side][pos][1]
                            if len(self.fpdf.multi_cell(32.5, 5, nam,
                                    border="LRB", split_only=True)) == 1:
                                nam = "%s\n " % nam
                            self.fpdf.drawText(nam, x=x, y=yaxis+12, w=32.5,
                                h=6, align="C", border="LRB", ctyp="M")
                        if side == 0:
                            txt = "%s%s" % (grn, rnk)
                            self.fpdf.drawText(txt, x=x+32.5, y=yaxis,
                                w=15, h=24, font=["Arial", "BI", 24],
                                fill=1, border="TLB", ln=0)
        key = "%s_%s" % (self.conum, self.date)
        pdfnam = getModName(self.mf.rcdic["wrkdir"], "draw", key, ext="pdf")
        self.fpdf.output(pdfnam, "F")
        doPrinter(mf=self.mf, conum=self.conum, pdfnam=pdfnam,
            repprt=self.repprt)

    def pageHeading(self, htyp="A", grn="A"):
        if not self.cdes:
            txt = "TABS for the %s of %s (T-%s, H-%s, E-%s)" % (self.timed,
                self.dated, self.dtype, self.dhist, self.dedit)
        else:
            txt = "%s for the %s of %s" % (self.cdes, self.timed, self.dated)
        if htyp == "C":
            self.fpdf.add_page(orientation="L")
            self.def_orientation = "L"
            self.fpdf.setFont("Arial", "B", 24)
            self.fpdf.drawText(txt, align="C")
            self.fpdf.drawText(" ")
            self.fpdf.drawText(" ")
            self.fpdf.drawText("%s - Green" % grn, align="C")
            self.fpdf.drawText(" ")
            text = copyList(self.position)
            for x in range(2):
                if x == 1 and self.drfmat == "B":
                    text = ["Lead", "Second", "Third", "Skip"]
                for txt in text:
                    if x == 1 and txt == text[3]:
                        ln = 1
                    else:
                        ln = 0
                    self.fpdf.drawText(txt, w=32.5, h=10, align="C", fill=1,
                        font=["Arial", "B", 24], border="TLRB", ln=ln)
                    if x == 0 and txt == "Lead":
                        self.fpdf.drawText("RK", w=15, h=10, align="C", fill=1,
                            font=["Arial", "BI", 24], border="TLRB", ln=ln)
            return
        self.fpdf.add_page()
        if htyp == "A":
            pad = " " * (55 - len(txt))
        elif htyp == "B":
            pad = " " * (57 - len(txt))
        self.fpdf.drawText("%s %s" % (txt, pad), font=["courier", "B", 14])
        self.fpdf.drawText()
        if htyp == "A":
            self.fpdf.setFont(style="B", size=8)
            self.fpdf.drawText("%2s     %-20s     %-20s     %-20s     "\
                "%-20s" % ("RK", "Skip", "Third", "Second", "Lead"),
                border="TBLR", fill=1)
        else:
            txt = " TAB  RK  P  Position "
            w = self.fpdf.get_string_width(txt)
            for x in range(2):
                self.fpdf.drawText(w=w,h=5,txt=txt,border="TBLR",fill=1,ln=0)
            self.fpdf.drawText(w=w,h=5,txt=txt,border="TBLR",fill=1)
            self.fpdf.setFont(style="", size=14)

class PwdConfirm(object):
    """
    mf          The mainframe class
    screen      The screen to use if not mf.body
    conum       Company Number
    system      The system code i.e.
                    MST - Control Modules
                    ASS - Asset's Ledger
                    BKM - Bookings Master
                    BWL - Bowling Clubs
                    CRS - Creditor's Ledger
                    DRS - Debtor's Ledger
                    GEN - General Ledger
                    INV - Sales Invoicing
                    MEM - Member's Ledger
                    RCA - Rental's Ledger (Extended)
                    RTL - Rental's Ledger (Basic)
                    SEC - Sectional Competitions
                    STR - Store's Ledger
                    USR - The user's password
    user        The username for system USR.
    code        The routine as per the 'pwctrl' tuple in tartanWork.py e.g.
                    NoPost - Allow Postings to Locked Accounts
    desc        A description to use as the title
    passwd      The non interactive password
    product     In the case of STR, the group, code and description of the
                product

    Examples:
    =========
    GEN NoPost Allows Postings to Locked G/Ledger Accounts
    DRS NewAcc Allows you to Create a New Debtor's Account
    STR ExQty  Allows you to Override a Quantity Issue
    """
    def __init__(self, mf, **args):
        self.mf = mf
        defaults = {
            "screen": None,
            "conum": None,
            "system": None,
            "user": None,
            "code": None,
            "desc": None,
            "passwd": None,
            "product": None}
        for nam in args:
            defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        self.setVariables()
        self.loadPasswords()
        if not self.pwd:
            self.flag = "ok"
        elif self.passwd:
            if self.passwd in (self.pwd, self.sup):
                self.flag = "ok"
            else:
                self.flag = "no"
        elif not mf.window:
            self.flag = "no"
        else:
            self.mainProcess()
            if self.pc.mstFrame.winfo_toplevel().state() == "withdrawn":
                self.pc.mstFrame.winfo_toplevel().deiconify()
            self.pc.mstFrame.wait_window()

    def setVariables(self):
        self.sql = Sql(self.mf.dbm, "ctlpwr", prog=__name__)
        self.tries = 0

    def loadPasswords(self):
        pwd = self.sql.getRec("ctlpwr", cols=["pwd_pass", "pwd_desc"],
            where=[("pwd_cono", "=", 0), ("pwd_sys", "=", self.system),
            ("pwd_code", "=", self.code)], limit=1)
        if not pwd:
            pwd = self.sql.getRec("ctlpwr", cols=["pwd_pass",
                "pwd_desc"], where=[("pwd_cono", "=", self.conum), ("pwd_sys",
                "=", self.system), ("pwd_code", "=", self.code)], limit=1)
        if not pwd:
            self.pwd = None
            if not self.desc:
                self.desc = self.code
        elif self.system == "INV" and self.code == "UserPwd" and pwd[0]:
            pwd = self.sql.getRec("ctlpwu", cols=["usr_pwd"],
                where=[("usr_name", "=", self.user)], limit=1)
            self.pwd = b64Convert("decode", pwd[0])
            self.sup = None
            return
        else:
            self.pwd = b64Convert("decode", pwd[0])
            if not self.desc:
                self.desc = pwd[1]
        sup = self.sql.getRec("ctlpwr", cols=["pwd_pass"],
            where=[("pwd_cono", "=", 0), ("pwd_sys", "=", self.system),
            ("pwd_code", "=", "Super")], limit=1)
        if not sup:
            sup = self.sql.getRec("ctlpwr", cols=["pwd_pass"],
                where=[("pwd_cono", "=", self.conum), ("pwd_sys", "=",
                self.system), ("pwd_code", "=", "Super")], limit=1)
        if not sup:
            self.sup = None
        else:
            self.sup = b64Convert("decode", sup[0])

    def mainProcess(self):
        tit = ("%s" % self.desc,)
        if self.product:
            fld = (
                (("T",0,0,0),"ONA",3,"Product Group"),
                (("T",0,0,0),"ONA",20," Code "),
                (("T",0,1,0),"ONA",30,"Description"),
                (("T",0,2,0),"IHA",30,"Password","",
                    "","Y",self.doPwd,None,None,None))
        else:
            fld = (
                (("T",0,0,0),"IHA",30,"Password","",
                    "","Y",self.doPwd,None,None,None),)
        but = (("Cancel",None,self.doExit,1,None, None),)
        tnd = ((self.doEnd, "n"), )
        txt = (self.doExit, )
        self.pc = TartanDialog(self.mf, screen=self.screen, tops=True,
            title=tit, eflds=fld, butt=but, tend=tnd, txit=txt)
        if self.product:
            for x in range(3):
                self.pc.loadEntry("T", 0, x, data=self.product[x])

    def doPwd(self, frt, pag, r, c, p, i, w):
        if w not in (self.pwd, self.sup, self.mf.override):
            self.tries += 1
            if self.tries == 3:
                return "xt"
            else:
                return "Invalid Password"
        self.flag = "ok"
        return "nc"

    def doEnd(self):
        self.pc.closeProcess()

    def doExit(self):
        self.flag = "no"
        self.pc.closeProcess()

class BankImport(object):
    """
    Used for importing statement files

        mf     - The Mainframe object.
        impfmt - The format of the file to import:
                     O - OFX Open Finacial Exchange Format
                     Q - QIF format for Quiken.
                     S - Standard Bank Best file 133 bytes.
        bankac - The bank account number.
        dtefmt - The date format of the file:
                     A - CCYYMMDD
                     B - DDMMCCYY
                     C - MMDDCCYY or MMDDYY
    """
    def __init__(self, mf, impfmt, bankac, dtefmt):
        self.mf = mf
        self.impfmt = impfmt
        self.bankac = bankac
        self.dtefmt = dtefmt
        self.setVariables()
        self.selectFile()
        if self.fname:
            for fname in self.fname:
                if self.impfmt == "O":
                    if OFX:
                        try:
                            self.importOfxTool(fname)
                        except:
                            self.importOfxFile(fname)
                    else:
                        self.importOfxFile(fname)
                elif self.impfmt == "Q":
                    self.importQifFile(fname)
                elif self.impfmt == "S":
                    self.importStdFile(fname)

    def setVariables(self):
        self.trans = []
        self.fname = []

    def selectFile(self):
        if self.impfmt == "O":
            ftype = [("OFX Files", "*.ofx")]
        elif self.impfmt == "Q":
            ftype = [("QIF Files", "*.qif")]
        else:
            ftype = []
        if sys.platform == "win32":
            lastdir = os.path.join(self.mf.rcdic["wrkdir"], "bankdir")
        else:
            lastdir = os.path.join(self.mf.rcdic["wrkdir"], ".bankdir")
        if os.path.exists(lastdir):
            infle = open(lastdir, "r")
            name = infle.readline().rstrip()
            infle.close()
        else:
            name = self.mf.rcdic["wrkdir"]
        dialog = FileDialog(**{
            "parent": self.mf.body,
            "title": "Bank Account Number %s" % self.bankac,
            "initd": name,
            "ftype": ftype,
            "multi": True})
        self.fname = dialog.askopenfilename()
        if not self.fname:
            return
        if type(self.fname) == str:
            self.fname = list(self.fname)
        try:
            infle = open(lastdir, "w")
            infle.write(os.path.dirname(os.path.normpath(self.fname[0])))
            infle.close()
        except:
            pass
        if self.impfmt == "Q":
            for fname in self.fname:
                if fname and self.bankac and not fname.count(self.bankac):
                    showError(self.mf.body, "File Error %s" % fname,
                        "The File Name Must Contain the Bank Account Number.")
                    self.fname = []
                    break

    def importOfxTool(self, fname):
        parser = OFXTree()
        parser.parse(fname)
        ofx = parser.convert()
        if not ofx.statements[0].account.acctid.count(self.bankac):
            showError(self.mf.body, "File Error %s" % fname,
                "The File Must Contain the Bank Account Number.")
            return
        self.doClear()
        for trans in ofx.statements[0].transactions:
            self.date = CCD(trans.dtposted.strftime("%Y%m%d"), "D1", 10)
            if trans.memo and trans.name:
                self.memo = trans.memo
                self.payee = trans.name
            elif trans.memo:
                self.memo = trans.memo
                self.payee = trans.memo
            elif trans.name:
                self.memo = trans.name
                self.payee = trans.name
            refno = trans.fitid
            if len(refno) > 9:
                refno = refno[-9:]
            self.refno = CCD(refno, "Na", 9)
            self.amount = CCD(trans.trnamt, "SD", 13.2)
            self.addTrans()
        self.trans.sort()

    def importOfxFile(self, fname):
        infle = open(fname, "r")
        data = infle.readlines()
        infle.close()
        account = None
        start = False
        self.doClear()
        for line in data:
            line = line.strip()
            if not account and line.count("<ACCTID>"):
                account = self.removeTag("ACCTID", line)
                if not account.count(self.bankac):
                    showError(self.mf.body, "File Error %s" % fname,
                        "The File Must Contain the Bank Account Number.")
                    return
                continue
            if line.count("<BANKTRANLIST>"):
                # Beginning of Transactions
                start = True
                continue
            if not start:
                continue
            if line.count("<DTEFF>"):
                # Version 102 Date
                if self.date.work:
                    self.addTrans()
                    self.doClear()
                if self.getDate(self.removeTag("DTEFF", line)):
                    return
                continue
            if line.count("<DTPOSTED>"):
                # Version 200 Date
                if self.date.work:
                    self.addTrans()
                    self.doClear()
                if self.getDate(self.removeTag("DTPOSTED", line)):
                    return
                continue
            if line.count("<FITID>"):
                # Version 200 Reference
                ref = self.removeTag("FITID", line)
                if len(ref) > 9:
                    ref = ref[-9:]
                self.refno = CCD(ref, "Na", 9)
            elif line.count("<MEMO>"):
                # Version 102 Description
                self.memo = self.removeTag("MEMO", line, compress=True)
            elif line.count("<NAME>"):
                # Version 200 Description
                self.memo = self.removeTag("NAME", line, compress=True)
            elif line.count("<TRNAMT>"):
                # Transaction Amount
                self.amount = CCD(self.removeTag("TRNAMT", line), "SD", 13.2)
        if self.date.work:
            self.addTrans()
        self.trans.sort()

    def removeTag(self, tag, data, compress=False):
        data = data.replace("<%s>" % tag, "").replace("</%s>" % tag, "")
        if compress:
            newdata = data.split()
            data = ""
            for d in newdata:
                if not data:
                    data = d
                else:
                    data = "%s %s" % (data, d)
        return data

    def addTrans(self):
        if self.payee and not self.memo:
            self.memo = self.payee
        if self.memo and not self.payee:
            self.payee = self.memo
        self.trans.append([self.date.work, self.memo, self.refno.work,
            self.payee, self.amount.work])

    def importQifFile(self, fname):
        infle = open(fname, "r")
        data = infle.readlines()
        infle.close()
        self.doClear()
        for line in data:
            line = line.strip()
            if line[:1] == "^":
                self.trans.append([self.date.work, self.memo,
                    self.refno.work, self.payee, self.amount.work])
                self.doClear()
                continue
            if line[:1] == "!" and line[1:10] != "Type:Bank":
                showError(self.mf.body, "Import Error",
                    "Invalid Type %s" % line)
                self.trans = []
                return
            if line[:1] == "D":
                if self.getDate(line[1:]):
                    return
            elif line[:1] == "M":
                memo = line[1:].split()
                self.memo = ""
                for m in memo:
                    if not self.memo:
                        self.memo = m.strip()
                    else:
                        self.memo = "%s %s" % (self.memo, m.strip())
            elif line[:1] == "N":
                self.refno = CCD(line[1:], "Na", 9)
            elif line[:1] == "P":
                if line[1:2] == "#":
                    data = line.split()
                    self.refno = CCD(data[0][2:], "Na", 9)
                    for num, dat in enumerate(data[1:]):
                        if not num:
                            self.payee = dat
                        else:
                            self.payee = "%s %s" % (self.payee, dat)
                else:
                    self.payee = line[1:]
            elif line[:1] == "T":
                self.amount = CCD(line[1:].replace(",", ""), "SD", 13.2)
        self.trans.sort()

    def doClear(self):
        self.date = CCD(0, "d1", 10)
        self.memo = ""
        self.refno = CCD("", "Na", 9)
        self.payee = ""
        self.amount = CCD(0, "SD", 13.2)

    def getDate(self, dt):
        if dt.count("-"):
            dt = dt.split("-")
        elif dt.count("/"):
            dt = dt.split("/")
        error = False
        if type(dt) == list:
            if self.dtefmt == "A":                           # CCYYMMDD
                dte = (int(dt[0])*10000)+(int(dt[1])*100)+int(dt[2])
            elif self.dtefmt == "B":                         # DDMMCCYY
                dte = (int(dt[2])*10000)+(int(dt[1])*100)+int(dt[0])
            elif self.dtefmt == "C" and len(dt[2]) == 4:     # MMDDCCYY
                dte = (int(dt[2])*10000)+(int(dt[0])*100)+int(dt[1])
            elif self.dtefmt == "C" and len(dt[2]) == 2:     # MMDDYY
                dte = ((int(dt[2])+2000)*10000)+(int(dt[0])*100)+int(dt[1])
            else:
                error = True
        else:
            if self.dtefmt == "A" and len(dt) in (8, 14):    # CCYYMMDD
                dte = (int(dt[:4])*10000)+(int(dt[4:6])*100)+int(dt[6:8])
            elif self.dtefmt == "B" and len(dt) in (8, 14):  # DDMMCCYY
                dte = (int(dt[4:8])*10000)+(int(dt[2:4])*100)+int(dt[:2])
            elif self.dtefmt == "C" and len(dt) in (8, 14):  # MMDDCCYY
                dte = (int(dt[4:8])*10000)+(int(dt[:2])*100)+int(dt[2:4])
            elif self.dtefmt == "C" and len(dt) in (6, 12):  # MMDDYY
                dte = ((int(dt[4:6])+2000)*10000)+(int(dt[:2])*100)+int(dt[2:4])
            else:
                error = True
        if not error:
            self.date = CCD(dte, "D1", 10)
        if error or self.date.err:
            showError(self.mf.body, "Import Error", "Invalid Date Format")
            self.trans = []
            return "error"

    def importStdFile(self, fname):
        infle = open(fname, "r")
        data = csv.reader(infle, quoting=csv.QUOTE_MINIMAL)
        err = None
        for dat in data:
            if len(dat) != 8:
                err = "Number of Fields Error (S/B %s is %s)" % (8, len(dat))
                break
            if dat[2] in ("BRANCH", "OPEN", "CLOSE"):
                continue
            if dat[2] in ("ACC-NO", "ACC NO"):
                if int(self.bankac) == int(dat[1]):
                    continue
                err = "Invalid Bank Account Number"
                break
            date = CCD(int(dat[1]), "D1", 10)
            err = date.err
            if err:
                break
            text = dat[4] + dat[5]
            text = text.split()
            memo = ""
            for m in text:
                if not memo:
                    memo = m.strip()
                else:
                    memo = "%s %s" % (memo, m.strip())
            try:
                ref = int(dat[2])
            except:
                ref = ""
            refno = CCD(ref, "Na", 9)
            err = refno.err
            if err:
                break
            payee = dat[5].strip()
            amount = CCD(dat[3], "SD", 17.2)
            err = amount.err
            if err:
                break
            self.trans.append([date.work, memo, refno.work, payee,
                amount.work])
        if err:
            showError(self.mf.body, "Import Error", err)
            self.trans = []

class RepPrt(object):
    """
    This is a Report Printer, following are the required parameters:

    name    :   The report or program name
    tables  :   A list of lists of tables or select statements with as clause
                e.g. ["ombank", "omlegr"] or
                     [select sum(glt_tramt) from gentrn as value"]
                                        OR
                A list of lists of data. 'ttype' option below must be 'D'
    joins   :   A list of tables to join as follows:
                    Type   : I = Inner join
                             L = Left Outer join
                             R = Right Outer join
                             F = Full Outer join
                    Table  : Table to join
                    Conditions: String statement of columns to use for the join
                e.g. ["L", "gentrn", "glt_cono=glm_cono and glt_acno=glm_acno"]
    heads   :   A list of up to 3 Headings, if heading is a list or tuple then
                it comprises text, alignment and a font size but only for
                heads[0] if no page numbering.
                Examples:
                    ["Old Mutual Bank Account", "L"]
                    [("Old Mutual Bank Account", 12, "C"),]
    cols    :   A list of lists of columns to be printed, any expression MUST
                be renamed using the "as" clause, as follows:
                    name - The column or expression name
                    type - The data type as per CCD
                    size - The data display size as per CCD
                    head - The column heading
                    disp - The disp field, either y or n indicating whether or
                         not to display the column or just use it for selection
                         purposes. It defaults to y.
                    chgs - Whether or not the sign must be changed.
                e.g. [["omb_state",             "NA",  5,   "State", "y"],
                      ["omb_date",              "D1", 10,   "Date",  "y"],
                      ["sum(omb_amt) as value", "SD", 13.2, "Value", "y"]]
                Use a blank list to print all columns of a table.
                                        OR
                A single list of columns to be printed where the column data
                is extracted from the ffield table.
                                        OR
                A combination of the two options.
    stots   :   A list of lists of columns, in least significant order, to be
                breaks for sub-totalling as follows:
                [] = No sub-totals
                                        OR
                    column or "as" name
                    description or another column's contents
                    flag denoting new page after each sub-total (Y/N)
                e.g. [["omb_state","Total Statement Value","Y"],]
                Note: stots depends on gtots, without gtots, stots are ignored
    gtots   :   A list of columns to be totalled containing:
                    column or "as" name e.g. ["value"]
                                            OR
                    A blank list for no totals.
    summ    :   Whether or not to print a summary of stots. Defaults to True
    where   :   A list of tuples of where statements e.g.
                [("omb_recon", "=", "Y"), ("omb_recon", "=", "Y")]
                None = No where statement
    group   :   A string for the "group" statement
                e.g. "omb_date, omb_state, omb_recon"
                None = No group statement
    order   :   A string of the "order by" statement e.g.
                "omb_state asc, omb_date desc"
                None = No order statement
    pdffl   :   The name to override the default name of the pdf file
    opts    :   A string indicating the options selected e.g. "Sort-D Csv-Y"
                None = No Options
    conum   :   An integer for the Company Number e.g. 001
                None = No Company
    conam   :   A string indicating the Company Name e.g. "Tartan Systems"
                None = No Company Name
    trtp    :   A list containing transaction types e.g. ["crt_type", crtrtp]
    ttype   :   A string indicating what the tables variable contains:
                    "T" = Containing Table Names
                    "D" = Containing lists of Data
    prtdia  :   Whether to have a printer and mail dialog, default is False or
                   output and mail parameters e.g. [("Y","V"), ("Y","N")]
    repprt  :   A list as derived from TartanDialog for print options
    repeml  :   A list as derived from TartanDialog for email options
    fromad  :   The senders email address or None for the default company one
    lines   :   A number of body lines to Print Between Page Breaks
    margin  :   A number of lines to use as a Bottom Margin for Page Breaks
    pbar    :   Progress Bar Message:
                    "P" = Generating a Report OR
                    "F" = Updating Files OR
                    Any Other String OR
                    None for No Progressbar
    refmt   :   Reformat the data as per the column information, default True
    blank   :   Blank zero values (only if refmt = True)
    pages   :   Whether to number pages, defaults to True
    tails   :   List of text to print at the end of the report
    wrkdir  :   The directory to use for temporary files
    """
    def __init__(self, mf, **args):
        self.mf = mf
        defaults = {
            "blank": False,
            "cols": [],
            "conam": None,
            "conum": None,
            "fromad": None,
            "group": None,
            "gtots": [],
            "heads": [],
            "joins": [],
            "lines": None,
            "margin": 5,
            "name": "report",
            "opts": None,
            "order": None,
            "pages": True,
            "pbar": "P",
            "pdffl": None,
            "prtdia": False,
            "refmt": True,
            "repeml": None,
            "repprt": None,
            "stots": [],
            "summ": True,
            "tables": [],
            "tails": [],
            "trtp": None,
            "ttype": "T",
            "where": [],
            "wrkdir": self.mf.rcdic["wrkdir"]}
        for nam in args:
            defaults[nam] = args[nam]
        for nam in defaults:
            setattr(self, nam, defaults[nam])
        if not self.prtdia and not self.repprt:
            self.repprt = ["N", "V", "view"]
        self.fonts = []
        self.align = []
        if self.heads:
            heads = []
            for hdr in self.heads:
                if type(hdr) in (tuple, list):
                    font = False
                    align = False
                    for i, h in enumerate(hdr):
                        if not i:
                            heads.append(h)
                        elif type(h) == int:
                            self.fonts.append(h)
                            font = True
                        else:
                            self.align.append(h)
                            align = True
                    if not font:
                        self.fonts.append(None)
                    if not align:
                        self.align.append("L")
                else:
                    heads.append(hdr)
                    self.fonts.append(None)
                    self.align.append("L")
            self.heads = heads
            if self.prtdia:
                self.doPrtDialog()
            else:
                self.xits = False
            if not self.xits:
                self.setVariables()
                self.processReport()

    def doPrtDialog(self):
        tit = ("Printer/Email Selection",)
        self.pd = TartanDialog(self.mf, tops=True, title=tit, eflds=[],
            tend=((self.doPrtEnd,"y"),), txit=(self.doPrtExit,),
            view=self.prtdia[0], mail=self.prtdia[1])
        self.pd.mstFrame.wait_window()

    def doPrtEnd(self):
        self.xits = False
        self.repprt = copyList(self.pd.repprt)
        self.repeml = copyList(self.pd.repeml)
        self.pd.closeProcess()

    def doPrtExit(self):
        self.xits = True
        self.pd.closeProcess()

    def setVariables(self):
        if not self.gtots:
            self.stots = []
        if self.ttype == "T":
            alltab = copyList(self.tables)
            if self.joins:
                for join in self.joins:
                    alltab.append(join[1])
            if not self.cols:
                for tab in alltab:
                    tb = self.removeTables(tab)
                    sql = Sql(self.mf.dbm, tb, prog=__name__)
                    dic = getattr(sql, "%s_dic" % tb)
                    col = getattr(sql, "%s_col" % tb)
                    for c in col:
                        self.cols.append([c, dic[c][2], dic[c][3], dic[c][5]])
            else:
                dic = {}
                for tab in alltab:
                    tb = self.removeTables(tab)
                    sql = Sql(self.mf.dbm, tb, prog=__name__)
                    dic = dict(list(dic.items()) + \
                        list(getattr(sql, "%s_dic" % tb).items()))
                for n, c in enumerate(self.cols):
                    if type(c) == str:
                        self.cols[n] = [c, dic[c][2], dic[c][3], dic[c][5]]
        else:
            self.joins = []

    def processReport(self):
        self.setConstants()
        self.setTableFields()
        if self.ttype == "T":
            recs = self.getData()
        else:
            recs = self.tables
        if self.stots:
            # Sub total descriptions
            self.sdic = {}
            for s in self.stots:
                self.sdic[s[0]] = {}
                for rec in recs:
                    idx = rec[self.l3.index(s[0])]
                    if s[1] in self.l3:
                        des = rec[self.l3.index(s[1])]
                    else:
                        des = s[1]
                    self.sdic[s[0]][idx] = des
        if recs:
            self.formLayout()
            self.printReport(recs)
        else:
            showError(self.mf.body, "Error", "No Available Records")
            self.xits = True

    def setConstants(self):
        self.tot = "s"
        self.subsum = []
        self.gndsum = []

    def setTableFields(self):
        self.fd = {}                # dictionary of column type, size & head
        self.l1 = ""                # string of all columns
        self.l3 = []                # list of all columns
        self.o1 = ""                # string of all order columns
        self.pr = []                # list of all display columns
        self.cg = []                # list of all sign change columns
        self.gp = []                # list of all group columns
        func = 0
        for num, col in enumerate(self.cols):
            agg = chkAggregate(col[0])
            if agg and col[1][1:] in ("D", "d"):
                if type(col) == tuple:
                    col = list(col)
                if col[0].count("Sum(") and not col[0].count("Round("):
                    col[0] = col[0].replace("Sum(", "Round(Sum(")
                elif col[0].count("sum(") and not col[0].count("round("):
                    col[0] = col[0].replace("sum(", "round(sum(")
                b = col[0].find(")") + 1
                d = int(round(col[2] * 10) % 10)
                col[0] = col[0][:b] + ",%s)" % d + col[0][b:]
            else:
                d = 0
            if num:
                self.l1 = self.l1 + "," + col[0]
            else:
                self.l1 = self.l1 + col[0]
            nam = removeFunctions(col[0], d)
            c = nam.split(" as ")
            nam = c[len(c) - 1]
            if len(c) == 1:
                func += 1
                if func != 1:
                    self.o1 = self.o1 + "," + nam
                else:
                    self.o1 = self.o1 + nam
            self.l3.append(nam)
            self.fd[nam] = [col[1], col[2], col[3]]
            if len(col) == 4:
                self.pr.append(nam)
            if len(col) > 4 and col[4].lower() == "y":
                self.pr.append(nam)
            if len(col) > 5 and col[5].lower() == "y":
                self.cg.append(nam)
            if not agg:
                self.gp.append(nam)

    def formLayout(self):
        fn = self.pr
        cf = ""
        hf = ""
        tf = ""
        hd = []
        sd = []
        ds = 0
        self.dend = False
        self.expforms = []
        self.expcolsh = [[]]
        for x, nam in enumerate(fn):
            if self.trtp and nam == self.trtp[0]:
                fsiz = 3
                d = "Typ"
                cf = cf + '"%s",'
                hf = hf + "%-3s "
                hd.append(d)
                self.expcolsh[0].append(("Typ", ))
                self.expforms.append(("NA", fsiz))
            else:
                fsiz = int(self.fd[nam][1])
                t = self.fd[nam][0][1]
                if t == "A":
                    cf = cf + '"%s"'
                else:
                    cf = cf + "%s"
                hf = hf + "%-" + "%d" % fsiz + "s"
                if x != (len(fn) - 1):
                    cf = cf + ","
                    hf = hf + " "
                else:
                    cf = cf + "\n"
                d = str(self.fd[nam][2])
                if not d:
                    n = nam.split("_", 1)
                    d = n[len(n)-1]
                d = d[:fsiz]
                if self.fd[nam][0] == "UI" or self.fd[nam][0] == "UD":
                    d = d.rjust(fsiz)
                elif self.fd[nam][0] == "SI" or self.fd[nam][0] == "SD":
                    d = d.rjust(fsiz - 1)
                hd.append(d)
                self.expcolsh[0].append((d,))
                self.expforms.append((self.fd[nam][0], self.fd[nam][1]))
            if self.gtots:
                cnt = 0
                for c in range(0, len(self.gtots)):
                    if nam == self.gtots[c]:
                        sd.append(d)
                        if ds == 0:
                            if x == 0:
                                # First field is a Total therefore the
                                # Description 'Grand Total' must be
                                # appended after the last field.
                                self.dend = True
                        else:
                            tf = tf + "%-" + "%d" % (ds - 1) + "s "
                            ds = 0
                        tf = tf + "%-" + "%d" % fsiz + "s"
                        if c != (len(self.gtots) - 1):
                            tf = tf + " "
                        cnt += 1
                if not cnt:
                    sd.append("")
                    ds = ds + fsiz + 1
        self.h5hf = hf
        self.head4 = hf % tuple(hd)
        if sd:
            self.head5 = hf % tuple(sd)
        else:
            self.head5 = self.head4
        self.tf = tf
        self.cf = cf
        if len(self.heads) == 4:
            self.head0 = self.heads.pop(0)
        elif self.conum:
            self.head0 = "%03i %-30s" % (self.conum, self.conam)
        else:
            self.head0 = None
        self.head1 = "%-s" % self.heads[0]
        self.head2 = ""
        self.head3 = ""
        if len(self.heads) > 1:
            self.head2 = (self.heads[1])
        if len(self.heads) == 3:
            self.head3 = (self.heads[2])
        self.expheads = []
        for x in range(4):
            if x == 1:
                h = self.heads[0]
            else:
                h = getattr(self, "head%s" % x)
            if not h:
                continue
            if self.expheads:
                self.expheads.append("")
            self.expheads.append(h)
        if self.opts:
            self.expheads.append("")
            self.expheads.append("Options: %s" % self.opts)
        self.expheads.append("")

    def getData(self):
        # Check table relationships
        rels = []
        for t1 in self.tables:
            for t2 in self.tables:
                if t2 == t1:
                    continue
                sql = Sql(self.mf.dbm, "frelat", prog=__name__)
                recs = sql.getRec("frelat", cols=["rel_col1",
                    "rel_col2"], where=[("rel_tab1", "=", t1),
                    ("rel_tab2", "=", t2)])
                if not recs:
                    recs = sql.getRec("frelat", cols=["rel_col1",
                        "rel_col2"], where=[("rel_tab1", "=", t2),
                        ("rel_tab2", "=", t1)])
                if recs:
                    for rel in recs:
                        if rels.count(rel) == 0:
                            rels.append(rel)
        if len(self.tables) > 1 and not rels:
            showError(self.mf.body, "Relationship Error",
                "Some Relationships Between Tables %s Missing" % self.tables)
            return "error"
        for rel in rels:
            self.where.append(("%s=%s" % (rel[0], rel[1]),))
        # Create cols
        cols = []
        for col in self.cols:
            cols.append(col[0])
        # Create join statements
        join = ""
        for jon in self.joins:
            if jon[0] == "I":
                join = "%s inner join %s on %s," % (join, jon[1], jon[2])
            elif jon[0] == "L":
                join = "%s left outer join %s on %s," % (join, jon[1], jon[2])
            elif jon[0] == "R":
                join = "%s right outer join %s on %s," % (join, jon[1], jon[2])
            elif jon[0] == "F":
                join = "%s full outer join %s on %s," % (join, jon[1], jon[2])
        if not self.group and self.stots and self.o1:
            self.group = self.o1
        if self.group:
            for col in self.gp:
                if not self.group.count(col):
                    self.group = "%s, %s" % (self.group, col)
        if not self.order and self.o1:
            order = self.o1
        else:
            order = self.order
        sql = Sql(self.mf.dbm, self.tables, prog=__name__)
        return sql.getRec(tables=self.tables, cols=cols, join=join,
            where=self.where, group=self.group, order=order)

    def printReport(self, recs):
        if self.pbar:
            pb = ProgressBar(self.mf.body, mxs=len(recs), typ=self.pbar,
                esc=True)
        if self.pdffl:
            self.pdfnam = os.path.join(self.wrkdir, self.pdffl + ".pdf")
            self.expnam = os.path.join(self.wrkdir, self.pdffl)
        else:
            if not self.conum:
                coy = 0
            else:
                coy = self.conum
            self.pdfnam = getModName(self.wrkdir, self.name, coy, ext="pdf")
            self.expnam = self.pdfnam.replace(".pdf", "")
        for sub in self.stots:
            setattr(self, "%s_str" % sub[0], "")
        for tot in self.gtots:
            setattr(self, "%s_gtot" % tot, 0)
            idx = 0
            for sub in self.stots:
                setattr(self, "%s_%s_stot" % (tot, idx), 0)
                idx += 1
        self.pglin = 999
        self.expdatas = []
        if self.repprt[2] != "export":
            self.fpdf = MyFpdf(name=self.name, head=len(self.head4))
            for num, font in enumerate(self.fonts):
                if not font:
                    self.fonts[num] = self.fpdf.font[1]
            if self.lines:
                lin = 4                         # head0 and head4
                for h in self.heads:
                    if h:
                        lin = lin + 2
                lpp = self.lines + lin          # lin = Headings
                if self.margin:
                    lpp = lpp + self.margin
                if lpp <= self.fpdf.lpp:
                    self.fpdf.lpp = lpp
        for x, rec in enumerate(recs):
            if self.pbar:
                pb.displayProgress(x)
                if pb.quit:
                    break
            if self.ttype == "D" and rec == ["BLANK"]:
                if self.repprt[2] == "export":
                    self.expdatas.append(rec)
                else:
                    self.fpdf.drawText()
                    self.pglin += 1
                continue
            fn = self.l3
            num = 0
            fld = []
            cld = []
            ldic = {}
            for nam in fn:
                t = self.fd[nam][0]
                if t == "TX" and self.repprt[2] != "export":
                    t = "NA"
                s = self.fd[nam][1]
                if not rec[num]:
                    d = ""
                else:
                    d = rec[num]
                if t == "HA":
                    d = b64Convert("decode", rec[num])
                if self.refmt:
                    if self.trtp and nam == self.trtp[0]:
                        d = self.trtp[1][int(d) - 1][0]
                        ldic[nam] = CCD(d, "NA", 3)
                    else:
                        if nam in self.cg:
                            if not d:
                                d = 0
                            elif type(d) == int:
                                d = 0 - d
                            else:
                                d = float(ASD(0) - ASD(d))
                        if self.blank and t[1] in ("D", "I") and not d:
                            ldic[nam] = CCD("", "NA", s)
                        else:
                            ldic[nam] = CCD(d, t, s)
                    if nam in self.pr:
                        fld.append(ldic[nam].disp)
                        cld.append(ldic[nam].work)
                elif nam in self.pr:
                    fld.append(d)
                    cld.append(d)
                num += 1
            st = "n"
            for sq1 in range((len(self.stots) - 1), -1, -1):
                w = getattr(self, "%s_str" % self.stots[sq1][0])
                f = ldic[self.stots[sq1][0]].work
                if not w:
                    w = f
                if f != w and st == "n":
                    st = "y"
                    for sq2 in range(0, sq1+1):
                        self.subTotal(self.stots[sq2])
                setattr(self, "%s_str" % self.stots[sq1][0], f)
            if self.repprt[2] != "export" and self.pglin >= self.fpdf.lpp:
                self.pageHeading()
            for gt in self.gtots:
                gtot = getattr(self, "%s_gtot" % gt)
                if self.fd[gt][0][1] == "D":
                    gtot = float(ASD(gtot) + ASD(ldic[gt].work))
                else:
                    gtot = gtot + ldic[gt].work
                setattr(self, "%s_gtot" % gt, gtot)
                idx = 0
                for s in self.stots:
                    stot = getattr(self, "%s_%s_stot" % (gt, idx))
                    if self.fd[gt][0][1] == "D":
                        stot = float(ASD(stot) + ASD(ldic[gt].work))
                    else:
                        stot = stot + ldic[gt].work
                    setattr(self, "%s_%s_stot" % (gt, idx), stot)
                    idx += 1
            if self.repprt[2] == "export":
                self.expdatas.append(["BODY", cld])
            else:
                self.fpdf.drawText(txt=self.h5hf % tuple(fld))
                self.pglin += 1
        if self.pbar:
            pb.closeProgress()
            if pb.quit:
                return
        self.grandTotal()
        try:
            if self.repprt[2] == "export":
                doWriteExport(rcdic=self.mf.rcdic, xtype=self.repprt[1],
                    name=self.expnam, heads=self.expheads, colsh=self.expcolsh,
                    forms=self.expforms, datas=self.expdatas)
            else:
                if self.tails:
                    self.fpdf.setFont(style="B")
                    for tail in self.tails:
                        self.fpdf.drawText(txt=tail)
                self.fpdf.output(self.pdfnam)
                doPrinter(mf=self.mf, conum=self.conum, pdfnam=self.pdfnam,
                    header=self.heads[-1], fromad=self.fromad,
                    repprt=self.repprt, repeml=self.repeml)
        except:
            pass

    def pageHeading(self, htype=None):
        if self.repprt[2] == "export":
            return
        self.pglin = 0
        self.fpdf.add_page()
        self.fpdf.setFont(style="B")
        if self.head0:
            self.fpdf.drawText(txt=self.head0)
            self.fpdf.drawText()
            self.pglin += 2
        self.fpdf.setFont(style="B", size=self.fonts[0])
        self.fpdf.drawText(txt=self.head1, align=self.align[0])
        self.fpdf.drawText()
        self.pglin += 2
        if self.head2:
            self.fpdf.setFont(style="B", size=self.fonts[1])
            self.fpdf.drawText(txt=self.head2, align=self.align[1])
            self.fpdf.drawText()
            self.pglin += 2
        if self.head3:
            self.fpdf.setFont(style="B", size=self.fonts[2])
            self.fpdf.drawText(txt=self.head3, align=self.align[2])
            self.fpdf.drawText()
            self.pglin += 2
        if self.opts:
            self.fpdf.setFont(style="B")
            self.fpdf.drawText(txt="%-10s%s" % ("Options:", self.opts))
            self.fpdf.drawText()
            self.pglin += 2
        self.fpdf.setFont(style="B")
        if not htype:
            self.fpdf.drawText(txt=self.head4)
        elif self.head5[:20] == " " * 20:
            head = "%s%s" % ("Totals Summary", self.head5[14:])
            self.fpdf.drawText(txt=head)
        else:
            self.fpdf.drawText(txt="Totals Summary")
            self.fpdf.drawText(txt=self.head5)
            self.pglin += 1
        self.pglin += 1
        self.fpdf.underLine(t="S")
        self.pglin += 1
        self.fpdf.setFont()

    def subTotal(self, col):
        if self.repprt[2] == "export":
            self.expdatas.append(["ULINES"])
        else:
            self.fpdf.setFont(style="B")
            self.fpdf.underLine(t="S")
            self.fpdf.setFont()
            self.pglin += 1
        snum = self.stots.index(col)
        sdet = getattr(self, "%s_str" % col[0])
        self.printTotals("s", snum, sdet)
        if self.repprt[2] != "export" and col[2] == "Y":
            self.pglin = 999
        elif self.tot != "g":
            if self.repprt[2] == "export":
                self.expdatas.append(["BLANK", [""]])
            else:
                self.fpdf.drawText()
                self.pglin += 1

    def grandTotal(self):
        self.tot = "g"
        for stot in self.stots:
            self.subTotal(stot)
        if self.gtots:
            if self.repprt[2] == "export":
                self.expdatas.append(["ULINES"])
            else:
                self.fpdf.setFont(style="B")
                self.fpdf.underLine(t="S")
            self.printTotals("g")
            if self.repprt[2] == "export":
                self.expdatas.append(["ULINED"])
            else:
                self.fpdf.underLine(t="D")
                self.fpdf.setFont()
                self.pglin += 2
        if self.repprt[2] != "export" and self.stots and self.summ:
            self.printSummary()

    def printTotals(self, typ, snum=None, sdet=None):
        if typ == "h":
            clr = flg = "n"
            if self.stots:
                typ = "s"
            else:
                typ = "g"
        else:
            clr = flg = "y"
        fld = []
        cld = []
        seq = 0
        for s, n in enumerate(self.pr):
            cldf = False
            for c in self.gtots:
                if n == c:
                    cldf = True
                    if typ == "s":
                        j = CCD(getattr(self, "%s_%s_stot" % (n, snum)),
                            self.fd[n][0], float(self.fd[n][1]))
                    else:
                        j = CCD(getattr(self, "%s_gtot" % n),
                            self.fd[n][0], float(self.fd[n][1]))
                    if not j.err:
                        fld.append(j.disp)
                        cld.append(j.work)
                        if j.work != 0:
                            flg = "y"
                    else:
                        fld.append("0.00")
                        cld.append(0.00)
                    if clr == "y":
                        if typ == "s":
                            setattr(self, "%s_%s_stot" % (n, snum), 0)
                        else:
                            setattr(self, "%s_gtot" % n, 0)
                    if c == self.gtots[-1]:
                        seq = 0
                    else:
                        seq = s
            if not cldf:
                if seq:
                    fld.append("")
                cld.append("")
        if self.tf and flg == "y":
            # Only print Totals if there are some Non Zero Totals
            if self.dend:
                # Total description at end of line
                if typ == "s":
                    dsc = self.sdic[self.stots[snum][0]][sdet] + " %s" % sdet
                else:
                    dsc = "Grand Total"
                siz = len(dsc)
                self.tf = self.tf + " %-" + "%d" % siz + "s"
                fld.append(dsc)
                cld.append(dsc)
            elif typ == "s":
                dsc = self.sdic[self.stots[snum][0]][sdet] + " %s" % sdet
                fld.insert(0, dsc)
                cld[0] = dsc
            else:
                fld.insert(0, "Grand Total")
                cld[0] = "Grand Total"
            if typ == "s":
                self.subsum.append(fld)
            else:
                self.gndsum.append(fld)
            if self.repprt[2] == "export":
                self.expdatas.append(["TOTAL", cld])
            else:
                self.fpdf.drawText(txt=self.tf % tuple(fld))
                self.pglin += 1

    def printSummary(self):
        if not self.subsum:
            return
        self.pageHeading("summary")
        for summ in self.subsum:
            self.fpdf.drawText(txt=self.tf % tuple(summ))
            self.pglin += 1
            if self.pglin >= self.fpdf.lpp:
                self.pageHeading("summary")
        self.fpdf.setFont(style="B")
        self.fpdf.underLine(t="S")
        self.fpdf.drawText(txt=self.tf % tuple(self.gndsum[0]))
        self.fpdf.underLine(t="D")
        self.fpdf.setFont()

    def removeTables(self, tab):
        n = tab.split(" as ")
        if len(n) == 1:
            return n[0]
        else:
            return n[1]

class TabPrt(object):
    """
    This is a front-end to the RepPrt class for printing table details
        mf     = The mainframe class
        conum  = The company number
        conam  = The company name
        name   = The calling program
        head   = The report heading
        tabs   = The table name or list of tables which must also have cols
        cols   = The columns to print else None for all
        where  = The where statement
        keys   = The Balances class key
        pdia   = The Print Dialog, defaults to True
        repprt = The Printer Details
        repeml = The Email Details
        export = Return the details
    """
    def __init__(self, mf, conum=None, conam=None, name=None, head=None, tabs=None, cols=None, where=None, keys=None, pdia=True, repprt=None, repeml=None, export=False):
        self.mf = mf
        self.conum = conum
        self.conam = conam
        self.name = name
        self.head = head
        if type(tabs) == str:
            self.tabs = [tabs]
        else:
            self.tabs = tabs
        self.cols = cols
        self.where = where
        self.keys = keys
        self.pdia = pdia
        self.repprt = repprt
        self.repeml = repeml
        self.export = export
        self.setVariables()
        self.printReport()

    def setVariables(self):
        self.sql = Sql(self.mf.dbm, self.tabs + ["ftable"], prog=__name__)
        self.dics = {}
        for tab in self.tabs:
            dic = getattr(self.sql, "%s_dic" % tab)
            for key in dic:
                self.dics[key] = dic[key]
        if not self.cols:
            self.cols = []
            for tab in self.tabs:
                col = getattr(self.sql, "%s_col" % tab)
                self.cols.extend(col)
        t = time.localtime()
        self.sysdt = ((t[0] * 10000) + (t[1] * 100) + t[2])
        self.curdt = int(self.sysdt / 100)
        if not self.head:
            self.head = self.sql.getRec("ftable", cols=["ft_desc"],
                where=[("ft_tabl", "=", self.tabs[0])], limit=1)[0]

    def printReport(self):
        recs = self.sql.getRec(tables=self.tabs, cols=self.cols,
            where=self.where, limit=1)
        if not recs:
            return
        max1, max2 = 0, 0
        fld = []
        for c in self.cols:
            if max1 < len(self.dics[c][4]):
                max1 = len(self.dics[c][4])
            if max2 < int(self.dics[c][3]):
                max2 = int(self.dics[c][3])
            if self.tabs.count("ctlmst") and c == "ctm_modules":
                mods = recs[self.cols.index("ctm_modules")]
                newm = ""
                for x in range(0, len(mods), 2):
                    if not newm:
                        newm = mods[x:x+2]
                    else:
                        newm += ",%s" % mods[x:x+2]
                fld.append([self.dics[c][4], newm])
            else:
                fld.append([self.dics[c][4], CCD(recs[self.cols.index(c)],
                    self.dics[c][2], self.dics[c][3]).disp])
        if self.tabs.count("crsmst") and self.keys:
            bals = Balances(self.mf, "CRS", self.conum, self.curdt, self.keys)
            more = bals.doAllBals()
        elif self.tabs.count("drsmst") and self.keys:
            bals = Balances(self.mf, "DRS", self.conum, self.curdt, self.keys)
            more = bals.doAllBals()
        elif self.tabs[0] == "strmf1" and self.keys:
            bals = Balances(self.mf, "STR", self.conum, self.curdt, self.keys)
            more = bals.doStrBals()
        else:
            more = None
        if more and (self.tabs.count("drsmst") or self.tabs.count("crsmst")):
            fld.append(["", ""])
            fld.append(["Current",  CCD(more[2][0], "SD", 13.2).disp])
            fld.append([" 30-Days", CCD(more[2][1], "SD", 13.2).disp])
            fld.append([" 60-Days", CCD(more[2][2], "SD", 13.2).disp])
            fld.append([" 90-Days", CCD(more[2][3], "SD", 13.2).disp])
            fld.append(["120-Days", CCD(more[2][4], "SD", 13.2).disp])
            fld.append(["", ""])
            fld.append(["Balance", CCD(more[1], "SD", 13.2).disp])
        elif more and self.tabs[0] == "strmf1":
            fld.append(["", ""])
            fld.append(["Last    Cost    ", CCD(more[6],    "SD", 13.2).disp])
            fld.append(["Average Cost    ", CCD(more[7],    "SD", 13.2).disp])
            fld.append(["", ""])
            fld.append(["Quantity Balance", CCD(more[5][0], "SD", 13.2).disp])
            fld.append(["Value    Balance", CCD(more[5][1], "SD", 13.2).disp])
        cols = [
            ["a", "NA", (max1+5), "Field Name"],
            ["b", "NA",     max2,    "Details"]]
        if not self.export and (self.pdia or self.repprt or self.repeml):
            if not self.pdia or type(self.pdia) in (list, tuple):
                prtdia = self.pdia
            else:
                prtdia = (("Y","V"),("Y","N"))
            rp = RepPrt(self.mf, name=self.name, tables=fld, heads=[self.head],
                cols=cols, conum=self.conum, conam=self.conam, prtdia=prtdia,
                ttype="D", repprt=self.repprt, repeml=self.repeml, refmt=False)
            self.repprt = rp.repprt
            self.repeml = rp.repeml
            self.xits = rp.xits
        else:
            self.repprt = None
            self.repeml = None
            self.xits = False
            self.mess = ""
            for f in fld:
                if not self.mess:
                    self.mess = "%-20s %30s" % (f[0], f[1])
                else:
                    self.mess = "%s\n%-20s %30s" % (self.mess, f[0], f[1])
            if not self.export:
                ScrollText(scrn=self.mf.body, title=self.head, mess=self.mess,
                    width=(max1 + 5 + max2))

class TartanUser(object):
    """
    This class is used to add, delete and amend Tartan users
    """
    def __init__(self, mf, usr, sys, mod):
        self.mf = mf
        self.user = usr
        self.syss = sys
        self.mod = mod
        if self.setVariables():
            self.doUser()
            self.mf.startLoop()

    def setVariables(self):
        self.sql = Sql(self.mf.dbm, ["ctlsys", "ctlmst", "ctlpwu", "ctlpwm"],
            prog=__name__)
        ctlsys = self.sql.getRec("ctlsys", limit=1)
        self.pwmust, self.pwsize, self.pwlife = ctlsys[1:4]
        return True

    def doUser(self):
        susr = {
            "stype": "R",
            "tables": ("ctlpwu",),
            "cols": (
                ("usr_name", "", 0, "User-Name"),
                ("usr_fnam", "", 0, "Full-Name", "Y"),
                ("usr_acoy", "", 0, "Allowed-Companies"),
                ("usr_dcoy", "", 0, "Disallowed-Companies"),
                ("usr_lvl", "", 0, "L"))}
        scoy = {
            "stype": "R",
            "tables": ("ctlmst",),
            "cols": (
                ("ctm_cono", "", 0, "Coy"),
                ("ctm_name", "", 0, "Name", "Y"))}
        lvls = [
            [0, "= Enquiries Only"],
            [1, "= 0 plus Reporting"],
            [2, "= 1 plus Data Capture"],
            [3, "= 2 plus File Maintenance"],
            [4, "= 3 plus Month and Year End Routines"],
            [5, "= 4 plus Control Routines"],
            [6, "= 5 plus Financial Year End Routine"],
            [7, "= 6 plus User and Module Password Maintenance"],
            [8, "= 7 plus Update File Formats"],
            [9, "= Supervisor level, Everything"]]
        slvl = {
            "stype": "C",
            "titl": "Available Levels",
            "head": ("L", "Permissions"),
            "data": lvls}
        ssys = {
            "stype": "C",
            "titl": "Available Systems",
            "head": ("System", "Cd"),
            "data": self.syss,
            "index": 1}
        sprg = {
            "stype": "C",
            "titl": "Available Modules",
            "head": ("L", "Code", "Description"),
            "data": [],
            "index": 1}
        smod = {
            "stype": "R",
            "tables": ("ctlpwm",),
            "cols": (
                ("mpw_coy", "", 0, "Coy"),
                ("mpw_sys", "", 0, "Sy"),
                ("mpw_prg", "", 0, "Prog")),
            "whera": (("T", "mpw_usr", 0),),
            "order": "mpw_coy, mpw_sys, mpw_prg",
            "comnd": self.doSelRet}
        r1s = (("Yes", "Y"), ("No", "N"))
        fld = (
            (("T",0,0,0),"INA",20,"User Name","",
                "","Y",self.doUserName,susr,None,("notblank",)),
            (("T",0,1,0),"INA",30,"Full Name","",
                "","N",self.doFullName,susr,self.doDelUser,("notblank",)),
            (("T",0,2,0),"ITX",50,"Email Address","",
                "","N",self.doEmailAddr,None,None,("efld",)),
            (("T",0,3,0),"ITX",20,"Mobile Number","",
                "","N",self.doMobile,None,None,("efld",)),
            (("T",0,4,0),"IHA",30,"User Password","",
                "","N",self.doUserPwd1,None,None,None),
            (("T",0,5,0),"IHA",30,"Check Password ","",
                "","N",self.doUserPwd2,None,None,None),
            (("T",0,6,0),("IRB",r1s),0,"Copy Existing User","",
                "N","N",self.doCopyUser1,None,None,None),
            (("T",0,7,0),"INA",20,"Copy User Name","User Name to Copy",
                "","N",self.doCopyUser2,susr,None,None),
            (("T",0,8,0),"ITX",30,"Allowed Companies","",
                "All","N",self.doValCoy,None,None,None,None,
                "Comma Seperated List of Allowed Companies e.g. 1,2,3"),
            (("T",0,9,0),"ITX",30,"Disallowed Companies","",
                "None","N",self.doDisCoy,None,None,None,None,
                "Comma Seperated List of Disallowed Companies e.g. 1,2,3"),
            (("T",0,10,0),"IUI",1,"Security Level","",
                0,"N",self.doSecLvl,slvl,None,("between",0,9)),
            (("C",0,0,0),"IUI",3,"Coy","Company",
                "r","N",self.doCoy,scoy,None,None),
            (("C",0,0,1),"INA",2,"SS","System Code",
                "r","N",self.doSys,ssys,None,None),
            (("C",0,0,2),"INA",4,"Prog","Module",
                "","N",self.doMod,sprg,self.doDelSys,None),
            (("C",0,0,3),"IHA",20,"Password","",
                "","N",self.doPw1,None,self.doDelMod,None),
            (("C",0,0,4),"IHA",20,"Check Password","Re-Enter Password",
                "","N",self.doPw2,None,None,None))
        row = (15,)
        but = (
            ("Show All Modules",smod,None,0,("C",0,1),(("T",0,1),("C",0,2))),
            ("Quit",None,self.doUserExit,1,None,None))
        tnd = ((self.doWriteUser,"y"),)
        txt = (self.doUserExit,)
        cnd = ((self.doWriteModule,"y"),)
        cxt = (self.doUserNext,)
        self.df = TartanDialog(self.mf, eflds=fld, rows=row, butt=but,
            tend=tnd, txit=txt, cend=cnd, cxit=cxt)

    def doUserName(self, frt, pag, r, c, p, i, w):
        self.unam = w
        usr = self.sql.getRec("ctlpwu", where=[("usr_name", "=", w)], limit=1)
        if not usr:
            self.new_usr = "y"
            if self.unam == "admin":
                self.fnam = "System Administrator"
                self.df.loadEntry(frt, pag, p+1, data=self.fnam)
        else:
            self.new_usr = "n"
            self.fnam = usr[self.sql.ctlpwu_col.index("usr_fnam")]
            self.emadd = usr[self.sql.ctlpwu_col.index("usr_emadd")]
            self.celno = usr[self.sql.ctlpwu_col.index("usr_celno")]
            self.pwd = usr[self.sql.ctlpwu_col.index("usr_pwd")]
            self.acoy = usr[self.sql.ctlpwu_col.index("usr_acoy")]
            self.dcoy = usr[self.sql.ctlpwu_col.index("usr_dcoy")]
            self.lvl = usr[self.sql.ctlpwu_col.index("usr_lvl")]
            self.df.loadEntry(frt, pag, p+1, data=self.fnam)
            self.df.loadEntry(frt, pag, p+2, data=self.emadd)
            self.df.loadEntry(frt, pag, p+3, data=self.celno)
            self.df.loadEntry(frt, pag, p+4, data=b64Convert("decode",self.pwd))
            self.df.loadEntry(frt, pag, p+5, data=b64Convert("decode",self.pwd))
            if not self.acoy:
                self.df.loadEntry(frt, pag, p+8, data="All")
            else:
                self.df.loadEntry(frt, pag, p+8, data=self.acoy)
            if not self.dcoy:
                self.df.loadEntry(frt, pag, p+9, data="None")
            else:
                self.df.loadEntry(frt, pag, p+9, data=self.dcoy)
            self.df.loadEntry(frt, pag, p+10, data=self.lvl)
            self.doLoadMods(focus=False)
        if self.unam == "admin":
            return "sk1"

    def doDelUser(self):
        if self.unam == "admin" or self.unam == self.user:
            return "Not Allowed to Delete yourself"
        self.sql.delRec("ctlpwu", where=[("usr_name", "=", self.unam)])
        self.sql.delRec("ctlpwm", where=[("mpw_usr", "=", self.unam)])

    def doFullName(self, frt, pag, r, c, p, i, w):
        self.fnam = w

    def doEmailAddr(self, frt, pag, r, c, p, i, w):
        self.emadd = w

    def doMobile(self, frt, pag, r, c, p, i, w):
        self.celno = w

    def doUserPwd1(self, frt, pag, r, c, p, i, w):
        if not w and self.pwmust == "Y":
            return "You Must Have a Password"
        if self.pwsize and len(w) < self.pwsize:
            return "Minimum of %s Characters Required" % self.pwsize
        self.pwd = w

    def doUserPwd2(self, frt, pag, r, c, p, i, w):
        if w != self.pwd:
            return "Passwords Do Not Match"
        if self.unam == "admin":
            self.cpy = "N"
            self.acoy = ""
            self.dcoy = ""
            self.lvl = 9
            return "sk5"

    def doCopyUser1(self, frt, pag, r, c, p, i, w):
        self.cpy = w
        if self.cpy == "N":
            return "sk1"

    def doCopyUser2(self, frt, pag, r, c, p, i, w):
        usr = self.sql.getRec("ctlpwu", where=[("usr_name", "=", w)],
            limit=1)
        if not usr:
            return "Invalid User Name"
        self.cnam = w
        self.acoy = usr[self.sql.ctlpwu_col.index("usr_acoy")]
        self.dcoy = usr[self.sql.ctlpwu_col.index("usr_dcoy")]
        self.lvl = usr[self.sql.ctlpwu_col.index("usr_lvl")]
        if not self.acoy:
            self.df.loadEntry(frt, pag, p+1, data="All")
        else:
            self.df.loadEntry(frt, pag, p+1, data=self.acoy)
        if not self.dcoy:
            self.df.loadEntry(frt, pag, p+2, data="None")
        else:
            self.df.loadEntry(frt, pag, p+2, data=self.dcoy)
        self.df.loadEntry(frt, pag, p+3, data=self.lvl)

    def doValCoy(self, frt, pag, r, c, p, i, w):
        self.acoy = ""
        if w and w != "All":
            for ck in w.replace(" ", "").split(","):
                try:
                    chk = self.sql.getRec("ctlmst", where=[("ctm_cono",
                        "=", int(ck))], limit=1)
                    if not chk:
                        raise Exception
                    if self.acoy:
                        self.acoy = "%s,%s" % (self.acoy, ck)
                    else:
                        self.acoy = "%s" % ck
                except:
                    return "Invalid Company Number (%s)" % ck
            self.df.loadEntry(frt, pag, i, data=self.acoy)
        else:
            self.df.loadEntry(frt, pag, i, data="All")

    def doDisCoy(self, frt, pag, r, c, p, i, w):
        self.dcoy = ""
        if w and w != "None":
            for ck in w.replace(" ", "").split(","):
                try:
                    chk = self.sql.getRec("ctlmst", where=[("ctm_cono",
                        "=", int(ck))], limit=1)
                    if not chk:
                        raise Exception
                    if self.dcoy:
                        self.dcoy = "%s,%s" % (self.dcoy, ck)
                    else:
                        self.dcoy = "%s" % ck
                except:
                    return "Invalid Company Number (%s)" % ck
            self.df.loadEntry(frt, pag, i, data=self.dcoy)
        else:
            self.df.loadEntry(frt, pag, i, data="None")

    def doSecLvl(self, frt, pag, r, c, p, i, w):
        self.lvl = w

    def doWriteUser(self):
        if self.pwd:
            pwd = b64Convert("encode", self.pwd)
            t = time.localtime()
            dte = (t[0] * 10000) + (t[1] * 100) + t[2]
        else:
            pwd = ""
            dte = 0
        data = [self.unam, self.fnam, self.emadd, self.celno, pwd, dte,
            self.acoy, self.dcoy, self.lvl]
        if self.new_usr == "y":
            self.sql.insRec("ctlpwu", data=data)
        else:
            self.sql.updRec("ctlpwu", data=data, where=[("usr_name",
                "=", self.unam)])
        if self.unam == "admin":
            self.mf.dbm.commitDbase()
            self.df.focusField("T", 0, 1)
        else:
            if self.cpy == "Y":
                self.sql.delRec("ctlpwm", where=[("mpw_usr", "=", self.unam)])
                recs = self.sql.getRec("ctlpwm", where=[("mpw_usr", "=",
                    self.cnam)], order="mpw_coy, mpw_sys, mpw_prg")
                for rec in recs:
                    rec[0] = self.unam
                    self.sql.insRec("ctlpwm", data=rec)
            self.doLoadMods()

    def doCoy(self, frt, pag, r, c, p, i, w):
        if w and self.acoy:
            coys = self.acoy.split(",")
            if str(w) not in coys:
                return "Invalid Company for This User"
        if w and self.dcoy:
            coys = self.dcoy.split(",")
            if str(w) in coys:
                return "Invalid Company for This User"
        self.con = w

    def doSys(self, frt, pag, r, c, p, i, w):
        s = False
        for sss in self.syss:
            if w == sss[1]:
                s = True
        if not s:
            return "Invalid System"
        self.sys = w
        self.chk = []
        data = []
        for mod in self.mod[self.sys]:
            new = copyList(list(mod))
            if new[0] <= self.lvl:
                if new not in data:
                    data.append(new)
                    self.chk.append(new[1])
        self.df.colf[pag][2][8]["data"] = data

    def doMod(self, frt, pag, r, c, p, i, w):
        if w and w not in self.chk:
            return "Invalid Module"
        self.prg = w
        if not self.prg:
            if self.con:
                return "Invalid Module, Company Exists"
            else:
                self.pwd = ""
                self.df.loadEntry(frt, pag, p+1, data=self.pwd)
                self.df.loadEntry(frt, pag, p+2, data=self.pwd)
                return "sk2"
        pwd = self.getPassword()
        if pwd:
            self.df.loadEntry(frt, pag, p+1, data=pwd)
            self.df.loadEntry(frt, pag, p+2, data=pwd)

    def doPw1(self, frt, pag, r, c, p, i, w):
        self.pwd = w
        if not self.pwd:
            return "sk1"

    def doPw2(self, frt, pag, r, c, p, i, w):
        if w != self.pwd:
            return "Passwords do Not Match"

    def doSelRet(self, frt, pag, r, c, p, i, w):
        self.con = w[0]
        self.sys = w[1]
        self.prg = w[2]
        self.pwd = self.getPassword()
        self.df.loadEntry(frt, pag, p, data=self.con)
        self.df.loadEntry(frt, pag, p+1, data=self.sys)
        self.df.loadEntry(frt, pag, p+2, data=self.prg)
        self.df.loadEntry(frt, pag, p+3, data=self.pwd)
        self.df.loadEntry(frt, pag, p+4, data=self.pwd)

    def getPassword(self):
        pwd = self.sql.getRec("ctlpwm", cols=["mpw_pwd"],
            where=[("mpw_usr", "=", self.unam), ("mpw_coy", "=", self.con),
            ("mpw_sys", "=", self.sys), ("mpw_prg", "=", self.prg)],
            limit=1)
        if pwd:
            return b64Convert("decode", pwd[0])

    def doDelSys(self):
        self.sql.delRec("ctlpwm", where=[("mpw_usr", "=", self.unam),
            ("mpw_coy", "=", self.con), ("mpw_sys", "=", self.sys)])
        self.doLoadMods()

    def doDelMod(self):
        self.sql.delRec("ctlpwm", where=[("mpw_usr", "=", self.unam),
            ("mpw_coy", "=", self.con), ("mpw_sys", "=", self.sys),
            ("mpw_prg", "=", self.prg)])
        self.doLoadMods()

    def doWriteModule(self):
        whr = [("mpw_usr", "=", self.unam), ("mpw_coy", "=", self.con),
            ("mpw_sys", "=", self.sys), ("mpw_prg", "=", self.prg)]
        usr = self.sql.getRec("ctlpwm", where=whr, limit=1)
        if self.pwd:
            pwd = b64Convert("encode", self.pwd)
        else:
            pwd = ""
        dat = [self.unam, self.con, self.sys, self.prg, pwd]
        if not usr:
            self.sql.insRec("ctlpwm", data=dat)
        else:
            dat.append(usr[self.sql.ctlpwm_col.index("mpw_xflag")])
            self.sql.updRec("ctlpwm", data=dat, where=whr)
        self.doLoadMods()

    def doLoadMods(self, focus=True):
        self.df.clearFrame("C", 0)
        recs = self.sql.getRec("ctlpwm", where=[("mpw_usr", "=",
            self.unam)], order="mpw_coy, mpw_sys, mpw_prg")
        self.pos = 0
        for row, rec in enumerate(recs):
            for num, dat in enumerate(rec):
                if num in (0, 5):
                    continue
                self.df.loadEntry("C", 0, self.pos, data=dat)
                if num == 4:
                    self.df.loadEntry("C", 0, self.pos, data=dat)
                    self.pos += 1
                self.pos += 1
                if self.pos == (self.df.rows[0] * self.df.colq[0]):
                    self.df.scrollScreen(0)
                    self.pos = self.pos - self.df.colq[0]
        if focus:
            self.df.focusField("C", 0, (self.pos+1))

    def doUserNext(self):
        self.mf.dbm.commitDbase()
        self.df.focusField("T", 0, 1)

    def doUserExit(self):
        self.mf.dbm.commitDbase()
        self.df.closeProcess()
        self.mf.closeLoop()

class TarBckRes(object):
    """
    Backup and Restore routines for Tartan Systems

        mode - B(ackup), R(estore).
        csys - The ctlsys detail.
        ver  - The tartan version.
        pbar - Whether to display a progressbar,
    """
    def __init__(self, mf, mode="B", csys=None, ver=None, pbar=True):
        self.mf = mf
        self.mode = mode
        self.budays = 0
        self.smtp = False
        if csys:
            self.budays = csys[0]
            if csys[1] and sendMail(csys[1:], "", "", "", check=True,
                    wrkdir=self.mf.rcdic["wrkdir"]):
                self.smtp = csys[1:]
        self.ver = ver
        self.pbar = pbar
        if self.setVariables():
            if self.mode == "B":
                if self.mf.window and self.smtp and self.smtp[0]:
                    self.startBck()
                    self.mf.startLoop()
                else:
                    self.endBckRes()
            else:
                self.startRes()
                self.mf.startLoop()

    def setVariables(self):
        t = time.localtime()
        self.sysdtw = ((t[0] * 10000) + (t[1] * 100) + t[2])
        self.bupdir = os.path.join(self.mf.rcdic["bupdir"],
            self.mf.rcdic["dbname"])
        if self.mode == "R" and not os.path.isdir(self.bupdir):
            return
        self.archdir = os.path.join(self.bupdir, "arch")
        if self.mode == "R" and not os.path.isdir(self.archdir):
            return
        self.tmpdir = os.path.join(self.bupdir, "temp")
        # Get all company and non-company tables
        sql = Sql(self.mf.dbm, ["ftable", "ffield"])
        tab = sql.getRec("ftable", cols=["ft_tabl"], group="ft_tabl")
        self.yess = []
        self.nons = []
        for nam in tab:
            coy = sql.getRec("ffield", cols=["ff_name"], where=[("ff_tabl",
                "=", nam[0]), ("ff_seq", "=", 0)], limit=1)
            if coy[0].count("_cono"):
                self.yess.append(nam[0])
            else:
                self.nons.append(nam[0])
        self.yess.sort()
        self.nons.sort()
        # Get all companies
        sql = Sql(self.mf.dbm, "ctlmst", prog=__name__)
        if sql.error:
            return
        self.coys = sql.getRec("ctlmst", cols=["ctm_cono",
            "ctm_name", "ctm_email"], order="ctm_cono")
        self.incoac = False
        self.slsinv = False
        if self.coys:
            # Check for Intercompany Facility
            sql = Sql(self.mf.dbm, "genint", prog=__name__)
            if not sql.error:
                itg = sql.getRec("genint", cols=["count(*)"], limit=1)
                if itg[0]:
                    self.incoac = True
            # Check for Sales Invoicing
            sql = Sql(self.mf.dbm, "slsiv1", prog=__name__)
            if not sql.error:
                inv = sql.getRec("slsiv1", cols=["count(*)"], limit=1)
                if inv[0]:
                    self.slsinv = True
            self.cnum = []
            for coy in self.coys:
                self.cnum.append(coy[0])
            self.cnum.sort()
        # Load System Names and Check for Integration
        sysd = {
            "ASS": ["Asset's Register", "assctl", "cta_glint"],
            "BKM": ["Bookings Master", "bkmctl", "cbk_glint"],
            "BKS": ["Book Clubs"],
            "BWL": ["Bowling Clubs"],
            "CRS": ["Creditor's Ledger", "crsctl", "ctc_glint"],
            "CSH": ["Cash Analysis"],
            "DRS": ["Debtor's Ledger", "drsctl", "ctd_glint"],
            "GEN": ["General Ledger"],
            "LON": ["Loans Ledger", "lonctl", "cln_glint"],
            "MEM": ["Member's Ledger", "memctl", "mcm_glint"],
            "RCA": ["Rental's Ledger (Extended)", "rcactl", "cte_glint"],
            "RTL": ["Rental's Ledger (Basic)", "rtlctl", "ctr_glint"],
            "SCP": ["Sectional Competitions"],
            "STR": ["Store's Ledger", "strctl", "cts_glint"],
            "WAG": ["Salaries and Wages", "wagctl", "ctw_glint"]}
        syss = []
        self.sysi = []
        keys = list(sysd.keys())
        keys.sort()
        for s in keys:
            det = sysd[s]
            syss.append([s, det[0]])
            if len(det) == 1:
                continue
            sql = Sql(self.mf.dbm, det[1], prog=__name__)
            if sql.error:
                continue
            itg = sql.getRec(tables=det[1], cols=[det[2]])
            if not itg:
                continue
            for i in itg:
                if i[0] == "Y":
                    self.sysi.append(s)
                    break
        self.csys = []
        self.syst = []
        for sss in syss:
            self.csys.append(sss[0])
            self.syst.append([sss[0], sss[1]])
        self.full = "N"
        self.icoy = "Y"
        self.isys = "Y"
        self.sys = []
        return True

    def startBck(self):
        self.bu = TartanDialog(self.mf, eflds=[],
            tend=((self.endBckRes, "y"),), txit=(self.noBckRes,),
            mail=("Y","N","N","E-Mail Archive"))

    def startRes(self):
        dat = []
        dat = glob.glob(os.path.join(self.archdir, "*.tar"))
        dat.sort()
        fst = os.path.join(self.archdir, dat[-1])
        arc = {
            "stype": "F",
            "types": "fle",
            "initd": self.archdir,
            "ftype": (("Archives", "*.tar"),)}
        data = []
        for coy in self.coys:
            data.append(coy[:2])
        coy = {
            "stype": "C",
            "titl": "",
            "head": ("Num", "Name"),
            "typs": (("UI",3), ("NA",30)),
            "data": data,
            "mode": "M",
            "comnd": self.doCoyCmd}
        sys = {
            "stype": "C",
            "titl": "",
            "head": ("Cod", "Description"),
            "typs": (("UA",3), ("NA",30)),
            "data": self.syst,
            "mode": "M",
            "comnd": self.doSysCmd}
        r1s = (("Full","Y"),("Partial","N"))
        r2s = (("Yes","Y"),("Include","I"),("Exclude","E"))
        fld = (
            (("T",0,0,0),("IRB",r1s),0,"Type","",
                "Y","N",self.doType,None,None,None),
            (("T",0,1,0),"IFF",50,"Archive","",
                fst,"Y",self.doPlace,arc,None,None),
            (("T",0,2,0),("IRB",r2s),0,"All Companies","",
                "Y","N",self.doAllCoy,None,None,None),
            (("T",0,3,0),"INA",(30,100),"Companies","",
                "","N",self.doCoySel,coy,None,None),
            (("T",0,4,0),("IRB",r2s),0,"All Systems","",
                "Y","N",self.doAllSys,None,None,None),
            (("T",0,5,0),"INA",(50,100),"Systems","",
                "","N",self.doSysSel,sys,None,None))
        self.bu = TartanDialog(self.mf, eflds=fld,
            tend=((self.endBckRes, "y"),), txit=(self.noBckRes,))

    def doType(self, frt, pag, r, c, p, i, w):
        self.full = w

    def doPlace(self, frt, pag, r, c, p, i, w):
        if not os.path.isfile(os.path.join(self.archdir, w)):
            return "Invalid Archive"
        arcfle = w
        try:
            # Delete existing directory and recreate it
            if os.path.exists(self.tmpdir):
                shutil.rmtree(self.tmpdir)
            os.makedirs(self.tmpdir)
        except:
            showException(self.mf.body, self.mf.rcdic["wrkdir"],
                "Permission Error", dbm=self.mf.dbm)
            return "Invalid Restore Directory"
        cwd = os.getcwd()
        os.chdir(self.tmpdir)
        tarfle = tarfile.open(os.path.join(self.archdir, arcfle), "r")
        try:
            # Only available since 2.5
            tarfle.extractall()
        except:
            for member in tarfle.getmembers():
                tarfle.extract(member)
        tarfle.close()
        os.chdir(cwd)
        if not os.path.isfile(os.path.join(self.tmpdir, "verupd_0.dat")):
            return
        zipfle = gzip.open(os.path.join(self.tmpdir, "verupd_0.dat"), mode="rb")
        data = zipfle.readlines()
        data = data[0].decode("utf-8").replace("[[", "").replace("]]", "")
        data = data.split("], [")
        data = list(eval(data[0]))
        if data[0] != self.ver:
            showError(self.mf.window, "Version Error", "The Data "\
                "to be Restored is Version %s which is Not the Same as "\
                "this Version of TARTAN, %s.\n\nOnly Matching Versions "\
                "Can Be Restored." % (data[0], self.ver))
            return "Invalid Version"
        if self.full == "Y":
            return "nd"

    def doAllCoy(self, frt, pag, r, c, p, i, w):
        self.icoy = w
        if self.icoy == "Y":
            self.coy = copyList(self.cnum)
            self.bu.loadEntry(frt, pag, i+1, data="")
            return "sk1"
        elif self.icoy == "I":
            self.bu.topf[pag][p+1][8]["titl"] = "Select Companies to Include"
        else:
            self.bu.topf[pag][p+1][8]["titl"] = "Select Companies to Exclude"

    def doCoyCmd(self, frt, pag, r, c, p, i, w):
        c = ""
        for co in w:
            if not c:
                c = co[0]
            else:
                c = "%s,%i" % (c, co[0])
        self.bu.loadEntry(frt, pag, p, data=c)

    def doCoySel(self, frt, pag, r, c, p, i, w):
        if not w:
            return "Invalid List of Companies"
        w = w.split(",")
        self.coy = []
        for coy in w:
            if not coy:
                continue
            try:
                c = int(coy)
                if c not in self.cnum:
                    raise Exception
                self.coy.append(c)
            except:
                return "Invalid Company %s" % coy

    def doAllSys(self, frt, pag, r, c, p, i, w):
        self.isys = w
        if self.isys == "Y":
            self.bu.loadEntry(frt, pag, i+1, data="")
            return "sk1"
        cf = PwdConfirm(self.mf, conum=0, system="MST", code="Systems")
        if cf.flag == "no":
            return "No Authority for Single Systems"
        if self.isys == "I":
            self.bu.topf[pag][p+1][8]["titl"] = "Select Systems to Include"
        else:
            self.bu.topf[pag][p+1][8]["titl"] = "Select Systems to Exclude"

    def doSysCmd(self, frt, pag, r, c, p, i, w):
        s = ""
        for sy in w:
            if not s:
                s = str(sy[0])
            else:
                s = "%s,%s" % (s, sy[0])
        self.bu.loadEntry(frt, pag, p, data=s)

    def doSysSel(self, frt, pag, r, c, p, i, w):
        if not w:
            return "Invalid List of Systems"
        w = w.split(",")
        self.sys = []
        for sss in w:
            if not sss:
                continue
            if sss in self.csys:
                self.sys.append(sss)
            else:
                return "Invalid System %s" % sss

    def endBckRes(self):
        if self.mode == "B":
            if self.mf.window and self.smtp and self.smtp[0]:
                self.bu.closeProcess()
            self.doBackup()
            if self.mf.window and self.smtp and self.smtp[0]:
                if self.bu.repeml[2]:
                    dated = CCD(self.sysdtw, "D1", 10).disp
                    sendMail(self.smtp, self.coys[0][2], self.bu.repeml[2],
                        "Tartan Backup for %s at %s" % (self.coys[0][1],
                        dated), self.bu.repeml[3], attach=self.arcfle,
                        wrkdir=self.mf.rcdic["wrkdir"], err=self.mf.body)
                self.mf.closeLoop()
            return
        self.bu.closeProcess()
        ok = askQuestion(self.mf.body, "Restore", "Are You Certain this is "\
            "what you want to do?\n\nExisting Data in the Database Will be "\
            "Replaced and will Not be Recoverable.", default="no")
        if ok == "no":
            self.mf.closeLoop()
            return
        if self.full == "Y":
            self.doFullRestore()
            self.mf.closeLoop()
            return
        if self.icoy == "Y":
            self.con = copyList(self.coy)
            self.con.sort()
        elif self.icoy == "I":
            # Included companies
            self.con = copyList(self.coy)
            self.con.sort()
        elif self.icoy == "E":
            # Excluded companies
            self.con = copyList(self.cnum)
            for coy in self.coy:
                self.con.remove(coy)
            self.con.sort()
        if self.isys == "Y":
            self.syn = copyList(self.csys)
        elif self.isys == "I":
            # Included systems
            self.syn = copyList(self.sys)
        elif self.isys == "E":
            # Excluded systems
            self.syn = copyList(self.csys)
            for sss in self.sys:
                self.syn.remove(sss)
        if len(self.con) == len(self.cnum) and len(self.syn) == len(self.csys):
            self.doFullRestore()
            self.mf.closeLoop()
            return
        if self.isys == "Y":
            self.syn.extend(["CTL", "RPT"])
        # Check if drs or str and slsinv, slspos
        if self.slsinv:
            if "DRS" in self.syn:
                if "STR" not in self.syn:
                    self.syn.append("STR")
                self.syn.extend(["POS", "SLS"])
            elif "str" in self.syn:
                if "DRS" not in self.syn:
                    self.syn.append("DRS")
                self.syn.extend(["POS", "SLS"])
        # Check for integrated systems and exit if not all included
        if self.sysi:
            itg = False
            chk = ["GEN"] + self.sysi
            for c in chk:
                if c in self.syn:
                    itg = True
                elif itg:
                    showError(self.mf.window, "System Error",
                        "Some Systems Are Integrated. You Will Have To "\
                        "Restore the General Ledger System as well as "\
                        "All Other Integrated Systems!")
                    self.mf.closeLoop()
                    return
        self.doPartialRestore()
        self.mf.closeLoop()

    def doBackup(self):
        try:
            os.umask(0000)
        except:
            pass
        try:
            # Delete existing temp directory and recreate it
            if os.path.exists(self.tmpdir):
                shutil.rmtree(self.tmpdir)
            os.makedirs(self.tmpdir)
            # Create the backup archive directory if not existing
            if not os.path.exists(self.archdir):
                os.makedirs(self.archdir)
        except:
            showError(self.mf.window, "Backup Error", "The Backup Directory "\
                "%s is Either Not Available or Write Protected! Please "\
                "Rectify this Condition and Retry the Backup!" % self.bupdir)
            return
        if self.budays:
            # Remove old archives
            try:
                for arc in os.listdir(self.archdir):
                    date = arc.split("_")[1].split(".")[0][:8]
                    if dateDiff(date, self.sysdtw, ptype="days") > self.budays:
                        os.remove(os.path.join(self.archdir, arc))
            except:
                pass
        tabs = []
        sql = Sql(self.mf.dbm, "ftable", prog=__name__)
        tbs = sql.getRec("ftable", cols=["ft_tabl"], group="ft_tabl")
        for tb in tbs:
            tabs.append(tb[0])
        tabs.sort()
        coys = [[0, ""]] + self.coys
        if self.mf.window:
            p1 = ProgressBar(self.mf.body, mxs=len(coys))
            p2 = ProgressBar(self.mf.body, mxs=len(tabs), inn=p1)
        names = []
        for cn, coy in enumerate(coys):
            if self.mf.window:
                p1.txtlab.configure(text="Backing Up Company %s" % coy[0])
                p1.displayProgress(cn)
            elif self.pbar and TBAR:
                p2 = Bar("Company %s" % coy[0], max=len(tabs))
            for tn, tab in enumerate(tabs):
                sql = Sql(self.mf.dbm, ["ffield", tab], prog=__name__)
                if self.mf.window:
                    p2.txtlab.configure(text="Backing Up Table %s" % tab)
                    p2.displayProgress(tn)
                elif self.pbar and TBAR:
                    p2.next()
                fld = sql.getRec("ffield", cols=["ff_name"],
                    where=[("ff_tabl", "=", tab), ("ff_seq", "=", 0)],
                    limit=1)
                if fld[0].count("_cono"):
                    whr = [("%s" % fld[0], "=", coy[0])]
                elif coy[0]:
                    continue
                else:
                    whr = None
                flenam = "%s_%s.dat" % (tab, coy[0])
                names.append(flenam)
                zipfle = gzip.open(os.path.join(self.tmpdir, flenam),
                    mode="wb", compresslevel=5)
                sql.getRec(tables=tab, where=whr, fetch=True)
                data = sql.sqlRec(fetch=True, limit=100000)
                while data:
                    zipfle.write(str(data).encode("utf-8"))
                    data = sql.sqlRec(fetch=True, limit=100000)
                zipfle.close()
            if not self.mf.window and self.pbar and TBAR:
                p2.finish()
        cwd = os.getcwd()
        os.chdir(self.tmpdir)
        tme = "%04i%02i%02i%02i%02i%02i" % time.localtime()[:-3]
        self.arcfle = os.path.join(self.archdir,
            "%s_%s.tar" % (self.mf.rcdic["dbname"], tme))
        tarfle = tarfile.open(self.arcfle, "w")
        for name in names:
            tarfle.add(name)
        tarfle.close()
        os.chdir(cwd)
        if self.mf.window:
            p1.closeProgress()
            p2.closeProgress()

    def doFullRestore(self):
        # Restore the Files
        files = glob.glob(os.path.join(self.tmpdir, "??????_*.dat"))
        files.sort()
        self.p1 = ProgressBar(self.mf.body, mxs=len(files),
            typ="Restoring Database %s" % self.mf.rcdic["dbname"])
        for num1, fle in enumerate(files):
            self.p1.displayProgress(num1)
            base = os.path.basename(fle)
            name = base.split("_")[0]
            if name in ("ffield", "ftable"):
                continue
            comp = int(base.split("_")[1].split(".")[0])
            sql = Sql(self.mf.dbm, name, prog=__name__)
            if not comp:
                sql.delRec(name)
            zipfle = gzip.open(fle, mode="rb")
            data = zipfle.readlines()
            zipfle.close()
            if data:
                newd = self.extractData(name, data)
                p2 = ProgressBar(self.mf.body, inn=self.p1, mxs=len(newd),
                    typ="Restoring Table (%s)" % name)
                sql.insRec(name, data=newd, dofmt=False, pbar=p2)
                p2.closeProgress()
        self.p1.closeProgress()
        self.mf.dbm.commitDbase()

    def doPartialRestore(self):
        self.p1 = ProgressBar(self.mf.body, mxs=len(self.con),
            typ="Restoring Database %s" % self.mf.rcdic["dbname"])
        for cn, coy in enumerate(self.con):
            self.p1.displayProgress(cn)
            sql = Sql(self.mf.dbm, self.yess, prog=__name__)
            for tab in self.yess:
                if tab[:3].upper() not in self.syn:
                    continue
                flenam = os.path.join(self.tmpdir, "%s_%s.dat" % (tab, coy))
                if os.path.exists(flenam):
                    # Delete existing records
                    if not coy:
                        whr = None
                    else:
                        fld = sql.sqlRec(("Select ff_name from ffield "\
                            "where ff_tabl = %s and ff_seq = %s" % \
                            (self.mf.dbm.dbf, self.mf.dbm.dbf), (tab, 0)),
                            limit=1)
                        # Ignore tables without _cono columns
                        if not fld[0].count("_cono"):
                            continue
                        whr = [("%s" % fld[0], "=", coy)]
                    sql.delRec(tab, where=whr)
                    # Restore backed up records
                    zipfle = gzip.open(flenam, mode="rb")
                    data = zipfle.readlines()
                    zipfle.close()
                    if data:
                        newd = self.extractData(tab, data)
                        p2 = ProgressBar(self.mf.body, inn=self.p1,
                            mxs=len(newd), typ="Restoring Table (%s)" % tab)
                        sql.insRec(tab, data=newd, dofmt=False, pbar=p2)
                        p2.closeProgress()
        self.p1.closeProgress()
        self.mf.dbm.commitDbase()

    def noBckRes(self):
        self.bu.closeProcess()
        self.mf.closeLoop()

    def extractData(self, tab, data):
        data = data[0].decode("utf-8").replace("]][[", "], [")
        return list(eval(data))

class CreateChart(object):
    """
    mf       - The mainframe class
    conum    - The company number
    conam    - The company name
    periods  - A list of the starting and ending period
    title    - A list of the chart's title
    achart   - A list of automatic chart values
    mchart   - A list of manual chart values
    xlab     - The x-axis's label
    ylab     - The y-axis's label
    xcol     - The x-axis's column labels
    """
    def __init__(self, mf, conum, conam, periods, title, achart, mchart, xlab=None, ylab=None, xcol=None):
        if not PYGAL:
            return
        self.mf = mf
        self.conum = conum
        self.conam = conam
        self.periods = periods
        self.title = title
        self.achart = achart
        self.mchart = mchart
        if xlab:
            self.xlab = xlab
        else:
            self.xlab = "Months"
        if ylab:
            self.ylab = ylab
        else:
            self.ylab = "Values"
        self.xcol = xcol
        if self.setVariables():
            self.doProcess()

    def setVariables(self):
        self.s_per = CCD(self.periods[0], "D2", 7)
        self.e_per = CCD(self.periods[1], "D2", 7)
        return True

    def doProcess(self):
        tit = ("Pie, Bar and Line Charts",)
        if self.achart:
            r1s = [("Automatic","A")]
        else:
            r1s = []
        r1s.extend([("Manual","M"),("Exit","E")])
        r2s = (
            ("Line","L"),
            ("Bar Normal","B"),
            ("Bar Stacked","S"),
            ("Pie Chart","P"))
        r3s = (("PDF","P"),("SVG","S"))
        r4s = (("Yes","Y"),("No","N"))
        fld = [
            (("T",0,0,0),("IRB",r1s),0,"Action","Select Action",
                "M","N",self.doAct,None,None,None,None),
            (("T",0,1,0),("IRB",r2s),0,"Select Chart","",
                "L","N",self.doCht,None,None,None,None)]
        if PYGAL and CVTSVG:
            fld.append((("T",0,2,0),("IRB",r3s),0,"View As","",
                "P","N",self.doVwr,None,None,None,None))
            idx = 3
        else:
            self.vwr = "S"
            idx = 2
        fld.append((("T",0,idx,0),("IRB",r4s),0,"Show Labels","",
            "Y","N",self.doLab,None,None,None,None))
        fld.append((("T",0,idx+1,0),("IRB",r4s),0,"Show Legends","",
            "Y","N",self.doLeg,None,None,None,None))
        self.dc = TartanDialog(self.mf, tops=True, title=tit, eflds=fld,
            tend=((self.doChtEnd, "y"),), txit=(self.doChtExit,),
            mail=("Y","N"))
        self.dc.mstFrame.wait_window()

    def doAct(self, frt, pag, r, c, p, i, w):
        self.act = w
        if self.act == "E":
            return "xt"
        elif self.act == "A":
            self.chart = copyList(self.achart)
        elif len(self.mchart) == 1:
            self.chart = copyList(self.mchart)
        else:
            titl = "Available Values"
            head = ["", "T", "Description"]
            lines = []
            for line in self.mchart:
                lines.append([line[0], line[1]])
            state = self.dc.disableButtonsTags()
            sc = SChoice(self.mf, scrn=self.mf.body, titl=titl, head=head,
                data=lines, rows=None, retn="I", mode="M")
            self.dc.enableButtonsTags(state=state)
            if sc.selection in (None, []):
                return "rf"
            self.chart = []
            for idx in sc.selection:
                self.chart.append(self.mchart[idx])

    def doCht(self, frt, pag, r, c, p, i, w):
        if w == "P" and len(self.chart) > 1:
            return "Only One Range of Values Allowed for a Pie Chart"
        self.cht = w

    def doVwr(self, frt, pag, r, c, p, i, w):
        self.vwr = w

    def doLab(self, frt, pag, r, c, p, i, w):
        self.lab = bool(w == "Y")

    def doLeg(self, frt, pag, r, c, p, i, w):
        self.leg = bool(w == "Y")

    def doChtEnd(self):
        self.repprt = ["Y", "V", "view"]
        self.repeml = copyList(self.dc.repeml)
        labs = []
        zers = True
        for line in self.chart:
            labs.append(line[1])
            for num in line[2:]:
                if num:
                    zers = False
                if num < 0 and self.cht == "P":
                    showError(self.mf.body, "Chart Error",
                        "You Cannot Have Negative Values in a Pie Chart!\n\n"\
                        "Please Select an Alternate Chart Type.")
                    self.dc.focusField("T",0,1)
                    return
        if zers:
            showError(self.mf.body, "Chart Error",
                "All the Items Selected Have Zero Values!\n\n"\
                "Please Select Items which Have Non Zero Values.")
            self.dc.focusField("T",0,1)
            return
        if not self.xcol:
            self.xcol = []
            y = int(self.s_per.work / 100) % 100
            m = self.s_per.work % 100
            for x in range(0, 12):
                self.xcol.append(tartanWork.mthnam[m][0])
                if m == (self.e_per.work % 100):
                    break
                m += 1
                if m > 12:
                    y += 1
                    m = 1
        # Generate Chart
        flenam = getModName(self.mf.rcdic["wrkdir"], "chart",
            self.conum, ext="svg")
        if self.cht == "B":
            chart = pygal.Bar(print_labels=self.lab, show_legend=self.leg)
        elif self.cht == "S":
            chart = pygal.StackedBar(print_labels=self.lab,
                show_legend=self.leg)
        elif self.cht == "L":
            chart = pygal.Line(print_labels=self.lab, show_legend=self.leg)
        elif self.cht == "P":
            chart = pygal.Pie(print_labels=self.lab, show_legend=self.leg)
        if self.cht == "P":
            titl = "%s - %s" % tuple(self.title[0])
            titl = "%s\n%s for Period %s to %s (%s)" % (titl, self.chart[0][1],
                self.s_per.disp, self.e_per.disp, self.title[1])
        else:
            titl = self.title[0][0]
            titl = "%s\n%s for Period %s to %s (%s)" % (titl, self.title[0][1],
                self.s_per.disp, self.e_per.disp, self.title[1])
        chart.title = titl
        if self.cht == "P":
            for seq, dat in enumerate(self.xcol):
                chart.add(dat, [{"value": self.chart[0][seq+2], "label": dat}])
        else:
            chart.x_labels = self.xcol
            if self.xlab:
                chart.x_title = self.xlab
            if self.ylab:
                chart.y_title = self.ylab
            for l in self.chart:
                chart.add(l[1], l[2:])
        chart.render_to_file(flenam)
        if CVTSVG and self.vwr == "P":
            # Convert to PDF
            self.pdfnam = flenam.replace("svg", "pdf")
            renderPDF.drawToFile(svg2rlg(flenam), self.pdfnam)
        else:
            self.pdfnam = flenam
        # Print or Display the Chart
        doPrinter(mf=self.mf, conum=self.conum, pdfnam=self.pdfnam,
            header=self.title[0], repprt=self.repprt, repeml=self.repeml)
        # Return focus
        self.dc.focusField("T", 0, 1)

    def doChtExit(self):
        self.dc.closeProcess()

class NotesCreate(object):
    def __init__(self, mf, conum, conam, user, sys, key, commit=True):
        self.mf = mf
        self.conum = conum
        self.conam = conam
        self.user = user
        self.sys = sys
        self.key = str(key)
        self.commit = commit
        self.setVariables()
        self.mainProcess()

    def setVariables(self):
        if not self.mf.dbm.dbopen:
            self.mf.dbm.openDbase()
            self.opened = True
        else:
            self.opened = False
        self.sql = Sql(self.mf.dbm, ["ctlnot", "ctlpwu"], prog=__name__)
        t = time.localtime()
        self.sysdtw = ((t[0] * 10000) + (t[1] * 100) + t[2])

    def mainProcess(self):
        tit = "%s Notes" % tartanWork.allsys[self.sys][0]
        self.ntm = {
            "stype": "R",
            "tables": ("ctlnot",),
            "title": "Notes",
            "cols": (
                ("not_date", "", 0, "Cap-Date"),
                ("not_user", "", 0, "User"),
                ("not_aflag", "", 0, "F"),
                ("not_adate", "", 0, "Action-Dte"),
                ("not_auser", "", 0, "Action-User"),
                ("not_desc", "", 0, "Description"),
                ("not_seq", "", 0, "Sequence")),
            "where": [("not_cono", "=", self.conum), ("not_sys", "=",
                self.sys), ("not_key", "=", self.key)],
            "order": "not_seq",
            "comnd": self.selectNote}
        r1s = (("Normal", "N"), ("Urgent", "U"))
        fld = (
            (("T",0,0,0),"ITV",(50,10),"Details","",
                "","N",self.doNote,None,None,("notblank",)),
            (("T",0,1,0),("IRB",r1s),0,"Action Flag","",
                "N","N",self.doActFlag,None,None,None),
            (("T",0,2,0),"I@not_adate",0,"","",
                self.sysdtw,"N",self.doActDate,None,None,("efld",)),
            (("T",0,3,0),"I@not_auser",0,"","",
                self.user,"N",self.doActUser,None,None,("efld",)))
        tnd = ((self.doEnd,"y"), )
        txt = (self.doExit, )
        but = (
            ("Show",None,self.doShow,0,("T",0, 1),("T",0,2)),
            ("Cancel",None,self.doCancel,0,("T",0, 2),None),
            ("Quit",None,self.doExit,1,None,None))
        self.df = TartanDialog(self.mf, tops=True, title=tit, eflds=fld,
            tend=tnd, txit=txt, butt=but)
        self.df.mstFrame.wait_window()

    def doNote(self, frt, pag, r, c, p, i, w):
        self.note = w

    def doActFlag(self, frt, pag, r, c, p, i, w):
        self.actflag = w

    def doActDate(self, frt, pag, r, c, p, i, w):
        if w and w < self.sysdtw:
            return "Action Date in the Past"
        self.actdate = w

    def doActUser(self, frt, pag, r, c, p, i, w):
        usr = self.sql.getRec("ctlpwu", where=[("usr_name", "=", w)],
            limit=1)
        if not usr:
            return "Invalid User Name"
        self.actuser = w

    def doShow(self):
        state = self.df.disableButtonsTags()
        self.df.setWidget(self.df.mstFrame, "hide")
        self.df.selectItem(0, self.ntm)
        self.df.setWidget(self.df.mstFrame, "show")
        self.df.enableButtonsTags(state=state)
        self.df.focusField("T", 0, 1)

    def selectNote(self, frt, pag, r, c, p, i, w):
        self.chgflag = w[2]
        self.chgdate = w[3]
        self.chguser = w[4]
        self.nseq = w[-1:][0]
        tit = ("Notes Editing",)
        r1s = (("Normal","N"),("Urgent","U"),("Completed","C"))
        fld = (
            (("T",0,0,0),"O@not_date",0,""),
            (("T",0,1,0),"O@not_user",0,""),
            (("T",0,2,0),"OTV",(50,10),"Details"),
            (("T",0,3,0),("IRB",r1s),0,"Action Flag (C/N/U)","Action Flag",
                "N","N",self.doChgFlag,None,None,None),
            (("T",0,4,0),"I@not_adate",0,"","",
                "","N",self.doChgDate,None,None,("efld",)),
            (("T",0,5,0),"I@not_auser",0,"","",
                "","N",self.doChgUser,None,None,("efld",)))
        but = (
            ("Accept",None,self.doNEnd,1,None,None),
            ("Cancel",None,self.doNExit,1,None, None))
        tnd = ((self.doNEnd, "n"), )
        txt = (self.doNExit, )
        self.nf = TartanDialog(self.mf, tops=True, title=tit, eflds=fld,
            butt=but, tend=tnd, txit=txt, focus=False)
        self.nf.loadEntry("T", 0, 0, data=w[0])
        self.nf.loadEntry("T", 0, 1, data=w[1])
        mess = w[5]
        if len(w) > 7:
            for x in range(6, len(w) - 1):
                mess = "%s\n%s" % (mess, w[x])
        self.nf.loadEntry("T", 0, 2, data=mess)
        self.nf.loadEntry("T", 0, 3, data=w[2])
        self.nf.loadEntry("T", 0, 4, data=w[3])
        self.nf.loadEntry("T", 0, 5, data=w[4])
        if self.chgflag == "C":
            self.nf.setWidget(self.nf.B0, "normal")
            self.nf.setWidget(self.nf.B1, "normal")
        else:
            self.nf.focusField("T", 0, 4)
        self.nf.mstFrame.wait_window()

    def doChgFlag(self, frt, pag, r, c, p, i, w):
        self.chgflag = w
        if self.chgflag == "C":
            return "nd"

    def doChgDate(self, frt, pag, r, c, p, i, w):
        self.chgdate = w

    def doChgUser(self, frt, pag, r, c, p, i, w):
        usr = self.sql.getRec("ctlpwu", where=[("usr_name", "=", w)],
            limit=1)
        if not usr:
            return "Invalid User Name"
        self.chguser = w
        return "nd"

    def doNEnd(self):
        self.nf.closeProcess()
        self.sql.updRec("ctlnot", cols=["not_aflag", "not_adate", "not_auser"],
            data=[self.chgflag, self.chgdate, self.chguser], where=[("not_seq",
            "=", self.nseq)])
        self.mf.dbm.commitDbase()
        if self.opened:
            self.mf.dbm.closeDbase()

    def doNExit(self):
        self.nf.closeProcess()
        if self.opened and self.mf.dbm.dbopen:
            self.mf.dbm.closeDbase()

    def doEnd(self):
        self.df.closeProcess()
        self.sql.insRec("ctlnot", data=[self.conum, self.sys, self.key,
            self.sysdtw, self.user, self.note, self.actflag, self.actdate,
            self.actuser, 0])
        if self.commit:
            self.mf.dbm.commitDbase()
        if self.opened:
            self.mf.dbm.closeDbase()

    def doCancel(self):
        self.df.clearFrame("T", 0)
        self.df.focusField("T", 0, 1)

    def doExit(self):
        self.df.closeProcess()
        if self.opened and self.mf.dbm.dbopen:
            self.mf.dbm.closeDbase()

class NotesPrint(object):
    def __init__(self, mf, conum, conam, sys, loop=True):
        self.mf = mf
        self.conum = conum
        self.conam = conam
        self.sys = sys
        self.loop = loop
        self.setVariables()
        self.mainProcess()
        if self.loop:
            self.mf.startLoop()
        else:
            self.df.mstFrame.wait_window()

    def setVariables(self):
        self.sql = Sql(self.mf.dbm, "ctlnot", prog=__name__)
        t = time.localtime()
        self.sysdtw = ((t[0] * 10000) + (t[1] * 100) + t[2])
        self.data = []

    def mainProcess(self):
        r1s = (("All", "A"), ("Normal", "N"), ("Urgent", "U"))
        r2s = (("Account", "A"), ("Capture", "B"), ("Action", "C"))
        fld = (
            (("T",0,0,0),("IRB",r1s),0,"Action Flag","",
                "A","N",self.doFlag,None,None,None),
            (("T",0,1,0),"Id1",10,"From Capture Date","",
                0,"N",self.doFDate,None,None,("efld",)),
            (("T",0,2,0),"I@not_date",0,"To   Capture Date","",
                self.sysdtw,"N",self.doTDate,None,None,("efld",)),
            (("T",0,3,0),"I@not_adate",0,"From Action Date","",
                "","N",self.doFADate,None,None,("efld",)),
            (("T",0,4,0),"I@not_adate",0,"To   Action Date","",
                "","N",self.doTADate,None,None,("efld",)),
            (("T",0,5,0),("IRB",r2s),0,"Order By","",
                "A","N",self.doOrder,None,None,None))
        tnd = ((self.doEnd,"y"), )
        txt = (self.doExit, )
        self.df = TartanDialog(self.mf, eflds=fld, tend=tnd, txit=txt,
            view=("N","V"), mail=("Y","N"))

    def doFlag(self, frt, pag, r, c, p, i, w):
        if w == "A":
            self.flag = ""
        else:
            self.flag = w

    def doFDate(self, frt, pag, r, c, p, i, w):
        self.fdate = w

    def doTDate(self, frt, pag, r, c, p, i, w):
        if w < self.fdate:
            return "Invalid Date, Earlier than From Date"
        self.tdate = w

    def doFADate(self, frt, pag, r, c, p, i, w):
        self.fadate = w

    def doTADate(self, frt, pag, r, c, p, i, w):
        if w < self.fadate:
            return "Invalid Date, Earlier than From Date"
        self.tadate = w

    def doOrder(self, frt, pag, r, c, p, i, w):
        self.order = w

    def doEnd(self):
        self.df.closeProcess()
        tab = ["ctlnot"]
        whr = [("not_cono", "=", self.conum), ("not_sys", "=", self.sys)]
        if self.flag:
            whr.extend([("not_aflag", "=", self.flag), ("not_date", ">=",
                self.fdate)])
        else:
            whr.append(("not_date", ">=", self.fdate))
        if self.tdate:
            whr.append(("not_date", "<=", self.tdate))
        whr.append(("not_adate", ">=", self.fadate))
        if self.tadate:
            whr.append(("not_adate", "<=", self.tadate))
        if self.order == "A":
            odr = "not_key, not_date desc, not_seq"
        elif self.order == "B":
            odr = "not_date desc, not_key, not_seq"
        else:
            odr = "not_adate desc, not_key, not_seq"
        self.data = self.sql.getRec(tables=tab, where=whr, order=odr)
        if not self.data:
            showError(self.mf.body, "Error", "No Notes Selected")
        if self.loop:
            self.mf.closeLoop()

    def doExit(self):
        self.df.closeProcess()
        if self.loop:
            self.mf.closeLoop()

class FileImport(object):
    """
    This is a file import class.

    mf   - The Mainframe class.
    args - Must contain the following:

            imptab  - A table to import

                                or

            impcol - A list of lists of field parameters e.g.
                    [[Label, Column Number, Format, Size]] where:
                        Label           - The column name
                        Column Number   - The default column number from 0
                        Format          - The data format as per CCD
                        Size            - The data size as per CCD

           And additionally can contain the following:

            impskp - Fields to skip with table option.
            impfld - Any additional dialog fields.
            impfle - The file name to import else blank to select.
            impsht - The sheet name to import else blank for the first one.
            impdlg - Whether to have a dialog, True or False. Default is True.
    """
    def __init__(self, mf, **args):
        self.mf = mf
        self.impdat = []
        self.impskp = []
        self.impfld = []
        self.impfle = ""
        self.impsht = None
        self.impdlg = True
        self.impign = "N"
        self.chgcol = {}
        for x in range(52):
            if x > 25:
                self.chgcol["A" + chr(x + 39)] = x
            else:
                self.chgcol[chr(x + 65)] = x
        if "impcol" in args:
            self.impcol = args["impcol"]
            sql = Sql(self.mf.dbm, "ffield")
            for num, col in enumerate(self.impcol):
                if col[2][0] == "@":
                    dat = sql.getRec("ffield", cols=["ff_type",
                        "ff_size"], where=[("ff_name", "=", col[2][1:])],
                        limit=1)
                    self.impcol[num][2] = dat[0]
                    self.impcol[num][3] = dat[1]
        elif "imptab" in args:
            self.imptab = args["imptab"]
            self.impcol = []
            if "impskp" in args:
                self.impskp = args["impskp"]
            sql = Sql(self.mf.dbm, "ffield")
            acc = sql.getRec("ffield", where=[("ff_tabl", "=",
                self.imptab)], order="ff_seq")
            if not acc:
                showError(self.mf.body, head="Import Error",
                    mess="Invalid table %s in List of Variables" % self.imptab)
                return
            count = 0
            for a in acc:
                if a[2].count("_xflag"):
                    continue
                if self.impskp and a[2] in self.impskp:
                    continue
                self.impcol.append([a[5], count, a[3], a[4]])
                count += 1
        else:
            showError(self.mf.body, head="Import Error",
                mess="Missing 'impcol' and/or 'table' in List of Variables")
            return
        if "impfld" in args:
            self.impexe = {}
            self.impfld = []
            for fld in args["impfld"]:
                fld = list(fld)
                if fld[7]:
                    self.impexe[fld[0][2]] = fld[7]
                    fld[7] = self.doExeFld
                self.impfld.append(list(fld))
            self.index = len(self.impfld)
        else:
            self.index = 0
        if "impfle" in args:
            self.impfle = args["impfle"]
        if "impsht" in args:
            self.impsht = args["impsht"]
        if "impdlg" in args:
            self.impdlg = args["impdlg"]
        if self.impdlg or not self.impfle:
            self.impdlg = True
            self.doDialog()
        else:
            err = self.getFileType(args["impfle"])
            if not err or err == "sk1":
                self.doImpEnd()

    def doDialog(self):
        maxs = 0
        for field in self.impcol:
            if len(field[0]) > maxs:
                maxs = len(field[0])
        tit = ("Import File Parameters",)
        if sys.platform == "win32":
            self.lastdir = os.path.join(self.mf.rcdic["wrkdir"], "impdir")
        else:
            self.lastdir = os.path.join(self.mf.rcdic["wrkdir"], ".impdir")
        if os.path.exists(self.lastdir):
            infle = open(self.lastdir, "r")
            name = infle.readline().rstrip()
            infle.close()
        else:
            name = self.mf.rcdic["wrkdir"]
        typ = []
        if XLSX:
            typ.append(("XlSX Files", "*.xlsx"))
        if XLS:
            typ.append(("XlS Files", "*.xls"))
        if ODS:
            typ.append(("ODS Files", "*.ods"))
        typ.append(("CSV Files", "*.csv"))
        typ.append(("All Files", "*.*"))
        fle = {
            "stype": "F",
            "types": "fle",
            "initd": name,
            "ftype": typ}
        self.sht = {
            "stype": "C",
            "titl": "Available Sheets",
            "head": ("Name",),
            "typs": (("NA",30),),
            "data": []}
        r1s = (("Yes", "Y"), ("No", "N"))
        idx = onn = len(self.impfld)
        if self.impfle:
            self.impfld.append((("T",0,idx,0),"OFF",50,"File Name"))
        else:
            self.impfld.append((("T",0,idx,0),"IFF",50,"File Name","",
                "","N",self.doImpFle,fle,None,("file",)))
        idx += 1
        if self.impsht is None:
            self.impfld.append((("T",0,idx,0),"INA",50,"Sheet Name","",
                0,"N",self.doImpNum,self.sht,None,("efld",)))
        else:
            self.impfld.append((("T",0,idx,0),"ONA",50,"Sheet Name"))
        idx += 1
        self.impfld.append((("T",0,idx,0),("IRB",r1s),0,"Ignore Invalid Lines",
            "","N","N",self.doImpIgn,None,None,None))
        idx += 1
        for num in range(0, len(self.impcol), 2):
            pad = " " * (maxs - len(self.impcol[num][0]))
            self.impfld.append([("T", 0, idx, 0),"IUA",2,
                "%s%s - Column" % (self.impcol[num][0], pad),
                "%s Column" % self.impcol[num][0],0,"N",
                self.doImpCol,None,None,("efld",),None,
                "This is the Column in the File Starting with A"])
            if num != (len(self.impcol) - 1):
                pad = " " * (maxs - len(self.impcol[num+1][0]))
                self.impfld.append([("T", 0, idx, 46),"IUA",2,
                    "%s%s - Column" % (self.impcol[num+1][0], pad),
                    "%s Column" % self.impcol[num+1][0],0,"N",
                    self.doImpCol,None,None,("efld",)])
            idx += 1
        but = (
            ("Accept",None,self.doImpEnd,0,("T",0,onn+4),("T",0,onn+3),None),)
        self.ip = TartanDialog(self.mf, tops=True, title=tit,
            eflds=self.impfld, tend=((self.doImpEnd,"y"),),
            txit=(self.doImpExit,), butt=but, focus=False)
        if self.impfle:
            self.ip.loadEntry("T", 0, 0, data=self.impfle)
            self.ip.focusField("T", 0, 2)
        else:
            self.ip.focusField("T", 0, 1)
        self.ip.mstFrame.wait_window()

    def doExeFld(self, frt, pag, r, c, p, i, w):
        return self.impexe[p](self.ip, w)

    def doImpFle(self, frt, pag, r, c, p, i, w):
        self.impfle = w
        return self.getFileType(self.impfle)

    def getFileType(self, name):
        if name.split(".")[-1].lower() == "csv":
            self.ftype = "csv"
            try:
                self.csvfle = open(name, "r")
                return "sk1"
            except Exception as err:
                return "Invalid csv File (%s)" % err
        elif name.split(".")[-1].lower() == "ods":
            self.ftype = "ods"
            try:
                self.workbk = getods(name)
                self.sht["data"] = list(self.workbk.keys())
                self.sht["data"].sort()
            except Exception as err:
                return "Invalid ods File (%s)" % err
        elif name.split(".")[-1].lower() == "xls" and XLS:
            self.ftype = "xls"
            try:
                self.workbk = getxls(name)
                self.sht["data"] = list(self.workbk.keys())
                self.sht["data"].sort()
            except Exception as err:
                return "Invalid xls File (%s)" % err
        elif name.split(".")[-1].lower() == "xlsx" and XLSX:
            self.ftype = "xlsx"
            try:
                self.workbk = openpyxl.load_workbook(filename=name)
                self.sht["data"] = self.workbk.sheetnames
            except Exception as err:
                return "Invalid xlsx File (%s)" % err
        else:
            return "Invalid File Type"

    def doImpNum(self, frt, pag, r, c, p, i, w):
        if w not in self.sht["data"]:
            return "Invalid Sheet Name"
        try:
            if self.ftype == "ods":
                self.worksh = self.workbk[w]
            elif self.ftype == "xls":
                self.worksh = self.workbk[w]
            else:
                self.worksh = self.workbk[w].values
        except:
            return "Invalid Sheet"

    def doImpIgn(self, frt, pag, r, c, p, i, w):
        self.impign = w
        for col in range(len(self.impcol)):
            if self.ip.topf[pag][p+1+col][1][0] == "O":
                dat = ""
            elif col > 25:
                dat = ("A" + chr(col + 39))
            else:
                dat = chr(col + 65)
            self.ip.loadEntry("T", 0, p+1+col, data=dat)

    def doImpCol(self, frt, pag, r, c, p, i, w):
        if w:
            self.impcol[p-3-self.index][1] = self.chgcol[w.strip()]
        else:
            self.impcol[p-3-self.index][1] = None

    def doImpEnd(self):
        if self.impdlg:
            self.ip.closeProcess()
        self.impdat = []
        try:
            if self.ftype == "csv":
                data = csv.reader(self.csvfle, quoting=csv.QUOTE_MINIMAL)
            else:
                data = self.worksh
            for row, rdd in enumerate(data):
                if not rdd:
                    continue
                try:
                    lin = []
                    for col, cdd in enumerate(self.impcol):
                        if cdd[1] is None:
                            dat = ""
                        else:
                            dat = rdd[cdd[1]]
                        if self.ftype in ("xls", "xlsx"):
                            if cdd[2] in ("D1", "d1"):
                                if isinstance(dat, datetime.date):
                                    dat = int(dat.strftime("%Y%m%d"))
                                elif isinstance(dat, datetime.datetime):
                                    dat = int(rdd.strftime("%Y%m%d"))
                                else:
                                    try:
                                        dat = int(dat)
                                    except:
                                        dat = 0
                                        cdd[2] = "d1"
                        d = CCD(dat, cdd[2], cdd[3])
                        if d.err:
                            raise Exception(d.err)
                        lin.append(d.work)
                    self.impdat.append(lin)
                except Exception as err:
                    if self.impign == "N":
                        showError(self.mf.body, "Column Error",
                            "Row %s, Column %s (%s)\n\nDoes Not Exist "\
                            "or is Invalid\n\n%s" % (row + 1,
                            self.ip.t_work[0][0][col+3], cdd[0], err))
                        raise Exception
        except:
            self.impdat = []
        try:
            infle = open(self.lastdir, "w")
            infle.write(os.path.dirname(os.path.normpath(self.impfle)))
            infle.close()
        except:
            pass

    """
    def doImpEnd(self):
        if self.impdlg:
            self.ip.closeProcess()
        imperr = False
        if self.ftype == "csv":
            try:
                data = csv.reader(self.csvfle, quoting=csv.QUOTE_MINIMAL)
                self.impdat = []
                for row, dat in enumerate(data):
                    if not dat:
                        continue
                    try:
                        lin = []
                        for col in self.impcol:
                            d = CCD(dat[col[1]], col[2], col[3])
                            if d.err:
                                raise Exception
                            lin.append(d.work)
                        self.impdat.append(lin)
                    except:
                        if self.impign == "N":
                            showError(self.mf.body, "Column Error",
                                "Row %s Column %s %s\n\n%s is Invalid" %
                                (row, col[1], col[0], dat[col[1]]))
                            raise Exception
            except:
                imperr = True
        elif self.ftype == "ods":
            try:
                self.impdat = []
                for row, dat in enumerate(self.worksh):
                    if not dat:
                        continue
                    try:
                        lin = []
                        for col in self.impcol:
                            d = CCD(dat[col[1]], col[2], col[3])
                            if d.err:
                                raise Exception
                            lin.append(d.work)
                        self.impdat.append(lin)
                    except:
                        if self.impign == "N":
                            showError(self.mf.body, "Column Error",
                                "Row %s, Column %s\n\n%s is Invalid" %
                                (row, col[1], col[0]))
                            raise Exception
            except:
                imperr = True
        elif self.ftype in ("xls", "xlsx"):
            try:
                self.impdat = []
                for row, dat in enumerate(self.worksh):
                    try:
                        lin = []
                        for col, rec in enumerate(self.impcol):
                            cdd = dat[rec[1]]
                            if rec[2] in ("D1", "d1"):
                                if isinstance(cdd, datetime.date):
                                    cdd = int(cdd.strftime("%Y%m%d"))
                                elif isinstance(cdd, datetime.datetime):
                                    cdd = int(dat.strftime("%Y%m%d"))
                                else:
                                    try:
                                        cdd = int(cdd)
                                    except:
                                        cdd = 0
                                        rec[2] = "d1"
                            d = CCD(cdd, rec[2], rec[3])
                            if d.err:
                                raise Exception
                            lin.append(d.work)
                        self.impdat.append(lin)
                    except Exception as err:
                        if self.impign == "N":
                            showError(self.mf.body, "Column Error",
                                "Row %s, Column %s %s\n\n%s is Invalid" %
                                (row + 1, col + 1, rec[0], cdd))
                            raise Exception(err)
            except:
                imperr = True
        else:
            imperr = True
        if imperr:
            self.impdat = []
        else:
            try:
                infle = open(self.lastdir, "w")
                infle.write(os.path.dirname(os.path.normpath(self.impfle)))
                infle.close()
            except:
                pass
    """

    def doImpExit(self):
        self.impdat = []
        self.ip.closeProcess()

class MyFpdf(fpdf.FPDF):
    """
    This class generates a subclass of fpdf with a self.font variable i.e.
        self.font = (family, size, height)

    orientation - Portrait or Landscape
    unit        - mm or cm
    fmat        - Page format e.g. A4
    font        - A family string or a tuple (family, style, size)
    name        - Document title
    head        - Integer / String or List / Tuple as folows:
                    Integer - Number of characters or
                    String - Heading or
                    List / Tuple - (Heading, ("Family", "Style", Size))
    auto        - Automatic page breaks
    foot        - Turn footer on or off
    """
    def __init__(self, orientation="P", unit="mm", fmat="A4", font="courier", name="", head="", auto=False, foot=True):
        self.unit = unit
        self.fmat = fmat
        if self.fmat == "A4":
            self.foot = foot
        else:
            self.foot = False
        self.suc = chr(151)
        t = time.localtime()
        self.sysdt = time.strftime("%d %B %Y %H:%M:%S", t)
        try:
            super().__init__(orientation, self.unit, self.fmat)
            self.setValues(name, head, font)
            self.set_author("Tartan Systems")
            self.set_fill_color(220)
            self.set_title(name)
            if not auto:
                self.set_auto_page_break(False, margin=0)
            if foot:
                self.lpp -= round(15 / self.chgt, 0)
                self.alias_nb_pages()
        except Exception as err:
            print(err)

    def setValues(self, name, head, font="", border=""):
        # Add TTF Fonts
        try:
            for pth in glob.glob(os.path.join(getPrgPath(), "fnt/*.ttf")):
                nam = os.path.basename(pth)
                self.add_font(nam, "", pth, uni=True)
        except:
            pass
        # Defaults
        self.portrait = (210 - (int(self.l_margin) + int(self.r_margin)),
            297 - (2 * int(self.t_margin)))
        if not name or not head or type(head) in (list, tuple):
            if type(font) == str:
                font = [font, "", 10]
            self.setFont(font[0], font[1], font[2], default=True)
            self.chgt = round(font[2] * .4, 1)
            if self.def_orientation == "P":
                self.lpp = int(self.portrait[1] / self.chgt)
            else:
                self.lpp = int(self.portrait[0] / self.chgt)
            return
        # Adjust font to heading width
        if type(head) == int:
            head = "X" * head
        if type(font) == str:
            family = font
        else:
            family = font[0]
        if family.lower() == "courier" and len(head) > 120:
            self.def_orientation = "L"
        self.font = None
        while not self.font:
            if self.def_orientation == "P":
                mm = self.portrait[0]
            else:
                mm = self.portrait[1]
            for size in range(mm, 30, -1):
                siz = round(size / 10.0, 1)
                self.chgt = round(siz * .4, 2)
                if border:
                    self.chgt += 1
                self.setFont(family, "", siz, default=True)
                self.width = self.get_string_width(head)
                if self.def_orientation == "P":
                    if self.width > self.portrait[0]:
                        continue
                    self.font = [family, siz, self.chgt]
                    self.lpp = int(self.portrait[1] / self.chgt)
                    break
                if self.def_orientation == "L":
                    if self.width > self.portrait[1]:
                        continue
                    self.font = [font, siz, self.chgt]
                    self.lpp = int(self.portrait[0] / self.chgt)
                    break
            if not self.font:
                if self.def_orientation == "P":
                    self.def_orientation = "L"
                else:
                    print("Invalid head length", len(head))
                    sys.exit()

    def drawText(self, txt="", x=0, y=0, w=0, h=0, font=None, border=0, ln=1, align="", fill=0, link="", ctyp="S"):
        if type(txt) is not str:
            txt = str(txt)
        if x and y:
            self.set_xy(x, y)
        elif x:
            self.set_x(x)
        elif y:
            self.set_y(y)
        if font is not None:
            if type(font) in (list, tuple) and len(font) == 3:
                family, style, size = font
                self.setFont(family, style, size)
            elif type(font) in (list, tuple) and len(font) == 2:
                family = self.font_family
                style, size = font
                self.setFont(family, style, size)
            else:
                family = self.font_family
                if type(font) == str:
                    style = font
                else:
                    style = font[0]
                self.setFont(family, style)
        if not h:
            h = self.font[2]
        if ctyp == "S":
            self.cell(w=w, h=h, ln=ln, txt=txt, border=border, align=align,
                fill=fill)
        else:
            self.multi_cell(w=w, h=h, txt=txt, border=border, align=align,
                fill=fill)

    def setFont(self, family="", style="", size=0, default=False):
        if not family:
            if self.font:
                family = self.font[0]
            else:
                family = self.font_family
        if not size:
            if self.font:
                size = self.font[1]
            else:
                size = 10
        try:
            self.set_font(family, style, size)
            self.cwth = self.get_string_width("X")
            if default:
                self.font = [family, size, self.font_size]
        except Exception as err:
            print(err, family, style, size)

    def underLine(self, t="S", x=0, y=0, h=0, w=0, txt=""):
        if txt.count(self.suc):
            if t == "D":
                y = self.get_y()
            self.cell(w=0, h=self.font[2], ln=1, txt=txt)
            if t == "D":
                self.set_y(y + .5)
                self.cell(w=0, h=self.font[2], ln=1, txt=txt)
        else:
            if not x:
                x = self.x
            if not y:
                y = self.y
            if not h:
                h = self.font[2]
            if not w:
                if txt:
                    w = self.get_string_width(txt+"X")
                else:
                    w = self.width
            if t == "D":
                self.set_line_width(.5)
            self.line(x + 1, y + int(h / 2), x + w, y + int(h / 2))
            if t == "D":
                self.set_line_width(0)
            self.set_xy(x, y + h)

    def footer(self):
        if not self.foot:
            return
        self.set_y(-15)
        if self.font[1] > 8:
            self.set_font("Arial", "I", 8)
        else:
            self.set_font("Arial", "I", self.font[1])
        if self.title:
            txt = "Tartan Systems (%s) %s" % (self.title, self.sysdt)
        else:
            txt = "Tartan Systems %s" % self.sysdt
        self.cell(w=0, h=10, txt=txt, border=0, ln=0, align="L")
        self.cell(w=0, h=10, txt="Page " + str(self.page_no()) + "/{nb}",
            border=0, ln=0, align="R")

    def newPage(self, lines=1, lhgt=None):
        if self.page:
            if lhgt is None:
                lhgt = self.font[2]
            y = self.get_y()
            if self.def_orientation == "P":
                pd = self.portrait[1]
            else:
                pd = self.portrait[0]
            h = lines * lhgt
            m = math.ceil(float(ASD(y) + ASD(h)))
        if not self.page or m >= pd:
            return True

class TartanLabel(MyFpdf):
    def __init__(self, label, unit="mm", posY=1, posX=1):
        super().__init__(name=__name__, head=90, font="arial", foot=False)
        self.setFont(self.font[0], "", self.font[1])
        if label in tartanWork.labels:
            type_format = tartanWork.labels[label]
        else:
            raise NameError("Model %s is not in the database" % label)
        self.margin_left = type_format["marginLeft"]
        self.margin_top = type_format["marginTop"]
        self.space_x = type_format["SpaceX"]
        self.space_y = type_format["SpaceY"]
        self.number_x = type_format["NX"]
        self.number_y = type_format["NY"]
        self.lab_width = type_format["width"]
        self.lab_height = type_format["height"]
        self.line_height = self.font[2]
        self.set_margins(0, 0)
        self.countY = posY - 1
        self.countX = posX - 2
        self.padding = 3

    def add_label(self, text):
        """Print a label"""
        self.countX += 1
        if self.countX == self.number_x:
            # Row full, we start a new one
            self.countX = 0
            self.countY += 1
            if self.countY == self.number_y:
                # End of page reached, we start a new one
                self.countY = 0
                self.add_page()
        width = self.lab_width + self.space_x
        height = self.lab_height + self.space_y
        posX = self.margin_left + (self.countX * width) + self.padding
        posY = self.margin_top + (self.countY * height) + self.padding
        self.set_xy(posX, posY)
        line_width = self.lab_width - self.padding
        self.multi_cell(line_width, self.line_height, text, 0, "L")

    def _putcatalog(self):
        self._out("/Type /Catalog")
        # Disable the page scaling option in the printing dialog
        self._out("/ViewerPreferences [/PrintScaling /None]")
        self._out("/Pages 1 0 R")
        self._out("/OpenAction [3 0 R /XYZ null null 1]")
        self._out("/PageLayout /OneColumn")

class DrawForm(MyFpdf):
    """
    Draw a form from a template created using tp1010.py

    dbm    - A database class
    tname  - A template name or a list having template data
    foot   - Whether to print a footer on each form, default True
    wrkdir - A directory to be used as a working directory
    """
    def __init__(self, dbm, tname, foot=True, wrkdir=None):
        self.dbm = dbm
        self.tname = tname
        self.wrkdir = wrkdir
        if self.setVariables():
            super().__init__(orientation=self.ortn, fmat=self.pgsz, foot=foot)
            self.set_title(self.titl)
            self.set_author("Tartan Systems")
            self.set_auto_page_break(False, margin=0)
            self.set_fill_color(220)

    def setVariables(self):
        self.bcode = {"x1": "tpd_x1", "y1": "tpd_y1", "x2": "tpd_x2",
            "y2": "tpd_y2", "text": "tpd_text", "size": "tpd_size"}
        self.ccode = {"mrg_x1": "tpd_mrg_x1", "mrg_y1": "tpd_mrg_y1",
            "mrg_x2": "tpd_mrg_x2", "mrg_y2": "tpd_mrg_y2", "text": "tpd_text",
            "mrg_font": "tpd_mrg_font", "mrg_size": "tpd_mrg_size",
            "mrg_colour": "tpd_mrg_colour", "mrg_bold": "tpd_mrg_bold",
            "mrg_italic": "tpd_mrg_italic", "mrg_uline": "tpd_mrg_uline",
            "mrg_align": "tpd_mrg_align", "mrg_border": "tpd_mrg_border",
            "mrg_fill": "tpd_mrg_fill"}
        self.icode = {"x1": "tpd_x1", "y1": "tpd_y1", "x2": "tpd_x2",
            "y2": "tpd_y2", "text": "tpd_text"}
        self.lcode = {"x1": "tpd_x1", "y1": "tpd_y1", "x2": "tpd_x2",
            "y2": "tpd_y2", "font": "tpd_font", "size": "tpd_size",
            "colour": "tpd_colour", "thick": "tpd_thick"}
        self.rcode = {"x1": "tpd_x1", "y1": "tpd_y1", "x2": "tpd_x2",
            "y2": "tpd_y2", "thick": "tpd_thick"}
        self.tcode = {"x1": "tpd_x1", "y1": "tpd_y1", "x2": "tpd_x2",
            "y2": "tpd_y2", "text": "tpd_text", "font": "tpd_font",
            "size": "tpd_size", "colour": "tpd_colour", "bold": "tpd_bold",
            "italic": "tpd_italic", "uline": "tpd_uline", "align": "tpd_align",
            "border": "tpd_border", "fill": "tpd_fill"}
        self.sql = Sql(self.dbm, ["ffield", "tplmst", "tpldet"],
            prog=__name__)
        if type(self.tname) in (list, tuple):
            tplmst = self.tname[0]
            self.tpldet = self.tname[1]
        else:
            tplmst = self.sql.getRec("tplmst", where=[("tpm_tname",
                "=", self.tname)], limit=1)
            if not tplmst:
                showError(None, "Template Error", "Invalid Template Name")
                return
            self.tpldet = self.sql.getRec("tpldet",
                where=[("tpd_tname", "=", self.tname)], order="tpd_detseq")
        self.titl = tplmst[self.sql.tplmst_col.index("tpm_title")]
        self.tptyp = tplmst[self.sql.tplmst_col.index("tpm_type")]
        self.pgsz = tplmst[self.sql.tplmst_col.index("tpm_pgsize")]
        if self.pgsz == "CC":
            self.pgsz = (86, 54)
        elif self.pgsz == "S8":
            self.pgsz = (80, 500)
        elif self.pgsz == "S6":
            self.pgsz = (57, 500)
        elif self.pgsz == "A6":
            self.pgsz = (105, 148)
        self.ortn = tplmst[self.sql.tplmst_col.index("tpm_orient")]
        self.tptp = {}
        self.head = []
        self.body = []
        self.total = []
        self.tail = []
        self.newkey = []
        self.newdic = {}
        for line in self.tpldet:
            mrgcod = line[self.sql.tpldet_col.index("tpd_mrgcod")]
            if mrgcod in tartanWork.tptrtp["G"]["codes"]:
                self.tptp[mrgcod] = copyList(
                    tartanWork.tptrtp["G"]["codes"][mrgcod])
            elif mrgcod in tartanWork.tptrtp[self.tptyp]["codes"]:
                self.tptp[mrgcod] = copyList(
                    tartanWork.tptrtp[self.tptyp]["codes"][mrgcod])
            else:
                fld = self.sql.getRec("ffield", cols=["ff_tabl", "ff_type",
                    "ff_size"], where=[("ff_name", "=", mrgcod)], limit=1)
                if fld:
                    self.tptp[mrgcod] = [[fld[0], fld[1], fld[2], ""], []]
        return True

    def doText(self, x1=0, y1=0, x2=0, y2=0, text="", font="courier", size=10, colour=0, bold=False, italic=False, uline=False, align="", border=0, ln=1, fill=0):
        if text:
            if text != str:
                text = str(text)
            font = font.strip().lower()
            if self.text_color != rgb(colour):
                self.set_text_color(*rgb(colour))
            style = ""
            if bold:
                style += "B"
            if italic:
                style += "I"
            if uline:
                style += "U"
            if align == "C":
                text = text.rstrip()
            self.setFont(font, style, size)
            if border and not self.line_width:
                self.set_line_width(0.2)
            self.set_xy(x1, y1)
            try:
                self.cell(w=x2-x1, h=y2-y1, txt=text, border=border, ln=ln,
                    align=align, fill=fill)
            except Exception as err:
                print(err)

    def doMultiText(self, x1=0, y1=0, x2=0, y2=0, text="", font="courier", size=10, colour=0, bold=False, italic=False, uline=False, align="", border=0, ln=1, fill=0):
        if text:
            font = font.strip().lower()
            if self.text_color != rgb(colour):
                self.set_text_color(*rgb(colour))
            style = ""
            if bold:
                style += "B"
            if italic:
                style += "I"
            if uline:
                style += "U"
            self.setFont(font, style, size)
            if border and not self.line_width:
                self.set_line_width(0.2)
            self.set_xy(x1, y1)
            self.multi_cell(w=x2-x1, h=y2-y1, txt=text, border=border,
                fill=fill)

    def doLine(self, x1=0, y1=0, x2=0, y2=0, font="courier", colour=0, thick=0):
        font = font.strip().lower()
        if self.draw_color != rgb(colour):
            self.set_draw_color(*rgb(colour))
        self.setFont(font)
        self.set_line_width(thick)
        self.line(x1, y1, x2, y2)
        self.set_line_width(0)

    def doRect(self, x1=0, y1=0, x2=0, y2=0, thick=0, colour=0):
        if self.draw_color != rgb(colour):
            self.set_draw_color(*rgb(colour))
        self.set_line_width(thick)
        self.rect(x1, y1, x2-x1, y2-y1)
        self.set_line_width(0)

    def doImage(self, x1=0, y1=0, x2=0, y2=0, text=""):
        if x2:
            w = x2 - x1
        else:
            w = 0
        if y2:
            h = y2 - y1
        else:
            h = 0
        try:
            self.image(text, x1, y1, w, h)
        except Exception as err:
            print(err)

    def doBarcode(self, x1=0, y1=0, x2=0, y2=0, text="", font="interleaved 2of5 nt", size=1, colour=0):
        if self.draw_color != rgb(colour):
            self.set_draw_color(*rgb(colour))
        font = font.lower().strip()
        if font == "interleaved 2of5 nt":
            self.interleaved2of5(text, x1, y1, w=size, h=y2-y1)

    def doDrawDetail(self, line, fmat=True):
        lic = {}
        tdc = self.sql.tpldet_col
        if line[tdc.index("tpd_type")] == "B":
            for key in self.bcode:
                dat = line[tdc.index(self.bcode[key])]
                lic[key] = dat
            self.doBarcode(x1=lic["x1"], y1=lic["y1"], x2=lic["x2"],
                y2=lic["y2"], text=lic["text"], size=lic["size"])
        elif line[tdc.index("tpd_type")] == "C":
            for key in self.ccode:
                dat = line[tdc.index(self.ccode[key])]
                if key == "text":
                    if type(dat) == str and dat.count("BLANK"):
                        dat = dat.replace("BLANK", " ")
                    elif fmat:
                        tp = self.tptp[line[tdc.index("tpd_mrgcod")]][0]
                        if tp[1] == "BL":
                            dat = " "
                        else:
                            dat = CCD(dat, tp[1], tp[2]).disp
                    if not dat and line[tdc.index("tpd_mrg_border")]:
                        dat = " "
                lic[key] = dat
            if line[tdc.index("tpd_lines")] == 1:
                self.doText(x1=lic["mrg_x1"], y1=lic["mrg_y1"],
                    x2=lic["mrg_x2"], y2=lic["mrg_y2"], text=lic["text"],
                    font=lic["mrg_font"], size=lic["mrg_size"],
                    colour=lic["mrg_colour"], bold=lic["mrg_bold"],
                    italic=lic["mrg_italic"], uline=lic["mrg_uline"],
                    align=lic["mrg_align"], border=lic["mrg_border"],
                    ln=1, fill=lic["mrg_fill"])
            else:
                self.doMultiText(x1=lic["mrg_x1"], y1=lic["mrg_y1"],
                    x2=lic["mrg_x2"], y2=lic["mrg_y2"], text=lic["text"],
                    font=lic["mrg_font"], size=lic["mrg_size"],
                    colour=lic["mrg_colour"], bold=lic["mrg_bold"],
                    italic=lic["mrg_italic"], uline=lic["mrg_uline"],
                    align=lic["mrg_align"], border=lic["mrg_border"],
                    ln=0, fill=lic["mrg_fill"])
        elif line[tdc.index("tpd_type")] == "I":
            for key in self.icode:
                dat = line[tdc.index(self.icode[key])]
                lic[key] = dat
            if lic["text"]:
                lic["text"] = getFileName(lic["text"], wrkdir=self.wrkdir)
            if lic["text"] and os.path.isfile(lic["text"]):
                self.doImage(x1=lic["x1"], y1=lic["y1"], x2=lic["x2"],
                    y2=lic["y2"], text=lic["text"])
        elif line[tdc.index("tpd_type")] == "L":
            for key in self.lcode:
                dat = line[tdc.index(self.lcode[key])]
                lic[key] = dat
            self.doLine(x1=lic["x1"], y1=lic["y1"], x2=lic["x2"], y2=lic["y2"],
                font=lic["font"], colour=lic["colour"], thick=lic["thick"])
        elif line[tdc.index("tpd_type")] == "R":
            for key in self.rcode:
                dat = line[tdc.index(self.rcode[key])]
                lic[key] = dat
            self.doRect(x1=lic["x1"], y1=lic["y1"], x2=lic["x2"], y2=lic["y2"],
                thick=lic["thick"])
        elif line[tdc.index("tpd_type")] == "T":
            for key in self.tcode:
                dat = line[tdc.index(self.tcode[key])]
                if key == "text":
                    if type(dat) == str and dat.count("BLANK"):
                        dat = dat.replace("BLANK", " ")
                    elif line[tdc.index("tpd_ttyp")] == "H":
                        mrgcod = line[tdc.index("tpd_mrgcod")]
                        if mrgcod and self.tptp[mrgcod][0][1][0] == "S":
                            dat = "%s " % dat.strip()
                    if not dat and line[tdc.index("tpd_border")]:
                        dat = " "
                lic[key] = dat
            if self.get_string_width(lic["text"]) > (lic["x2"] - lic["x1"]):
                self.doMultiText(x1=lic["x1"], y1=lic["y1"], x2=lic["x2"],
                    y2=lic["y2"], text=lic["text"], font=lic["font"],
                    size=lic["size"], colour=lic["colour"], bold=lic["bold"],
                    italic=lic["italic"], uline=lic["uline"],
                    align=lic["align"], border=lic["border"],
                    ln=0, fill=lic["fill"])
            else:
                self.doText(x1=lic["x1"], y1=lic["y1"], x2=lic["x2"],
                    y2=lic["y2"], text=lic["text"], font=lic["font"],
                    size=lic["size"], colour=lic["colour"], bold=lic["bold"],
                    italic=lic["italic"], uline=lic["uline"],
                    align=lic["align"], border=lic["border"],
                    ln=1, fill=lic["fill"])

    def doNewDetail(self):
        tdc = self.sql.tpldet_col
        self.maxlines = 0
        self.y2 = 0
        self.mrg_y2 = 0
        for line in self.tpldet:
            mrgcod = line[tdc.index("tpd_mrgcod")]
            if mrgcod and line[tdc.index("tpd_place")] == "A":
                self.head.append(mrgcod)
            elif line[tdc.index("tpd_place")] == "B":
                self.body.append(mrgcod)
                if not self.maxlines and line[-1] > 1:
                    self.maxlines = line[-1]
            elif line[tdc.index("tpd_place")] == "C":
                if mrgcod:
                    if mrgcod == "carried_forward":
                        self.cfwd = self.doToggleStyle(line)
                        continue
                    self.total.append(mrgcod)
                else:
                    self.total.append(line[tdc.index("tpd_detseq")])
            elif line[tdc.index("tpd_place")] == "D":
                if mrgcod:
                    self.tail.append(mrgcod)
                else:
                    self.tail.append(line[tdc.index("tpd_detseq")])
            self.doNewLines(line)

    def doNewLines(self, line):
        nl = copyList(line)
        tdc = self.sql.tpldet_col
        mrgcod = nl[tdc.index("tpd_mrgcod")]
        if mrgcod:
            if mrgcod not in self.tptp:
                return
            nl[tdc.index("tpd_detseq")] = mrgcod
        detseq = nl[tdc.index("tpd_detseq")]
        repeat = nl[tdc.index("tpd_repeat")]
        if not repeat:
            repeat = 1
        nl = self.doToggleStyle(nl)
        if nl[tdc.index("tpd_type")] in ("I", "R"):
            if mrgcod:
                nl[tdc.index("tpd_text")] = self.doGetData(mrgcod)
                nl[tdc.index("tpd_mrgcod")] = ""
            self.newkey.append(detseq)
            self.newdic[detseq] = nl
        elif nl[tdc.index("tpd_type")] == "C":
            if nl[tdc.index("tpd_ttyp")] in ("H", "L"):
                tl = copyList(nl)
                tl[tdc.index("tpd_type")] = "T"
                if tl[tdc.index("tpd_place")] == "B" and \
                        tl[tdc.index("tpd_ttyp")] == "H":
                    tl[tdc.index("tpd_place")] = "A"
                if not tl[tdc.index("tpd_y1")]:
                    height = tl[tdc.index("tpd_y2")]
                    t_y1, t_y2 = self.getY(tl[tdc.index("tpd_y2")])
                    tl[tdc.index("tpd_y1")] = t_y1
                    tl[tdc.index("tpd_y2")] = t_y2
                else:
                    height = tl[tdc.index("tpd_y2")] - \
                        tl[tdc.index("tpd_y1")]
                if tl[tdc.index("tpd_ttyp")] == "H":
                    times = 1
                else:
                    times = repeat
                for x in range(times):
                    newseq = "%s_T%02i" % (detseq, x)
                    tl[tdc.index("tpd_detseq")] = newseq
                    if tl[tdc.index("tpd_border")] and x and x == times-1:
                        tl[tdc.index("tpd_border")] += "B"
                    self.newkey.append(newseq)
                    self.newdic[newseq] = copyList(tl)
                    t_y2 = tl[tdc.index("tpd_y2")]
                    tl[tdc.index("tpd_y1")] = t_y2
                    tl[tdc.index("tpd_y2")] = t_y2 + height
            nl[tdc.index("tpd_text")] = self.doGetData(mrgcod)
            if not nl[tdc.index("tpd_mrg_y1")]:
                height = nl[tdc.index("tpd_mrg_y2")]
                c_y1, c_y2 = self.getY(height)
                nl[tdc.index("tpd_mrg_y1")] = c_y1
                nl[tdc.index("tpd_mrg_y2")] = c_y2
            else:
                height = nl[tdc.index("tpd_mrg_y2")] - \
                    nl[tdc.index("tpd_mrg_y1")]
            if nl[tdc.index("tpd_ttyp")] in ("H", "L"):
                self.y2 = t_y2
            for x in range(repeat):
                newseq = "%s_C%02i" % (detseq, x)
                nl[tdc.index("tpd_detseq")] = newseq
                if nl[tdc.index("tpd_mrg_border")] and x and x == repeat-1:
                    nl[tdc.index("tpd_mrg_border")] += "B"
                self.newkey.append(newseq)
                self.newdic[newseq] = copyList(nl)
                c_y2 = nl[tdc.index("tpd_mrg_y2")]
                nl[tdc.index("tpd_mrg_y1")] = c_y2
                nl[tdc.index("tpd_mrg_y2")] = c_y2 + height
            self.mrg_y2 = c_y2
        else:
            if not nl[tdc.index("tpd_y1")]:
                y1, y2 = self.getY(nl[tdc.index("tpd_y2")])
                nl[tdc.index("tpd_y1")] = y1
                nl[tdc.index("tpd_y2")] = y2
            self.y2 = nl[tdc.index("tpd_y2")]
            if nl[tdc.index("tpd_type")] == "T":
                if len(nl[tdc.index("tpd_text")]) > nl[tdc.index("tpd_chrs")]:
                    nl[tdc.index("tpd_lines")] = 2
            self.newkey.append(detseq)
            self.newdic[detseq] = nl

    def getY(self, height):
        y1 = 0
        if self.y2:
            y1 = self.y2
        if self.mrg_y2 and self.mrg_y2 > y1:
            y1 = self.mrg_y2
        y2 = y1 + height
        return y1, y2

    def doToggleStyle(self, nl):
        for typ in ("bold", "italic", "uline", "fill"):
            if nl[self.sql.tpldet_col.index("tpd_%s" % typ)] in ("N", ""):
                nl[self.sql.tpldet_col.index("tpd_%s" % typ)] = 0
            else:
                nl[self.sql.tpldet_col.index("tpd_%s" % typ)] = 1
            if nl[self.sql.tpldet_col.index("tpd_mrg_%s" % typ)] in ("N", ""):
                nl[self.sql.tpldet_col.index("tpd_mrg_%s" % typ)] = 0
            else:
                nl[self.sql.tpldet_col.index("tpd_mrg_%s" % typ)] = 1
        return nl

    def doSplitText(self, key, text):
        tdc = self.sql.tpldet_col
        x1 = self.newdic[key][tdc.index("tpd_mrg_x1")]
        x2 = self.newdic[key][tdc.index("tpd_mrg_x2")]
        y1 = self.newdic[key][tdc.index("tpd_mrg_y1")]
        y2 = self.newdic[key][tdc.index("tpd_mrg_y2")]
        return self.multi_cell(w=x2-x1, h=y2-y1, txt=text, split_only=True)

    def doGetData(self, mrgcod):
        if mrgcod and mrgcod in self.tptp and self.tptp[mrgcod][1]:
            return self.tptp[mrgcod][1]
        return ""

    # Standard Routines
    def letterhead(self, cmc, ctm, fld, img):
        if not img and "LETTERHEAD" in os.environ:
            img = os.environ["LETTERHEAD"]
        if not img:
            img = getFileName(ctm[cmc.index("ctm_logo")], wrkdir=self.wrkdir)
        else:
            img = getFileName(img, wrkdir=self.wrkdir)
        if not img:
            del self.tptp[fld]
            for col in ("ctm_regno", "ctm_taxno", "ctm_tel", "ctm_fax"):
                if col in self.tptp:
                    dat = ctm[cmc.index(col)]
                    if dat:
                        self.tptp[col][1] = dat
                    else:
                        del self.tptp[col]
            return
        self.tptp[fld][1] = img
        for col in ("ctm_name", "ctm_add1", "ctm_add2", "ctm_add3",
                "ctm_pcode", "ctm_regno", "ctm_taxno", "ctm_tel", "ctm_fax"):
            if col in self.tptp:
                del self.tptp[col]

    def document_number(self, number):
        if "document_number" in self.tptp:
            self.tptp["document_number"][1] = number

    def document_date(self, date):
        if "document_date" in self.tptp:
            self.tptp["document_date"][1] = date

    def bank_details(self, col, mst, seq=0):
        if "bank_details" in self.tptp:
            if mst[col.index("ctm_b_name")]:
                dat = "Name:    %s" % mst[col.index("ctm_b_name")]
                dat = "%s\nBranch:  %s" % (dat, mst[col.index("ctm_b_branch")])
                dat = "%s\nCode:    %s" % (dat, mst[col.index("ctm_b_ibt")])
                dat = "%s\nAccount: %s\n " % (dat, mst[col.index("ctm_b_acno")])
                if not seq:
                    self.tptp["bank_details"][1] = dat
                else:
                    tdc = self.sql.tpldet_col
                    self.newdic["bank_details_C00"][tdc.index("tpd_text")] = dat
            elif not seq:
                del self.tptp["bank_details"]

    def account_details(self, cod, col, mst, seq=0):
        if "account_details" in self.tptp:
            dat = mst[col.index("%s_name" % cod)]
            try:
                dat = "%1s\n%1s" % (dat, mst[col.index("%s_add1" % cod)])
                dat = "%1s\n%1s" % (dat, mst[col.index("%s_add2" % cod)])
                dat = "%1s\n%1s" % (dat, mst[col.index("%s_add3" % cod)])
                try:
                    dat = "%1s\n%1s" % (dat, mst[col.index("%s_pcod" % cod)])
                except:
                    dat = "%1s\n%1s" % (dat, mst[col.index("%s_pcode" % cod)])
            except:
                dat = "%1s\n%1s" % (dat, mst[col.index("%s_addr1" % cod)])
                dat = "%1s\n%1s" % (dat, mst[col.index("%s_addr2" % cod)])
                dat = "%1s\n%1s" % (dat, mst[col.index("%s_addr3" % cod)])
                dat = "%1s\n%1s" % (dat, mst[col.index("%s_pcode" % cod)])
            if not seq:
                self.tptp["account_details"][1] = dat
            else:
                tdc = self.sql.tpldet_col
                self.newdic["account_details_C00"][tdc.index("tpd_text")] = dat

    def name_init(self, col, mst, seq=0):
        if "name_init" in self.tptp:
            dat = mst[col.index("mlm_names")].split()[0][0].upper()
            dat = "%s %s" % (dat, mst[col.index("mlm_surname")])
            if not seq:
                self.tptp["name_init"][1] = dat
            else:
                tdc = self.sql.tpldet_col
                self.newdic["name_init"][tdc.index("tpd_text")] = dat

    def changeSize(self, pdfnam):
        doc = fitz.open(pdfnam)
        try:
            mbox = doc[0].mediabox
        except:
            mbox = doc[0].MediaBox
        mbox[1] = float(mbox[3] - (self.get_y() * 3))
        try:
            doc[0].set_mediabox(mbox)
        except:
            doc[0].setMediaBox(mbox)
        doc2 = fitz.open()
        try:
            doc2.insert_pdf(doc, from_page=0, to_page=0)
        except:
            doc2.insertPDF(doc, from_page=0, to_page=0)
        doc2.save(pdfnam)
        doc.close()
        doc2.close()

class ToolTip(object):
    def __init__(self, widget, text, font=None, color=("black","light yellow")):
        self.widget = widget
        self.text = text
        if not font:
            self.font = "TkTooltipFont"
        else:
            self.font = font
        self.color = color
        if widget.winfo_class() == "TButton":
            self.pause = None
        else:
            self.pause = len(text) * 40
        self.tipwindow = None
        self.x = self.y = 0
        self.widget.bind("<Enter>", self.showTip)
        self.widget.bind("<Leave>", self.hideTip)
        self.widget.bind("<FocusIn>", self.showTip)
        self.widget.bind("<FocusOut>", self.hideTip)

    def showTip(self, event):
        # Display text in tooltip window if widget active
        if self.tipwindow:
            return
        try:
            if str(self.widget.cget("state")) == "disabled":
                return
        except:
            pass
        x = self.widget.winfo_rootx()
        y = self.widget.winfo_rooty() + self.widget.winfo_height()
        self.tipwindow = tw = tk.Toplevel(self.widget)
        self.tipwindow.withdraw()
        tw.overrideredirect(True)
        label = MyLabel(tw, color=self.color, text=self.text, justify="left",
            relief="solid", borderwidth=2, font=self.font, wraplength=400)
        label.pack(ipadx=1, ipady=2)
        tw.update_idletasks()
        ww = label.winfo_reqwidth()
        if x + ww > label.winfo_screenwidth():
            x = x - ww
        tw.geometry("+%d+%d" % (x, y))
        self.tipwindow.deiconify()
        if self.pause and event.type == "9":
            label.after(self.pause, self.hideTip)

    def hideTip(self, event=None):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

class AboutTartan(object):
    def __init__(self, mf, ver):
        self.mf = mf
        self.mf.destroyChildren()
        self.about = MyFrame(self.mf.window, borderwidth=2, bg="black")
        self.about.place(anchor="center", relx=0.5, rely=0.5)
        info = """
Tartan Systems %s

%s 2004-%s Paul Malherbe.
All Rights Reserved.

Tartan is free software distributed under the terms
of the GNU General Public License.

For information and or support for this product please contact:

Paul Malherbe

Phone:             27-28-3165036
Mobile:            27-82-9005260
""" % (ver, chr(169), time.localtime().tm_year)
        label = MyText(self.about, font=("Helvetica", 14), width=65,
            height=19, takefocus=False)
        label.insert("insert", info)
        label.insert("end", "Email:         ")
        link = HyperlinkManager(label)
        label.insert("end", "paul@tartan.co.za",
            link.add(self.linkMail))
        label.insert("end", "\nWeb:            ")
        label.insert("end", "www.tartan.co.za",
            link.add(self.linkWeb))
        label.pack()
        label.tag_add("text", "1.0", "end")
        label.tag_config("text", justify="center")
        label.configure(state="disabled")
        bbox = MyButtonBox(self.about, row=1)
        self.b0 = bbox.addButton("Licence", self.licenceAbout)
        if changes:
            self.b1 = bbox.addButton("Changes", self.changesAbout)
        else:
            self.b1 = None
        self.b2 = bbox.addButton("Exit", self.exitAbout)
        self.mf.startLoop()

    def linkMail(self, *args):
        try:
            web = "mailto:paul@tartan.co.za"
            if sys.platform == "win32":
                os.startfile(web)
            else:
                subprocess.call(["xdg-open", web])
            self.exitAbout()
        except:
            showError(self.mf.window, "Browser Error",
                "Cannot Load Browser or URL")

    def linkWeb(self, *args):
        try:
            web = "https://www.tartan.co.za"
            webbrowser.open_new(web)
            self.exitAbout()
        except:
            showError(self.mf.window, "Browser Error",
                "Cannot Load Browser or URL")

    def changesAbout(self, *args):
        self.b0.configure(state="disabled")
        self.b1.configure(state="disabled")
        self.b2.configure(state="disabled")
        self.about.place_forget()
        ScrollText(scrn=self.mf.window, mess=changes)
        self.about.place(anchor="center", relx=0.5, rely=0.5)
        self.b0.configure(state="normal")
        self.b1.configure(state="normal")
        self.b2.configure(state="normal")

    def licenceAbout(self, *args):
        self.b0.configure(state="disabled")
        if self.b1:
            self.b1.configure(state="disabled")
        self.b2.configure(state="disabled")
        self.about.place_forget()
        try:
            docdir = os.path.join(getPrgPath(), "doc")
            if HTML and os.path.isfile(os.path.join(docdir, "gnugpl.html")):
                fle = "gnugpl.html"
            elif os.path.isfile(os.path.join(docdir, "gnugpl.md")):
                fle = "gnugpl.md"
            elif os.path.isfile(os.path.join(docdir, "gnugpl.txt")):
                fle = "gnugpl.txt"
            else:
                raise Exception
            doPublish(self.mf.window, os.path.join(docdir, fle))
        except:
            pass
        self.about.place(anchor="center", relx=0.5, rely=0.5)
        self.b0.configure(state="normal")
        if self.b1:
            self.b1.configure(state="normal")
        self.b2.configure(state="normal")

    def exitAbout(self, *args):
        self.about.destroy()
        self.mf.createChildren()
        self.mf.closeLoop()

class HyperlinkManager(object):
    def __init__(self, text):
        self.text = text
        self.text.tag_config("hyper", foreground="blue", underline=1)
        self.text.tag_bind("hyper", "<Enter>", self._enter)
        self.text.tag_bind("hyper", "<Leave>", self._leave)
        self.text.tag_bind("hyper", "<ButtonRelease-1>", self._click)
        self.reset()

    def reset(self):
        self.links = {}

    def add(self, action):
        tag = "hyper-%d" % len(self.links)
        self.links[tag] = action
        return "hyper", tag

    def _enter(self, event):
        self.text.config(cursor="hand2")

    def _leave(self, event):
        self.text.config(cursor="arrow")

    def _click(self, event):
        for tag in self.text.tag_names("current"):
            if tag[:6] == "hyper-":
                self.links[tag]()
                return

class SimpleDialog(object):
    """
    This is a simple input dialog for simple applications.

    parent  -   The parent widget or None
    trans   -   The transient window
    modal   -   If modal, True or False
    decor   -   Decorate, True or False
    style   -   The style to apply to the frame
    cols    -   A list of lists/tuples of entry fields as follows:
                    ("name", "label details", size, "type") e.g.
                cols=[
                    ("a", "Number", 7, "UI"),
                    ("b", "Surname", 20, "TX"),
                    ("c", "Names", 20, "TX"),
                    ("d", "Group", 30, "TX"),
                    ("e", "Arrival", 10, "d1")]
    butt    -  A list of lists/tuples of buttons as follows:
                    ("text", "command") e.g.
                butt=[
                    ("Continue", abcd),
                    ("Cancel", wxyz)]
    title   - The widget's title
    conf    - Whether or not to ask for confirmation
    """
    def __init__(self, parent=None, trans=None, modal=False, decor=True, style="TFrame", pad=2, bd=2, cols=None, butt=None, title=None, conf=False):
        self.parent = parent
        if not self.parent:
            mkw = MkWindow(trans=trans, decor=decor, remov=False)
            self.window = mkw.newwin
        else:
            self.window = self.parent
        self.cols = cols
        self.conf = conf
        self.data = []
        self.ents = {}
        if not parent and not style:
            style = ttk.Style()
            style.theme_use("clam")
            fonts = (
                ("TkDefaultFont", "Courier", 12, "normal"),
                ("TkTextFont", "Courier", 12, "normal"),
                ("TkFixedFont", "Courier", 12, "normal"),
                ("TkMenuFont", "Courier", 12, "normal"),
                ("TkHeadingFont", "Courier", 12, "normal"))
            for font in fonts:
                nf = tkfont.nametofont(font[0])
                if len(font) != 4:
                    nf.configure(family=font[1], size=font[2])
                else:
                    nf.configure(family=font[1], size=font[2], weight=font[3])
                self.sframe.option_add(font[0], nf)
        self.sframe = MyFrame(self.window, bg="black", borderwidth=bd,
            style=style, padding=pad)
        self.sframe.place(anchor="center", relx=0.5, rely=0.5)
        if title:
            head = MyLabel(self.sframe, text=title, color=True, borderwidth=2,
                anchor="c", padding=0, relief="ridge")
            head.pack(fill="x", expand="yes")
        nframe = MyFrame(self.sframe, padding=2)
        nframe.pack(fill="both", expand="yes")
        for num, col in enumerate(self.cols):
            lab = MyLabel(nframe, color=False, text=col[1])
            lab.grid(row=num, column=0, sticky="w")
            if col[3] == "TV":
                wid = MyText(nframe, width=col[2][0], height=col[2][1])
                wid.bind("<F9>", functools.partial(self.goNext, num))
            elif col[3] == "FF":
                wid = MyEntry(nframe, maxsize=0, width=int(col[2]),
                    cmd=(self.goNext, num))
                wid.bind("<F1>", functools.partial(self.goNext, num))
            elif col[3] == "TX":
                wid = MyEntry(nframe, width=int(col[2]), cmd=(self.goNext, num))
            else:
                wid = MyEntry(nframe, maxsize=int(col[2]), width=int(col[2]),
                    cmd=(self.goNext, num))
                if col[3][0] == "H":
                    wid.configure(show="*")
            wid.bind("<Escape>", functools.partial(self.goBack, num, col))
            wid.grid(row=num, column=1, sticky="w")
            self.ents[col[0]] = wid
        txt = "Enter %s" % self.cols[0][1]
        if self.cols[0][3] == "TV":
            txt += " then <F9> to Accept or <Esc> to Exit"
        else:
            txt += " or <Esc> to Exit"
        self.stat = MyLabel(nframe, text=txt, color=False, borderwidth=2,
            anchor="w", background="white", foreground="black", padding=0,
            relief="ridge")
        self.stat.grid(row=num+1, column=0, columnspan=2, pady=2, sticky="ew")
        if butt:
            bbox = MyButtonBox(self.sframe, row=1)
            for but in butt:
                bbox.addButton(but[0], but[1])
        if not parent:
            self.window.update_idletasks()
            geo = (self.sframe.winfo_reqwidth(), self.sframe.winfo_reqheight())
            placeWindow(self.window, parent=trans, size=geo, expose=True)
            if modal:
                self.window.wait_visibility()
                self.window.grab_set()
        self.fld = self.ents[self.cols[0][0]]
        self.fld.focus_set()

    def goBack(self, num, col, event=None):
        if not num:
            self.sframe.destroy()
            if not self.parent:
                self.window.destroy()
        else:
            self.fld = self.ents[self.cols[num - 1][0]]
            self.fld.focus_set()
            if self.fld.winfo_class() == "Text":
                txt = "Enter %s then <F9> to Accept or <Esc> to" % \
                    self.cols[num - 1][1]
            else:
                self.fld.selection_range(0, "end")
                txt = "Enter %s or <Esc> to" % self.cols[num - 1][1]
            if num == 1:
                txt += " Exit"
            else:
                txt += " go Back"
            self.stat.configure(text=txt)

    def goNext(self, num, event=None):
        col = self.cols[num]
        if event and event.keysym == "F1":
            dia = FileDialog(**{
                "parent": self.sframe,
                "title": "Attachment",
                "initd": ".", "multi": True})
            fle = dia.askopenfilename()
            if not fle:
                return
            dat = ""
            for f in fle:
                if not dat:
                    dat = f
                else:
                    dat = "%s,%s" % (dat, f)
            self.ents[col[0]].insert(0, dat)
        typ = col[3]
        if typ == "TV":
            siz = col[2][0]
            dat = CCD(self.ents[col[0]].get("1.0", "end"), typ, siz)
        else:
            siz = col[2]
            dat = CCD(self.ents[col[0]].get(), typ, siz)
        if dat.err:
            self.ents[col[0]].focus_set()
            if self.ents[col[0]].winfo_class() != "Text":
                self.ents[col[0]].selection_range(0, "end")
            self.stat.configure(text="Error, %s" %
                self.stat.cget("text").replace("Error, ", ""))
        else:
            if not dat.work:
                dsp = ""
            else:
                dsp = dat.disp
            if typ == "TV":
                self.ents[col[0]].delete("1.0", "end")
                self.ents[col[0]].insert("1.0", dsp)
            else:
                self.ents[col[0]].delete(0, "end")
                self.ents[col[0]].insert(0, dsp)
            if num == len(self.cols) - 1:
                self.doEnd()
            else:
                self.fld = self.ents[self.cols[num + 1][0]]
                self.fld.focus_set()
                if self.fld.winfo_class() == "Text":
                    txt = "Enter %s then <F9> to Accept or "\
                        "<Esc> to Go Back" % self.cols[num + 1][1]
                else:
                    self.fld.selection_range(0, "end")
                    txt = "Enter %s or <Esc> to Go Back" % \
                        self.cols[num + 1][1]
                self.stat.configure(text=txt)

    def doEnd(self):
        if self.conf:
            ok = askQuestion(self.sframe, "Confirm",
                "Do You wish to Accept All Entries.",
                default="yes")
            if ok == "no":
                self.fld.focus_set()
                return
        for col in self.cols:
            typ = col[3]
            if typ == "TV":
                siz = col[2][0]
                dat = CCD(self.ents[col[0]].get("1.0", "end"), typ, siz)
            else:
                siz = col[2]
                dat = CCD(self.ents[col[0]].get(), typ, siz)
            self.data.append(dat.work)
        self.sframe.destroy()
        if not self.parent:
            self.window.destroy()
        else:
            self.window.update()

class ExportDbase(object):
    def __init__(self, **opts):
        self.opts = opts
        if self.doVariables():
            if "args" in opts:
                self.cono = eval(self.opts["args"][0])
                self.dir = self.opts["args"][1]
                self.nam = self.opts["args"][2]
                self.doEnd()
            else:
                self.doMain()
                self.opts["mf"].startLoop()

    def doVariables(self):
        self.noco = []
        self.coys = []
        self.skip = ("ldraws", "lentry", "lfixed", "ltrans", "lprize")
        for tab in tartanWork.tabdic:
            if tab in self.skip:
                continue
            if tartanWork.tabdic[tab]["fld"][0][1].count("_cono"):
                self.coys.append(tab)
            else:
                self.noco.append(tab)
        self.noco.extend(["ctlpwr", "genrpt"])
        self.coys.remove("ctlpwr")
        self.coys.remove("genrpt")
        self.sqf = Sql(self.opts["mf"].dbm, self.noco + self.coys,
            prog=__name__)
        if self.sqf.error:
            return
        self.cv = self.sqf.getRec("verupd", cols=["ver_version"],
            limit=1)[0]
        self.wrkdir = self.opts["mf"].rcdic["wrkdir"]
        return True

    def doMain(self):
        tit = "Database Export Routine"
        coy = {
            "stype": "R",
            "tables": ("ctlmst",),
            "cols": (
                ("ctm_cono", "", 0, "Coy"),
                ("ctm_name", "", 0, "Name"))}
        pth = {
            "stype":  "F",
            "types":  "dir",
            "initd":  ""}
        r1s = (("Single", "S"), ("Multi", "M"))
        fld = (
            (("T",0,0,0),("IRB",r1s),0,"Company(s)","",
                "S","N",self.doTyp,None,None,None),
            (("T",0,1,0),"I@ctm_cono",0,"","",
                "","N",self.doCoy,coy,None,("notzero",)),
            (("T",0,1,0),"ONA",46,""),
            (("T",0,2,0),"ITX",50,"Directory","",
                self.wrkdir,"N",self.doDir,pth,None,("dir",)),
            (("T",0,3,0),"ITX",50,"File Name","",
                "","N",self.doNam,None,None,("efld",)))
        but = (("Exit", None, self.doExit, 0, ("T",0,1), ("T",0,0)),)
        tnd = ((self.doEnd,"y"),)
        txt = (self.doExit,)
        self.df = TartanDialog(self.opts["mf"], tops=True, title=tit,
            eflds=fld, butt=but, tend=tnd, txit=txt)

    def doTyp(self, frt, pag, r, c, p, i, w):
        if w == "M":
            self.cono = []
            self.df.setWidget(self.df.mstFrame, state="hide")
            coys = getSingleRecords(self.opts["mf"], "ctlmst",
                ("ctm_cono", "ctm_name"))
            self.df.setWidget(self.df.mstFrame, state="show")
            for coy in coys:
                self.cono.append(coy[0])
            if not self.cono:
                return "No Companies Selected"
            if not self.doChkInco():
                return "rf"
            return "sk1"

    def doCoy(self, frt, pag, r, c, p, i, w):
        acc = self.sqf.getRec("ctlmst", where=[("ctm_cono", "=", w),
            ("ctm_cono", "=", w)], limit=1)
        if not acc:
            return "Invalid Company"
        self.cono = [w]
        self.df.loadEntry(frt, pag, p+1,
            data=acc[self.sqf.ctlmst_col.index("ctm_name")])
        if not self.doChkInco():
            return "rf"

    def doChkInco(self):
        mis = []
        if len(self.cono) == 1:
            tx1 = "Company."
        else:
            tx1 = "Companies."
        for coy in self.cono:
            cti = self.sqf.getRec("genint", cols=["cti_inco"],
                where=[("cti_cono", "=", coy)], order="cti_inco")
            for c in cti:
                if c[0] not in self.cono:
                    mis.append(c[0])
        if mis:
            if len(mis) == 1:
                tx2 = "This Company %s is " % mis
            else:
                tx2 = "These Companies %s are " % mis
            self.df.setWidget(self.df.mstFrame, state="hide")
            ok = askQuestion(self.opts["mf"].body, "Intercompany",
                """%s Linked to the Selected %s

Do you want to Export All Linked Companies?""" % (tx2, tx1), default="no")
            self.df.setWidget(self.df.mstFrame, state="show")
            if ok == "no":
                return
            for c in mis:
                self.cono.append(c)
        self.cono.sort()
        return True

    def doDir(self, frt, pag, r, c, p, i, w):
        self.dir = w
        nam = "%s%03i.db" % (self.opts["mf"].dbm.dbname, self.cono[0])
        self.df.topf[0][4][5] = nam

    def doNam(self, frt, pag, r, c, p, i, w):
        self.nam = w

    def doEnd(self):
        if "args" not in self.opts:
            self.df.closeProcess()
        if os.path.exists(self.nam):
            os.remove(self.nam)
        rcdic = loadRcFile(self.nam, default=True)
        rcdic["dbname"] = self.nam
        rcdic["dbdir"] = self.dir
        self.dbm = Dbase(rcdic=rcdic, screen=self.opts["mf"].body)
        opts = [
            ("-c", "i"),
            ("-d", True),
            ("-l", self.opts["mf"].body),
            ("-u", rcdic["dbuser"]),
            ("-p", rcdic["dbpwd"]),
            ("-v", self.cv),
            ("-x", True)]
        DBCreate(dbm=self.dbm, opts=opts)
        self.dbm.openDbase()
        self.dbm.cu.execute("PRAGMA JOURNAL_MODE=OFF")
        self.dbm.cu.execute("PRAGMA SYNCHRONOUS=OFF")
        self.sqt = Sql(self.dbm, self.noco + self.coys, prog=__name__)
        mxs = len(self.noco) + len(self.coys)
        num = 0
        if "args" not in self.opts:
            pb = ProgressBar(self.opts["mf"].body, mxs=mxs,
                typ="Exporting ... Please Wait")
        for tab in self.noco:
            if "args" not in self.opts:
                pb.displayProgress(num)
            if tab in self.skip:
                continue
            self.doTable(tab)
            num += 1
        self.only = True
        if 1 not in self.cono:
            self.only = True
        for tab in self.coys:
            if "args" not in self.opts:
                pb.displayProgress(num)
            if tab in self.skip:
                continue
            for cnt, coy in enumerate(self.cono):
                self.doTable(tab, coy, cnt)
            num += 1
        if "args" not in self.opts:
            pb.closeProgress()
        coy = ""
        for c in self.cono:
            if not coy:
                coy = str(c)
            else:
                coy = "%s,%s" % (coy, str(c))
        self.sqt.updRec("ctlpwu", cols=["usr_acoy", "usr_lvl"], data=[coy, 6])
        # Create Deletion table
        self.sqt.sqlRec("Create table if not exists delrec ("\
            "dd_tab varchar(6), dd_dat %s)" % self.dbm.txt)
        self.dbm.commitDbase()
        self.dbm.closeDbase()
        if "args" not in self.opts:
            self.opts["mf"].closeLoop()

    def doTable(self, tab, cono=None, num=0):
        self.opts["mf"].updateStatus(tab)
        if not num:
            self.sqt.delRec(tab)
        cols = getattr(self.sqf, "%s_col" % tab)
        flds = getattr(self.sqf, "%s_fld" % tab)
        pfx = cols[0].split("_")[0]
        xfl = "%s_xflag" % pfx
        if cono and tab != "genrpt":
            if tab == "ctlmst" and self.only:
                self.only = False
                whr = [("ctm_cono", "in", (1, cono))]
            else:
                whr = [(cols[0], "=", cono)]
        else:
            whr = None
        recs = self.sqf.getRec(tables=tab, where=whr)
        for rec in recs:
            try:
                if xfl in cols:
                    rec[cols.index(xfl)] = "Y"
                for chg in ("mlm_idnum", "wgm_idno"):
                    if chg in cols:
                        rec[cols.index(chg)] = int(rec[cols.index(chg)])
                self.dbm.cu.execute("Insert into %s (%s) values %s" %
                    (tab, flds, tuple(rec)))
            except:
                pass

    def doExit(self):
        self.df.closeProcess()
        self.opts["mf"].closeLoop()

class MergeDbase(object):
    def __init__(self, **opts):
        self.opts = opts
        if self.doVariables():
            if "args" in self.opts:
                self.fle = self.opts["args"]
                self.doEnd()
            else:
                self.doMain()
                self.opts["mf"].startLoop()

    def doVariables(self):
        self.tabs = list(tartanWork.tabdic.keys())
        self.sqt = Sql(self.opts["mf"].dbm, self.tabs, prog=__name__)
        if self.sqt.error:
            return
        self.dbnam = self.opts["mf"].dbm.dbname
        self.cv = self.sqt.getRec("verupd", cols=["ver_version"],
            limit=1)[0]
        self.wrkdir = self.opts["mf"].rcdic["wrkdir"]
        return True

    def doMain(self):
        tit = "Database Merge Routine"
        fle = {
            "stype": "F",
            "types": "fle",
            "initd": self.wrkdir,
            "ftype": [["Database File", "*.db"]]}
        fld = (
            (("T",0,0,0),"ITX",50,"Merge File","",
                "","N",self.doFle,fle,None,("fle",)),)
        but = (("Exit", None, self.doExit, 0, ("T",0,1), ("T",0,0)),)
        tnd = ((self.doEnd,"y"),)
        txt = (self.doExit,)
        self.df = TartanDialog(self.opts["mf"], tops=True, title=tit,
            eflds=fld, butt=but, tend=tnd, txit=txt)

    def doFle(self, frt, pag, r, c, p, i, w):
        self.fle = w

    def doEnd(self):
        if "args" not in self.opts:
            self.df.closeProcess()
        dbnam = os.path.basename(self.fle)
        dbdir = os.path.dirname(self.fle)
        rcdic = loadRcFile(dbnam, default=True)
        rcdic["dbname"] = dbnam
        rcdic["dbdir"] = dbdir
        txt = None
        tab = None
        try:
            self.dbm = Dbase(rcdic=rcdic, screen=self.opts["mf"].body)
            self.dbm.openDbase()
            self.sqf = Sql(self.dbm, self.tabs, prog=__name__)
            if not self.dbm.checkTable("delrec"):
                txt = "This is Not a Valid Exported Database"
                raise Exception
            cv = self.sqf.getRec("verupd", cols=["ver_version"],
                limit=1)[0]
            if cv != self.cv:
                txt = """The Versions of the Two Databases are Not the Same.

Database: %s %s
Database: %s %s

Only Databases of the Same Version can be Merged.""" % (self.dbnam,
                    self.cv, dbnam, cv)
                raise Exception
            err = []
            c1 = self.sqf.getRec("ctlmst", cols=["ctm_cono", "ctm_name"],
                order="ctm_cono")
            for c in c1:
                c2 = self.sqt.getRec("ctlmst", cols=["ctm_name"],
                    where=[("ctm_cono", "=", c[0])], limit=1)
                if not c2:
                    err.append((c[0], c[1], "Not Found"))
                elif c2[0] != c[1]:
                    err.append((c[0], c[1], c2[0]))
            if err:
                txt = "Different Company Names for Company %03i\n" % err[0][0]
                txt += "---------------------------------------\n"
                txt += "Original:  %-s\n" % err[0][1]
                txt += "Importing: %-s" % err[0][2]
                raise Exception
            if self.dbm.checkTable("delrec"):
                self.tabs.insert(0, "delrec")
            if "args" not in self.opts:
                pb = ProgressBar(self.opts["mf"].body, mxs=len(self.tabs),
                    typ="Merging ... Please Wait")
            for num, tab in enumerate(self.tabs):
                if "args" not in self.opts:
                    pb.displayProgress(num)
                self.doTable(tab)
            if "args" not in self.opts:
                pb.closeProgress()
            if "args" in self.opts:
                ask = False
            else:
                ask = True
            self.opts["mf"].dbm.commitDbase(ask=ask)
            if self.opts["mf"].dbm.commit == "yes":
                # If db 'to' is committed then commit db 'from'
                self.dbm.dropTable("delrec")
                self.dbm.commitDbase()
            self.dbm.closeDbase()
            if "args" not in self.opts:
                self.opts["mf"].closeLoop()
        except Exception as err:
            if not txt:
                txt = "Merge Error (%s), Cannot Merge this File\n\n(%s)" % \
                    (tab, err)
            showError(self.opts["mf"].body, "Error", txt)
            self.dbm.closeDbase()
            if "args" not in self.opts:
                self.opts["mf"].closeLoop()

    def doTable(self, tab):
        self.opts["mf"].updateStatus(tab)
        if tab == "delrec":
            # Deleted Records
            recs = self.sqf.sqlRec("Select * from delrec")
            for rec in recs:
                cols = getattr(self.sqf, "%s_col" % rec[0])
                pfx = cols[0].split("_")[0]
                xfl = "%s_xflag" % pfx
                data = eval(rec[1])
                if xfl in cols:
                    data[cols.index(xfl)] = ""
                self.sqt.delRec(rec[0], data=data)
        else:
            # New Records
            cols = getattr(self.sqf, "%s_col" % tab)
            pfx = cols[0].split("_")[0]
            xfl = "%s_xflag" % pfx
            if xfl in cols:
                whr = [(xfl, "==", "")]
                recs = self.sqf.getRec(tables=tab, where=whr)
                if recs:
                    self.sqt.insRec(tab, data=recs)
                self.sqf.sqlRec("Delete from %s" % tab)

    def doExit(self):
        self.df.closeProcess()
        self.opts["mf"].closeLoop()

class MakeManual(object):
    """
    This class takes a restructured document file and produces a pdf document.
    """
    def __init__(self, docfle, vwr=""):
        if self.setVariables():
            self.fpdf.add_page()
            self.pmax = 200
            head1 = False
            head2 = False
            head3 = False
            head4 = False
            head5 = False
            table = False
            skips = False
            tcont = False
            note = False
            conts = []
            paras = []
            if type(docfle) == str:
                lines = open(docfle, "r").readlines()
            else:
                lines = docfle.readlines()
            for num, line in enumerate(lines):
                if skips:
                    skips = False
                    continue
                # Remove suffixes
                line = line.rstrip()
                # Remove hyperlinks
                if line and line[:4] == ".. _":
                    continue
                # Replace links with blank
                line = line.replace("_", "")
                # Remove | lines
                if line and line[0] == "|":
                    skips = True
                    continue
                if line and line.count("PageBreak"):
                    self.fpdf.add_page()
                    continue
                if line.count(".. NOTE::"):
                    note = True
                    continue
                if note:
                    if not line:
                        note = False
                        self.fpdf.drawText()
                    else:
                        line = line.strip()
                        self.fpdf.drawText(line, font=self.fonts["bodyb"],
                            fill=True, h=6, border="TLRB", ctyp="M")
                    continue
                if line.count(".. contents::"):
                    if vwr:
                        link = self.fpdf.add_link()
                        tcont = line.split("::")[1].strip()
                        tcont = (tcont.replace("**", ""), link)
                        self.fpdf.drawText(tcont[0], w=20 * self.fpdf.cwth,
                            h=10, fill=True, font=self.fonts["head3"],
                            border="TLRB", link=link)
                    else:
                        txt = "F2 for Table of Contents"
                        self.fpdf.drawText(txt, w=25 * self.fpdf.cwth,
                            h=10, fill=True, font=self.fonts["head3"],
                            border="TLRB")
                    continue
                if line.count(".. csv-table::"):
                    table = True
                    self.title = line.split("::")[1].lstrip().replace("**", "")
                    self.heads = []
                    self.widths = []
                    self.table = []
                    continue
                if table:
                    chk = line.strip().split()
                    if chk and chk[0] == ":header:":
                        self.heads = chk[1:]
                        continue
                    if chk and chk[0] == ":widths:":
                        self.widths = chk[1:]
                        continue
                    if chk and self.heads and self.widths:
                        self.table.append(line.strip())
                        continue
                    if self.table:
                        self.printTable()
                        table = False
                    else:
                        continue
                if head2:
                    head2 = False
                    continue
                if head3:
                    head3 = False
                    continue
                if head4:
                    head4 = False
                    continue
                if head5:
                    head5 = False
                    continue
                if head1 and line.count("===="):
                    head1 = False
                    continue
                if line.count("===="):
                    head1 = True
                    continue
                if head1:
                    self.fpdf.drawText(line, font=self.fonts["head1"],
                        align="C")
                    self.fpdf.drawText()
                    continue
                if (num + 1) < len(lines) and lines[num + 1].count("===="):
                    head1 = True
                    self.fpdf.drawText(line, font=self.fonts["head1"],
                        align="C")
                    self.fpdf.drawText()
                    continue
                if (num + 1) < len(lines) and lines[num + 1].count("----"):
                    head2 = True
                    link = self.fpdf.add_link()
                    self.fpdf.set_link(link, -1, -1)
                    conts.append((line.strip(), self.fpdf.page_no(), 1, link))
                    self.fpdf.drawText(line, font=self.fonts["head2"],
                        align="L", link=link)
                    self.fpdf.drawText()
                    continue
                if (num + 1) < len(lines) and lines[num + 1].count("...."):
                    head3 = True
                    link = self.fpdf.add_link()
                    self.fpdf.set_link(link, -1, -1)
                    conts.append((line.strip(), self.fpdf.page_no(), 2, link))
                    self.fpdf.drawText(line, font=self.fonts["head3"],
                        align="L", link=link)
                    self.fpdf.drawText()
                    continue
                if (num + 1) < len(lines) and lines[num + 1].count("++++"):
                    head4 = True
                    link = self.fpdf.add_link()
                    self.fpdf.set_link(link, -1, -1)
                    conts.append((line.strip(), self.fpdf.page_no(), 3, link))
                    self.fpdf.drawText(line, font=self.fonts["head4"],
                        align="L", link=link)
                    self.fpdf.drawText()
                    continue
                if (num + 1) < len(lines) and lines[num + 1].count("~~~~"):
                    head5 = True
                    link = self.fpdf.add_link()
                    self.fpdf.set_link(link, -1, -1)
                    conts.append((line.strip(), self.fpdf.page_no(), 4, link))
                    self.fpdf.drawText(line, font=self.fonts["head5"],
                        align="L", link=link)
                    self.fpdf.drawText()
                    continue
                if line.lstrip() and line.lstrip()[0] in ("+", ":"):
                    paras.append(line)
                    continue
                if paras:
                    if not line.strip() and lines[num + 1].strip() and \
                            lines[num + 1].lstrip()[0] in ("+", ":"):
                        # Blank line in a paragraph
                        paras.append(line)
                        continue
                    self.printPara(paras)
                    paras = []
                if not line.strip() and not lines[num + 1].strip():
                    # Two blank lines
                    continue
                # Other text
                self.printText(line)
            if paras:
                self.printPara(paras)
            if tcont and vwr:
                # Table of Contents
                self.fpdf.add_page()
                self.fpdf.set_link(tcont[1], -1, -1)
                self.fpdf.drawText(tcont[0], font=self.fonts["head1"],
                    align="C", link=tcont[1])
                self.fpdf.drawText()
                for line in conts:
                    fnt = self.fonts["bodyn"]
                    if line[2] == 1:
                        txt = line[0]
                        fnt = self.fonts["bodyb"]
                    elif line[2] == 2:
                        txt = "    " + line[0]
                    else:
                        txt = "        " + line[0]
                    self.fpdf.drawText(txt, w=100, font=fnt,
                        ln=0, link=line[3])
                    self.fpdf.drawText("%3s" % line[1],
                        font=self.fonts["bodyn"], align="R")

    def setVariables(self):
        self.fonts = {
            "head1": ("Arial", "B", 18),
            "head2": ("Arial", "B", 16),
            "head3": ("Arial", "B", 14),
            "head4": ("Arial", "B", 12),
            "head5": ("Arial", "B", 10),
            "bodyb": ("Arial", "B", 10),
            "bodyi": ("Arial", "I", 10),
            "bodyn": ("Arial", "", 10)}
        self.fpdf = MyFpdf(name="Manual", head=80, font=self.fonts["bodyn"],
            auto=True)
        return True

    def printPara(self, paras):
        indent = {}
        lines = []
        self.setFont("bodyb")
        for line in paras:
            if not line:
                lines.append([])
                continue
            words = None
            bold = False
            try:
                level = line.index("+")
            except:
                level = 0
            if level not in indent:
                indent[level] = 0
            if line.count(":") and line.index(":") == 0:
                words = line.split(":", 2)
                bold = True
            elif line.count("**") and line.index("**") in (2, 6, 10, 14, 18):
                words = line.split("**", 2)
                if level == 0:
                    bold = True
            elif line.count("*") and line.index("*") in (2, 6, 10, 14, 18):
                words = line.split("*", 2)
            elif line.count("#"):
                words = line.split("#")
            elif line.count("+ "):
                match = re.compile("[^\W\d]").search(line)
                words = [line[:match.start()], line[match.start():]]
            if words:
                if len(words) == 3:
                    if words[0]:
                        text = "%s%s" % (words[0], words[1].strip())
                    else:
                        text = words[1].rstrip()
                    wdth = self.fpdf.get_string_width(text) + 1
                    if words[2] and wdth > indent[level]:
                        indent[level] = wdth
                    lines.append((text, words[2], level, bold))
                else:
                    text = words[0].rstrip()
                    wdth = self.fpdf.get_string_width(text) + 1
                    if words[1] and wdth > indent[level]:
                        indent[level] = wdth
                    lines.append((text, words[1].strip(), level, bold))
            else:
                lines.append((line, level, bold))
        keys = list(indent.keys())
        keys.sort()
        for num, key in enumerate(keys):
            if key == keys[-1]:
                continue
            if indent[keys[num + 1]] and indent[keys[num + 1]] < indent[key]:
                indent[keys[num + 1]] = indent[key]
        for line in lines:
            if len(line) == 0:
                self.fpdf.drawText()
            elif len(line) == 4:
                x = self.fpdf.l_margin + \
                    self.fpdf.get_string_width((line[2] + 2) * "X")
                if line[3]:
                    self.setFont("bodyb")
                else:
                    self.setFont("bodyn")
                text = line[0].strip().replace("+", "%s " % chr(127))
                self.fpdf.drawText(text, x=x, w=indent[line[2]], ln=0)
                if line[1]:
                    self.printText(line[1])
                else:
                    self.fpdf.drawText()
            else:
                x = self.fpdf.l_margin + \
                    self.fpdf.get_string_width((line[2] + 2) * "X")
                text = line[0].strip().replace("+", "%s " % chr(127))
                self.printText(text, x=x)

    def printTable(self):
        self.fpdf.drawText(self.title, font="B")
        self.fpdf.drawText()
        for num, dat in enumerate(self.heads):
            txt = dat.replace('",', "").replace('"', "")
            wid = int(self.widths[num].replace(',', "")) * self.fpdf.cwth
            self.fpdf.drawText(txt, w=wid, border="TLRB", fill=True, ln=0)
        self.fpdf.drawText()
        for line in self.table:
            data = line.replace('"', "").split(", ")
            for num, dat in enumerate(data):
                wid = int(self.widths[num].replace(',', "")) * self.fpdf.cwth
                self.fpdf.drawText(dat, w=wid, border="TLRB", ln=0)
            self.fpdf.drawText()

    def printText(self, line):
        line = line.strip()
        spc = False
        for tst in ("- ", "%s " % chr(127)):
            if line[:2] == tst:
                spc = True
                break
        x1 = x2 = self.fpdf.get_x()
        y1 = self.fpdf.get_y()
        line = line.split()
        bold = False
        spcc = False
        font = "bodyn"
        for num, word in enumerate(line):
            if word.startswith(("**", "*")):
                bold = True
                font = "bodyb"
            if word.endswith(("*", "*.", "*,")):
                bold = False
            word = word.replace("*", "")
            if word.startswith("`"):
                bold = True
                font = "bodyi"
            if word.endswith(("`", "`.", "`,")):
                bold = False
            word = word.replace("`", "")
            if x2 + self.fpdf.get_string_width(word) > self.pmax:
                x2 = x1
                if spc:
                    x2 += self.fpdf.get_string_width(tst)
                y1 += 5
            elif spcc:
                word = " %s" % word
            self.fpdf.drawText(word, x=x2, y=y1, font=self.fonts[font], ln=0)
            self.setFont(font)
            x2 += self.fpdf.get_string_width(word)
            y1 = self.fpdf.get_y()
            if not bold:
                font = "bodyn"
            spcc = True
        self.fpdf.drawText()

    def setFont(self, font):
        fam, sty, siz = self.fonts[font]
        self.fpdf.setFont(fam, sty, siz)

class ViewPDF(object):
    """
    This class is used to view pdf files using either the system default
    pdf viewer, a viewer set up in the rcfile or pymupdf.
    """
    def __init__(self, mf=None, pdfnam=None):
        self.mf = mf
        if self.mf and self.mf.window:
            self.mf.window.withdraw()
        try:
            if pdfnam is None:
                pdfnam = self.doGetFile()
            if pdfnam and os.path.isfile(pdfnam):
                if self.mf:
                    vwr = mf.rcdic["vwr"]
                else:
                    vwr = ""
                self.pdfnam = pdfnam
                if vwr and os.path.exists(vwr):
                    # default to the selected viewer from the rcfile
                    exe, cmd = parsePrg(vwr)
                    cmd.append(pdfnam)
                    subprocess.call(cmd)
                elif not FITZ:
                    # Try and use the default pdf viewer
                    if sys.platform == "win32":
                        os.startfile(pdfnam)
                    else:
                        subprocess.call(["xdg-open", pdfnam])
                else:
                    # Use the Tartan pdf viewer
                    if self.mf and self.mf.window:
                        self.win = MkWindow(remov=False).newwin
                    else:
                        self.win = MkWindow().newwin
                    try:
                        self.win.tk.call("wm", "iconphoto", self.win._w,
                            "-default", getImage("pdfimg"))
                    except:
                        pass
                    self.win.title(pdfnam)
                    self.doDisplay()
            elif pdfnam:
                raise Exception("Invalid File Name\n\n%s" % pdfnam)
        except Exception as err:
            showError(None, "Error", "Cannot Display Document.\n\n%s" % err)
        if self.mf:
            self.mf.window.deiconify()

    def doGetFile(self):
        dialog = FileDialog(**{
            "title": "Select PDF File",
            "ftype": (("PDF Files", "*.pdf"),),
            "multi": False})
        return dialog.askopenfilename()

    def doDisplay(self):
        # Window dimensions and Image sizes
        self.win.resizable(0, 0)
        self.win.configure(borderwidth=2)
        self.sw = self.win.winfo_screenwidth()
        self.sh = int(self.win.winfo_screenheight() * .90)
        self.doc = fitz.open(self.pdfnam)
        pwd = self.doc.needs_pass
        if pwd and not self.doPassword():
            return
        self.lastpg = self.doc.page_count
        rect = self.doc[0].mediabox
        if not self.doc[0].rotation:
            self.siz = [rect[2], rect[3]]
        else:
            self.siz = [rect[3], rect[2]]
        self.rotate = 0
        # Theme and fonts
        self.style = ttk.Style()
        if not self.mf:
            self.style.theme_use("clam")
        self.font = ["Helvetica", 12]
        self.style.configure("pdf.TFrame", font=self.font)
        self.style.configure("pdf.TLabel", font=self.font)
        self.style.configure("pdf.TButton", font=self.font, relief="flat")
        self.style.configure("pdf.TEntry", font=self.font, height=4)
        self.style.configure("pdfbold.TLabel", font=self.font + ["bold"])
        fg = self.style.lookup("pdf.TButton", "foreground")
        bg = self.style.lookup("pdf.TButton", "background")
        self.style.configure("pdf.TRadiobutton", width=5, font=self.font)
        # Create arrowless scrollbars
        self.style.layout("h.TScrollbar", [
            ("Horizontal.Scrollbar.trough", {
                "children": [("Horizontal.Scrollbar.thumb", {
                    "expand": "1",
                    "sticky": "nswe"})],
                "sticky": "ew"})])
        # Widgets
        fr1 = MyFrame(self.win)
        fr1.pack(side="top", fill="x")
        fr2 = MyFrame(fr1, borderwidth=1, relief="raised")
        fr2.pack(fill="x", expand="yes")
        # Buttons and Entries
        self.bt1 = MyButton(fr2, text="Goto", cmd=self.gotoPage,
            style="pdf.TButton", underline=0)
        ToolTip(self.bt1, "Jump To Page Number")
        self.bt1.pack(padx=3, pady=3, side="left")
        self.entsiz = len(str(self.lastpg))
        self.pgd = MyEntry(fr2, width=self.entsiz, maxsize=self.entsiz,
            style="pdf.TEntry")
        self.pgd.bind("<Return>", self.enterPage)
        self.pgd.bind("<KP_Enter>", self.enterPage)
        self.pgd.pack(padx=3, pady=3, side="left")
        lab = MyLabel(fr2, text="of %s" % self.lastpg, color=False,
            style="pdf.TLabel")
        lab.pack(padx=3, pady=3, side="left")
        self.bt2 = MyButton(fr2, txt=False, text="Former",
            cmd=self.priorPage, style="pdf.TButton", underline=0)
        ToolTip(self.bt2, "Show Previous Page")
        self.bt2.pack(padx=3, pady=3, side="left")
        self.bt3 = MyButton(fr2, txt=False, text="Next", cmd=self.nextPage,
            style="pdf.TButton", underline=getUnderline(fr2, "Next")[1])
        ToolTip(self.bt3, "Show Next Page")
        self.bt3.pack(padx=3, pady=3, side="left")
        # Draw menu
        imgm = getImage("menu", siz=(20, 20))
        self.bt4 = MyMenuButton (fr2, text="Menu", relief="flat", fg=fg,
            bg=bg, font=self.font, image=imgm, compound="left", underline=0)
        self.bt4.pack(exp="no", padx=3, pady=3, side="right")
        self.bt4.menu = tk.Menu(self.bt4, font=self.font, tearoff=0)
        if fg and bg:
            self.bt4.menu.configure(fg=fg, bg=bg)
        self.bt4["menu"] = self.bt4.menu
        mods = [
            ("Print", self.doPrint),
            ("Save as..", self.doSave),
            ("Send to..", self.doSend),
            ("Help", self.doHelp)]
        if self.mf and self.mf.dbm:
            try:
                gc = GetCtl(self.mf)
                ctlsys = gc.getCtl("ctlsys", error=False)
                if ctlsys and ctlsys["sys_msvr"]:
                    self.server = [ctlsys["sys_msvr"], ctlsys["sys_mprt"],
                        ctlsys["sys_msec"], ctlsys["sys_maut"],
                        ctlsys["sys_mnam"], ctlsys["sys_mpwd"]]
                    if sendMail(self.server, "", "", "", check=True,
                            errwid=self.mf.window,
                            wrkdir=self.mf.rcdic["wrkdir"]):
                        mods.insert(0, ("Email", self.doEmail))
            except:
                pass
        img = {}
        for num, text in enumerate(mods):
            img[num] = getImage(text[0], siz=(20, 20))
            self.bt4.menu.add_command(label=text[0], image=img[num],
                compound="left", command=text[1], underline=0)
        self.bt4.menu.add_separator()
        imgr = getImage("exit", (20, 20))
        self.bt4.menu.add_command(label="Exit", image=imgr, compound="left",
            command=self.doClose, underline=1)
        # Rest of Buttons
        self.bt5 = MyButton(fr2, text="Zoom", style="pdf.TButton")
        self.bt5.bind("<Button-1>", self.doZoom)
        self.bt5.bind("<Button-3>", self.doZoom)
        ToolTip(self.bt5, "Left Button to Zoom in Increments of 25%. Right "\
            "Button to UnZoom. Ctrl plus the Numeric Keypad +- Keys can "\
            "also be used. Use F11 to toggle full screen.")
        self.bt5.pack(padx=3, pady=3, side="right")
        fr1.update_idletasks()
        # Canvas
        self.cv = tk.Canvas(self.win, highlightthickness=0)
        self.horz = ttk.Scrollbar(self.win, orient="horizontal",
            style="h.TScrollbar")
        self.horz.config(command=self.cv.xview)
        self.cv.config(xscrollcommand=self.horz.set)
        self.horz.pack(fill="x", expand="no", side="bottom")
        self.cv.pack(fill="both", expand="yes")
        if sys.platform == "win32":
            self.win.bind("<MouseWheel>", self.doWheel)
        else:
            self.win.bind("<Button-4>", self.doWheel)
            self.win.bind("<Button-5>", self.doWheel)
        self.win.bind("<Escape>", self.doClose)
        self.win.bind("<F1>", self.doHelp)
        self.win.bind("<F2>", self.doContents)
        self.win.bind("<F11>", self.doMaxi)
        self.win.bind("<Home>", self.homePage)
        self.win.bind("<Next>", self.nextPage)
        self.win.bind("<Prior>", self.priorPage)
        self.win.bind("<End>", self.lastPage)
        self.win.bind("<Up>", self.doKey)
        self.win.bind("<Down>", self.doKey)
        self.win.bind("<Left>", self.doKey)
        self.win.bind("<Right>", self.doKey)
        self.win.bind("<Control-f>", self.doSearch)
        self.win.bind("<Control-n>", self.nextSearch)
        self.win.bind("<Control-e>", self.endSearch)
        self.win.bind("<Control-r>", self.doRotate)
        self.win.bind("<Control-KP_Add>", self.doZoom)
        self.win.bind("<Control-KP_Subtract>", self.doZoom)
        self.win.update_idletasks()
        # Scale settings
        self.scale = .75
        self.zoom = 1.25
        if self.mf:
            try:
                chk = os.path.join(self.mf.rcdic["wrkdir"], "pdfview.conf")
                if os.path.isfile(chk):
                    cnf = open(chk, "r")
                    self.zoom = float(cnf.read())
                    cnf.close()
            except:
                pass
        self.matrix = list(fitz.Matrix(self.zoom, self.zoom))
        # Other settings
        self.pgno = 1
        self.pags = []
        self.cont = False
        self.help = False
        self.prec = {}
        self.search = ""
        self.wsiz = []
        # Display 1st page
        self.showPage()
        self.maxi = False
        self.win.wait_window()

    def doPassword(self):
        def enterPwd(event=None):
            self.pwd = ent.get().strip()
            frm.destroy()
        def exitFrm(event=None):
            self.pwd = None
            frm.destroy()
        frm = MyFrame()
        frm.pack()
        lab = MyLabel(frm, text="Enter Password")
        lab.pack(side="left")
        ent = MyEntry(frm, show="*")
        ent.pack(side="left")
        ent.bind("<Escape>", exitFrm)
        ent.bind("<Return>", enterPwd)
        ent.bind("<KP_Enter>", enterPwd)
        placeWindow(self.win, place="M", expose=True)
        ent.focus_set()
        frm.wait_window()
        if not self.pwd or not self.doc.authenticate(self.pwd):
            self.win.destroy()
            if self.pwd:
                showError(None, "Error", "Invalid Password")
            return
        return True

    def doWheel(self, event=None):
        if sys.platform == "win32":
            scroll = -1 if event.delta > 0 else 1
        else:
            scroll = -1 if event.num == 4 else 1
        if scroll == 1:
            self.doScroll("Down", scroll)
        else:
            self.doScroll("Up", scroll)

    def doKey(self, event=None):
        scroll = -1 if event.keysym in ("Up", "Left") else 1
        self.doScroll(event.keysym, scroll)

    def doScroll(self, key, scroll):
        self.cv.update_idletasks()
        if key == "Down" and self.cv.yview()[1] == 1:
            self.nextPage()
        elif key == "Up" and self.cv.yview()[0] == 0:
            self.priorPage(pos="end")
        elif key in ("Up", "Down") and self.cv.yview() != (0.0, 1.0):
            self.cv.yview_scroll(scroll, "units")
        elif self.cv.xview() != (0.0, 1.0):
            self.cv.xview_scroll(scroll, "units")

    def homePage(self, event=None):
        self.pgno = 1
        self.showPage(pos="top")

    def lastPage(self, event=None):
        self.pgno = self.lastpg
        self.showPage(pos="top")

    def nextPage(self, event=None):
        if self.pgno == self.lastpg:
            return
        self.pgno += 1
        self.showPage(pos="top")

    def priorPage(self, event=None, pos="top"):
        if self.pgno == 1:
            return
        self.pgno -= 1
        self.showPage(pos=pos)

    def gotoPage(self, event=None):
        self.pgd.configure(state="normal")
        self.pgd.selection_range(0, "end")
        self.pgd.focus_set()

    def enterPage(self, event=None):
        try:
            pgno = int(self.pgd.get())
            if pgno < 1 or pgno > self.lastpg:
                raise Exception
        except:
            self.pgd.selection_range(0, "end")
            return
        self.pgno = pgno
        self.showPage(pos="top")

    def doMaxi(self, event=None):
        if not self.maxi:
            self.maxi = self.matrix[:]
            self.win.geometry("%sx%s+0+0" % (self.sw, self.sh))
            self.matrix[0] = self.matrix[3] = self.sw / self.siz[0]
        else:
            self.matrix = self.maxi
            self.maxi = False
        self.wsiz = []
        self.showPage()

    def doRotate(self, event=None):
        if self.rotate == 270:
            self.rotate = 0
        else:
            self.rotate += 90
        for page in self.doc:
            page.set_rotation(self.rotate)
        self.showPage()

    def doZoom(self, event=None):
        if event.num == 3 or event.keysym == "KP_Subtract":
            if self.zoom > self.scale:
                self.zoom = round((self.zoom - .25), 2)
        elif self.zoom < (self.scale + 3):
            while self.zoom <= self.matrix[0]:
                self.zoom = round((self.zoom + .25), 2)
        self.wsiz = []
        self.matrix[0] = self.matrix[3] = self.zoom
        self.showPage()

    def doSearch(self, event=None):
        def getSearch(event=None):
            self.search = ent.get()
            self.found = False
            self.pags = []
            self.prec = {}
            for page in self.doc:
                annot = page.first_annot
                while annot:
                    annot = page.delete_annot(annot)
                found = page.search_for(self.search)
                if found:
                    self.found = True
                    for inst in found:
                        numb = (page.number + 1)
                        if numb not in self.pags:
                            self.pags.append(numb)
                            self.prec[numb] = inst
                        page.add_highlight_annot(inst)
            if self.pags:
                self.pgno = self.pags[0]
            frm.destroy()

        self.doUnbind(key=True)
        frm = MyFrame(self.cv, bg="black", borderwidth=5)
        tit = MyLabel(frm, text="Search For", anchor="c", relief="raised")
        tit.pack(fill="x")
        ent = MyEntry(frm, width=30)
        if self.search:
            ent.insert(0, self.search)
            ent.selection_range(0, "end")
        ent.pack(fill="x")
        tt = ToolTip(ent, "Once you are in Search Mode these Keys Apply:\n"\
            "<Ctrl-n> Scroll to Next Occurrence of Search String\n"\
            "<Ctrl-e> Clear All Highlights and Exit Search Mode.")
        tt.pause = None
        ent.bind("<Return>", getSearch)
        ent.bind("<KP_Enter>", getSearch)
        frm.place(anchor="center", relx=0.5, rely=0.5)
        ent.focus_set()
        frm.wait_window()
        if self.search is not None and not self.found:
            sp = SplashScreen(self.win, "Sorry, Not Found")
            time.sleep(2)
            sp.closeSplash()
        self.doUnbind(False, key=True)
        self.showPage()

    def nextSearch(self, event=None):
        if self.pags:
            if self.pgno < self.pags[-1]:
                for pag in self.pags:
                    if pag > self.pgno:
                        break
            else:
                pag = self.pags[0]
            self.pgno = pag
            self.showPage()

    def endSearch(self, event=None):
        self.pags = []
        self.prec = {}
        self.search = ""
        for page in self.doc:
            annot = page.first_annot
            while annot:
                annot = page.delete_annot(annot)
        self.showPage()

    def doUnbind(self, unbind=True, key=True, exc=None):
        if unbind:
            if key:
                self.cvbinds = []
                for bind in self.win.bind():
                    if exc is None or bind != exc:
                        self.cvbinds.append((bind, self.win.bind(bind)))
                        self.win.unbind(bind)
            for x in range(1, 6):
                bt = getattr(self, "bt%s" % x)
                bt.configure(state="disabled")
        else:
            if key:
                for bind in self.cvbinds:
                    self.win.bind(bind[0], bind[1])
            for x in range(1, 6):
                bt = getattr(self, "bt%s" % x)
                bt.configure(state="normal")
        self.win.update_idletasks()

    def showPage(self, pos=None):
        self.ltime = 0
        self.cv.delete("all")
        page = self.doc[self.pgno - 1]
        self.pgd.configure(state="normal")
        self.pgd.delete(0, "end")
        self.pgd.insert(0, "%s" % CCD(self.pgno, "UI", self.entsiz).disp)
        self.pgd.configure(state="disabled")
        dlist = page.get_displaylist()
        pix = dlist.get_pixmap(matrix=self.matrix, alpha=False)
        self.ti = tk.PhotoImage(data=pix.tobytes("ppm"))
        self.cv.create_image(0, 0, image=self.ti, anchor="nw", tags="img")
        self.cv.configure(width=self.ti.width(), height=self.ti.height())
        # Limit display size to %-tage of screen size
        self.win.update_idletasks()
        wsiz = [self.win.winfo_reqwidth(), self.win.winfo_reqheight()]
        if wsiz[0] > self.sw:
            wsiz[0] = self.sw
        if wsiz[1] > self.sh:
            wsiz[1] = self.sh
        if wsiz != self.wsiz:
            # Re-size and place window
            placeWindow(self.win, place="C", size=wsiz, expose=True)
            self.cv.configure(scrollregion=self.cv.bbox("all"))
            self.wsiz = wsiz
        if pos and pos == "top":
            # Scroll to top
            self.cv.yview_moveto("0.0")
        elif pos and pos == "end":
            # Scroll to bottom
            self.cv.yview_moveto("1.0")
        elif self.pags or self.cont:
            # Scroll to next highlighted text
            self.cv.yview_moveto("0.0")
            rect = self.prec[self.pgno]
            if self.zoom != self.scale:
                c1 = rect[1] * self.zoom
            else:
                c1 = rect[1]
            y1, y2 = self.getVisibleArea()
            while not y1 <= c1 <= y2:
                self.doScroll("Down", 1)
                y1, y2 = self.getVisibleArea()
            if self.cont:
                self.cont = False
        self.win.lift()
        self.win.grab_set()
        self.win.focus_set()
        self.win.update_idletasks()

    def getVisibleArea(self):
        y1 = self.cv.canvasy(0)
        if y1 == -1:
            y1 = 0
        y2 = self.cv.canvasy(self.cv.winfo_height())
        if self.zoom != self.scale:
            y2 = y2 - (self.zoom * 20)
        return (y1, y2)

    def doHelp(self, event=None):
        cols = (
            (0, "Keys", 6, "NA", None),
            (1, "Action", 41, "NA", None))
        data = (
            ("F1", "This help"),
            ("F2", "Generate a Table of Contents, if possible"),
            ("F11", "Toggle Full Screen"),
            ("Alt m", "Show the Menu"),
            ("Alt g", "Go To Page"),
            ("Alt f", "Former Page"),
            ("Alt n", "Next Page"),
            ("Ctrl +", "Zoom"),
            ("Ctrl -", "Reverse Zoom"),
            ("Ctrl f", "Search for Text"),
            ("Ctrl n", "Next occurrence of Text"),
            ("Ctrl e", "Clear highlighted Text"),
            ("Ctrl r", "Rotate Document"),
            ("Arrows", "Scroll up, down, left and right"),
            ("Esc", "Close the Current View"))
        self.doUnbind(exc="<Key-F1>")
        SelectChoice(self.cv, titl="Keyboard Bindings", deco=False, modal=True,
            cols=cols, data=data, font=self.font, lines=len(data),
            sort=False, scrl=False, live=False)
        self.doUnbind(False)
        self.cv.focus_force()

    def doContents(self, event=None):
        def getData(page, siz=None):
            def addText(txt, tsz, bbx):
                txts.append((tsz, txt, page.number + 1, bbx))
            txts = []
            blocks = page.get_text("dict")["blocks"]
            for b in blocks:
                if b["type"] == 0:
                    ts = 0
                    txt = ""
                    bbx = [0, 0, 0, 0]
                    for l in b["lines"]:
                        if bbx[1] != l["bbox"][1]:
                            if txt:
                                addText(txt, ts, bbx)
                                txt = ""
                            bbx = list(l["bbox"])
                        for s in l["spans"]:
                            if s["size"] <= siz:
                                continue
                            ts = s["size"]
                            tt = s["text"]
                            if tt and not txt.count("Table of Contents"):
                                bbx[2] = l["bbox"][2]
                                if not txt:
                                    txt = tt
                                else:
                                    txt = "%s %s" % (txt, tt)
                    if siz and txt:
                        addText(txt, ts, bbx)
            return txts
        sp = SplashScreen(self.win, "Generating Table of Contents\n\n"\
            "Please Wait...")
        # Try to use embedded toc
        toc = self.doc.get_toc()
        if toc:
            tabs = []
            cdata = {}
            mxss = 0
            indx = 0
            for item in toc:
                if not item[1].strip():
                    continue
                pgno = int(item[2]) - 1
                page = self.doc[pgno]
                text = item[1].replace("\u2003", " ")
                rect = page.search_for(text)
                xxx = []
                for rec in rect:
                    if not xxx:
                        xxx = rec
                        continue
                    if rec.y0 == xxx.y0:
                        xxx.x1 = rec.x1
                text = "%s%s" % ("    " * (item[0] - 1), text)
                if len(text) > mxss:
                    mxss = len(text)
                tabs.append((text, item[2]))
                cdata[indx] = xxx
                indx += 1
        else:
            # Try to Generate own toc
            sizs = {}
            for page in self.doc:
                blocks = page.get_text("dict")["blocks"]
                for b in blocks:
                    if b["type"] == 0:
                        for l in b["lines"]:
                            for s in l["spans"]:
                                tst = s["size"]
                                if tst not in sizs:
                                    sizs[tst] = 1
                                else:
                                    sizs[tst] += 1
            if sizs:
                cnts = [0, 0]
                cdata = {}
                for siz in sizs:
                    if sizs[siz] > cnts[1]:
                        cnts = [siz, sizs[siz]]
                indx = 0
                mxss = 0
                tabs = []
                spcs = list(sizs.keys())
                spcs.remove(cnts[0])
                spcs.sort(reverse=True)
                for page in self.doc:
                    for data in getData(page, cnts[0]):
                        if not data[1].strip():
                            continue
                        text = "%s%s" % ("  " * (spcs.index(data[0]) - 1),
                            data[1])
                        if text.count("F2"):
                            continue
                        tabs.append((text, data[2]))
                        cdata[indx] = data[3]
                        if len(text) > mxss:
                            mxss = len(text)
                        indx += 1
            if not sizs or not tabs:
                sp.refreshSplash(text="Sorry, No Table Found")
                time.sleep(2)
                sp.closeSplash()
                self.cv.focus_force()
                return
        sp.closeSplash()
        # Display toc
        cols = (
            (0, "Description", mxss, "NA", "Y"),
            (1, "Page", 4, "UI", None))
        self.doUnbind()
        sc = SelectChoice(self.cv, titl="Table of Contents", deco=False,
            modal=True, cols=cols, data=tabs, font="Courier", sort=False)
        self.doUnbind(False)
        if sc.selection:
            for page in self.doc:
                annot = page.first_annot
                while annot:
                    annot = page.delete_annot(annot)
            bbox = cdata[sc.selection[0]]
            rect = fitz.Rect(bbox[0], bbox[1], bbox[2], bbox[3])
            self.pgno = int(sc.selection[2])
            page = self.doc[self.pgno - 1]
            self.prec[self.pgno] = rect
            page.add_highlight_annot(rect)
            self.cont = True
            self.showPage()
        self.cv.focus_force()

    def doSave(self):
        if self.mf:
            if sys.platform == "win32":
                lastdir = os.path.join(self.mf.rcdic["wrkdir"], "savedir")
            else:
                lastdir = os.path.join(self.mf.rcdic["wrkdir"], ".savedir")
            if os.path.exists(lastdir):
                infle = open(lastdir, "r")
                name = infle.readline().rstrip()
                infle.close()
            else:
                name = self.mf.rcdic["wrkdir"]
        else:
            name = os.getcwd()
        path = tkfile.asksaveasfilename(title="Enter Filename",
            defaultextension=".pdf", filetypes=[("pdf Files", "*.pdf")],
            initialdir=name, initialfile=os.path.basename(self.pdfnam))
        if path:
            try:
                shutil.copyfile(self.pdfnam, path)
                if self.mf:
                    infle = open(lastdir, "w")
                    infle.write(os.path.dirname(os.path.normpath(path)))
                    infle.close()
            except Exception as err:
                showError(None, "Error", err)
        self.doClose()

    def doSend(self):
        if sys.platform == "win32":
            os.startfile(self.pdfnam)
        else:
            subprocess.Popen(["xdg-open", self.pdfnam])
        self.doClose()

    def doEmail(self):
        self.win.withdraw()
        # From address
        if sys.platform == "win32":
            self.lasteml = os.path.join(self.mf.rcdic["wrkdir"], "fromeml")
        else:
            self.lasteml = os.path.join(self.mf.rcdic["wrkdir"], ".fromeml")
        if os.path.exists(self.lasteml):
            infle = open(self.lasteml, "r")
            self.efrom = infle.readline().rstrip()
            infle.close()
        else:
           self.efrom = ""
        tit = "Email Document"
        fld = (
            (("T",0,0,0),"ITX",30,"From Address","",
                self.efrom,"N",None,None,None,("email", False)),)
        self.ed = TartanDialog(self.mf, tops=True, tile=tit,
            eflds=fld, mail=("Y","N", "S"), tend=((self.doEEnd, "y"),),
            txit=(self.doEExit,))
        if self.mf and self.mf.window:
            self.mf.window.deiconify()
        self.ed.mstFrame.wait_window()
        if self.mf and self.mf.window:
            self.mf.window.withdraw()
        self.win.deiconify()

    def doEEnd(self):
        self.ed.closeProcess()
        fromad = self.ed.t_work[0][0][0]
        toad = self.ed.repeml[2].split(",")
        subj = "PDF Report"
        att = []
        if self.ed.repeml[3]:
            body = self.ed.repeml[3][1]
            if len(self.ed.repeml[3]) > 2:
                att = [self.ed.repeml[3][2]]
        else:
            body = ""
        att.append(self.pdfnam)
        sql = Sql(self.mf.dbm, "emllog", prog=__name__)
        for add in toad:
            ok = False
            while not ok:
                sp = SplashScreen(self.mf.window.focus_displayof(),
                    "E-Mailing the Report to:\n\n%s\n\nPlease Wait....." % add)
                ok = sendMail(self.server, fromad, add, subj, mess=body,
                    attach=att, wrkdir=self.mf.rcdic["wrkdir"])
                sp.closeSplash()
                if not ok:
                    ok = askQuestion(self.mf.window.focus_displayof(),
                        "E-Mail Error", "Problem Delivering This "\
                        "Message.\n\nTo: %s\nSubject: %s\n\nWould "\
                        "You Like to Retry?" % (add, subj))
                    if ok == "yes":
                        ok = None
                    else:
                        ok = "Failed"
                else:
                    ok = "OK"
                if ok:
                    try:
                        tim = time.localtime()[0:  5]
                        tim = "%04i-%02i-%02i %02i:%02i" % tim
                        sql.insRec("emllog", data=[fromad.strip(),
                            add.strip(), subj, tim, ok])
                    except:
                        pass
                    break
        self.mf.dbm.commitDbase()
        infle = open(self.lasteml, "w")
        infle.write("%s\n" % self.ed.t_work[0][0][0])
        infle.close()
        self.ed.mstFrame.destroy()

    def doEExit(self):
        self.ed.closeProcess()
        self.ed.mstFrame.destroy()

    def doPrint(self):
        def doDisable(event=None):
            ent.delete(0, "end")
            ent.configure(state="disabled")

        def doPages(event=None):
            ent.configure(state="normal")
            ent.focus_set()

        def doCancel(event=None):
            win.destroy()
            self.cv.focus_force()

        def doExec(event=None):
            try:
                nam = lbx.get("active")
                sel = var.get()
                if sel == "A":
                    fle = self.pdfnam
                else:
                    if sel == "P":
                        wrk = ent.get()
                        if wrk.count(",") and wrk.count("-"):
                            raise Exception("Invalid Range Selected")
                        if wrk.count(","):
                            wrk = wrk.split(",")
                            wrk = list(dict.fromkeys(wrk))
                        elif wrk.count("-"):
                            rng = wrk.split("-")
                            rng = list(dict.fromkeys(rng))
                            wrk = []
                            for x in range(int(rng[0]), int(rng[1]) + 1):
                                wrk.append(x)
                        else:
                            wrk = [wrk]
                        pag = []
                        for w in wrk:
                            p = int(w) - 1
                            if p in self.doc:
                                pag.append(p)
                        pag.sort()
                    else:
                        pag = [self.pgno - 1]
                    if not pag:
                        raise Exception("No Valid Pages Selected")
                    tme = "%04i%02i%02i%02i%02i%02i" % time.localtime()[:-3]
                    fle = os.path.join(tempfile.gettempdir(), "%s.pdf" % tme)
                    doc2 = fitz.open()
                    for pg in pag:
                        doc2.insert_pdf(self.doc, from_page=pg, to_page=pg)
                    doc2.save(fle)
                    doc2.close()
                cpy = int(spn.get())
                printPDF(nam, fle, cpy)
                doCancel()
            except Exception as err:
                doCancel()
                if self.mf:
                    wrk = self.mf.rcdic["wrkdir"]
                else:
                    wrk = tempfile.gettempdir()
                showException(self.cv, wrk, err)

        prts = getPrinters()
        if not prts:
            return
        self.doUnbind()
        win = MkWindow(trans=self.cv, decor=False).newwin
        win.configure(bg="black", borderwidth=2)
        fr1 = MyFrame(win, relief="ridge", borderwidth=2,
            style="pdf.TFrame")
        fr1.pack(fill="both", expand="yes")
        lb1 = MyLabel(fr1, color=False, text="Available Printers",
            relief="raised", style="pdfbold.TLabel")
        lb1.pack(fill="x")
        lbx = tk.Listbox(fr1, font=self.font, width=10,
            selectbackground="blue", selectforeground="white")
        lbx.bind("<Return>", doExec)
        lbx.bind("<KP_Enter>", doExec)
        lbx.bind("<Escape>", doCancel)
        for prt in prts:
            lbx.insert("end", prt)
        lbx.pack(fill="both", expand="yes")
        fr2 = MyFrame(win, relief="ridge", borderwidth=2,
            style="pdf.TFrame")
        fr2.pack(fill="both", expand="yes")
        fr2 = MyFrame(win, relief="ridge", borderwidth=2,
            style="pdf.TFrame")
        fr2.pack(fill="both", expand="yes")
        fr2.columnconfigure(0, weight=1)
        fr2.columnconfigure(1, weight=1)
        fr2.columnconfigure(2, weight=1)
        lb2 = MyLabel(fr2, color=False, text="Range", style="pdfbold.TLabel")
        lb2.grid(row=0, column=0, sticky="w")
        var = tk.StringVar()
        var.set("A")
        rb1 = MyRadioButton(fr2, variable=var, text="All Pages",
            value="A", command=doDisable, style="pdf.TRadiobutton",
            width=12)
        rb1.grid(row=1, column=0, sticky="nsew")
        rb2 = MyRadioButton(fr2, variable=var, text="Current Page",
            value="C", command=doDisable, style="pdf.TRadiobutton",
            width=12)
        rb2.grid(row=2, column=0, sticky="nsew")
        rb3 = MyRadioButton(fr2, variable=var, text="Pages",
            value="P", command=doPages, style="pdf.TRadiobutton")
        rb3.grid(row=3, column=0, sticky="nsew")
        ent = MyEntry(fr2, style="pdf.TEntry")
        ToolTip(ent, "Enter page numbers separated by commas or one dash.")
        ent.bind("<Return>", doExec)
        ent.bind("<KP_Enter>", doExec)
        ent.configure(state="disabled")
        ent.grid(row=3, column=1, sticky="nsew", columnspan=2)
        lb3 = MyLabel(fr2, text="Copies", style="pdfbold.TLabel", color=False)
        lb3.grid(row=0, column=2, sticky="e")
        spn = tk.Spinbox(fr2, from_=1, to=10, font=self.font, width=5)
        spn.grid(row=1, column=2, sticky="e")
        fr3 = MyFrame(win, style="pdf.TFrame")
        fr3.pack(fill="x", expand="yes")
        bt1 = MyButton(fr3, text="Cancel", cmd=doCancel, underline=0,
            style="pdf.TButton")
        bt1.pack(side="left", fill="x", expand="yes")
        bt2 = MyButton(fr3, text="Print", cmd=doExec, underline=0,
            style="pdf.TButton")
        bt2.pack(side="left", fill="x", expand="yes")
        placeWindow(win, self.cv, expose=True)
        lbx.selection_set(0)
        lbx.activate(0)
        lbx.focus_set()
        win.grab_set()
        win.wait_window()
        self.win.update_idletasks()
        self.doUnbind(False)
        self.showPage()

    def doClose(self, event=None):
        if self.mf:
            try:
                chk = os.path.join(self.mf.rcdic["wrkdir"], "pdfview.conf")
                cnf = open(chk, "w")
                cnf.write("%s" % self.zoom)
                cnf.close()
            except:
                pass
        self.doc.close()
        self.win.destroy()

# vim:set ts=4 sw=4 sts=4 expandtab:
